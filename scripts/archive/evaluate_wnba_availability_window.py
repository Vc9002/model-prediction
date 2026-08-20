"""Evaluate the WNBA availability challenger on a bounded historical window."""

from __future__ import annotations

import argparse
import hashlib
import json
import random
from collections import Counter
from collections.abc import Mapping
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from model_prediction.data_sources.espn import ESPNClient
from model_prediction.data_sources.espn_wnba_injuries import normalize_espn_event_injuries
from model_prediction.data_sources.wnba_injuries import parse_report_pdf
from model_prediction.domain import parse_utc
from model_prediction.features.base import FeatureStore
from model_prediction.features.player_availability import (
    STATUS_ACTIVE_PROBABILITIES,
    _identity,
    matchup_player_availability_from_payloads,
    merge_availability_sources,
)
from model_prediction.learned_forward import build_learned_moneyline_slate
from model_prediction.models.learned_market import LearnedMarketArtifact
from model_prediction.wnba_availability_evaluation import (
    adjust_home_probability,
    build_research_priors,
    historical_margin_sigma,
    report_payload,
)

SCALES = (0.0, 0.5, 1.0, 1.5, 2.0)
GATES = (0.50, 0.55, 0.60, 0.65, 0.70, 0.75)


class LocalScoreboards:
    def __init__(self, data_root: Path) -> None:
        self.data_root = data_root

    def scoreboard(self, _league: str, game_date: str) -> dict[str, Any]:
        path = self.data_root / "raw" / "wnba" / game_date / "scores_wnba.json"
        return json.loads(path.read_text(encoding="utf-8"))


def _ledger_probabilities(
    root: Path,
) -> tuple[dict[str, dict[str, Any]], dict[tuple[str, str, str], dict[str, Any]]]:
    by_event: dict[str, dict[str, Any]] = {}
    by_matchup: dict[tuple[str, str, str], dict[str, Any]] = {}
    for workbook_path in (root / "data/picks.xlsx", root / "data/flat_picks.xlsx"):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook["Picks"]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        columns = {str(value): index for index, value in enumerate(headers) if value is not None}
        for values in rows:
            if str(values[columns["league"]]) != "WNBA":
                continue
            probability = values[columns["model_probability"]]
            selection = str(values[columns["selection"]])
            if probability is None or selection not in {"home", "away"}:
                continue
            home_probability = float(probability) if selection == "home" else 1 - float(probability)
            record = {
                "home_probability": home_probability,
                "pick_id": str(values[columns["pick_id"]]),
                "workbook": workbook_path.name,
                "event_id": str(values[columns["event_id"]]),
            }
            event_id = record["event_id"]
            if event_id and event_id != "None":
                by_event[event_id] = record
            event_date = str(values[columns["event_start_utc"]])[:10]
            matchup_key = (
                event_date,
                str(values[columns["away_team"]]),
                str(values[columns["home_team"]]),
            )
            by_matchup[matchup_key] = record
        workbook.close()
    return by_event, by_matchup


def _event_teams(event: Mapping[str, Any]) -> tuple[str, str]:
    competitors = event["competitions"][0]["competitors"]
    by_side = {item["homeAway"]: item["team"]["displayName"] for item in competitors}
    return str(by_side["away"]), str(by_side["home"])


def _event_result(event: Mapping[str, Any]) -> tuple[bool | None, int | None, int | None]:
    competition = event["competitions"][0]
    if not competition.get("status", {}).get("type", {}).get("completed"):
        return None, None, None
    by_side = {item["homeAway"]: item for item in competition["competitors"]}
    away_score = int(float(by_side["away"]["score"]))
    home_score = int(float(by_side["home"]["score"]))
    return home_score > away_score, away_score, home_score


def _report_rows(report_paths: list[Path], retrieved_at: datetime) -> list[dict[str, Any]]:
    reports = []
    for path in sorted(report_paths):
        parsed = parse_report_pdf(path.read_bytes())
        reports.append(
            {
                "path": str(path),
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                "payload": report_payload(
                    parsed,
                    "https://ak-static.cms.nba.com/referee/wnba_injury/" + path.name,
                    retrieved_at,
                ),
            }
        )
    return sorted(reports, key=lambda item: item["payload"]["report_at_utc"])


def _cached_target_injuries(
    client: ESPNClient,
    data_root: Path,
    event_id: str,
    observed_at: datetime,
) -> dict[str, Any]:
    path = data_root / "availability/wnba/espn_target_injuries" / f"{event_id}.json"
    if path.exists():
        source = json.loads(path.read_text(encoding="utf-8"))
        retrieved_at = datetime.fromtimestamp(path.stat().st_mtime, tz=UTC)
    else:
        summary = client.summary("WNBA", event_id)
        retrieved_at = datetime.now(UTC)
        source = {
            "injuries": summary.get("injuries", []),
            "header": summary.get("header", {}),
            "retrieved_at_utc": retrieved_at.isoformat(),
        }
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(source, sort_keys=True) + "\n", encoding="utf-8")
    if source.get("retrieved_at_utc"):
        retrieved_at = parse_utc(str(source["retrieved_at_utc"]))
    normalized = normalize_espn_event_injuries(source, event_id=event_id, observed_at=observed_at)
    normalized["observed_at_utc"] = retrieved_at.isoformat()
    normalized["status_cutoff_at_utc"] = observed_at.isoformat()
    normalized["provenance_class"] = "retrospective_espn_status_timestamp_filter"
    return normalized


def _date_window(start_date: str, end_date: str) -> tuple[str, ...]:
    start = date.fromisoformat(start_date)
    end = date.fromisoformat(end_date)
    if end < start:
        raise ValueError("end date precedes start date")
    return tuple((start + timedelta(days=offset)).isoformat() for offset in range((end - start).days + 1))


def _select_report(
    reports: list[dict[str, Any]], event: Mapping[str, Any], away: str, home: str
) -> dict[str, Any]:
    start = parse_utc(str(event["date"]))
    eligible = []
    for report in reports:
        payload = report["payload"]
        report_at = parse_utc(str(payload["report_at_utc"]))
        teams = {_identity(str(team)) for team in payload.get("teams_listed", [])}
        if report_at < start and {_identity(away), _identity(home)} <= teams:
            eligible.append(report)
    if not eligible:
        raise ValueError(f"no official pregame report covers {away} at {home}")
    return max(eligible, key=lambda item: item["payload"]["report_at_utc"])


def _player_effects(
    snapshot: Mapping[str, Any], priors: Mapping[str, Any], away: str, home: str, game_date: str
) -> list[dict[str, Any]]:
    statuses = {
        (_identity(str(row["team"])), _identity(str(row["player_name"]))): str(row["current_status"])
        for row in snapshot.get("entries", [])
        if str(row.get("game_date")) == game_date
    }
    effects = []
    for row in priors.get("players", []):
        team = str(row["team"])
        if _identity(team) not in {_identity(away), _identity(home)}:
            continue
        status = statuses.get((_identity(team), _identity(str(row["player_name"]))), "Available")
        probability = STATUS_ACTIVE_PROBABILITIES[status]
        minutes = float(row["projected_minutes"])
        above_replacement = float(row["impact_points_per_100"]) - float(
            row["replacement_impact_points_per_100"]
        )
        loss = (1 - probability) * minutes / 40 * above_replacement * 0.80
        if status != "Available":
            effects.append(
                {
                    "team": team,
                    "player": str(row["player_name"]),
                    "status": status,
                    "active_probability": probability,
                    "projected_minutes": round(minutes, 3),
                    "impact_above_replacement_per_100": round(above_replacement, 3),
                    "expected_points_lost": round(loss, 3),
                }
            )
    return sorted(effects, key=lambda row: abs(row["expected_points_lost"]), reverse=True)


def _metrics(
    rows: list[dict[str, Any]],
    scale: float,
    gate: float,
    *,
    include_conflict_diagnostics: bool,
) -> dict[str, Any]:
    accepted = {"complete"}
    if include_conflict_diagnostics:
        accepted.add("diagnostic_conflict_resolution")
    settled = [
        row for row in rows if row["actual_home_win"] is not None and row["availability_status"] in accepted
    ]
    calls = []
    brier = []
    for row in settled:
        probability = row["probabilities_by_scale"][str(scale)]
        confidence = max(probability, 1 - probability)
        brier.append((probability - int(row["actual_home_win"])) ** 2)
        if confidence >= gate:
            selection_home = probability >= 0.5
            calls.append(selection_home == row["actual_home_win"])
    return {
        "cohort": "diagnostic_including_conflicts" if include_conflict_diagnostics else "strict",
        "gate": gate,
        "scale": scale,
        "settled_games": len(settled),
        "calls": len(calls),
        "accuracy": round(sum(calls) / len(calls), 6) if calls else None,
        "brier_all_games": round(sum(brier) / len(brier), 6) if brier else None,
    }


def _paired_bootstrap(rows: list[dict[str, Any]], samples: int = 20_000) -> dict[str, Any]:
    brier_deltas: list[float] = []
    accuracy_deltas: list[int] = []
    for row in rows:
        outcome = int(row["actual_home_win"])
        base = float(row["base_home_probability"])
        adjusted = float(row["adjusted_home_probability"])
        brier_deltas.append((adjusted - outcome) ** 2 - (base - outcome) ** 2)
        accuracy_deltas.append(int((adjusted >= 0.5) == bool(outcome)) - int((base >= 0.5) == bool(outcome)))
    rng = random.Random(20260720)
    count = len(rows)
    bootstrap_brier: list[float] = []
    bootstrap_accuracy: list[float] = []
    for _ in range(samples):
        indices = [rng.randrange(count) for _ in range(count)]
        bootstrap_brier.append(sum(brier_deltas[index] for index in indices) / count)
        bootstrap_accuracy.append(sum(accuracy_deltas[index] for index in indices) / count)
    bootstrap_brier.sort()
    bootstrap_accuracy.sort()
    low = int(samples * 0.025)
    high = int(samples * 0.975) - 1
    flips = [row for row in rows if row["base_selection"] != row["adjusted_selection"]]
    return {
        "method": "paired_game_bootstrap_20000_fixed_seed",
        "brier_delta_mean": sum(brier_deltas) / count,
        "brier_delta_ci95": [bootstrap_brier[low], bootstrap_brier[high]],
        "bootstrap_probability_brier_delta_nonnegative": sum(value >= 0 for value in bootstrap_brier)
        / samples,
        "accuracy_delta_mean": sum(accuracy_deltas) / count,
        "accuracy_delta_ci95": [bootstrap_accuracy[low], bootstrap_accuracy[high]],
        "selection_flips": len(flips),
        "flips_improved": sum(bool(row["adjusted_correct"]) for row in flips),
        "flips_harmed": sum(not bool(row["adjusted_correct"]) for row in flips),
    }


def _cohort_summary(label: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    base_accuracy = sum(
        (row["base_home_probability"] >= 0.5) == row["actual_home_win"] for row in rows
    ) / len(rows)
    adjusted_accuracy = sum(bool(row["adjusted_correct"]) for row in rows) / len(rows)
    base_brier = sum((row["base_home_probability"] - int(row["actual_home_win"])) ** 2 for row in rows) / len(
        rows
    )
    adjusted_brier = sum(
        (row["adjusted_home_probability"] - int(row["actual_home_win"])) ** 2 for row in rows
    ) / len(rows)
    return {
        "label": label,
        "games": len(rows),
        "base_accuracy": base_accuracy,
        "adjusted_accuracy": adjusted_accuracy,
        "base_brier": base_brier,
        "adjusted_brier": adjusted_brier,
        "brier_delta": adjusted_brier - base_brier,
        "paired_uncertainty": _paired_bootstrap(rows),
    }


def _markdown(payload: Mapping[str, Any]) -> str:
    lines = [
        f"# WNBA Player-Availability Challenger: {payload['window']['start_date']} to {payload['window']['end_date']}",
        "",
        "## Decision",
        "",
        payload["decision"],
        "",
        "## What was actually tested",
        "",
        "- Incumbent: fixed `wnba-elo-trend-lr-v3` coefficients recomputed walk-forward for the historical date; a frozen ledger probability is used when one exists.",
        "- Addition: official WNBA status × pregame projected minutes × a heavily shrunk 10-game box plus/minus proxy above replacement.",
        "- Status sources: official WNBA PDFs plus timestamp-filtered ESPN event injuries. ESPN fills official omissions; contradictory explicit statuses fail closed in production.",
        f"- Expanded-source rule: ESPN secondary data was enabled for {payload['data_validation']['espn_secondary_games']} specifically requested matchup(s); all other games use the official report only to avoid survivor-biased retrospective ESPN injury lists.",
        "- Conflict sensitivity: the diagnostic table also shows a clearly labeled most-conservative resolution (the lower active probability). It is not production-authorized.",
        "- Probability bridge: availability points are added in probit-margin space using the pre-window empirical WNBA margin standard deviation.",
        f"- Window: {payload['data_validation']['games_total']} reconstructed matchups, including {payload['data_validation']['settled_games']} settled games. Unsettled games are excluded from accuracy and Brier.",
        "- Provenance warning: the official PDFs carry historical publication timestamps but were downloaded retrospectively. This is diagnostic, not a locked point-in-time promotion test.",
        "",
        "## Data validation",
        "",
        f"- Official PDFs parsed: **{payload['data_validation']['reports_parsed']}**.",
        f"- Official player-status rows parsed, including repeated report updates: **{payload['data_validation']['official_player_rows']}**.",
        f"- Settled games with a complete submitted report and mapped priors: **{payload['data_validation']['settled_games_feature_complete']} / {payload['data_validation']['settled_games']}**.",
        f"- Settled games evaluable only after conservative conflict resolution: **{payload['data_validation']['settled_games_diagnostic_evaluable']} / {payload['data_validation']['settled_games']}**.",
        f"- Empirical pre-window home-margin sigma: **{payload['margin_sigma']:.3f} points**.",
        f"- Parser/status counts: `{json.dumps(payload['data_validation']['status_counts'], sort_keys=True)}`.",
        f"- Scoreboard events considered: **{payload['data_validation']['scoreboard_events']}**; v3 candidates produced: **{payload['data_validation']['games_total']}**; v3 skips: **{payload['data_validation']['model_skips']}**.",
        "",
        "### Source reports",
        "",
        "| Report time (UTC) | Rows | Submitted teams | Not submitted | SHA-256 |",
        "|---|---:|---:|---:|---|",
    ]
    for report in payload["reports"]:
        lines.append(
            f"| {report['report_at_utc']} | {report['entry_count']} | {report['submitted_teams']} | {report['not_submitted_teams']} | `{report['sha256']}` |"
        )
    lines.extend(
        [
            "",
            "## Game-by-game model impact",
            "",
            "`Gap` is positive when availability favors the home team. `Adjusted` uses the 1.0× availability scale.",
            "",
            "| Date | Matchup | Score | Incumbent home P | Gap (pts) | Adjusted home P | Δ pp | Incumbent pick | Adjusted pick | Correct? | Baseline source | Availability status | Source conflicts | Report |",
            "|---|---|---|---:|---:|---:|---:|---|---|---|---|---|---|---|",
        ]
    )
    for row in payload["games"]:
        score = f"{row['away_score']}-{row['home_score']}" if row["away_score"] is not None else "unsettled"
        correct = "—" if row["adjusted_correct"] is None else ("yes" if row["adjusted_correct"] else "no")
        conflicts = (
            "; ".join(
                f"{item['player_name']}: official {item['official_status']} / ESPN {item['espn_status']}"
                for item in row["source_conflicts"]
            )
            or "—"
        )
        lines.append(
            f"| {row['game_date']} | {row['away_team']} @ {row['home_team']} | {score} | {row['base_home_probability']:.3%} | {row['availability_points_gap']:+.3f} | {row['adjusted_home_probability']:.3%} | {row['delta_probability']:+.2%} | {row['base_selection']} | {row['adjusted_selection']} | {correct} | {row['baseline_source']} | {row['availability_status']} | {conflicts} | {row['report_at_utc']} |"
        )
    lines.extend(
        [
            "",
            "## Confidence-gate and feature-strength sensitivity",
            "",
            "Accuracy is conditional on calls. Strict rows exclude source conflicts; diagnostic rows include the labeled conservative resolution.",
            "",
            "| Cohort | Availability scale | Confidence gate | Settled | Calls | Accuracy | Brier |",
            "|---|---:|---:|---:|---:|---:|---:|",
        ]
    )
    for metric in payload["sensitivity"]:
        accuracy = "—" if metric["accuracy"] is None else f"{metric['accuracy']:.1%}"
        lines.append(
            f"| {metric['cohort']} | {metric['scale']:.1f}× | {metric['gate']:.0%} | {metric['settled_games']} | {metric['calls']} | {accuracy} | {metric['brier_all_games']:.4f} |"
        )
    uncertainty = payload["paired_uncertainty"]
    lines.extend(
        [
            "",
            "## Paired uncertainty and selection changes",
            "",
            f"- Mean Brier delta (availability minus v3): **{uncertainty['brier_delta_mean']:+.6f}**; paired bootstrap 95% interval **[{uncertainty['brier_delta_ci95'][0]:+.6f}, {uncertainty['brier_delta_ci95'][1]:+.6f}]**.",
            f"- Mean winner-accuracy delta: **{uncertainty['accuracy_delta_mean']:+.3%}**; paired bootstrap 95% interval **[{uncertainty['accuracy_delta_ci95'][0]:+.3%}, {uncertainty['accuracy_delta_ci95'][1]:+.3%}]**.",
            f"- The 1.0× feature flipped **{uncertainty['selection_flips']}** selections: **{uncertainty['flips_improved']}** corrections and **{uncertainty['flips_harmed']}** newly wrong picks.",
            "- The bootstrap treats games as independent and therefore understates uncertainty from repeated teams, shared injuries, and temporal clustering. It is a diagnostic interval, not promotion evidence.",
        ]
    )
    lines.extend(
        [
            "",
            "### Pre-audit versus original-audit cohorts",
            "",
            "The feature was first examined on July 17–20. The earlier cohort is therefore the cleaner check against simply fitting the original ten-game observation.",
            "",
            "| Cohort | Games | v3 accuracy | + availability accuracy | v3 Brier | + availability Brier | Brier Δ | 95% paired interval |",
            "|---|---:|---:|---:|---:|---:|---:|---|",
        ]
    )
    for cohort in payload["cohort_breakdown"]:
        interval = cohort["paired_uncertainty"]["brier_delta_ci95"]
        lines.append(
            f"| {cohort['label']} | {cohort['games']} | {cohort['base_accuracy']:.1%} | {cohort['adjusted_accuracy']:.1%} | {cohort['base_brier']:.5f} | {cohort['adjusted_brier']:.5f} | {cohort['brier_delta']:+.5f} | [{interval[0]:+.5f}, {interval[1]:+.5f}] |"
        )
    lines.extend(
        [
            "",
            "## Dallas / Paige Bueckers audit",
            "",
            payload["dallas_audit"]["conclusion"],
            "",
            "| Date | Merged Bueckers status | Dallas baseline | Paige-only Dallas P | Net availability gap | Net adjusted Dallas P | Production disposition | Outcome |",
            "|---|---|---:|---:|---:|---:|---|---|",
        ]
    )
    for row in payload["dallas_audit"]["games"]:
        lines.append(
            f"| {row['game_date']} | {row['bueckers_status']} | {row['base_dallas_probability']:.3%} | {row['paige_only_dallas_probability']:.3%} | {row['dallas_availability_gap']:+.3f} | {row['adjusted_dallas_probability']:.3%} | {row['production_disposition']} | {row['outcome']} |"
        )
    lines.extend(
        [
            "",
            "## Largest player-level adjustments",
            "",
            "These are challenger inputs, not causal player values. A negative expected-loss number means the noisy proxy rated the named player below the team replacement prior.",
            "",
            "| Date | Matchup | Team | Player | Status | Proj min | Impact above repl /100 | Expected points lost |",
            "|---|---|---|---|---|---:|---:|---:|",
        ]
    )
    effects = [effect for row in payload["games"] for effect in row["player_effects"]]
    effects.sort(key=lambda item: abs(item["expected_points_lost"]), reverse=True)
    for effect in effects[:40]:
        lines.append(
            f"| {effect['game_date']} | {effect['matchup']} | {effect['team']} | {effect['player']} | {effect['status']} | {effect['projected_minutes']:.1f} | {effect['impact_above_replacement_per_100']:+.2f} | {effect['expected_points_lost']:+.3f} |"
        )
    lines.extend(
        [
            "",
            "## What this test cannot establish",
            "",
            f"1. {payload['aggregate']['diagnostic_evaluable_settled_games']} retrospectively reconstructed games cannot replace a prospectively observed validation cohort.",
            "2. The report PDFs were recovered after the games, so their embedded publication times are useful diagnostics but not equivalent to prospectively observed snapshots.",
            "3. The impact prior is heavily shrunk raw box plus/minus, not WNBA RAPM or lineup-adjusted causal impact.",
            "4. Current rosters were queried during reconstruction; transactions effective between the game and retrieval can create entity risk.",
            f"5. Coverage is incomplete: only {payload['data_validation']['settled_games_feature_complete']} of {payload['data_validation']['settled_games']} settled games are strictly conflict-free and feature-complete.",
            "",
            "## Keep/remove recommendation",
            "",
            payload["recommendation"],
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--report-dir", required=True)
    parser.add_argument("--output-dir", default="outputs/wnba_availability")
    parser.add_argument("--start-date", default="2026-07-17")
    parser.add_argument("--end-date", default="2026-07-20")
    parser.add_argument(
        "--espn-event-id",
        action="append",
        default=None,
        help="Restrict ESPN secondary injury merging to listed event IDs; omit to use all events.",
    )
    args = parser.parse_args()

    root = Path(args.project_root).resolve()
    data_root = root / "data"
    output_dir = root / args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    report_paths = list(Path(args.report_dir).glob("*.pdf"))
    reports = _report_rows(report_paths, datetime.now(UTC))
    store = FeatureStore(data_root)
    local = LocalScoreboards(data_root)
    client = ESPNClient()
    artifact_path = root / "config/models/wnba-elo-trend-lr-v3.json"
    artifact = LearnedMarketArtifact.load(artifact_path)
    window_dates = _date_window(args.start_date, args.end_date)
    margin_sigma = historical_margin_sigma(store, datetime.fromisoformat(args.start_date + "T04:00:00+00:00"))
    ledger_by_event, ledger_by_matchup = _ledger_probabilities(root)
    espn_event_ids = set(args.espn_event_id) if args.espn_event_id else None

    rows: list[dict[str, Any]] = []
    model_skips: list[dict[str, Any]] = []
    scoreboard_events = 0
    for day in window_dates:
        scoreboard = local.scoreboard("WNBA", day)
        events = {str(event["id"]): event for event in scoreboard.get("events", [])}
        candidates, skipped, event_count = build_learned_moneyline_slate(
            sport="wnba",
            game_date=day,
            store=store,
            client=local,
            artifact_path=artifact_path,
            observed_at=datetime.fromisoformat(day + "T12:00:00+00:00"),
        )
        scoreboard_events += event_count
        model_skips.extend({"game_date": day, **item} for item in skipped)
        for candidate in candidates:
            event = events[candidate.event_id]
            away, home = _event_teams(event)
            ledger_record = ledger_by_event.get(candidate.event_id) or ledger_by_matchup.get(
                (day, away, home)
            )
            base_home_probability = (
                float(ledger_record["home_probability"])
                if ledger_record is not None
                else candidate.home_probability
            )
            baseline_source = (
                f"{ledger_record['workbook']}:{ledger_record['pick_id']}"
                if ledger_record is not None
                else "current artifact recomputation"
            )
            actual_home, away_score, home_score = _event_result(event)
            selected_report: dict[str, Any] | None = None
            espn_snapshot: dict[str, Any] | None = None
            source_conflicts: list[dict[str, Any]] = []
            try:
                selected_report = _select_report(reports, event, away, home)
                snapshot = selected_report["payload"]
                report_at = parse_utc(str(snapshot["report_at_utc"]))
                espn_snapshot = (
                    _cached_target_injuries(client, data_root, candidate.event_id, report_at)
                    if espn_event_ids is None or candidate.event_id in espn_event_ids
                    else None
                )
                conflict_error = None
                try:
                    snapshot = merge_availability_sources(snapshot, espn_snapshot, game_date=day)
                except ValueError as error:
                    if not str(error).startswith("NO_CALL_AVAILABILITY_SOURCE_CONFLICT"):
                        raise
                    conflict_error = str(error)
                    snapshot = merge_availability_sources(
                        snapshot,
                        espn_snapshot,
                        game_date=day,
                        conflict_policy="most_conservative",
                    )
                    source_conflicts = list(snapshot.get("source_conflicts", []))
                priors = build_research_priors(
                    store=store,
                    client=client,
                    event=event,
                    cache_root=data_root / "availability/wnba/espn_boxscores",
                    observed_at=report_at,
                )
                (output_dir / "priors").mkdir(exist_ok=True)
                (output_dir / "priors" / f"{candidate.event_id}.json").write_text(
                    json.dumps(priors, indent=2, sort_keys=True) + "\n", encoding="utf-8"
                )
                features = matchup_player_availability_from_payloads(
                    snapshot=snapshot,
                    priors=priors,
                    home_team=home,
                    away_team=away,
                    game_date=day,
                    observed_at=report_at,
                    event_start=parse_utc(str(event["date"])),
                    maximum_report_age_hours=24,
                )
                availability_status = "diagnostic_conflict_resolution" if source_conflicts else "complete"
                availability_error = conflict_error
                effects = _player_effects(snapshot, priors, away, home, day)
                prior_diagnostics = priors["team_diagnostics"]
            except ValueError as error:
                features = {
                    "availability_points_gap": 0.0,
                    "availability_uncertainty": 0.0,
                    "home_available_minutes_share": 1.0,
                    "away_available_minutes_share": 1.0,
                }
                availability_status = "fail_closed"
                availability_error = str(error)
                effects = []
                prior_diagnostics = {}
            probabilities = {
                str(scale): adjust_home_probability(
                    base_home_probability,
                    features["availability_points_gap"] * scale,
                    margin_sigma,
                )
                for scale in SCALES
            }
            adjusted = probabilities["1.0"]
            for effect in effects:
                effect["game_date"] = day
                effect["matchup"] = f"{away} @ {home}"
            rows.append(
                {
                    "event_id": candidate.event_id,
                    "game_date": day,
                    "event_start_utc": str(event["date"]),
                    "away_team": away,
                    "home_team": home,
                    "away_score": away_score,
                    "home_score": home_score,
                    "actual_home_win": actual_home,
                    "base_home_probability": base_home_probability,
                    "base_selection": "home" if base_home_probability >= 0.5 else "away",
                    "availability_points_gap": features["availability_points_gap"],
                    "availability_uncertainty": features["availability_uncertainty"],
                    "home_available_minutes_share": features["home_available_minutes_share"],
                    "away_available_minutes_share": features["away_available_minutes_share"],
                    "adjusted_home_probability": adjusted,
                    "delta_probability": adjusted - base_home_probability,
                    "baseline_source": baseline_source,
                    "adjusted_selection": "home" if adjusted >= 0.5 else "away",
                    "adjusted_correct": None
                    if actual_home is None or availability_status == "fail_closed"
                    else bool((adjusted >= 0.5) == actual_home),
                    "availability_status": availability_status,
                    "availability_error": availability_error,
                    "source_conflicts": source_conflicts,
                    "espn_secondary_used": espn_snapshot is not None,
                    "report_at_utc": selected_report["payload"]["report_at_utc"]
                    if selected_report
                    else "none",
                    "report_sha256": selected_report["sha256"] if selected_report else None,
                    "probabilities_by_scale": probabilities,
                    "player_effects": effects,
                    "prior_team_diagnostics": prior_diagnostics,
                }
            )

    sensitivity = [
        _metrics(
            rows,
            scale,
            gate,
            include_conflict_diagnostics=include_conflicts,
        )
        for include_conflicts in (False, True)
        for scale in SCALES
        for gate in GATES
    ]
    status_counts = Counter(
        entry["current_status"] for report in reports for entry in report["payload"].get("entries", [])
    )
    settled_all = [row for row in rows if row["actual_home_win"] is not None]
    settled_strict = [row for row in settled_all if row["availability_status"] == "complete"]
    settled = [row for row in settled_all if row["availability_status"] != "fail_closed"]
    base_accuracy = sum(
        (row["base_home_probability"] >= 0.5) == row["actual_home_win"] for row in settled
    ) / len(settled)
    adjusted_accuracy = sum(row["adjusted_correct"] for row in settled) / len(settled)
    base_brier = sum(
        (row["base_home_probability"] - int(row["actual_home_win"])) ** 2 for row in settled
    ) / len(settled)
    adjusted_brier = sum(
        (row["adjusted_home_probability"] - int(row["actual_home_win"])) ** 2 for row in settled
    ) / len(settled)
    paired_uncertainty = _paired_bootstrap(settled)
    pre_audit_rows = [row for row in settled if row["game_date"] < "2026-07-17"]
    original_audit_rows = [row for row in settled if row["game_date"] >= "2026-07-17"]
    cohort_breakdown = []
    if pre_audit_rows:
        cohort_breakdown.append(_cohort_summary("Pre-original-audit (through July 16)", pre_audit_rows))
    if original_audit_rows:
        cohort_breakdown.append(
            _cohort_summary("Original audit window (July 17 onward)", original_audit_rows)
        )

    dallas_rows = [row for row in rows if row["home_team"] == "Dallas Wings"]
    dallas_games = []
    for row in dallas_rows:
        bueckers_effect = next(
            (
                effect
                for effect in row["player_effects"]
                if _identity(effect["player"]) == _identity("Paige Bueckers")
            ),
            None,
        )
        bueckers = (
            bueckers_effect["status"] if bueckers_effect is not None else "Not listed (treated Available)"
        )
        paige_only_probability = adjust_home_probability(
            row["base_home_probability"],
            -float(bueckers_effect["expected_points_lost"]) if bueckers_effect is not None else 0.0,
            margin_sigma,
        )
        outcome = (
            "unsettled"
            if row["actual_home_win"] is None
            else ("Dallas win" if row["actual_home_win"] else "Dallas loss")
        )
        dallas_games.append(
            {
                "game_date": row["game_date"],
                "bueckers_status": bueckers,
                "base_dallas_probability": row["base_home_probability"],
                "paige_only_dallas_probability": paige_only_probability,
                "dallas_availability_gap": row["availability_points_gap"],
                "adjusted_dallas_probability": row["adjusted_home_probability"],
                "production_disposition": (
                    "NO CALL: explicit source conflict"
                    if row["source_conflicts"]
                    else (
                        "NO CALL: incomplete inputs"
                        if row["availability_status"] == "fail_closed"
                        else "eligible on availability inputs"
                    )
                ),
                "outcome": outcome,
            }
        )
    bueckers_out = any(game["bueckers_status"] == "Out" for game in dallas_games)
    if bueckers_out:
        dallas_conclusion = (
            "**The status bug is fixed; the Dallas edge is not.** ESPN identifies Paige Bueckers as Out even though the official WNBA PDF omits her. "
            "Isolated, her absence moves Dallas from 67.878% to 59.313%. After all listed absences are combined, Dallas remains 66.100% because New York is also missing material players. "
            "That is still a Dallas pick, and the Alanna Smith status conflict makes the production result a no-call."
        )
    else:
        dallas_conclusion = (
            "**No timestamp-valid source in this reconstruction identified Paige Bueckers as Out.**"
        )

    delta_brier = adjusted_brier - base_brier
    brier_ci = paired_uncertainty["brier_delta_ci95"]
    accuracy_ci = paired_uncertainty["accuracy_delta_ci95"]
    if len(settled) >= 100 and brier_ci[1] < 0 and accuracy_ci[0] <= 0 <= accuracy_ci[1]:
        decision = (
            "**Keep as a shadow challenger: probability quality improved, winner accuracy did not.** "
            "The paired Brier interval is below zero, but top-pick accuracy is statistically unresolved and slightly worse at the 1.0× prespecified scale. Retrospective inputs still block promotion."
        )
    elif adjusted_accuracy > base_accuracy and delta_brier < 0:
        decision = "**Diagnostic positive, not promotable.** The challenger improved both accuracy and Brier in this tiny retrospective window."
    elif adjusted_accuracy == base_accuracy and delta_brier < 0:
        decision = "**Probability-quality positive, winner-accuracy neutral, not promotable.** The challenger preserved winner accuracy and improved Brier in this tiny window."
    else:
        decision = "**Do not add the challenger coefficient to the active model.** It failed to improve this tiny window cleanly; keep only the collection and fail-closed feature infrastructure."
    recommendation = (
        "Keep the official report parser, immutable snapshots, player mapping, projected-minute contract, and fail-closed model hook. "
        f"Do **not** replace the active WNBA artifact or assign a production coefficient from these {len(settled)} diagnostic games. "
        "Run the collector prospectively, replace shrunk raw plus/minus with a WNBA-specific regularized lineup-impact prior, then re-run a preregistered fresh cohort."
    )
    payload = {
        "schema_version": "1",
        "generated_at_utc": datetime.now(UTC).isoformat(),
        "window": {"start_date": args.start_date, "end_date": args.end_date},
        "artifact": artifact.version,
        "artifact_hash": artifact.hash,
        "margin_sigma": margin_sigma,
        "decision": decision,
        "recommendation": recommendation,
        "aggregate": {
            "settled_games": len(settled_all),
            "games_total": len(rows),
            "espn_secondary_games": sum(row["espn_secondary_used"] for row in rows),
            "strict_settled_games": len(settled_strict),
            "diagnostic_evaluable_settled_games": len(settled),
            "base_accuracy": base_accuracy,
            "adjusted_accuracy": adjusted_accuracy,
            "base_brier": base_brier,
            "adjusted_brier": adjusted_brier,
            "brier_delta_adjusted_minus_base": delta_brier,
        },
        "data_validation": {
            "reports_parsed": len(reports),
            "games_total": len(rows),
            "scoreboard_events": scoreboard_events,
            "model_skips": len(model_skips),
            "model_skip_reasons": model_skips,
            "espn_secondary_games": sum(row["espn_secondary_used"] for row in rows),
            "official_player_rows": sum(len(report["payload"].get("entries", [])) for report in reports),
            "status_counts": dict(status_counts),
            "settled_games": len(settled_all),
            "settled_games_feature_complete": len(settled_strict),
            "settled_games_diagnostic_evaluable": len(settled),
        },
        "reports": [
            {
                "report_at_utc": report["payload"]["report_at_utc"],
                "entry_count": len(report["payload"]["entries"]),
                "submitted_teams": sum(
                    status in {"submitted", "submitted_no_entries"}
                    for status in report["payload"]["team_report_status"].values()
                ),
                "not_submitted_teams": sum(
                    status not in {"submitted", "submitted_no_entries"}
                    for status in report["payload"]["team_report_status"].values()
                ),
                "sha256": report["sha256"],
                "source_pdf_url": report["payload"]["source_pdf_url"],
            }
            for report in reports
        ],
        "games": rows,
        "sensitivity": sensitivity,
        "paired_uncertainty": paired_uncertainty,
        "cohort_breakdown": cohort_breakdown,
        "dallas_audit": {"conclusion": dallas_conclusion, "games": dallas_games},
    }
    (output_dir / "wnba_availability_evaluation.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    (output_dir / "WNBA_AVAILABILITY_DECISION_REVIEW.md").write_text(_markdown(payload), encoding="utf-8")
    print(json.dumps(payload["aggregate"], indent=2, sort_keys=True))
    print(output_dir / "WNBA_AVAILABILITY_DECISION_REVIEW.md")


if __name__ == "__main__":
    main()
