"""Migrate all MLB v3 picks in both ledgers to v4 model lineage.

Strategy:
- Keep existing model_probability (decision-time value, not retroactively changed)
- Keep existing market prices (american_odds, market_implied_probability)
- Update model_version, model_artifact_hash, calibration fields to v4
- Recompute edge, confidence_score, units from existing model_prob + market price
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from model_prediction.config import load_config, ledger_path, unit_policy
from model_prediction.units import edge_scaled_units

V4_VERSION = "mlb-elo-trend-lr-v4"
V4_HASH = "5224fb6ffbc9ddf8fa517627b830ac851192da94d241963d556945393a42bb9d"


def _col_map(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value or ""): c for c in range(1, ws.max_column + 1)}


def migrate_ledger(path: Path, policy) -> dict:
    wb = openpyxl.load_workbook(path)
    ws = wb["Picks"]
    cols = _col_map(ws)

    updated = 0
    for r in range(2, ws.max_row + 1):
        league = str(ws.cell(r, cols.get("league", 1)).value or "").upper()
        mv = str(ws.cell(r, cols.get("model_version", 1)).value or "")
        if league != "MLB" or mv != "mlb-elo-trend-lr-v3":
            continue

        # ── Metadata swap ──────────────────────────────────────────
        ws.cell(r, cols["model_version"]).value = V4_VERSION
        ws.cell(r, cols["model_artifact_hash"]).value = V4_HASH
        if "calibration_version" in cols:
            ws.cell(r, cols["calibration_version"]).value = V4_VERSION
        if "calibration_artifact_hash" in cols:
            ws.cell(r, cols["calibration_artifact_hash"]).value = V4_HASH

        # ── Recompute edge from existing prices ─────────────────────
        model_prob = float(ws.cell(r, cols["model_probability"]).value or 0.5)
        market_implied = float(ws.cell(r, cols["market_implied_probability"]).value or 0.5)
        edge = model_prob - market_implied
        ws.cell(r, cols["edge"]).value = round(edge, 6)

        # ── Recompute confidence score ──────────────────────────────
        confidence = min(100, max(0, int(edge * 1000))) if edge > 0 else max(1, int(100 + edge * 1000))
        ws.cell(r, cols["confidence_score"]).value = confidence

        # ── Recompute units via edge_scaled_units ───────────────────
        american_odds = int(ws.cell(r, cols["american_odds"]).value or -110)
        new_units = edge_scaled_units(model_prob, 0.05, american_odds, policy)
        ws.cell(r, cols["units"]).value = new_units

        # ── Update rationale version string ─────────────────────────
        if "rationale" in cols:
            old = str(ws.cell(r, cols["rationale"]).value or "")
            if "mlb-elo-trend-lr-v3" in old:
                ws.cell(r, cols["rationale"]).value = old.replace(
                    "mlb-elo-trend-lr-v3", V4_VERSION
                )

        updated += 1

    wb.save(path)
    return {"updated": updated}


def main():
    config = load_config()
    policy = unit_policy(config)
    data_root = Path(ledger_path(config)).parent

    print("=== MLB v3 → v4 Migration ===")
    print(f"Version: {V4_VERSION}  Hash: {V4_HASH[:16]}...")
    print()

    for fname in ["picks.xlsx", "flat_picks.xlsx"]:
        path = data_root / fname
        print(f"--- {fname} ---")
        stats = migrate_ledger(path, policy)
        print(f"  Updated {stats['updated']} picks")
        print()

    print("=== Done ===")


if __name__ == "__main__":
    main()
