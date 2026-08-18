"""Systematic post-loss review workflow (ops brainstorm, 2026-08-17).

Payloads already carry loss_classification/loss_cause/review_status on
settled loss rows, but that data only gets looked at one row at a time via
the manual `review_loss` workflow -- a run of consecutive losses on the SAME
model can silently accumulate without ever being flagged as a DISTINCT
signal from "here's another individually-reviewable loss". This is a
read-only report: for each (tier, sport, model_version), walk settled rows
chronologically and surface any active streak of >= --min-streak
consecutive losses, whether each loss in it has been reviewed, and how the
reviewed ones were classified (variance-vs-signal split, using this
project's own LOSS_CLASSIFICATIONS taxonomy: "bad_luck"/"missing_
information" read as variance-leaning, "model_error"/"market_or_rule_error"/
"process_error"/"bad_data" read as signal-leaning).

Reads the xlsx workbooks directly via openpyxl in read-only mode -- NOT
through PickLedger (which takes a file lock and can run schema migration
on open) -- to guarantee this can never contend with or mutate a live
ledger the daily/production scheduler might be writing to concurrently.
Never writes anything back.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from model_prediction.config import PROJECT_ROOT

DATA_ROOT = Path("/Users/vincentc9002/model-prediction/data")
TIERS = ("main", "flat")
VARIANCE_LEANING = {"bad_luck", "missing_information"}
SIGNAL_LEANING = {"model_error", "market_or_rule_error", "process_error", "bad_data"}


def _read_picks(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Picks" not in wb.sheetnames:
            return []
        ws = wb["Picks"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        idx = {name: i for i, name in enumerate(header) if name}
        needed = (
            "model_version",
            "event_start_utc",
            "status",
            "result",
            "review_status",
            "loss_classification",
        )
        if not all(name in idx for name in needed):
            return []
        out = []
        for row in rows_iter:
            if row is None:
                continue
            entry = {name: row[idx[name]] for name in needed}
            out.append(entry)
        return out
    finally:
        wb.close()


def _streaks(rows: list[dict], min_streak: int) -> list[dict]:
    settled = [r for r in rows if r["status"] == "settled" and r["result"] in ("win", "loss")]
    settled.sort(key=lambda r: str(r["event_start_utc"] or ""))

    by_model: dict[str, list[dict]] = {}
    for row in settled:
        by_model.setdefault(str(row["model_version"] or "unknown"), []).append(row)

    findings = []
    for model_version, model_rows in by_model.items():
        current_streak: list[dict] = []
        streaks: list[list[dict]] = []
        for row in model_rows:
            if row["result"] == "loss":
                current_streak.append(row)
            else:
                if current_streak:
                    streaks.append(current_streak)
                current_streak = []
        if current_streak:
            streaks.append(current_streak)  # active streak, still ongoing as of the most recent settled pick

        for streak in streaks:
            if len(streak) < min_streak:
                continue
            is_active = streak is streaks[-1] and streak[-1] is model_rows[-1]
            reviewed = [r for r in streak if r["review_status"] == "complete"]
            classifications = [r["loss_classification"] for r in reviewed if r["loss_classification"]]
            variance_count = sum(1 for c in classifications if c in VARIANCE_LEANING)
            signal_count = sum(1 for c in classifications if c in SIGNAL_LEANING)
            findings.append(
                {
                    "model_version": model_version,
                    "streak_length": len(streak),
                    "start_date": str(streak[0]["event_start_utc"])[:10],
                    "end_date": str(streak[-1]["event_start_utc"])[:10],
                    "is_active_as_of_latest_settled_pick": is_active,
                    "reviewed_count": len(reviewed),
                    "unreviewed_count": len(streak) - len(reviewed),
                    "variance_leaning_reviews": variance_count,
                    "signal_leaning_reviews": signal_count,
                    "needs_operator_attention": is_active
                    and (len(reviewed) < len(streak) or signal_count > 0),
                }
            )
    return findings


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--min-streak", type=int, default=3)
    parser.add_argument("--out", default=str(PROJECT_ROOT / "outputs/research/loss_streak_review.json"))
    args = parser.parse_args()

    all_findings = []
    for tier in TIERS:
        tier_dir = DATA_ROOT / tier
        if not tier_dir.exists():
            continue
        for xlsx_path in sorted(tier_dir.glob("*.xlsx")):
            rows = _read_picks(xlsx_path)
            if not rows:
                continue
            findings = _streaks(rows, args.min_streak)
            for f in findings:
                f["tier"] = tier
                f["sport"] = xlsx_path.stem
            all_findings.extend(findings)

    flagged = [f for f in all_findings if f["needs_operator_attention"]]
    report = {
        "min_streak": args.min_streak,
        "total_streaks_found": len(all_findings),
        "flagged_for_operator_attention": len(flagged),
        "flagged": flagged,
        "all_streaks": all_findings,
    }
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(report, indent=2, default=str) + "\n", encoding="utf-8")

    print(f"{len(all_findings)} loss streak(s) >= {args.min_streak} found across {len(TIERS)} tiers")
    print(f"{len(flagged)} flagged for operator attention (active + unreviewed-or-signal-leaning):")
    for f in flagged:
        print(
            f"  {f['tier']}/{f['sport']}/{f['model_version']}: {f['streak_length']} losses "
            f"({f['start_date']}..{f['end_date']}), {f['unreviewed_count']} unreviewed, "
            f"{f['signal_leaning_reviews']} signal-leaning review(s)"
        )
    print(f"\nwritten to {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
