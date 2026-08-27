import fcntl
from datetime import UTC, datetime

import pytest
from openpyxl import Workbook, load_workbook

from model_prediction.domain import (
    League,
    MarketType,
    ModelOrigin,
    ModelState,
    PickRequest,
    PickResult,
    RecordType,
)
from model_prediction.eligibility import evaluate_eligibility
from model_prediction.ledger import FIELDNAMES, LEGACY_FIELDNAMES, PickLedger, _settlement_pnl
from model_prediction.units import Exposure, UnitPolicy
from model_prediction.xlsx_ledger import read_xlsx_rows

NOW = datetime(2026, 7, 13, 12, tzinfo=UTC)


def request(
    event_id: str, probability: float = 0.6235, state: ModelState = ModelState.SHADOW_QUALIFIED
) -> PickRequest:
    return PickRequest(
        event_start_utc="2026-07-14T00:00:00Z",
        event_id=event_id,
        league=League.MLB,
        away_team="BOS",
        home_team="BAL",
        market_type=MarketType.MONEYLINE,
        selection="home",
        line=None,
        sportsbook="Book",
        american_odds=-110,
        model_probability=probability,
        model_uncertainty=0.004,
        model_version="stat-v1",
        rationale="fixture",
        risks="",
        model_origin=ModelOrigin.STATISTICAL_MODEL,
        model_state=state,
        observed_at_utc="2026-07-13T11:00:00Z",
        calibration_version="cal-v1",
        model_artifact_hash="model-hash",
        calibration_artifact_hash="calibration-hash",
        code_revision="abc123",
    )


def test_research_is_zero_unit_excluded_from_exposure_roi_but_in_calibration(
    registry, ban_list, tmp_path
) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    # SUSPENDED, not RESEARCH -- operator directive, 2026-08-02: promotion
    # tier no longer gates qualified calls, so RESEARCH now produces a real
    # QUALIFIED_SHADOW_CALL too. SUSPENDED remains a hard stop, so it's what
    # this test uses to reliably exercise a genuine RESEARCH_OBSERVATION row.
    research_request = request("research", state=ModelState.SUSPENDED)
    research_gate = evaluate_eligibility(research_request, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    research_row = ledger.append_evaluated(research_request, research_gate, NOW)
    assert research_row["record_type"] == RecordType.RESEARCH_OBSERVATION.value
    # Every logged pick carries a real paper size now (operator directive,
    # 2026-07-31) -- what stays excluded is exposure/ROI accounting below,
    # which filters on record_type, not on units being zero.
    assert float(research_row["units"]) > 0
    ledger.settle(research_row["pick_id"], 2, 3)

    qualified_request = request("qualified")
    qualified_gate = evaluate_eligibility(
        qualified_request, registry, ban_list, Exposure(), UnitPolicy(), NOW
    )
    qualified_row = ledger.append_evaluated(qualified_request, qualified_gate, NOW)
    ledger.settle(qualified_row["pick_id"], 3, 2)

    exposure = ledger.exposure(qualified_request, NOW, ("mlb-bos", "mlb-bal"))
    assert exposure.daily_units == float(qualified_row["units"])
    report = ledger.report()
    assert report["qualified_shadow_calls"] == 1
    assert report["research_observations"] == 1
    assert report["qualified_losses"] == 1
    assert report["qualified_pnl_units"] < 0
    assert report["calibration"]["sample_size"] == 2
    assert report["by_model_version"] == {"stat-v1": 2}


def test_research_can_auto_score_one_hypothetical_unit_on_settlement(registry, ban_list, tmp_path) -> None:
    ledger = PickLedger(
        tmp_path / "picks.xlsx",
        tmp_path / "events.jsonl",
        research_score_units=1.0,
        research_scoring_note="one-unit hypothetical policy",
    )
    req = request(
        "research-one-unit", state=ModelState.SUSPENDED
    )  # see the note above on why SUSPENDED, not RESEARCH
    gate = evaluate_eligibility(req, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    row = ledger.append_evaluated(req, gate, NOW)

    settled = ledger.settle(row["pick_id"], 2, 3)

    # units/pnl_units are the row's own real paper size now (operator
    # directive, 2026-07-31); research_score_units/research_pnl_units remain
    # the separate, independently-configured hypothetical-scoring columns.
    assert float(settled["units"]) > 0
    assert settled["research_score_units"] == "1.0000"
    assert float(settled["research_pnl_units"]) > 0
    assert settled["research_scoring_note"] == "one-unit hypothetical policy"


def test_research_can_keep_model_recommended_units_on_settlement(registry, ban_list, tmp_path) -> None:
    ledger = PickLedger(
        tmp_path / "picks.xlsx",
        tmp_path / "events.jsonl",
        research_score_units=1.0,
        research_scoring_mode="model_recommended",
        research_scoring_note="decision-time model size",
    )
    req = request("research-model-size", probability=0.672, state=ModelState.SUSPENDED)
    gate = evaluate_eligibility(req, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    row = ledger.append_evaluated(req, gate, NOW)

    settled = ledger.settle(row["pick_id"], 3, 2)

    assert float(settled["units"]) > 0
    assert settled["research_score_units"] == "2.0000"
    assert settled["research_pnl_units"] == "-2.000000"
    assert settled["research_scoring_note"] == "decision-time model size"


def test_research_model_recommended_skips_scoring_without_uncertainty(registry, ban_list, tmp_path) -> None:
    ledger = PickLedger(
        tmp_path / "picks.xlsx",
        tmp_path / "events.jsonl",
        research_score_units=1.0,
        research_scoring_mode="model_recommended",
        research_scoring_note="decision-time model size",
    )
    req = request("research-missing-uncertainty", probability=0.672, state=ModelState.SUSPENDED)
    req = PickRequest(**{**req.__dict__, "model_uncertainty": None})
    gate = evaluate_eligibility(req, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    row = ledger.append_evaluated(req, gate, NOW)
    assert row["model_uncertainty"] == ""

    settled = ledger.settle(row["pick_id"], 3, 2)

    # A row with no recorded uncertainty must not be scored as if uncertainty
    # were 0 (which would manufacture an overconfident unit size) -- it stays
    # unscored instead.
    assert not settled["research_score_units"]
    assert not settled["research_pnl_units"]


def test_banned_total_is_recorded_as_positive_unit_research_paper_call(registry, ban_list, tmp_path) -> None:
    ban_list.add(League.MLB, "BAL")
    req = request("banned-total")
    req = PickRequest(**{**req.__dict__, "market_type": MarketType.TOTAL, "selection": "over", "line": 8.5})
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    gate = evaluate_eligibility(req, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    row = ledger.append_evaluated(req, gate, NOW)
    assert row["decision"] == "CALL"
    assert row["reason_code"] == "PAPER_CALL_TEAM_BANNED"
    assert row["banned_team_id"] == "mlb-bal"
    assert float(row["units"]) > 0
    assert ledger.report()["paper_call_team_banned_count"] == 1


def test_older_excel_schema_migrates_with_backup_and_preserved_units(tmp_path) -> None:
    path = tmp_path / "picks.xlsx"
    old = {field: "" for field in LEGACY_FIELDNAMES}
    old.update(
        {
            "pick_id": "legacy",
            "created_at_utc": "2026-01-01T00:00:00Z",
            "event_start_utc": "2026-01-02T00:00:00Z",
            "event_id": "old",
            "league": "MLB",
            "away_team": "NYY",
            "home_team": "BOS",
            "market_type": "total",
            "selection": "over",
            "line": "8.5",
            "sportsbook": "Book",
            "american_odds": "-110",
            "decimal_odds": "1.909091",
            "market_implied_probability": "0.523810",
            "model_probability": "0.55",
            "model_uncertainty": "0.03",
            "edge": "0.02619",
            "confidence_score": "50",
            "units": "0.25",
            "model_version": "mlb-analyst-v0",
            "status": "open",
            "call_type": "forced_call",
            "rationale": "legacy",
            "review_status": "not_applicable",
        }
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Picks"
    worksheet.append(LEGACY_FIELDNAMES)
    worksheet.append([old[field] for field in LEGACY_FIELDNAMES])
    workbook.save(path)
    ledger = PickLedger(path, tmp_path / "events.jsonl")
    row = ledger.rows()[0]
    assert path.with_suffix(".xlsx.bak-v1").exists()
    assert row["record_type"] == "RESEARCH_OBSERVATION"
    assert row["units"] == "0.00" and row["legacy_units"] == "0.25"
    headers, _ = read_xlsx_rows(path)
    assert headers == FIELDNAMES
    assert ledger.audit.events()[-1]["event_type"] == "ledger_migrated"


def test_settlement_and_void_are_idempotent(tmp_path) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    first = ledger.append_call(request("settle"), 0.25, 60, now=NOW)
    settled = ledger.settle(first["pick_id"], 2, 3)
    assert ledger.settle(first["pick_id"], 2, 3) == settled
    second = ledger.append_call(request("void"), 0.25, 60, now=NOW)
    voided = ledger.void(second["pick_id"], "postponed")
    assert voided["result"] == "push" and float(voided["pnl_units"]) == 0
    assert ledger.void(second["pick_id"], "postponed") == voided


def test_tied_binary_contract_uses_half_value_instead_of_zero_pnl(tmp_path) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    row = ledger.append_call(request("half-value-tie"), 1.0, 60, now=NOW)

    settled = ledger.settle(
        row["pick_id"],
        3,
        3,
        binary_contract_settlement_value=0.5,
    )

    entry_probability = float(row["decision_raw_implied_probability"])
    assert settled["result"] == "push"
    assert float(settled["pnl_units"]) == round(0.5 / entry_probability - 1, 4)
    assert ledger.audit.events()[-1]["payload"]["binary_contract_settlement_value"] == 0.5


def test_prediction_market_pricing_is_selected_by_explicit_venue_not_minus_110() -> None:
    sportsbook_pnl = _settlement_pnl(
        result=PickResult.WIN,
        units=1.0,
        sportsbook="consensus",
        decimal_odds=1.909091,
        entry_probability=0.40,
        binary_contract_settlement_value=None,
    )
    contract_pnl = _settlement_pnl(
        result=PickResult.WIN,
        units=1.0,
        sportsbook="polymarket_us",
        decimal_odds=1.909091,
        entry_probability=0.40,
        binary_contract_settlement_value=None,
    )

    assert sportsbook_pnl == pytest.approx(0.909091)
    assert contract_pnl == pytest.approx(1.5)
    assert (
        _settlement_pnl(
            result=PickResult.PUSH,
            units=1.0,
            sportsbook="polymarket_us",
            decimal_odds=1.909091,
            entry_probability=0.40,
            binary_contract_settlement_value=None,
        )
        == 0.0
    )


def test_prediction_market_settlement_prefers_exact_recorded_entry_price(tmp_path) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    req = request("exact-contract-entry")
    req = PickRequest(
        **{
            **req.__dict__,
            "sportsbook": "polymarket_us",
            "market_probability_at_decision": 0.40,
        }
    )
    row = ledger.append_call(req, 1.0, 60, now=NOW)

    settled = ledger.settle(row["pick_id"], 2, 3)

    assert settled["decision_raw_implied_probability"] == "0.523810"
    assert float(settled["market_probability_at_decision"]) == pytest.approx(0.40)
    assert float(settled["pnl_units"]) == pytest.approx(1.5)


def test_report_filters_are_version_aware(tmp_path) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    ledger.append_call(request("one"), 0.25, 60, now=NOW)
    second = request("two")
    second = PickRequest(**{**second.__dict__, "model_version": "stat-v2", "calibration_version": "cal-v2"})
    ledger.append_call(second, 0.25, 60, now=NOW)
    assert ledger.report({"model_version": "stat-v1"})["records"] == 1
    assert ledger.report({"calibration_version": "cal-v2"})["records"] == 1
    assert ledger.report({"record_type": "qualified"})["records"] == 2


def test_report_odds_range_includes_brier_flat_pnl_and_clv(tmp_path) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    first = ledger.append_call(request("bucket-win"), 0.25, 60, now=NOW)
    second_request = request("bucket-loss", probability=0.45)
    second_request = PickRequest(**{**second_request.__dict__, "american_odds": 150})
    second = ledger.append_call(second_request, 0.25, 60, now=NOW)
    ledger.settle(first["pick_id"], 2, 3, closing_american_odds=-105)
    ledger.settle(second["pick_id"], 3, 2, closing_american_odds=140)

    report = ledger.report(by_odds_range=True)
    populated = [bucket for bucket in report["by_odds_range"] if bucket["count"]]

    assert len(populated) == 2
    assert all(bucket["model_brier"] is not None for bucket in populated)
    assert all(bucket["market_brier"] is not None for bucket in populated)
    assert all(bucket["mean_clv"] is not None for bucket in populated)
    assert sum(bucket["flat_one_unit_pnl"] for bucket in populated) != 0
    assert report["by_market_evaluation"]["moneyline"]["count"] == 2


def test_research_scoring_is_separate_from_qualified_units(tmp_path) -> None:
    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    research = request("research-score", state=ModelState.RESEARCH)
    row = ledger.append_call(research, 1.0, 60, call_type="forced_call", now=NOW)
    ledger.settle(row["pick_id"], 2, 3)
    scored = ledger.score_research([row["pick_id"]], 1.0)[0]
    assert scored["units"] == "1.00"
    assert scored["research_score_units"] == "1.0000"
    assert float(scored["research_pnl_units"]) > 0
    report = ledger.report()
    assert report["qualified_pnl_units"] == 0
    assert report["research_staked_units"] == 1.0
    assert report["research_pnl_units"] > 0


def test_audit_append_happens_while_the_ledger_lock_is_still_held(
    monkeypatch, registry, ban_list, tmp_path
) -> None:
    """A ledger write and its audit event must commit in one held-lock
    critical section, not as two separately-lockable steps -- otherwise a
    crash between them leaves a row with no matching audit event."""
    from model_prediction.audit import AuditLog

    lock_depth = 0
    audit_calls_while_locked: list[bool] = []

    real_flock = fcntl.flock

    def tracking_flock(fd, cmd):
        nonlocal lock_depth
        real_flock(fd, cmd)
        if cmd & fcntl.LOCK_EX:
            lock_depth += 1
        elif cmd == fcntl.LOCK_UN:
            lock_depth -= 1

    real_append = AuditLog.append

    def tracking_append(self, *args, **kwargs):
        audit_calls_while_locked.append(lock_depth > 0)
        return real_append(self, *args, **kwargs)

    monkeypatch.setattr(fcntl, "flock", tracking_flock)
    monkeypatch.setattr(AuditLog, "append", tracking_append)

    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    req = request("lock-order")
    eligibility = evaluate_eligibility(req, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    row = ledger.append_evaluated(req, eligibility, NOW)
    ledger.settle(row["pick_id"], 2, 3)

    req2 = request("lock-order-void")
    eligibility2 = evaluate_eligibility(req2, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    row2 = ledger.append_evaluated(req2, eligibility2, NOW)
    ledger.void(row2["pick_id"], "test void")

    req3 = request("lock-order-remove")
    eligibility3 = evaluate_eligibility(req3, registry, ban_list, Exposure(), UnitPolicy(), NOW)
    ledger.append_evaluated(req3, eligibility3, NOW)
    ledger.remove_open_rows(["lock-order-remove"], "test removal")

    assert audit_calls_while_locked, "expected at least one audit.append call to be observed"
    assert all(audit_calls_while_locked), "every audit.append must fire while the ledger lock is still held"


def test_ledger_write_crash_leaves_a_recoverable_audit_event_not_a_silent_gap(
    monkeypatch, registry, ban_list, tmp_path
) -> None:
    """The 2026-07-28 atomicity fix: audit.append happens BEFORE
    self._write_rows, not after (see ledger.py's module docstring). This
    proves the actual failure mode changed. Simulate the ledger file write
    crashing (a corrupt disk, a killed process mid-write) immediately after
    the audit event committed: the audit chain must show the attempted
    mutation even though the row never landed.

    pick_id is a random uuid4 (ledger.py:458), not derived from the request,
    so a retry after a crash does NOT reuse the crashed attempt's pick_id --
    it's a new row under a new id, not a literal replay. The old, real
    improvement this proves: the crashed attempt is now honestly visible in
    the audit chain (a real, if orphaned, record that something was tried
    and didn't land) and the ledger itself is left in a consistent state a
    fresh append can build on -- not the old failure mode, where a crash in
    the same spot occurred AFTER a successful write, leaving a real ledger
    row with no audit event at all and no way to ever discover the gap.
    """
    from model_prediction.xlsx_ledger import write_xlsx_rows_atomic

    ledger = PickLedger(tmp_path / "picks.xlsx", tmp_path / "events.jsonl")
    ledger.initialize()  # creates the empty file for real, before the crash is armed
    req = request("crash-recovery")
    eligibility = evaluate_eligibility(req, registry, ban_list, Exposure(), UnitPolicy(), NOW)

    def crashing_write(*args, **kwargs):
        raise OSError("simulated crash mid-write")

    monkeypatch.setattr("model_prediction.ledger.write_xlsx_rows_atomic", crashing_write)
    try:
        ledger.append_evaluated(req, eligibility, NOW)
        raise AssertionError("expected the simulated write crash to propagate")
    except OSError:
        pass

    # Audit event committed despite the ledger write never landing.
    events = ledger.audit.events()
    crash_events = [
        event for event in events if event["event_type"] in ("pick_created", "research_observation_created")
    ]
    assert len(crash_events) == 1, "the crashed attempt's audit event must still exist"
    # The row genuinely isn't in the ledger -- confirms this is a detectable
    # gap (audit says something happened, ledger disagrees), not silently
    # invisible the way an unaudited ledger mutation would have been.
    assert read_xlsx_rows(ledger.path)[1] == []

    monkeypatch.setattr("model_prediction.ledger.write_xlsx_rows_atomic", write_xlsx_rows_atomic)
    row = ledger.append_evaluated(req, eligibility, NOW)
    assert any(r["pick_id"] == row["pick_id"] for r in read_xlsx_rows(ledger.path)[1])
    # The retry's pick_id differs from the crashed attempt's -- both audit
    # events are real and legitimately distinct, not a duplicate.
    assert row["pick_id"] != crash_events[0]["subject_id"]

    # "Detectable" (the docstring above and ledger.py's own comment) is only
    # a real property if the operator-facing tool that's supposed to catch
    # this actually does -- confirm _verify_chain flags the orphaned event,
    # not just that a human diffing raw audit/ledger data by hand could.
    from model_prediction.cli import _verify_chain

    result = _verify_chain(ledger.audit.path, ledger)
    assert result["reconciled"] is False
    assert result["created_but_absent_without_removal_event"] == 1


def test_excel_ledger_has_office_table_filter_and_frozen_header(tmp_path) -> None:
    path = tmp_path / "picks.xlsx"
    ledger = PickLedger(path, tmp_path / "events.jsonl")
    ledger.append_call(request("excel-format"), 0.25, 60, now=NOW)

    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = workbook["Picks"]
        summary = workbook["Summary"]
        assert sheet.freeze_panes == "A2"
        assert "PicksLedger" in sheet.tables
        assert sheet.tables["PicksLedger"].autoFilter.ref
        assert sheet.auto_filter.ref is None
        assert sheet["A1"].font.bold
        assert sheet["A1"].fill.fgColor.rgb.endswith("1F4E78")
        assert summary["E7"].value.startswith("=IFERROR(")
        assert summary["G7"].number_format == "0.0%"
    finally:
        workbook.close()
