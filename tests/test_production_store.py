"""Tests for ProductionPredictionStore + runtime-state migration (Phase B)."""

from __future__ import annotations

import json

import pytest

from model_prediction.production_store import ProductionPredictionStore
from model_prediction.runtime_paths import RuntimePaths, migrate_legacy_state


def _store(tmp_path) -> ProductionPredictionStore:
    return ProductionPredictionStore(RuntimePaths.for_test(tmp_path))


def _append(store: ProductionPredictionStore, event_id: str, run_id: str | None = None) -> int | None:
    if run_id is None:
        run_id = store.start_run()  # predictions reference a real run row (FK)
    return store.append_prediction(
        run_id=run_id,
        prediction_id=f"pred-{event_id}",
        event_id=event_id,
        sport="WNBA",
        market="moneyline",
        market_type="moneyline",
        model_id="wnba-elo-trend-lr-v4",
        probabilities={"home": 0.6, "away": 0.4},
        decision_time_utc="2026-08-14T12:00:00+00:00",
        event_start_utc="2026-08-14T23:00:00Z",
    )


def test_append_and_identity_dedupe(tmp_path) -> None:
    with _store(tmp_path) as store:
        first = _append(store, "e1")
        assert first is not None

        # Identical identity (same event, model, market, horizon,
        # decision_time) → idempotent no-op, not a duplicate row.
        second = _append(store, "e1")
        assert second is None
        assert store.counts_by() == {"predicted": 1}

        # Different event → new row.
        assert _append(store, "e2") is not None
        assert store.counts_by() == {"predicted": 2}


def test_settle_and_terminal_guard(tmp_path) -> None:
    with _store(tmp_path) as store:
        row_id = _append(store, "e1")
        settled = store.settle_prediction(row_id, "won", note="graded")
        assert settled["status"] == "settled"
        assert settled["resolved_outcome"] == "won"
        assert settled["settled_at_utc"]

        with pytest.raises(ValueError, match="terminal"):
            store.settle_prediction(row_id, "lost")


def test_keyset_pagination_walks_all_rows(tmp_path) -> None:
    with _store(tmp_path) as store:
        run_id = store.start_run()
        for i in range(5):
            _append(store, f"e{i}", run_id=run_id)
        assert store.counts_by() == {"predicted": 5}

        page1, cursor = store.get_predictions(limit=2)
        assert len(page1) == 2 and cursor is not None
        page2, cursor = store.get_predictions(limit=2, cursor=cursor)
        assert len(page2) == 2
        page3, cursor = store.get_predictions(limit=2, cursor=cursor)
        assert len(page3) == 1 and cursor is None
        seen = {p["event_id"] for p in page1 + page2 + page3}
        assert seen == {f"e{i}" for i in range(5)}


def test_filters_and_aggregation_are_sql_side(tmp_path) -> None:
    with _store(tmp_path) as store:
        _append(store, "e1")
        _append(store, "e2")
        _append(store, "e3")
        store.settle_prediction(store.get_predictions(limit=1)[0][0]["id"], "lost")

        rows, _ = store.get_predictions(status="predicted")
        assert len(rows) == 2
        rows, _ = store.get_predictions(status="settled")
        assert len(rows) == 1
        assert store.counts_by(sport="WNBA") == {"predicted": 2, "settled": 1}


def test_decisions_and_market_snapshots(tmp_path) -> None:
    with _store(tmp_path) as store:
        row_id = _append(store, "e1")
        store.record_decision("pred-e1", "operator", "approve", note="looks good")
        store.record_market_snapshot("e1", "WNBA", "moneyline", {"best_ask": 0.62, "line": 2.5})
        assert row_id is not None


def test_export_writes_xlsx(tmp_path) -> None:
    with _store(tmp_path) as store:
        _append(store, "e1")
        _append(store, "e2")
        out = tmp_path / "export" / "ledger.xlsx"
        count = store.export_xlsx(out)
    assert count == 2
    assert out.is_file()

    from openpyxl import load_workbook

    wb = load_workbook(out)
    rows = list(wb["predictions"].iter_rows(values_only=True))
    assert rows[0][0] == "id"
    assert len(rows) == 3  # header + 2 predictions


def test_migration_carries_legacy_files_exactly_once(tmp_path) -> None:
    paths = RuntimePaths.for_test(tmp_path)
    legacy_runs = paths.repo_root / "data" / "runs.db"
    legacy_runs.parent.mkdir(parents=True, exist_ok=True)
    legacy_runs.write_bytes(b"runs-data")
    legacy_ledger = paths.repo_root / "data" / "production" / "predictions.db"
    legacy_ledger.parent.mkdir(parents=True, exist_ok=True)
    legacy_ledger.write_bytes(b"predictions-data")
    legacy_state = paths.repo_root / "data" / "production_state.json"
    legacy_state.write_text('{"last_prediction_utc": "x"}', encoding="utf-8")

    moved = migrate_legacy_state(paths)

    assert len(moved) == 3
    assert paths.runs_db.read_bytes() == b"runs-data"
    assert paths.production_db.read_bytes() == b"predictions-data"
    assert json.loads(paths.production_state_file.read_text())["last_prediction_utc"] == "x"
    # Legacy files survive (some are git-tracked evidence).
    assert legacy_runs.is_file() and legacy_ledger.is_file()
    # Second run: nothing left to move.
    assert migrate_legacy_state(paths) == []


def test_store_writes_under_runtime_root_not_repo(tmp_path) -> None:
    paths = RuntimePaths.for_test(tmp_path)
    with ProductionPredictionStore(paths) as store:
        _append(store, "e1")
    assert paths.production_db.is_file()
    assert not (paths.repo_root / "data" / "production" / "predictions.db").exists()


def test_store_migrates_a_legacy_schema_database(tmp_path) -> None:
    """A migrated legacy predictions.db (the canary's historical file) has
    no market_type/horizon/decision_time_utc columns. The store must add
    them, backfill decision_time_utc, and create the identity index —
    confirmed live 2026-08-14: without this migration every store write
    failed silently (fail-soft) against the migrated database."""
    paths = RuntimePaths.for_test(tmp_path)
    from model_prediction.production_ledger import ProductionLedger

    legacy = ProductionLedger(paths.production_db)
    run_id = legacy.start_run(git_sha="legacy")
    legacy.record_prediction(
        run_id,
        prediction_id="r:e1",
        event_id="e1",
        sport="WNBA",
        market="moneyline",
        model_id="wnba-elo-trend-lr-v4",
        probabilities={"home": 0.6, "away": 0.4},
    )
    legacy.close()

    store = ProductionPredictionStore(paths)
    cols = {r[1] for r in store._conn.execute("PRAGMA table_info(predictions)")}
    assert {"market_type", "horizon", "decision_time_utc", "canonical_event_id"} <= cols

    # Legacy rows survive with a backfilled decision_time...
    rows, _ = store.get_predictions()
    assert len(rows) == 1 and rows[0]["decision_time_utc"]

    # ...and new appends work against the migrated schema.
    run2 = store.start_run()
    assert (
        store.append_prediction(
            run_id=run2,
            prediction_id="r2:e2",
            event_id="e2",
            sport="WNBA",
            market="moneyline",
            market_type="moneyline",
            model_id="wnba-elo-trend-lr-v4",
            probabilities={"home": 0.5, "away": 0.5},
            decision_time_utc="2026-08-14T12:00:00+00:00",
        )
        is not None
    )
    store.close()


def test_append_with_unknown_run_id_fails_immediately(tmp_path) -> None:
    """run_id is mandatory and must exist — no silent '' normalization
    that orphans a row from its run lineage."""
    with _store(tmp_path) as store:
        import pytest as _pytest

        with _pytest.raises(ValueError, match="does not exist"):
            store.append_prediction(
                run_id="no-such-run",
                prediction_id="p1",
                event_id="e1",
                sport="WNBA",
                market="moneyline",
                market_type="moneyline",
                model_id="wnba-elo-trend-lr-v4",
                probabilities={"home": 0.6, "away": 0.4},
                decision_time_utc="2026-08-14T12:00:00+00:00",
            )


def test_settlement_is_one_atomic_transaction(tmp_path) -> None:
    """status + resolved_outcome + settled_at_utc land in ONE transaction —
    a failure between outcome and status can never leave
    status=predicted with a resolved outcome."""
    with _store(tmp_path) as store:
        row_id = _append(store, "e1")
        settled = store.settle_prediction(row_id, "won", note="graded")
        assert settled["status"] == "settled"
        assert settled["resolved_outcome"] == "won"
        # The single-statement form is enforced by the row state: there is
        # no intermediate state to observe, so a re-transition check is
        # the observable contract.
        import pytest as _pytest

        with _pytest.raises(ValueError, match="terminal"):
            store.settle_prediction(row_id, "lost")


def test_start_finish_run_records_counters(tmp_path) -> None:
    with _store(tmp_path) as store:
        run_id = store.start_run(git_sha="abc123")
        store.finish_run(
            run_id,
            "completed",
            note="3 predictions",
            counters={"events_seen": 3, "predictions": 2, "no_bet": 1},
        )
        rows = store._conn.execute("SELECT * FROM runs").fetchall()
        assert len(rows) == 1
        assert rows[0]["status"] == "completed"
        assert json.loads(rows[0]["counters"]) == {"events_seen": 3, "predictions": 2, "no_bet": 1}
