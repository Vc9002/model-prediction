#!/usr/bin/env python3
# TODO(DD-5): ~5,100-line monolithic server with manual if/elif routing
# (~20 GET + ~15 POST routes) — see MASTER.md DD-5. Split into
# dashboard/routes.py + views.py + orders.py before the next feature.
"""Local operations dashboard server for the model-prediction system.

HACK(DD-5): 5,121-line monolithic file with ~35 routes dispatched via manual
if/elif path chains. Should be split into dashboard/routes.py, views.py,
orders.py, etc. Every new feature lands here. See MASTER.md §DD-5.

Serves dashboard.html plus a small JSON API computed from the project's data
files. View clearing and order status use local dashboard state. Real order
submission remains behind the model-prediction CLI's hard gate and a separate,
exact-ticket confirmation in the UI.

Run:  python3 dashboard_server.py  [--port 8765]
Then open http://127.0.0.1:8765/
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import secrets
import subprocess
import sys
import threading
import time
import urllib.request
from datetime import UTC, datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse
from zoneinfo import ZoneInfo

import yaml

from model_prediction.dashboard.data_service import handle as data_service_handle

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

# ── SECTION: Paths & Constants ───────────────────────────────────────────

ROOT = Path(__file__).resolve().parent
DASHBOARD_PORT = 8765

# Load .env file so subprocess CLI commands inherit Polymarket keys
_ENV_PATH = ROOT / ".env"
if _ENV_PATH.exists():
    with _ENV_PATH.open(encoding="utf-8") as _fh:
        for _line in _fh:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _key, _, _val = _line.partition("=")
                if _key.strip() and _key.strip() not in os.environ:
                    os.environ[_key.strip()] = _val.strip()

DATA = ROOT / "data"
OUTPUTS = ROOT / "outputs" / "latest"
# Main/Flat split from one shared file per tier into one file per sport,
# 2026-08-03 (see model_prediction.main_ledgers for the real module this
# mirrors -- duplicated here as a plain tuple rather than importing that
# module at load time, matching this file's existing pattern of keeping
# model_prediction imports lazy/local to the functions that need them).
_MAIN_LEDGER_SPORTS = ("mlb", "wnba", "soccer", "tennis")
DASH_DIR = ROOT / "dashboard"
PID_FILE = DASH_DIR / "server.pid"
LOG_FILE = DASH_DIR / "server.log"
JOBS_FILE = DASH_DIR / "jobs.json"
ARCHIVE_FILE = DASH_DIR / "archive.json"
ORDERS_FILE = DASH_DIR / "orders.json"
PORTFOLIO_HISTORY_FILE = DASH_DIR / "portfolio_history.json"
CONFIG_FILE = ROOT / "config" / "model.yaml"
FEATURE_REGISTRY_FILE = ROOT / "config" / "tested_features.json"


def _log(message: str) -> None:
    try:
        DASH_DIR.mkdir(exist_ok=True)
        stamp = datetime.now(timezone.utc).isoformat()[:19]
        with LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write(f"{stamp} {message}\n")
    except OSError:
        pass


# EASTERN and SPORTS defined here for self-contained startup, but the canonical
# source of truth is src/model_prediction/domain.py (League enum, EASTERN zone).
# Keep these in sync with domain.py.
EASTERN = ZoneInfo("America/New_York")
SPORTS = (
    "mlb",
    "nba",
    "wnba",
    "nfl",
    "soccer",
    "tennis",
    "lol",
    "cs2",
    "dota2",
    "valorant",
    "rainbow_six",
    "kbo",
    "npb",
)
GATEWAY = "https://gateway.polymarket.us"

_CACHE: dict[str, tuple[float, object]] = {}
_CACHE_LOCK = threading.Lock()
_CONFIG_LOCK = threading.Lock()
_ACTION_LOCK = threading.Lock()
_LAST_ACTION: dict[str, object] = {}
_JOBS: dict[str, dict] = {}
_JOBS_LOCK = threading.Lock()


def _hydrate_jobs() -> None:
    """Load persisted job history into memory at server start.

    _persist_jobs() serializes ONLY the in-memory _JOBS dict, so without
    this the first persist after a restart would overwrite jobs.json with
    just the post-restart jobs and wipe all history. A restart also means
    any job persisted as 'running' was interrupted (its thread is gone)
    and monotonic timestamps are meaningless across processes — normalize
    both rather than restore them as if they were still valid.
    """
    for job in _load_persisted_jobs().values():
        if job.get("status") == "running":
            job["status"] = "interrupted"
            job["error"] = (
                "the dashboard server restarted while this job was running; "
                "the underlying CLI run may still have completed — check the ledger/summary"
            )
        job.pop("started_monotonic", None)
        with _JOBS_LOCK:
            _JOBS[job.get("job_id", "")] = job
_RUNNER: list[str] | None = None
_ORDER_PREVIEWS: dict[str, dict] = {}
_ORDER_LOCK = threading.Lock()
_MARKET_QUESTION_CACHE: dict[str, str | None] = {}
_MARKET_QUESTION_LOCK = threading.Lock()

# Real gap fixed 2026-08-02: this dashboard has real order-execution
# capability (POST /api/order/submit shells out to `execute --execute`) but
# previously had no authentication at all -- only an Origin/Host CSRF check
# (blocks a malicious webpage tricking the browser) and a client-supplied
# confirm:true flag (not a credential; any caller can set it). Neither stops
# a different local process or user account on the same machine from
# curling the API directly and placing a real order. Generated fresh per
# process start (a restart naturally invalidates any leaked/old token); the
# served dashboard.html gets it injected server-side (see do_GET) and
# attaches it to every POST automatically, so the browser UI keeps working
# with no manual step -- only a caller who never loaded the real page (or
# is on a different machine, blocked by the 127.0.0.1 bind regardless)
# lacks it.
_DASHBOARD_TOKEN = secrets.token_urlsafe(32)


def _inject_dashboard_token(html: bytes) -> bytes:
    """Embed the session token and a fetch wrapper that auto-attaches it to
    every POST, so the served UI keeps working with no manual step. Falls
    back to serving the page unmodified (POSTs then need the token supplied
    some other way) if the expected <script> opening isn't found, rather
    than raising and breaking the whole page load over a missing feature."""
    marker = b'<script>\n"use strict";'
    injected = (
        b'<script>\n"use strict";\nwindow.__DASH_TOKEN__=' + json.dumps(_DASHBOARD_TOKEN).encode()
        + b";\nconst __nativeFetch__=window.fetch.bind(window);"
        b"\nwindow.fetch=(input,init)=>{"
        b'\n  if(init&&init.method==="POST"){'
        b"\n    init={...init,headers:{...(init.headers||{}),'X-Dashboard-Token':window.__DASH_TOKEN__}};"
        b"\n  }"
        b"\n  return __nativeFetch__(input,init);"
        b"\n};"
    )
    if marker not in html:
        return html
    return html.replace(marker, injected, 1)


def _resolve_runner() -> list[str]:
    """Find a Python that can actually import model_prediction.

    The dashboard is often started with the system python3, while the package
    lives in the project's .venv — so try the venv first, then whatever is
    running this server, then a `model-prediction` executable on PATH.
    """
    global _RUNNER
    if _RUNNER is not None:
        return _RUNNER
    candidates = [
        str(ROOT / ".venv" / "bin" / "python"),
        str(ROOT / ".venv" / "Scripts" / "python.exe"),  # windows
        sys.executable,
    ]
    for python in candidates:
        if not Path(python).exists():
            continue
        try:
            probe = subprocess.run(
                [python, "-c", "import model_prediction"],
                cwd=ROOT,
                capture_output=True,
                text=True,
                timeout=30,
                env=_runner_env(),
            )
        except (OSError, subprocess.TimeoutExpired):
            continue  # e.g. a venv built on another OS/architecture
        if probe.returncode == 0:
            _RUNNER = [python, "-m", "model_prediction.cli"]
            return _RUNNER
    import shutil

    binary = shutil.which("model-prediction")
    if binary:
        _RUNNER = [binary]
        return _RUNNER
    raise RuntimeError(
        "no Python environment with model_prediction installed was found. "
        "Create it with: python3 -m venv .venv && .venv/bin/pip install -e '.[dev]'"
    )


def _runner_env() -> dict[str, str]:
    import os

    env = dict(os.environ)
    src = str(ROOT / "src")
    env["PYTHONPATH"] = src + (":" + env["PYTHONPATH"] if env.get("PYTHONPATH") else "")
    return env


# ── SECTION: Cache Layer ────────────────────────────────────────────


def _cached(key: str, ttl: float, builder):
    now = time.time()
    with _CACHE_LOCK:
        hit = _CACHE.get(key)
        if hit and now - hit[0] < ttl:
            return hit[1]
    value = builder()
    with _CACHE_LOCK:
        _CACHE[key] = (time.time(), value)
    return value


# ---------------------------------------------------------------------------
# Readers (all read-only)
# ---------------------------------------------------------------------------


def _count_lines(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open("rb") as handle:
        return sum(1 for _ in handle)


def _read_json(path: Path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


_runtime_paths_cache: list = []


def _runtime_paths():
    """Resolve RuntimePaths once and cache it -- never per-request."""
    if not _runtime_paths_cache:
        from model_prediction.runtime_paths import RuntimePaths

        _runtime_paths_cache.append(
            RuntimePaths.resolve(repo_root=ROOT, require_external_runtime=True)
        )
    return _runtime_paths_cache[0]


def rebuild_view(name: str) -> dict:
    """Return one isolated rebuild projection without importing writer classes."""
    from dashboard.rebuild_status import read_rebuild_view

    try:
        paths = _runtime_paths()
    except RuntimeError:
        # Env-less contexts (CI smoke, dev without the launchd env) get
        # the read-only view's own repo-local fallback. The rebuild view
        # never writes, so there is no split-brain risk here.
        paths = None
    return read_rebuild_view(name, ROOT, paths=paths)


def _unit_value_usd() -> float:
    try:
        payload = yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8")) or {}
        value = float((payload.get("bankroll") or {}).get("unit_value_usd", 7.5))
    except (OSError, TypeError, ValueError, yaml.YAMLError):
        return 7.5
    return value if value > 0 else 7.5


def _set_unit_value_usd(raw_value: object) -> dict:
    """Atomically persist the one-unit dollar value used by UI and execution."""
    try:
        value = float(raw_value)
    except (TypeError, ValueError) as error:
        raise ValueError("1U must be a dollar amount") from error
    if not math.isfinite(value) or not 0.01 <= value <= 100_000:
        raise ValueError("1U must be between $0.01 and $100,000.00")

    with _CONFIG_LOCK:
        original = CONFIG_FILE.read_text(encoding="utf-8")
        previous = _unit_value_usd()
        updated, count = re.subn(
            r"(?m)^(\s+unit_value_usd:\s*).*$",
            rf"\g<1>{value:.2f}",
            original,
        )
        if count != 1:
            raise RuntimeError("expected exactly one bankroll.unit_value_usd setting")
        parsed = yaml.safe_load(updated) or {}
        persisted = float((parsed.get("bankroll") or {}).get("unit_value_usd"))
        if not math.isclose(persisted, value, rel_tol=0, abs_tol=0.000001):
            raise RuntimeError("unit value failed configuration validation")

        temporary = CONFIG_FILE.with_name(f".{CONFIG_FILE.name}.{os.getpid()}.{secrets.token_hex(4)}.tmp")
        try:
            with temporary.open("w", encoding="utf-8") as handle:
                handle.write(updated)
                handle.flush()
                os.fsync(handle.fileno())
            os.chmod(temporary, CONFIG_FILE.stat().st_mode)
            os.replace(temporary, CONFIG_FILE)
        finally:
            temporary.unlink(missing_ok=True)

    try:
        from model_prediction.audit import AuditLog

        AuditLog(DATA / "events.jsonl").append(
            "unit_value_updated",
            "bankroll.unit_value_usd",
            {"previous_usd": previous, "unit_value_usd": value, "source": "dashboard"},
        )
    except (ImportError, OSError, ValueError) as error:
        _log(f"unit value updated but audit append failed: {error}")
    with _CACHE_LOCK:
        _CACHE.clear()
    return {
        "status": "ok",
        "previous_unit_value_usd": previous,
        "unit_value_usd": value,
        "note": "Applies to future dollar displays and order sizing; historical units are unchanged.",
    }


def _config_payload() -> dict:
    try:
        payload = yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8")) or {}
    except (OSError, yaml.YAMLError):
        return {}
    # ── SECTION: Configuration ──────────────────────────────────────────
    return payload if isinstance(payload, dict) else {}


def _row_has_banned_team(row: dict) -> bool:
    config = _config_payload()
    configured = (config.get("team_ban_list") or {}).get("teams") or {}
    banned_ids = {
        str(item.get("canonical_team_id") or "")
        for values in configured.values()
        for item in (values or [])
        if isinstance(item, dict)
    }
    registry = _read_json(DATA / "entities" / "teams.json") or {}
    entries = registry.get("teams") if isinstance(registry, dict) else registry
    banned_names: set[str] = set()
    for entry in entries or []:
        if str(entry.get("canonical_team_id") or "") not in banned_ids:
            continue
        banned_names.add(str(entry.get("canonical_name") or "").strip().casefold())
        for alias in entry.get("aliases") or []:
            if isinstance(alias, dict):
                banned_names.add(str(alias.get("source_name") or "").strip().casefold())
            else:
                banned_names.add(str(alias or "").strip().casefold())
    return any(
        str(row.get(key) or "").strip().casefold() in banned_names for key in ("away_team", "home_team")
    )


def _manual_research_eligibility(row: dict) -> tuple[bool, str]:
    if row.get("record_type") != "RESEARCH_OBSERVATION":
        return False, "not a research observation"
    config = _config_payload()
    execution = config.get("execution") or {}
    if not execution.get("allow_manual_research_orders", False):
        return False, "manual research orders are disabled"
    league = str(row.get("league") or "").upper()
    active_version = (config.get("models", {}).get(league, {}) or {}).get("active_production_version")
    if execution.get("manual_research_require_active_model", True):
        if not active_version or row.get("model_version") != active_version:
            return False, "pick is not from the active production model"
    edge = _number(row.get("model_probability")) - _number(row.get("market_implied_probability"))
    if execution.get("manual_research_require_positive_edge", True) and edge <= 0:
        return False, "model has no positive edge at the logged market price"
    if _row_has_banned_team(row):
        return False, "a team in this matchup is permanently banned"
    return True, "manual active-model order"


_PICKS_CACHE: dict[str, object] = {"mtime": None, "rows": []}
_FLAT_PICKS_CACHE: dict[str, object] = {"mtime": None, "rows": []}
# SQLite-backed cache: mirrors Excel ledger data for ~100x faster dashboard reads.
# Falls back to Excel parsing if the SQLite cache hasn't been built yet.
_DASHBOARD_CACHE: object | None = None
# ── SECTION: Picks & Performance ────────────────────────────────────


def _main_ledger_paths() -> list[Path]:
    return [DATA / "main" / f"{sport}.xlsx" for sport in _MAIN_LEDGER_SPORTS]


def _flat_ledger_paths() -> list[Path]:
    return [DATA / "flat" / f"{sport}.xlsx" for sport in _MAIN_LEDGER_SPORTS]


def _read_split_picks(paths: list[Path], cache: dict[str, object]) -> list[dict]:
    """Aggregate rows across one tier's per-sport files (Main or Flat),
    keyed on the combined mtime of every file that actually exists so any
    single sport's file changing invalidates the cache -- same "only
    re-parse on change" principle read_picks() always used, just extended
    to N files instead of one.

    When the SQLite dashboard cache is available, reads from it instead of
    parsing Excel files (~100x faster).  Falls back to Excel parsing if the
    cache hasn't been built yet.
    """
    global _DASHBOARD_CACHE
    if _DASHBOARD_CACHE is None:
        try:
            from model_prediction.dashboard_cache import get_cache as _get_dc
            from model_prediction.runtime_paths import RuntimePaths

            # The cache DB is mutable runtime state (runtime root); the
            # Excel sources it mirrors stay in the repo checkout.
            _DASHBOARD_CACHE = _get_dc(
                DATA, db_path=RuntimePaths.resolve(repo_root=ROOT).runtime_root / "dashboard_cache.db"
            )
        except Exception:
            pass

    # Try SQLite cache first
    if _DASHBOARD_CACHE is not None and paths:
        first_path = str(paths[0])
        if "/flat/" in first_path or "\\flat\\" in first_path:
            tier = "flat"
        elif "/research/" in first_path or "\\research\\" in first_path:
            tier = "research"
        elif "/gated_research/" in first_path:
            tier = "gated_research"
        else:
            tier = "main"

        try:
            dc = _DASHBOARD_CACHE
            dc.refresh()  # no-op if mtimes unchanged, fast SQLite otherwise
            if tier in ("flat", "main", "research", "gated_research"):
                rows = dc.read_picks(tier)
            else:
                rows = dc.read_picks(tier)
            if rows:
                return rows
        except Exception:
            pass

    # Fallback to Excel parsing (original path)
    if load_workbook is None:
        return []
    existing = [path for path in paths if path.exists()]
    if not existing:
        return []
    mtime_key = tuple(sorted((str(path), path.stat().st_mtime) for path in existing))
    with _CACHE_LOCK:
        if cache["mtime"] == mtime_key:
            return cache["rows"]  # type: ignore[return-value]
    rows_out = [row for path in existing for row in _parse_picks(path)]
    with _CACHE_LOCK:
        cache["mtime"] = mtime_key
        cache["rows"] = rows_out
    return rows_out


def read_picks() -> list[dict]:
    """Parse every sport's Main ledger (data/main/<sport>.xlsx), only
    re-parsing a file whose mtime changed -- parsing is the single most
    expensive repeated operation the dashboard performs."""
    return _read_split_picks(_main_ledger_paths(), _PICKS_CACHE)


def read_flat_picks() -> list[dict]:
    """Parse every sport's Flat ledger (data/flat/<sport>.xlsx), same
    change-detection caching as read_picks()."""
    return _read_split_picks(_flat_ledger_paths(), _FLAT_PICKS_CACHE)


def _find_pick_by_id(pick_id: str) -> dict | None:
    """Search both main and flat ledgers for a pick by pick_id."""
    row = next((item for item in read_picks() if str(item.get("pick_id")) == pick_id), None)
    if row is not None:
        return row
    return next((item for item in read_flat_picks() if str(item.get("pick_id")) == pick_id), None)


def _parse_picks(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb["Picks"] if "Picks" in wb.sheetnames else wb.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    keep = [
        "pick_id",
        "created_at_utc",
        "event_start_utc",
        "event_id",
        "league",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "american_odds",
        "market_implied_probability",
        "model_probability",
        "model_uncertainty",
        "edge",
        "trade_candidate",
        "confidence_score",
        "units",
        "model_version",
        "status",
        "result",
        "away_score",
        "home_score",
        "probability_clv",
        "pnl_units",
        "settled_at_utc",
        "record_type",
        "decision",
        "reason_code",
        "research_score_units",
        "research_pnl_units",
        "sportsbook",
        "decision_no_vig_probability",
        "rationale",
        "risks",
        "unavailable_features",
        "elo_probability",
        "trend_gap",
        "defensive_trend_gap",
        "park_factor",
        "weather_factor",
        "pitcher_era_gap",
        "probable_starter_era_gap",
        "market_residual_probability",
    ]
    index = {name: headers.index(name) for name in keep if name in headers}
    picks = []
    for values in rows:
        if values is None or all(value is None for value in values):
            continue
        row = {name: values[position] for name, position in index.items()}
        if not row.get("pick_id"):
            continue
        picks.append(row)
    wb.close()
    return picks


def _research_ledger_paths(*, gated: bool = False) -> list[Path]:
    directory = DATA / ("gated_research" if gated else "research")
    paths = sorted(directory.glob("*.xlsx")) if directory.exists() else []
    if paths:
        return paths
    legacy = DATA / ("gated_research.xlsx" if gated else "research.xlsx")
    return [legacy] if legacy.exists() else []


def _parse_research_picks(*, gated: bool = False) -> list[dict]:
    rows: list[dict] = []
    for path in _research_ledger_paths(gated=gated):
        rows.extend(_parse_picks(path))
    return rows


def _number(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _pick_probability(row) -> float | None:
    value = _number(row.get("model_probability"), -1)
    return value if 0 < value < 1 else None


def _pick_pnl(row) -> float:
    """Return recorded unit P&L for a row that actually has a scoring basis."""
    units = _number(row.get("units"))
    if units > 0:
        return _number(row.get("pnl_units"))
    if _number(row.get("research_score_units")) > 0:
        return _number(row.get("research_pnl_units"))
    return 0.0


def _pick_is_scored(row: dict) -> bool:
    return _number(row.get("units")) > 0 or _number(row.get("research_score_units")) > 0


def _performance_game_key(row: dict) -> str:
    """Stable-enough event identity for dashboard-only game coverage metrics."""
    event_id = str(row.get("event_id") or "").strip()
    if event_id:
        return f"{str(row.get('league') or '').casefold()}:{event_id}"
    parts = (
        row.get("league"),
        row.get("event_start_utc"),
        row.get("away_team"),
        row.get("home_team"),
    )
    return "|".join(str(part or "").strip().casefold() for part in parts)


def _performance_breakdown(
    rows: list[dict],
    key_fn,
) -> dict[str, dict]:
    """Build the same result/risk/game metrics for any ledger-derived dimension."""
    groups: dict[str, dict] = {}
    game_keys: dict[str, set[str]] = {}
    for row in rows:
        key = str(key_fn(row) or "unknown").strip() or "unknown"
        won = row.get("result") == "win"
        entry = groups.setdefault(
            key,
            {"wins": 0, "calls": 0, "scored_calls": 0, "pnl": 0.0, "risked": 0.0},
        )
        entry["calls"] += 1
        entry["wins"] += won
        if _pick_is_scored(row):
            entry["scored_calls"] += 1
            entry["pnl"] += _pick_pnl(row)
            entry["risked"] += max(
                _number(row.get("units")),
                _number(row.get("research_score_units")),
            )
        game_keys.setdefault(key, set()).add(_performance_game_key(row))
    for key, entry in groups.items():
        entry["losses"] = entry["calls"] - entry["wins"]
        entry["win_rate"] = round(entry["wins"] / entry["calls"], 4) if entry["calls"] else None
        entry["pnl"] = round(entry["pnl"], 4) if entry["scored_calls"] else None
        entry["risked"] = round(entry["risked"], 4) if entry["scored_calls"] else None
        entry["roi"] = (
            round(entry["pnl"] / entry["risked"], 4)
            if entry["pnl"] is not None and entry["risked"]
            else None
        )
        entry["games"] = len(game_keys.get(key, set()))
    return dict(
        sorted(
            groups.items(),
            key=lambda item: (-item[1]["calls"], item[0].casefold()),
        )
    )


def performance(picks: list[dict]) -> dict:
    settled = [
        row for row in picks if row.get("status") == "settled" and row.get("result") in ("win", "loss")
    ]
    settled.sort(key=lambda row: str(row.get("settled_at_utc") or ""))
    wins = sum(1 for row in settled if row["result"] == "win")
    cumulative, curve = 0.0, []
    for row in settled:
        if not _pick_is_scored(row):
            continue
        cumulative += _pick_pnl(row)
        curve.append(
            {
                "t": str(row.get("settled_at_utc") or "")[:16],
                "pnl": round(cumulative, 4),
            }
        )
    by_sport = _performance_breakdown(settled, lambda row: row.get("league"))
    by_market = _performance_breakdown(settled, lambda row: row.get("market_type"))
    probability_rows = [row for row in settled if _pick_probability(row) is not None]

    def probability_bucket(row: dict) -> str:
        probability = _pick_probability(row)
        if probability is None:
            return "unknown"
        index = min(19, int(probability * 20))
        lower = index / 20
        upper = (index + 1) / 20
        return f"{lower:.2f}-{upper:.2f}"

    by_bucket = dict(
        sorted(_performance_breakdown(probability_rows, probability_bucket).items())
    )
    by_month = {}
    for row in settled:
        won = row["result"] == "win"
        month = str(row.get("settled_at_utc") or "")[:7]
        if month:
            entry = by_month.setdefault(
                month,
                {"pnl": 0.0, "calls": 0, "scored_calls": 0, "wins": 0},
            )
            entry["calls"] += 1
            entry["wins"] += won
            if _pick_is_scored(row):
                entry["scored_calls"] += 1
                entry["pnl"] = round(entry["pnl"] + _pick_pnl(row), 4)
    for entry in by_month.values():
        entry["losses"] = entry["calls"] - entry["wins"]
        entry["win_rate"] = round(entry["wins"] / entry["calls"], 4) if entry["calls"] else None
        if not entry["scored_calls"]:
            entry["pnl"] = None

    sport_market_groups: dict[tuple[str, str], list[dict]] = {}
    for row in settled:
        key = (
            str(row.get("league") or "unknown"),
            str(row.get("market_type") or "unknown"),
        )
        sport_market_groups.setdefault(key, []).append(row)
    by_sport_market = []
    for (sport, market_type), rows in sorted(
        sport_market_groups.items(),
        key=lambda item: (-len(item[1]), item[0][0].casefold(), item[0][1].casefold()),
    ):
        summary = next(iter(_performance_breakdown(rows, lambda _row: "all").values()))
        by_sport_market.append(
            {"sport": sport, "market_type": market_type, **summary}
        )

    market_contexts = []
    for market_type, market_rows in sorted(
        (
            (market_type, [row for row in settled if str(row.get("market_type") or "unknown") == market_type])
            for market_type in by_market
        ),
        key=lambda item: (-len(item[1]), item[0].casefold()),
    ):
        game_ids = {_performance_game_key(row) for row in market_rows}
        lines = [
            _number(row.get("line"), None)
            for row in market_rows
            if row.get("line") not in (None, "")
        ]
        lines = [line for line in lines if line is not None]
        market_contexts.append(
            {
                "market_type": market_type,
                "calls": len(market_rows),
                "games": len(game_ids),
                "sports": sorted(
                    {str(row.get("league") or "unknown") for row in market_rows},
                    key=str.casefold,
                ),
                "selections": _performance_breakdown(
                    market_rows,
                    lambda row: row.get("selection"),
                ),
                "line": {
                    "count": len(lines),
                    "mean": round(sum(lines) / len(lines), 3) if lines else None,
                    "mean_abs": round(sum(abs(line) for line in lines) / len(lines), 3)
                    if lines
                    else None,
                    "min": round(min(lines), 3) if lines else None,
                    "max": round(max(lines), 3) if lines else None,
                },
            }
        )

    sport_contexts = []
    for sport, metrics in by_sport.items():
        sport_rows = [row for row in settled if str(row.get("league") or "unknown") == sport]
        sport_contexts.append(
            {
                "sport": sport,
                **metrics,
                "markets": sorted(
                    {str(row.get("market_type") or "unknown") for row in sport_rows},
                    key=str.casefold,
                ),
            }
        )

    game_market_types: dict[str, set[str]] = {}
    for row in settled:
        game_market_types.setdefault(_performance_game_key(row), set()).add(
            str(row.get("market_type") or "unknown")
        )
    unique_games = len(game_market_types)
    multi_market_games = sum(1 for markets in game_market_types.values() if len(markets) > 1)

    # calibration: 10 buckets predicted vs actual
    calibration = []
    for index in range(10):
        lo, hi = index / 10, (index + 1) / 10
        members = [
            row
            for row in settled
            if (p := _pick_probability(row)) is not None and (lo <= p < hi or (hi == 1.0 and p == 1.0))
        ]
        if members:
            calibration.append(
                {
                    "bucket": f"{lo:.1f}-{hi:.1f}",
                    "mean_p": round(sum(_pick_probability(m) for m in members) / len(members), 4),
                    "hit_rate": round(sum(1 for m in members if m["result"] == "win") / len(members), 4),
                    "count": len(members),
                }
            )
    # streaks
    current = longest_w = longest_l = 0
    run_w = run_l = 0
    for row in settled:
        if row["result"] == "win":
            run_w, run_l = run_w + 1, 0
        else:
            run_l, run_w = run_l + 1, 0
        longest_w, longest_l = max(longest_w, run_w), max(longest_l, run_l)
    current = run_w if run_w else -run_l
    clv = [
        _number(row.get("probability_clv"), None)
        for row in settled
        if row.get("probability_clv") not in (None, "")
    ]
    clv = [value for value in clv if value is not None]
    brier_values = [
        (probability - (1.0 if row["result"] == "win" else 0.0)) ** 2
        for row in settled
        if (probability := _pick_probability(row)) is not None
    ]
    scored = [row for row in settled if _pick_is_scored(row)]
    staked = sum(
        max(_number(row.get("units")), _number(row.get("research_score_units")))
        for row in scored
    )
    total_pnl = round(sum(_pick_pnl(row) for row in scored), 4) if scored else None
    event_dates = sorted(
        str(row.get("event_start_utc") or "")[:10]
        for row in settled
        if row.get("event_start_utc")
    )
    return {
        "total_picks": len(picks),
        "settled": len(settled),
        "open": sum(1 for row in picks if row.get("status") == "open"),
        "wins": wins,
        "losses": len(settled) - wins,
        "scored_calls": len(scored),
        "win_rate": round(wins / len(settled), 4) if settled else None,
        "total_pnl": total_pnl,
        "roi": round(total_pnl / staked, 4) if total_pnl is not None and staked else None,
        "pnl_curve": curve,
        "by_sport": by_sport,
        "by_market": by_market,
        "by_sport_market": by_sport_market,
        "by_confidence": {name: table for name, table in by_bucket.items()},
        "by_month": dict(sorted(by_month.items())),
        "calibration": calibration,
        "streaks": {"longest_win": longest_w, "longest_loss": longest_l, "current": current},
        "mean_clv": round(sum(clv) / len(clv), 6) if clv else None,
        "brier": round(sum(brier_values) / len(brier_values), 6) if brier_values else None,
        "context": {
            "sports": sport_contexts,
            "markets": market_contexts,
            "unique_games": unique_games,
            "multi_market_games": multi_market_games,
            "max_markets_per_game": max(
                (len(markets) for markets in game_market_types.values()),
                default=0,
            ),
            "probability_rows": len(brier_values),
            "clv_rows": len(clv),
            "period_start": event_dates[0] if event_dates else None,
            "period_end": event_dates[-1] if event_dates else None,
        },
    }


def performance_for_sport(picks: list[dict], sport: str | None = None) -> dict:
    """Return one ledger's performance, optionally scoped to an actual ledger sport."""
    available = sorted(
        {
            str(row.get("league") or "").strip()
            for row in picks
            if row.get("status") == "settled"
            and row.get("result") in ("win", "loss")
            and str(row.get("league") or "").strip()
        },
        key=str.casefold,
    )
    requested = str(sport or "").strip()
    selected = next(
        (candidate for candidate in available if candidate.casefold() == requested.casefold()),
        None,
    )
    if selected:
        scoped = [
            row
            for row in picks
            if str(row.get("league") or "").strip().casefold() == selected.casefold()
        ]
    else:
        scoped = [] if requested else picks
    payload = performance(scoped)
    payload["available_sports"] = available
    payload["selected_sport"] = selected or (requested if requested else None)
    return payload


# ── SECTION: Production Evidence ───────────────────────────────────


def _artifact_hash(payload: dict) -> str:
    canonical = {key: value for key, value in payload.items() if key != "artifact_hash"}
    encoded = json.dumps(canonical, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(encoded).hexdigest()


def _artifact_evidence(path: Path, expected_version: str, expected_sport: str) -> tuple[dict, dict]:
    raw = _read_json(path)
    if not isinstance(raw, dict):
        return (
            {
                "path": str(path),
                "available": False,
                "valid": False,
                "health": "MISSING",
                "sha256": None,
                "hash_verified": False,
                "lineage": "UNVERIFIED",
                "declared_hash": None,
                "computed_hash": None,
                "hash_valid": False,
                "artifact_model_version": None,
                "version_matches_config": False,
                "artifact_identity": None,
                "lineage_matches_config": False,
                "mismatches": ["artifact_missing_or_invalid_json"],
            },
            {},
        )

    declared_hash = raw.get("artifact_hash")
    computed_hash = _artifact_hash(raw)
    artifact_version = str(raw.get("model_version") or "")
    artifact_identity = str(raw.get("sport") or raw.get("title") or raw.get("league") or "")
    hash_valid = bool(declared_hash) and declared_hash == computed_hash
    version_valid = bool(expected_version) and artifact_version == expected_version
    lineage_valid = artifact_identity.casefold() == expected_sport.casefold()
    mismatches = []
    if not declared_hash:
        mismatches.append("artifact_hash_missing")
    elif not hash_valid:
        mismatches.append("artifact_hash_mismatch")
    if not version_valid:
        mismatches.append("artifact_model_version_mismatch")
    if not lineage_valid:
        mismatches.append("artifact_sport_or_title_mismatch")
    return (
        {
            "path": str(path),
            "available": True,
            "valid": hash_valid and version_valid and lineage_valid,
            "health": "VERIFIED" if hash_valid and version_valid and lineage_valid else "FAILED",
            "sha256": declared_hash,
            "hash_verified": hash_valid,
            "lineage": "VERIFIED" if version_valid and lineage_valid else "MISMATCH",
            "declared_hash": declared_hash,
            "computed_hash": computed_hash,
            "hash_valid": hash_valid,
            "artifact_model_version": artifact_version or None,
            "version_matches_config": version_valid,
            "artifact_identity": artifact_identity or None,
            "lineage_matches_config": lineage_valid,
            "mismatches": mismatches,
        },
        raw,
    )


def _production_model_spec(raw: dict) -> dict:
    moneyline = (raw.get("market_models") or {}).get("moneyline") or {}
    if moneyline:
        names = list(moneyline.get("feature_names") or [])
        coefficients = list(moneyline.get("coefficients") or [])
        return {
            "kind": "logistic_regression",
            "feature_schema_status": "declared",
            "features": [
                {"name": name, "coefficient": coefficient}
                for name, coefficient in zip(names, coefficients, strict=False)
            ],
            "feature_names": names,
            "coefficients": coefficients,
            "coefficient_count_matches_features": len(names) == len(coefficients),
            "intercept": moneyline.get("intercept"),
            "confidence_threshold": moneyline.get("confidence_threshold"),
            "positive_class": moneyline.get("positive_class"),
        }
    if raw.get("league") and "tie_probability" in raw:
        return {
            "kind": "tie_aware_elo",
            "feature_schema_status": "model_family_declared",
            "features": [{"name": "tie_aware_elo_rating_difference", "coefficient": None}],
            "feature_names": ["tie_aware_elo_rating_difference"],
            "coefficients": [],
            "coefficient_count_matches_features": True,
            "parameters": {
                "initial_rating": raw.get("initial_rating"),
                "k": raw.get("k"),
                "home_advantage": raw.get("home_advantage"),
                "tie_probability": raw.get("tie_probability"),
                "target": raw.get("target"),
            },
        }
    return {
        "kind": "neutral_series_elo",
        "feature_schema_status": "model_family_declared",
        "features": [{"name": "neutral_elo_rating_difference", "coefficient": None}],
        "feature_names": ["neutral_elo_rating_difference"],
        "coefficients": [],
        "coefficient_count_matches_features": True,
        "parameters": {
            "initial_rating": raw.get("initial_rating"),
            "k": raw.get("k"),
            "home_or_order_advantage": raw.get("home_or_order_advantage"),
            "confidence_threshold": raw.get("confidence_threshold"),
            "target": raw.get("target"),
        },
    }


def _rolling_declared_hash(artifact_path: str | None) -> str | None:
    """Declared hash of the ROLLING artifact for a config path, if any.

    K split (2026-08-15): the external validation reports describe the
    rolling artifacts under the runtime root; the frozen config copy is
    the promoted snapshot. When a rolling copy exists, its hash is what
    the report must match.
    """
    if not artifact_path:
        return None
    from model_prediction.runtime_paths import rolling_models_root

    candidate = rolling_models_root() / Path(artifact_path).name
    if not candidate.is_file():
        return None
    try:
        return json.loads(candidate.read_text(encoding="utf-8")).get("artifact_hash")
    except (OSError, json.JSONDecodeError):
        return None


def _locked_backfill_evidence(
    sport: str,
    version: str,
    raw: dict,
    artifact: dict,
    esports_validation: dict,
    international_validation: dict,
) -> dict:
    if not artifact.get("valid"):
        return {
            "status": "rejected_artifact_integrity",
            "source": None,
            "model_version": version,
            "metrics": None,
            "pnl_label": None,
            "profitability_claim": False,
        }
    if raw.get("method") == "logistic_regression":
        metrics = raw.get("qualification") or {}
        if not metrics or metrics.get("locked_holdout") is not True:
            return {
                "status": "rejected_missing_locked_holdout_metrics",
                "source": artifact.get("path"),
                "model_version": version,
                "metrics": None,
                "pnl_label": None,
                "profitability_claim": False,
            }
        return {
            "status": "verified",
            "source": artifact.get("path"),
            "model_version": version,
            "metrics": metrics,
            "pnl_label": "hypothetical_at_minus_110",
            "profitability_claim": False,
        }

    if raw.get("league"):
        league = str(raw.get("league") or sport).casefold()
        report = (international_validation.get("leagues") or {}).get(league) or {}
        locked = report.get("locked_test") or {}
        report_version = str(report.get("model_version") or "")
        report_hash = str(report.get("artifact_hash") or "")
        exact = (
            bool(locked)
            and report_version == version
            and report_hash
            == (_rolling_declared_hash(artifact.get("path")) or artifact.get("declared_hash"))
        )
        if not exact:
            return {
                "status": "rejected_missing_or_mismatched_locked_metrics",
                "source": str(OUTPUTS / "international-baseball-baseline-validation.json"),
                "model_version": report_version or None,
                "artifact_hash": report_hash or None,
                "metrics": None,
                "pnl_label": None,
                "profitability_claim": False,
            }
        return {
            "status": "verified",
            "source": str(OUTPUTS / "international-baseball-baseline-validation.json"),
            "model_version": report_version,
            "artifact_hash": report_hash,
            "metrics": locked,
            "pnl_label": "hypothetical_at_minus_110",
            "profitability_claim": False,
        }

    report = (esports_validation.get("titles") or {}).get(sport.lower()) or {}
    report_version = str(report.get("model_version") or "")
    report_hash = report.get("artifact_hash")
    exact_version = report_version == version
    exact_hash = bool(report_hash) and report_hash == (
        _rolling_declared_hash(artifact.get("path")) or artifact.get("declared_hash")
    )
    locked = report.get("locked_test") or {}
    if not exact_version or not exact_hash or not locked:
        reasons = []
        if not exact_version:
            reasons.append("validation_model_version_mismatch")
        if not exact_hash:
            reasons.append("validation_artifact_hash_mismatch_or_missing")
        if not locked:
            reasons.append("locked_test_metrics_missing")
        return {
            "status": "rejected_external_validation_mismatch",
            "source": str(OUTPUTS / "esports-baseline-validation.json"),
            "model_version": report_version or None,
            "metrics": None,
            "mismatches": reasons,
            "pnl_label": None,
            "profitability_claim": False,
        }
    return {
        "status": "verified",
        "source": str(OUTPUTS / "esports-baseline-validation.json"),
        "model_version": report_version,
        "artifact_hash": report_hash,
        "metrics": locked,
        "pnl_label": "hypothetical_at_minus_110",
        "profitability_claim": False,
    }


def _backfill_aliases(backfill: dict, raw: dict) -> dict:
    """Flatten locked metrics for UI consumers while retaining full detail."""
    metrics = backfill.get("metrics") or {}
    if raw.get("method") == "logistic_regression":
        locked_training = (raw.get("training") or {}).get("locked_holdout") or {}
        obs = metrics.get("total_predictions", locked_training.get("observations"))
        calls = metrics.get("calls")
        hr = metrics.get("hit_rate")
        br = metrics.get("brier_score")
        q = metrics.get("qualified")
        aliases = {
            "observations": obs, "calls": calls, "hit_rate": hr,
            "brier_score": br, "qualified": q,
        }
        units = metrics.get("units_at_minus_110")
        hb = {"observations": obs, "calls": calls, "hit_rate": hr, "brier": br, "units_at_minus_110": units}
        aliases["holdout_backfill"] = {"all_calls": hb, "selection_calls": hb}
    elif raw.get("league"):
        calls = metrics.get("calls")
        hits = metrics.get("hits")
        obs = metrics.get("observations")
        hr = hits / calls if calls and hits is not None else None
        br = metrics.get("brier_settlement")
        q = raw.get("qualified_for_betting")
        aliases = {
            "observations": obs, "calls": calls, "hit_rate": hr,
            "brier_score": br, "qualified": q,
        }
        units = metrics.get("units_at_minus_110")
        hb = {"observations": obs, "calls": calls, "hit_rate": hr, "brier": br, "units_at_minus_110": units}
        aliases["holdout_backfill"] = {"all_calls": hb, "selection_calls": hb}
    else:
        selected = metrics.get("selected_matches") or {}
        all_m = metrics.get("all_matches") or {}
        aliases = {
            "observations": selected.get("observations"),
            "calls": selected.get("calls"),
            "hit_rate": selected.get("accuracy"),
            "brier_score": selected.get("brier"),
            "qualified": raw.get("qualified_for_betting"),
        }
        aliases["holdout_backfill"] = {
            "all_calls": {
                "observations": all_m.get("observations"),
                "calls": all_m.get("calls"),
                "hit_rate": all_m.get("accuracy"),
                "brier": all_m.get("brier"),
                "units_at_minus_110": all_m.get("units_at_minus_110"),
            },
            "selection_calls": {
                "observations": selected.get("observations"),
                "calls": selected.get("calls"),
                "hit_rate": selected.get("accuracy"),
                "brier": selected.get("brier"),
                "units_at_minus_110": selected.get("units_at_minus_110"),
            },
        }
    return {**backfill, **aliases}


def _read_evidence_ledger(path: Path) -> list[dict]:
    if load_workbook is None or not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Picks"] if "Picks" in workbook.sheetnames else workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value) if value is not None else "" for value in next(values)]
    except StopIteration:
        workbook.close()
        return []
    rows = []
    for raw_row in values:
        if not raw_row or all(value is None for value in raw_row):
            continue
        row = {header: raw_row[index] for index, header in enumerate(headers) if header}
        if row.get("pick_id") or row.get("event_id"):
            rows.append(row)
    workbook.close()
    return rows


def _normalized_line(value: object) -> str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value).strip().casefold()
    return f"{number:.12g}"


def _ledger_deduplication_key(row: dict) -> tuple[str, ...]:
    return (
        str(row.get("event_id") or "").strip(),
        str(row.get("league") or "").strip().casefold(),
        str(row.get("market_type") or "").strip().casefold(),
        str(row.get("selection") or "").strip().casefold(),
        _normalized_line(row.get("line")),
        str(row.get("model_version") or "").strip(),
        str(row.get("model_artifact_hash") or "").strip(),
    )


def _deduplicate_ledger_rows(rows: list[dict]) -> tuple[list[dict], int]:
    unique: dict[tuple[str, ...], dict] = {}
    for row in rows:
        unique.setdefault(_ledger_deduplication_key(row), row)
    return list(unique.values()), len(rows) - len(unique)


def _model_owns_row(sport: str, model_config: dict, row: dict) -> bool:
    version = str(row.get("model_version") or "").casefold()
    league = str(row.get("league") or "").casefold()
    configured_leagues = {str(value).casefold() for value in model_config.get("leagues") or []}
    return version.startswith(f"{sport.casefold()}-") or league in {
        sport.casefold(),
        *configured_leagues,
    }


def _feature_attribution(rows: list[dict], feature_names: list[str]) -> dict:
    if not rows:
        return {
            "status": "no_exact_version_rows",
            "rows_with_complete_values": 0,
            "rows_missing_values": 0,
            "missing_by_feature": {name: 0 for name in feature_names},
        }
    if not feature_names:
        return {
            "status": "unavailable_artifact_has_no_explicit_feature_schema",
            "rows_with_complete_values": 0,
            "rows_missing_values": len(rows),
            "missing_by_feature": {},
        }
    missing_by_feature = {name: 0 for name in feature_names}
    complete = 0
    for row in rows:
        feature_values = row.get("feature_values")
        if isinstance(feature_values, str):
            try:
                feature_values = json.loads(feature_values)
            except json.JSONDecodeError:
                feature_values = {}
        feature_values = feature_values if isinstance(feature_values, dict) else {}
        missing = []
        for name in feature_names:
            value = row.get(name, row.get(f"feature_{name}", feature_values.get(name)))
            if value in (None, ""):
                missing.append(name)
                missing_by_feature[name] += 1
        if not missing:
            complete += 1
    return {
        "status": "complete" if complete == len(rows) and rows else "missing",
        "rows_with_complete_values": complete,
        "rows_missing_values": len(rows) - complete,
        "missing_by_feature": missing_by_feature,
    }


def _pnl_evidence(rows: list[dict]) -> dict:
    if not rows:
        empty = {
            "rows": 0,
            "staked_units": None,
            "pnl_units": None,
            "roi": None,
        }
        return {
            "shadow": {"label": "shadow_not_executed", **empty},
            "hypothetical": {"label": "hypothetical_fixed_unit_research", **empty},
            "executed": {
                "label": "executed",
                **empty,
                "status": "not_available_no_execution_attribution_in_ledgers",
            },
        }
    shadow = [row for row in rows if row.get("record_type") == "QUALIFIED_SHADOW_CALL"]
    shadow_staked = sum(_number(row.get("units")) for row in shadow)
    shadow_pnl = sum(_number(row.get("pnl_units")) for row in shadow)
    hypothetical = [row for row in rows if _number(row.get("research_score_units")) > 0]
    hypothetical_staked = sum(_number(row.get("research_score_units")) for row in hypothetical)
    hypothetical_pnl = sum(_number(row.get("research_pnl_units")) for row in hypothetical)
    return {
        "shadow": {
            "label": "shadow_not_executed",
            "rows": len(shadow),
            "staked_units": round(shadow_staked, 6),
            "pnl_units": round(shadow_pnl, 6),
            "roi": round(shadow_pnl / shadow_staked, 6) if shadow_staked else None,
        },
        "hypothetical": {
            "label": "hypothetical_fixed_unit_research",
            "rows": len(hypothetical),
            "staked_units": round(hypothetical_staked, 6),
            "pnl_units": round(hypothetical_pnl, 6),
            "roi": round(hypothetical_pnl / hypothetical_staked, 6) if hypothetical_staked else None,
        },
        "executed": {
            "label": "executed",
            "rows": 0,
            "staked_units": None,
            "pnl_units": None,
            "roi": None,
            "status": "not_available_no_execution_attribution_in_ledgers",
        },
    }


def _version_ledger_evidence(
    version: str,
    rows: list[dict],
    pushes: int,
    source: str,
    source_rows_before_deduplication: int,
    duplicates_removed: int,
    artifact: dict,
    feature_names: list[str],
    predecessor_version_counts: dict[str, int],
) -> dict:
    wins = sum(str(row.get("result") or "").casefold() == "win" for row in rows)
    brier_values = []
    for row in rows:
        probability = _number(row.get("model_probability"), None)
        if probability is None or not 0 <= probability <= 1:
            continue
        outcome = 1.0 if str(row.get("result") or "").casefold() == "win" else 0.0
        brier_values.append((probability - outcome) ** 2)
    clv_values = [
        _number(row.get("probability_clv"), None)
        for row in rows
        if row.get("probability_clv") not in (None, "")
    ]
    clv_values = [value for value in clv_values if value is not None]
    expected_hash = artifact.get("declared_hash") if artifact.get("valid") else None
    row_hashes = [str(row.get("model_artifact_hash") or "") for row in rows]
    matching_hash_rows = sum(bool(expected_hash) and value == expected_hash for value in row_hashes)
    mismatching_hash_rows = sum(bool(value) and value != expected_hash for value in row_hashes)
    missing_hash_rows = sum(not value for value in row_hashes)
    pnl = _pnl_evidence(rows)
    if pnl["shadow"]["rows"]:
        pnl_basis = "shadow"
        pnl_units = pnl["shadow"]["pnl_units"]
    elif pnl["hypothetical"]["rows"]:
        pnl_basis = "hypothetical"
        pnl_units = pnl["hypothetical"]["pnl_units"]
    else:
        pnl_basis = None
        pnl_units = None
    clv_complete = bool(rows) and len(clv_values) == len(rows)
    profitability_allowed = bool(rows) and pnl["executed"]["roi"] is not None and clv_complete
    if not rows:
        lineage_status = "no_exact_version_rows"
    elif not expected_hash:
        lineage_status = "artifact_unverified"
    elif matching_hash_rows == len(rows):
        lineage_status = "exact"
    elif mismatching_hash_rows and missing_hash_rows:
        lineage_status = "mixed_mismatch_and_missing"
    elif mismatching_hash_rows:
        lineage_status = "mismatch"
    elif missing_hash_rows:
        lineage_status = "missing"
    else:
        lineage_status = "mixed"
    blockers = []
    if not rows:
        blockers.append("no_exact_model_version_settled_decisive_rows")
    if pnl["executed"]["roi"] is None:
        blockers.append("executed_roi_unavailable")
    if not clv_complete:
        blockers.append("clv_missing_or_incomplete")
    return {
        "source": source,
        "model_version": version,
        "exact_version_rows": len(rows),
        "settled": len(rows) + pushes if rows or pushes else None,
        "settled_decisive_rows": len(rows),
        "wins": wins if rows else None,
        "losses": len(rows) - wins if rows else None,
        "pushes": pushes if rows or pushes else None,
        "hit_rate": round(wins / len(rows), 6) if rows else None,
        "brier": round(sum(brier_values) / len(brier_values), 6) if brier_values else None,
        "brier_rows": len(brier_values),
        "source_rows_before_deduplication": source_rows_before_deduplication,
        "duplicates_removed": duplicates_removed,
        "predecessor_rows_excluded": sum(predecessor_version_counts.values()),
        "predecessor_version_counts": predecessor_version_counts,
        "artifact_lineage": {
            "artifact_path": artifact.get("path"),
            "expected_hash": expected_hash,
            "artifact_valid": bool(artifact.get("valid")),
            "matching_hash_rows": matching_hash_rows,
            "mismatching_hash_rows": mismatching_hash_rows,
            "missing_hash_rows": missing_hash_rows,
            "status": lineage_status,
        },
        "feature_value_attribution": _feature_attribution(rows, feature_names),
        "pnl_units": pnl_units,
        "pnl_basis": pnl_basis,
        "pnl": pnl,
        "clv": {
            "rows": len(clv_values),
            "total_exact_version_rows": len(rows),
            "coverage": round(len(clv_values) / len(rows), 6) if rows else None,
            "complete": clv_complete,
            "mean_probability_clv": round(sum(clv_values) / len(clv_values), 6) if clv_values else None,
        },
        "profitability_claim": {
            "allowed": profitability_allowed,
            "blockers": blockers,
        },
    }


def _ledger_evidence_for_source(
    sport: str,
    model_config: dict,
    version: str,
    source: str,
    source_rows: list[dict],
    artifact: dict,
    feature_names: list[str],
) -> dict:
    relevant = [
        row
        for row in source_rows
        if str(row.get("status") or "").casefold() == "settled"
        and str(row.get("result") or "").casefold() in {"win", "loss", "push"}
        and _model_owns_row(sport, model_config, row)
    ]
    deduplicated, _all_duplicates_removed = _deduplicate_ledger_rows(relevant)
    exact_settled = [row for row in deduplicated if str(row.get("model_version") or "") == version]
    exact_rows = [row for row in exact_settled if str(row.get("result") or "").casefold() in {"win", "loss"}]
    pushes = sum(str(row.get("result") or "").casefold() == "push" for row in exact_settled)
    exact_source_rows = sum(
        str(row.get("model_version") or "") == version
        and str(row.get("result") or "").casefold() in {"win", "loss"}
        for row in relevant
    )
    predecessor_counts: dict[str, int] = {}
    for row in deduplicated:
        row_version = str(row.get("model_version") or "")
        decisive = str(row.get("result") or "").casefold() in {"win", "loss"}
        if row_version and row_version != version and decisive:
            predecessor_counts[row_version] = predecessor_counts.get(row_version, 0) + 1
    return _version_ledger_evidence(
        version,
        exact_rows,
        pushes,
        source,
        exact_source_rows,
        exact_source_rows - len(exact_rows),
        artifact,
        feature_names,
        dict(sorted(predecessor_counts.items())),
    )


def _feature_registry_evidence() -> dict:
    """Return a validated, read-only view of the durable feature registry."""
    path = FEATURE_REGISTRY_FILE
    raw = _read_json(path)
    errors: list[str] = []
    if not isinstance(raw, dict):
        return {
            "status": "missing_or_invalid",
            "valid": False,
            "path": str(path.relative_to(ROOT)) if path.is_relative_to(ROOT) else str(path),
            "sha256": None,
            "errors": ["registry_missing_or_invalid_json"],
            "features": [],
            "production_ablation_summary": [],
            "counts_by_verdict": {},
        }
    features = raw.get("features")
    if not isinstance(features, list):
        errors.append("features_not_list")
        features = []
    normalized: list[dict] = []
    names: set[str] = set()
    for index, item in enumerate(features):
        if not isinstance(item, dict):
            errors.append(f"feature_{index}_not_object")
            continue
        name = str(item.get("name") or "").strip()
        if not name:
            errors.append(f"feature_{index}_missing_name")
            continue
        if name in names:
            errors.append(f"duplicate_feature:{name}")
            continue
        names.add(name)
        if not str(item.get("verdict") or "").strip():
            errors.append(f"feature_missing_verdict:{name}")
        normalized.append(dict(item))
    ablations = raw.get("production_ablation_summary")
    if not isinstance(ablations, list):
        errors.append("production_ablation_summary_not_list")
        ablations = []
    valid_ablations: list[dict] = []
    seen_ablations: set[tuple[str, str, str]] = set()
    for index, item in enumerate(ablations):
        if not isinstance(item, dict):
            errors.append(f"ablation_{index}_not_object")
            continue
        identity = (
            str(item.get("sport") or "").casefold(),
            str(item.get("model_version") or ""),
            str(item.get("feature") or ""),
        )
        if not all(identity):
            errors.append(f"ablation_{index}_missing_identity")
            continue
        if identity in seen_ablations:
            errors.append("duplicate_ablation:" + ":".join(identity))
            continue
        seen_ablations.add(identity)
        if identity[2] not in names:
            errors.append(f"ablation_feature_not_registered:{identity[2]}")
        valid_ablations.append(dict(item))
    counts: dict[str, int] = {}
    for item in normalized:
        verdict = str(item.get("verdict") or "missing")
        counts[verdict] = counts.get(verdict, 0) + 1
    return {
        "status": "verified" if not errors else "invalid",
        "valid": not errors,
        "path": str(path.relative_to(ROOT)) if path.is_relative_to(ROOT) else str(path),
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "schema_version": raw.get("schema_version"),
        "last_updated": raw.get("last_updated"),
        "authoritative_evidence": raw.get("authoritative_evidence") or {},
        "retention_policy": raw.get("retention_policy") or {},
        "evidence_grades": raw.get("evidence_grades") or {},
        "errors": errors,
        "features": normalized,
        "production_ablation_summary": valid_ablations,
        "counts_by_verdict": dict(sorted(counts.items())),
    }


def _read_model_ledger_rows(path: Path) -> list[dict]:
    """Plain read of a data/model_ledgers/<model-id>.xlsx file -- same
    shape as _parse_picks but for the new per-model schema's "Predictions"
    sheet. Deliberately not an import from the model_prediction package
    (dashboard_server.py has zero imports from it) -- a small, duplicated
    reader here, matching how _parse_picks already duplicates rather than
    imports PickLedger's own reading logic."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb["Predictions"] if "Predictions" in wb.sheetnames else wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        rows = []
        for values in rows_iter:
            if values is None or all(value is None for value in values):
                continue
            row = {header: ("" if value is None else str(value)) for header, value in zip(headers, values, strict=False)}
            if row.get("prediction_id"):
                rows.append(row)
        return rows
    finally:
        wb.close()


def _model_evidence_from_rows(model_id: str, rows: list[dict]) -> dict:
    """Lighter-weight duplicate of model_ledger.compute_model_evidence's
    Brier/log-loss/sample-size math for dashboard display -- no calibration
    buckets/ECE/logistic fit here, just what the design spec's evidence
    columns actually need. Same push-exclusion rule as ledger.py's own
    calibration_rows filter: settled win/loss only, a push is never folded
    in as a loss."""
    open_rows = [r for r in rows if r.get("status") == "open"]
    settled = [r for r in rows if r.get("status") == "settled"]
    failed = [r for r in rows if r.get("status") == "failed"]
    calibration_rows = [
        r for r in settled if r.get("result") in ("win", "loss") and r.get("model_probability")
    ]
    brier = log_loss = None
    if len(calibration_rows) >= 30:
        pairs = [
            (min(1 - 1e-12, max(1e-12, float(r["model_probability"]))), 1 if r["result"] == "win" else 0)
            for r in calibration_rows
        ]
        brier = sum((p - y) ** 2 for p, y in pairs) / len(pairs)
        log_loss = -sum(y * math.log(p) + (1 - y) * math.log(1 - p) for p, y in pairs) / len(pairs)
    clv_values = [float(r["probability_clv"]) for r in settled if r.get("probability_clv")]
    pnl_values = [float(r["pnl_units"]) for r in settled if r.get("pnl_units")]
    missing_input_rows = [r for r in rows if r.get("missing_inputs")]
    observed = sorted(r["observed_at_utc"] for r in rows if r.get("observed_at_utc"))
    return {
        "model_id": model_id,
        "total": len(rows),
        "open": len(open_rows),
        "settled": len(settled),
        "failed": len(failed),
        "wins": sum(1 for r in calibration_rows if r["result"] == "win"),
        "losses": sum(1 for r in calibration_rows if r["result"] == "loss"),
        "pushes": sum(1 for r in settled if r.get("result") not in ("win", "loss")),
        "pnl_units": round(sum(pnl_values), 4),
        "brier_score": brier,
        "log_loss": log_loss,
        "clv_coverage": round(len(clv_values) / len(settled), 4) if settled else None,
        "mean_clv": round(sum(clv_values) / len(clv_values), 6) if clv_values else None,
        "missing_input_rate": round(len(missing_input_rows) / len(rows), 4) if rows else None,
        "latest_observed_at_utc": observed[-1] if observed else None,
    }


def model_ledger_comparison() -> dict:
    """One row per event, one column per applicable model -- the
    operator's dashboard design spec. No qualified/research classification
    badges: evidence only (sample size, Brier, log loss, CLV, ROI,
    missing-input rate, data age), read straight from data/model_ledgers/.
    """
    ledgers_dir = DATA / "model_ledgers"
    if not ledgers_dir.exists():
        return {"generated_at": datetime.now(UTC).isoformat(), "events": [], "models": {}}
    evidence_by_model: dict[str, dict] = {}
    predictions_by_event: dict[str, list[dict]] = {}
    for path in sorted(ledgers_dir.glob("*.xlsx")):
        model_id = path.stem
        rows = _read_model_ledger_rows(path)
        evidence_by_model[model_id] = _model_evidence_from_rows(model_id, rows)
        for row in rows:
            if row.get("status") != "open":
                continue  # open predictions only -- settled history lives in the model's own evidence
            predictions_by_event.setdefault(row["event_id"], []).append(
                {
                    "model_id": model_id,
                    "prediction_id": row.get("prediction_id"),
                    "market_type": row.get("market_type"),
                    "selection": row.get("selection"),
                    "line": row.get("line") or None,
                    "model_probability": _number(row.get("model_probability"), None),
                    "decision_price": _number(row.get("decision_price"), None),
                    "model_market_difference": _number(row.get("model_market_difference"), None),
                    "input_availability": row.get("input_availability") or None,
                    "event_start_utc": row.get("event_start_utc"),
                    "operator_decision": row.get("operator_decision") or None,
                }
            )
    events = [
        {"event_id": event_id, "predictions": predictions}
        for event_id, predictions in sorted(
            predictions_by_event.items(),
            key=lambda item: min((p["event_start_utc"] or "" for p in item[1]), default=""),
        )
    ]
    return {
        "generated_at": datetime.now(UTC).isoformat(),
        "events": events,
        "models": evidence_by_model,
    }


def record_model_ledger_decision(payload: dict) -> dict:
    """Store the operator's own event-level decision for one prediction,
    separate from the model's own output. "Not model promotion. It is an
    event-level decision... must not change the model's ledger,
    classification, historical statistics, or dashboard evidence."

    Reuses the real, lock-protected ModelLedger.record_operator_decision
    (a local, function-scoped import -- see dedupe_ledger's identical
    "heavy import" precedent -- rather than duplicating a second,
    potentially unlocked write path for the same file here).
    """
    model_id = str(payload.get("model_id") or "")
    prediction_id = str(payload.get("prediction_id") or "")
    decision = str(payload.get("decision") or "")
    if not model_id or not prediction_id or not decision:
        return {"status": "refused", "error": "model_id, prediction_id, and decision are required"}
    path = DATA / "model_ledgers" / f"{model_id}.xlsx"
    if not path.exists():
        return {"status": "refused", "error": f"unknown model_id {model_id!r}"}

    from model_prediction.model_ledger import ModelLedger  # local: heavy import

    ledger = ModelLedger(path)
    units = payload.get("units")
    try:
        row = ledger.record_operator_decision(
            prediction_id,
            decision=decision,
            selected_model=payload.get("selected_model") or None,
            selected_market=payload.get("selected_market") or None,
            units=None if units in (None, "") else float(units),
            note=payload.get("note") or None,
        )
    except KeyError:
        return {"status": "refused", "error": f"unknown prediction_id {prediction_id!r} in {model_id}"}
    except (TypeError, ValueError) as error:
        return {"status": "refused", "error": str(error)}
    with _CACHE_LOCK:
        _CACHE.pop("model-ledgers", None)
    return {"status": "ok", "row": row}


def _production_canary_status() -> dict:
    """Production canary dashboard card — model health, timestamps, prediction counts."""
    try:
        from dashboard.production import get_production_status
        return get_production_status()
    except Exception as e:
        return {"status": "DOWN", "error": str(e)}


def production_evidence() -> dict:
    """Read-only, fail-closed evidence for every configured production artifact."""
    config = _config_payload()
    configured_models = config.get("models") or {}
    esports_validation = _read_json(OUTPUTS / "esports-baseline-validation.json") or {}
    international_validation = (
        _read_json(OUTPUTS / "international-baseball-baseline-validation.json") or {}
    )
    ledger_paths = (
        *_main_ledger_paths(),
        *_flat_ledger_paths(),
        *_research_ledger_paths(),
    )
    rows_by_source = {str(path.relative_to(ROOT)): _read_evidence_ledger(path) for path in ledger_paths}
    feature_registry = _feature_registry_evidence()
    registry_by_name = {str(item.get("name")): item for item in feature_registry["features"]}
    ablation_by_identity = {
        (
            str(item.get("sport") or "").casefold(),
            str(item.get("model_version") or ""),
            str(item.get("feature") or ""),
        ): item
        for item in feature_registry["production_ablation_summary"]
    }
    models = []
    for sport, model_config in configured_models.items():
        if not isinstance(model_config, dict) or not model_config.get("production_artifact"):
            continue
        version = str(model_config.get("active_production_version") or "")
        configured_path = Path(str(model_config["production_artifact"]))
        artifact_path = configured_path if configured_path.is_absolute() else ROOT / configured_path
        artifact, raw = _artifact_evidence(artifact_path, version, str(sport))
        spec = _production_model_spec(raw)
        locked = _backfill_aliases(
            _locked_backfill_evidence(
                str(sport),
                version,
                raw,
                artifact,
                esports_validation,
                international_validation,
            ),
            raw,
        )

        feature_names = list(spec.get("feature_names") or [])
        registry_features = []
        for feature_name in feature_names:
            registry_record = registry_by_name.get(str(feature_name))
            registry_features.append(
                {
                    "name": str(feature_name),
                    "registered": registry_record is not None,
                    "verdict": registry_record.get("verdict") if registry_record else None,
                    "status": registry_record.get("status") if registry_record else None,
                    "evidence_grade": registry_record.get("evidence_grade") if registry_record else None,
                    "sport_evidence": ablation_by_identity.get(
                        (str(sport).casefold(), version, str(feature_name))
                    ),
                }
            )
        # Main/Flat are per-sport files now (data/main/<sport>.xlsx,
        # data/flat/<sport>.xlsx) -- only sports in _MAIN_LEDGER_SPORTS
        # (mlb/wnba/soccer/tennis) have one; every other configured sport
        # (nba/nfl/esports/kbo/npb) correctly gets empty evidence here since
        # rows_by_source has no entry for a file that was never created.
        main_source = f"data/main/{str(sport).casefold()}.xlsx"
        main_ledger = _ledger_evidence_for_source(
            str(sport),
            model_config,
            version,
            main_source,
            rows_by_source.get(main_source, []),
            artifact,
            feature_names,
        )
        flat_source = f"data/flat/{str(sport).casefold()}.xlsx"
        flat_ledger = _ledger_evidence_for_source(
            str(sport),
            model_config,
            version,
            flat_source,
            rows_by_source.get(flat_source, []),
            artifact,
            feature_names,
        )

        sport_research_source = f"data/research/{str(sport).casefold()}.xlsx"
        legacy_research_source = "data/research.xlsx"
        research_source = (
            sport_research_source
            if sport_research_source in rows_by_source
            else legacy_research_source
        )
        research_ledger = _ledger_evidence_for_source(
            str(sport),
            model_config,
            version,
            research_source,
            rows_by_source.get(research_source, []),
            artifact,
            feature_names,
        )

        warnings = [{"code": code, "scope": "artifact"} for code in artifact["mismatches"]]
        if locked.get("status") != "verified":
            warnings.append({"code": locked["status"], "scope": "backfill"})
        if spec.get("coefficient_count_matches_features") is False:
            warnings.append({"code": "feature_coefficient_length_mismatch", "scope": "features"})
        if not feature_registry["valid"]:
            warnings.append({"code": "feature_registry_invalid", "scope": "features"})
        for feature in registry_features:
            if not feature["registered"]:
                warnings.append(
                    {
                        "code": "active_feature_not_registered",
                        "scope": "features",
                        "feature": feature["name"],
                    }
                )

        config_qualified = str(model_config.get("status") or "").casefold() in {
            "qualified",
            "shadow_qualified",
            "production",
        }
        artifact_qualified = (
            (raw.get("qualification") or {}).get("qualified")
            if raw.get("method") == "logistic_regression"
            else raw.get("qualified_for_betting")
        )
        qualification_overridden = model_config.get("qualification_override") is True
        if isinstance(artifact_qualified, bool) and config_qualified != artifact_qualified and not qualification_overridden:
            warnings.append(
                {
                    "code": "config_artifact_qualification_mismatch",
                    "scope": "qualification",
                    "config_status": model_config.get("status"),
                    "artifact_qualified": artifact_qualified,
                }
            )

        for ledger_name, ledger in (("main_ledger", main_ledger), ("flat_ledger", flat_ledger), ("research_ledger", research_ledger)):
            if not ledger["exact_version_rows"]:
                warnings.append(
                    {
                        "code": "no_exact_version_settled_decisive_rows",
                        "scope": ledger_name,
                    }
                )
                continue
            if ledger["artifact_lineage"]["status"] != "exact":
                warnings.append(
                    {
                        "code": "ledger_artifact_lineage_missing_or_mismatched",
                        "scope": ledger_name,
                    }
                )
            if ledger["feature_value_attribution"]["status"] != "complete":
                warnings.append({"code": "feature_value_attribution_missing", "scope": ledger_name})

        definition_valid = (
            artifact["valid"]
            and locked.get("status") == "verified"
            and spec.get("coefficient_count_matches_features") is not False
        )
        performance_complete = (
            main_ledger["profitability_claim"]["allowed"] and flat_ledger["profitability_claim"]["allowed"]
        )
        profitability = {
            "claim_allowed": False,
            "status": "not_established",
            "requires": [
                "exact_model_version_settled_decisive_rows",
                "executed_roi",
                "complete_clv",
            ],
            "blockers": sorted(
                {
                    blocker
                    for ledger in (main_ledger, flat_ledger, research_ledger)
                    for blocker in ledger["profitability_claim"]["blockers"]
                }
            ),
            "main_ledger_claim_allowed": main_ledger["profitability_claim"]["allowed"],
            "flat_ledger_claim_allowed": flat_ledger["profitability_claim"]["allowed"],
        }

        models.append(
            {
                "sport": str(sport).lower(),
                "model_version": version or None,
                "status": model_config.get("status"),
                "features": spec.get("features") or [],
                "feature_registry": registry_features,
                "backfill": locked,
                "main_ledger": main_ledger,
                "flat_ledger": flat_ledger,
                "research_ledger": research_ledger,
                "profitability": profitability,
                "warnings": warnings,
                "configured_status": model_config.get("status"),
                "active_model_version": version or None,
                "production_artifact": str(model_config["production_artifact"]),
                "artifact": artifact,
                "model_spec": spec,
                "locked_backfill": locked,
                "ledger_evidence": {
                    "main_ledger": main_ledger,
                    "flat_ledger": flat_ledger,
                },
                "model_definition_and_backfill_valid": definition_valid,
                "production_performance_evidence_complete": performance_complete,
                "evidence_valid": definition_valid and performance_complete,
                "issues": warnings,
            }
        )

    generated_at = datetime.now(timezone.utc).isoformat()
    return {
        "generated_at": generated_at,
        "generated_at_utc": generated_at,
        "read_only": True,
        "claim_policy": {
            "profitability_requires": [
                "exact_model_version_settled_decisive_rows",
                "executed_roi",
                "complete_clv",
            ],
            "shadow_and_hypothetical_pnl_are_not_profitability_evidence": True,
        },
        "sources": {
            "config": str(CONFIG_FILE.relative_to(ROOT)),
            "feature_registry": feature_registry["path"],
            "ledgers": list(rows_by_source),
            "esports_validation": "outputs/latest/esports-baseline-validation.json",
            "international_validation": "outputs/latest/international-baseball-baseline-validation.json",
        },
        "configured_production_models": len(models),
        "feature_registry": feature_registry,
        "all_model_definitions_and_backfills_valid": bool(models)
        and all(model["model_definition_and_backfill_valid"] for model in models),
        "all_production_performance_evidence_complete": bool(models)
        and all(model["production_performance_evidence_complete"] for model in models),
        "all_production_evidence_valid": bool(models) and all(model["evidence_valid"] for model in models),
        "models": models,
    }


# ── SECTION: Status & Health ────────────────────────────────────────


def _daily_pipeline_status() -> dict:
    """Split-pipeline staleness from data/logs/daily_*.log.

    The split pipeline runs: settle → polymarket slate → flat forecast →
    main forecast (which also forecasts every esports title and any
    non-production learned sport, routing them to research/gated_research
    ledgers). Each step writes its own exit code to the log. Parses the
    latest log to extract per-step status and
    overall staleness.
    """
    logs = sorted((DATA / "logs").glob("daily_*.log"))
    if not logs:
        return {"last_run_at_utc": None, "age_hours": None, "stale": True, "steps": {}}
    latest = logs[-1]
    mtime = datetime.fromtimestamp(latest.stat().st_mtime, tz=timezone.utc)
    age_hours = (datetime.now(timezone.utc) - mtime).total_seconds() / 3600
    # Parse exit codes from the log. Matched on an exact "0", not a substring
    # check, since e.g. exit code 130 (SIGINT) or 120 also contain the digit
    # "0" and would otherwise be misreported as success.
    steps: dict[str, bool] = {}
    unified_ok: bool | None = None
    try:
        text = latest.read_text(encoding="utf-8", errors="replace")
        for line in text.splitlines():
            if "Settlement exit code:" in line:
                steps["settle_ok"] = line.split(":")[-1].strip() == "0"
            elif "Ingestion exit code:" in line:
                steps["ingest_ok"] = line.split(":")[-1].strip() == "0"
            elif "Polymarket slate exit code:" in line:
                steps["slate_ok"] = line.split(":")[-1].strip() == "0"
            elif "Flat forecast exit code:" in line:
                steps["flat_ok"] = line.split(":")[-1].strip() == "0"
            elif "Main forecast exit code:" in line:
                steps["main_ok"] = line.split(":")[-1].strip() == "0"
            elif "Unified daily exit code:" in line:
                unified_ok = line.split(":")[-1].strip() == "0"
    except OSError:
        pass
    # Unified pipeline (2026-08-05): slate + flat + main run as one step.
    # When the unified exit code is present, use it as fallback for any
    # step the log didn't report individually.
    if unified_ok is not None:
        for key in ("slate_ok", "flat_ok", "main_ok"):
            if key not in steps:
                steps[key] = unified_ok
    return {
        "last_run_at_utc": mtime.isoformat(),
        "age_hours": round(age_hours, 1),
        "stale": age_hours > 6,
        "steps": steps,
    }


def _data_inventory() -> tuple[dict[str, int], dict[str, str | None], dict[str, str]]:
    """Count each sport from its real storage layout and report refresh provenance."""
    data_counts: dict[str, int] = {}
    last_ingest: dict[str, str | None] = {}
    data_sources: dict[str, str] = {}
    for sport in SPORTS:
        if sport in {"lol", "cs2", "dota2", "valorant", "rainbow_six"}:
            data_path = DATA / "esports" / sport / "matches.jsonl"
            manifest = _read_json(DATA / "esports" / sport / "manifest.json") or {}
            data_counts[sport] = _count_lines(data_path)
            last_ingest[sport] = str(manifest.get("extracted_at_utc") or "")[:10] or None
        elif sport in {"kbo", "npb"}:
            data_path = DATA / "international_baseball" / sport / "games.jsonl"
            manifest = _read_json(
                DATA / "international_baseball" / sport / "manifest.json"
            ) or {}
            data_counts[sport] = _count_lines(data_path)
            last_ingest[sport] = str(manifest.get("extracted_at_utc") or "")[:10] or None
        else:
            data_path = DATA / "historical" / f"{sport}_games_all.jsonl"
            data_counts[sport] = _count_lines(data_path)
            raw_dir = DATA / "raw" / sport
            dates = sorted(d.name for d in raw_dir.iterdir() if d.is_dir()) if raw_dir.exists() else []
            last_ingest[sport] = dates[-1] if dates else None
        data_sources[sport] = str(data_path.relative_to(ROOT))
    return data_counts, last_ingest, data_sources


def status() -> dict:
    validation, _source = _newest_validation()
    termination = _read_json(OUTPUTS / "termination-audit-2026-07-17.json") or {}
    audits = sorted(OUTPUTS.glob("termination-audit-*.json"))
    if audits:
        termination = _read_json(audits[-1]) or termination
    models = sorted(path.name for path in (ROOT / "config" / "models").glob("*.json"))
    data_counts, last_ingest, data_sources = _data_inventory()
    audit_events = _count_lines(DATA / "events.jsonl")
    last_event = None
    events_path = DATA / "events.jsonl"
    if events_path.exists():
        with events_path.open(encoding="utf-8") as handle:
            for line in handle:
                if line.strip():
                    last_event = line
        try:
            last_event = json.loads(last_event) if last_event else None
        except json.JSONDecodeError:
            last_event = None
    alerts = []
    today = datetime.now(timezone.utc).astimezone(EASTERN).date()
    for sport, day in last_ingest.items():
        if day:
            age = (today - datetime.strptime(day, "%Y-%m-%d").date()).days
            if age > 30:
                alerts.append(
                    {"level": "warn", "kind": "data_stale", "text": f"{sport.upper()} data is {age} days old"}
                )

    # Check qualification using the same artifact fallback as _ml_cell and matrix().
    # kbo/npb are zero-unit research baselines tracked via baseball_grid, checked
    # separately below. lol/cs2 are now shadow_qualified production.
    sports_meta = validation.get("sports") or {}
    for sport in ("mlb", "nba", "wnba", "nfl", "soccer", "lol", "cs2", "dota2", "valorant"):
        meta = sports_meta.get(sport) or {}
        artifact = _production_artifact(validation, sport)
        ml = _ml_cell(meta, artifact)
        if ml.get("state") == "tested_not_qualified":
            alerts.append(
                {
                    "level": "warn",
                    "kind": "not_qualified",
                    "text": f"{sport.upper()} moneyline below qualification gate",
                }
            )
        if ml.get("state") == "no_data":
            alerts.append(
                {
                    "level": "info",
                    "kind": "no_data",
                    "text": f"{sport.upper()} moneyline has no validation data",
                }
            )

    for sport, grid_key in (("kbo", "baseball_grid"), ("npb", "baseball_grid")):
        research_ml = ((validation.get(grid_key) or {}).get(sport) or {}).get("moneyline") or {}
        if not research_ml or research_ml.get("state") == "no_data":
            alerts.append(
                {
                    "level": "info",
                    "kind": "no_data",
                    "text": f"{sport.upper()} moneyline has no validation data",
                }
            )

    if not os.environ.get("POLYMARKET_KEY_ID") or not os.environ.get("POLYMARKET_SECRET_KEY"):
        alerts.append(
            {
                "level": "error",
                "kind": "no_api_key",
                "text": "Polymarket API key not configured — execution disabled",
            }
        )

    daily_pipeline = _daily_pipeline_status()
    if daily_pipeline["stale"]:
        detail = (
            f"{daily_pipeline['age_hours']}h since last daily run"
            if daily_pipeline["age_hours"] is not None
            else "no daily log found"
        )
        alerts.append(
            {"level": "warn", "kind": "daily_pipeline_stale", "text": f"Daily pipeline is stale — {detail}"}
        )

    tests = _LAST_ACTION.get("run_tests") or _latest_persisted_action("run_tests")
    # Real bug fixed 2026-08-02: promotion_allowed used to be read straight
    # off a static termination-audit-*.json snapshot (whatever the latest
    # dated file on disk happened to say, sometimes weeks stale), completely
    # independent of /api/production-evidence's own live computation --
    # so the two endpoints could (and did) directly contradict each other:
    # one screen said "allowed", the authoritative evidence calculation said
    # "not established". promotion_allowed is now the live intersection of
    # every model's real evidence (production_evidence()'s own
    # all_production_evidence_valid, itself an AND over each model's
    # artifact/lineage validity and prospective performance completeness)
    # and operational health (no error-level alert active) -- any incomplete
    # input yields False, not a stale flag sitting next to a contradicting panel.
    evidence = _cached("production-evidence", 30, production_evidence)
    operational_checks_green = not any(alert.get("level") == "error" for alert in alerts)
    promotion_allowed = bool(evidence.get("all_production_evidence_valid")) and operational_checks_green
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "models_loaded": len(models),
        "model_artifacts": models,
        "data_counts": data_counts,
        "last_ingest": last_ingest,
        "data_sources": data_sources,
        "audit_events": audit_events,
        "last_audit_event": (last_event or {}).get("event_type") if last_event else None,
        "daily_pipeline": daily_pipeline,
        "alerts": alerts,
        "tests": tests or {"status": "not_run_this_session"},
        "validation_status": termination.get("status") or validation.get("status"),
        "promotion_allowed": promotion_allowed,
        "polymarket_odds": odds_summary(),
        "polymarket_configured": bool(os.environ.get("POLYMARKET_KEY_ID") and os.environ.get("POLYMARKET_SECRET_KEY")),
        "edge_filter_min": 0.02,
        "unit_value_usd": _unit_value_usd(),
    }


# ── SECTION: Validation & Matrix ────────────────────────────────────


def _newest_validation() -> tuple[dict, str]:
    """Newest core validation merged with the newest artifact-backed soccer report."""
    candidates = sorted(
        OUTPUTS.glob("learned-model-validation*.json"),
        key=lambda path: path.stat().st_mtime,
    )
    merged: dict = {"sports": {}, "production_artifacts": {}}
    sources = []
    newest: dict = {}
    if candidates:
        newest = _read_json(candidates[-1]) or {}
        merged["sports"].update(newest.get("sports") or {})
        merged["production_artifacts"].update(newest.get("production_artifacts") or {})
        sources.append(candidates[-1].name)

    soccer_candidates: list[tuple[float, Path, dict]] = []
    for path in OUTPUTS.glob("soccer-*.json"):
        payload = _read_json(path) or {}
        if (payload.get("sports") or {}).get("soccer") and (payload.get("production_artifacts") or {}).get(
            "soccer"
        ):
            soccer_candidates.append((path.stat().st_mtime, path, payload))
    if soccer_candidates:
        _, path, soccer = max(soccer_candidates, key=lambda item: item[0])
        merged["sports"].update(soccer["sports"])
        merged["production_artifacts"].update(soccer.get("production_artifacts") or {})
        sources.append(path.name)

    # Preserve embedded grids, then fill them from their dedicated validation
    # reports. Core learned-model validation intentionally does not own these
    # research-only leagues and may omit both keys entirely.
    esports_grid = dict(newest.get("esports_grid") or {})
    esports_validation = _read_json(OUTPUTS / "esports-baseline-validation.json") or {}
    for sport, result in (esports_validation.get("titles") or {}).items():
        locked = (result.get("locked_test") or {}).get("selected_matches") or {}
        if not locked:
            continue
        esports_grid[str(sport)] = {
            "moneyline": {
                "state": "research_only",
                "hit_rate": locked.get("accuracy"),
                "calls": locked.get("calls"),
                "brier": locked.get("brier"),
                "units": 0.0,
                "diagnostic_units": locked.get("units_at_minus_110"),
                "threshold": (result.get("chosen") or {}).get("confidence_threshold"),
                "model_version": result.get("model_version"),
                "qualified_for_betting": False,
            }
        }
    if esports_validation.get("titles"):
        sources.append("esports-baseline-validation.json")

    baseball_grid = dict(newest.get("baseball_grid") or {})
    baseball_validation = _read_json(OUTPUTS / "international-baseball-baseline-validation.json") or {}
    for sport, result in (baseball_validation.get("leagues") or {}).items():
        locked = result.get("locked_test") or {}
        if not locked:
            continue
        baseball_grid[str(sport)] = {
            "moneyline": {
                "state": "research_only",
                "hit_rate": locked.get("accuracy_decisive"),
                "calls": locked.get("calls"),
                "brier": locked.get("brier_settlement"),
                "units": 0.0,
                "diagnostic_units": locked.get("units_at_minus_110"),
                "observations": locked.get("observations"),
                "ties": locked.get("ties"),
                "model_version": result.get("model_version"),
                "qualified_for_betting": False,
            }
        }
    if baseball_validation.get("leagues"):
        sources.append("international-baseball-baseline-validation.json")

    merged["esports_grid"] = esports_grid
    merged["baseball_grid"] = baseball_grid

    return merged, " + ".join(sources)


def _production_artifact(validation: dict, sport: str) -> dict:
    raw_path = str((validation.get("production_artifacts") or {}).get(sport) or "")
    if not raw_path:
        # The validation report can contain sport metrics without repeating
        # the active artifact path. model.yaml remains authoritative for which
        # version the dashboard labels as production/shadow.
        raw_path = _config_production_artifact_path(sport)
    if not raw_path:
        return {}
    path = Path(raw_path)
    if not path.is_absolute():
        path = ROOT / path
    return _read_json(path) or {}


def _config_production_artifact_path(sport: str) -> str:
    """Read model.yaml and return the production_artifact path for a sport."""
    try:
        import yaml

        if CONFIG_FILE.exists():
            config = yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8")) or {}
            models = config.get("models") or {}
            sport_config = models.get(sport.upper()) or {}
            return str(sport_config.get("production_artifact") or "")
    except Exception:
        pass
    return ""


def _ml_cell(sport_meta: dict, artifact: dict | None = None) -> dict:
    """Moneyline cell pinned to the active artifact's exact validated variant."""
    variants = sport_meta.get("variants") or {}
    artifact = artifact or {}

    # Esports: flat Elo artifact without market_models wrapper
    if "k" in artifact and "ratings" in artifact:
        qual = artifact.get("qualified_for_betting", False)
        return {
            "state": "qualified" if qual else "research_only",
            "hit_rate": None,
            "calls": len(artifact.get("ratings", {})),
            "variant": ["elo_neutral"],
            "variant_name": "elo_neutral_series",
            "model_version": artifact.get("model_version"),
        }

    market_model = (artifact.get("market_models") or {}).get("moneyline") or {}
    artifact_features = tuple(market_model.get("feature_names") or ())
    artifact_qualification = artifact.get("qualification") or {}
    variant_name = None
    variant = None
    for name, candidate in variants.items():
        if not isinstance(candidate, dict) or tuple(candidate.get("features") or ()) != artifact_features:
            continue
        holdout = (candidate.get("primary_65") or {}).get("locked_holdout") or {}
        calls_match = artifact_qualification.get("calls") in (None, holdout.get("calls"))
        artifact_rate = artifact_qualification.get("hit_rate")
        holdout_rate = holdout.get("hit_rate")
        rate_match = artifact_rate is None or (
            holdout_rate is not None and abs(float(artifact_rate) - float(holdout_rate)) < 1e-9
        )
        if calls_match and rate_match:
            variant_name, variant = name, candidate
            break

    if variant is None and not artifact:
        for name in ("elo_trend", "elo_trend_defense", "elo_trend_park", "elo_only"):
            candidate = variants.get(name) or {}
            if ((candidate.get("primary_65") or {}).get("locked_holdout") or {}).get("qualified"):
                variant_name, variant = name, candidate
                break
        if variant is None:
            match = next(
                (
                    (name, candidate)
                    for name, candidate in variants.items()
                    if isinstance(candidate, dict) and candidate.get("primary_65")
                ),
                (None, None),
            )
            variant_name, variant = match

    primary = (variant or {}).get("primary_65") or {}
    holdout = primary.get("locked_holdout") or {}
    if not holdout and artifact_qualification:
        holdout = artifact_qualification
        primary = {"learned_threshold": market_model.get("confidence_threshold")}
        variant_name = "artifact_pinned"
    if not holdout:
        # A boolean without calls/hit-rate evidence is not enough to render a
        # qualified cell. Show as tested (model exists, qualified flag set)
        # rather than untested, but make the missing metrics explicit.
        if artifact.get("qualified"):
            return {
                "state": "tested_not_qualified",
                "hit_rate": None,
                "calls": None,
                "threshold": market_model.get("confidence_threshold"),
                "variant": list(artifact_features),
                "variant_name": "artifact_pinned",
                "model_version": artifact.get("model_version"),
                "readiness": "ARTIFACT_QUALIFIED_FLAG_WITHOUT_LOCKED_HOLDOUT_METRICS",
            }
        return {"state": "no_data"}

    cell = {
        "state": "qualified" if holdout.get("qualified") else "tested_not_qualified",
        "hit_rate": holdout.get("hit_rate"),
        "calls": holdout.get("calls"),
        "units": holdout.get("units_at_minus_110"),
        "brier": holdout.get("brier_score"),
        "threshold": primary.get("learned_threshold"),
        "roi": holdout.get("roi"),
        "variant": list(artifact_features or tuple((variant or {}).get("features") or ())),
        "variant_name": variant_name,
        "model_version": artifact.get("model_version"),
    }

    # Soccer: also show 3-way variant if available
    three = variants.get("soccer_3way")
    if three and isinstance(three, dict):
        tp = three.get("primary_65", {})
        th = tp.get("locked_holdout", {})
        if th.get("qualified"):
            cell["three_way"] = {
                "hit_rate": th.get("hit_rate"),
                "calls": th.get("calls"),
                "units": th.get("units_at_minus_110"),
            }

    return cell


def matrix() -> dict:
    """Wiring status per sport/market: is it live in `daily`, what does it
    actually run on, and which ledger does it write to.

    Deliberately NOT hit rates, Brier scores, MAE, or promotion-gate status
    -- those questions live on the System/Evidence tabs. This mirrors the
    operator's standing instruction to discuss models in terms of wiring and
    features, not validation metrics (see docs/PROJECT_STATUS.md's operating
    note). Rows are maintained by hand alongside cli.py's actual dispatch
    logic -- when a sport's wiring changes, update this list in the same
    change. Active version labels are read from config/model.yaml so this
    operational surface cannot silently lag a model promotion.
    """
    try:
        config = yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8")) or {}
    except (OSError, yaml.YAMLError):
        config = {}
    configured_models = config.get("models") or {}

    def active_version(sport: str) -> str:
        sport_config = configured_models.get(sport.upper()) or {}
        return str(
            sport_config.get("active_production_version")
            or sport_config.get("active_research_version")
            or "version unavailable"
        )

    esports_versions = ", ".join(
        f"{sport.upper()} {active_version(sport)}"
        for sport in ("lol", "cs2", "dota2", "valorant", "rainbow_six")
    )
    rows = [
        {
            "sport": "MLB", "market": "Moneyline", "wired": True,
            "model": (
                "learned_forward.py -- Elo + trend logistic regression "
                f"({active_version('mlb')}); features: elo_probability, trend_gap, "
                "park_factor, weather_factor, pitcher_era_gap, bullpen_weakness_gap "
                "-- no probable_starter_era_gap (v6's contaminated ESPN-live-probables "
                "feature, retired 2026-07-30). No confidence-threshold or "
                "min-edge-vs-ask gate as of 2026-07-30 -- every real forecasted game "
                "becomes a real, sized Main-ledger call; both numbers are still "
                "recorded for manual review, not used to hide the row."
            ),
            "ledger": "Main + Flat",
        },
        {
            "sport": "MLB", "market": "Totals & Spread", "wired": True,
            "model": "models/mlb.py MeasuredEdgeTotalsModel/margin -- Gamma-Poisson mixture "
            "Monte-Carlo (20000 sims), priced against real Polymarket lines closest to "
            "50/50. Rebuilt 2026-07-30: real elasticities fit per factor (offense 0.035, "
            "starter weakness 0.211, park 0.222, weather 0.021) replace the prior "
            "assumed-1.0 multiplicative weight on each; bullpen elasticity fit "
            "consistently negative/implausible and is zeroed rather than trusted.",
            "ledger": "Flat only",
        },
        {
            "sport": "MLB", "market": "Moneyline (legacy)", "wired": False,
            "model": "MeasuredEdgeMarginModel via --model legacy-measured-edge -- intentionally "
            "retained as an explicit manual rollback, not part of daily",
            "ledger": "Main, only if manually invoked",
        },
        {
            "sport": "NBA", "market": "Moneyline", "wired": True,
            "model": (
                "learned_forward.py -- Elo + trend logistic regression "
                f"({active_version('nba')}); features: elo_probability, trend_gap, "
                "defensive_trend_gap. Not in PRODUCTION_SPORTS (domain.py) -- never "
                "reaches Main regardless of real-world strength."
            ),
            "ledger": "Flat only",
        },
        {
            "sport": "WNBA", "market": "Moneyline", "wired": True,
            "model": (
                "learned_forward.py -- Elo + trend logistic regression "
                f"({active_version('wnba')}); features: elo_probability, trend_gap, "
                "defensive_trend_gap"
            ),
            "ledger": "Main + Flat",
        },
        {
            "sport": "NFL", "market": "Moneyline", "wired": True,
            "model": (
                "learned_forward.py -- Elo + trend logistic regression "
                f"({active_version('nfl')}); features: elo_probability, trend_gap. "
                "Not in PRODUCTION_SPORTS (domain.py) -- never reaches Main regardless "
                "of real-world strength."
            ),
            "ledger": "Flat only",
        },
        {
            "sport": "Soccer", "market": "Totals (2.5)", "wired": True,
            "model": "models/soccer.py -- Poisson/Dixon-Coles score matrix",
            "ledger": "Flat Research + Gated Research",
        },
        {
            "sport": "Soccer", "market": "Moneyline", "wired": True,
            "model": "Same score matrix, matched against Polymarket's per-team team_win "
            "Yes/No markets (not a single combined moneyline market)",
            "ledger": "Flat Research + Gated Research",
        },
        {
            "sport": "Soccer", "market": "BTTS", "wired": True,
            "model": "Same score matrix, Platt-recalibrated (2026-07-31: raw joint-matrix "
            "probability was overconfident, 55.0% real accuracy; calibrated to 56.7%). "
            "Matching/pricing fully wired (soccer_forward.py), but no BTTS market has "
            "ever been observed live on Polymarket US (checked across all 19 configured "
            "leagues and 4 real captured days) -- activates automatically, no further "
            "code changes, once one appears and its raw market type is confirmed.",
            "ledger": "Flat Research + Gated Research (currently prices 0 -- no real market exists yet)",
        },
        {
            "sport": "Tennis", "market": "Moneyline", "wired": True,
            "model": "models/tennis.py -- surface-blended Elo, singles only, WTA only "
            f"({active_version('tennis')}; Polymarket US has no ATP market; "
            "ESPN has no ITF scoreboard)",
            "ledger": "Flat Research + Gated Research",
        },
        {
            "sport": "LOL / CS2 / DOTA2 / VALORANT / Rainbow Six", "market": "Moneyline", "wired": True,
            "model": "esports.py -- result-based neutral Elo, Platt-scaled, refreshed from "
            f"bo3.gg before every forecast; active config: {esports_versions}. Gated "
            "Research's research_confidence_gate raised 2026-07-31 (was 0.0 for every "
            "title, barely filtering anything -- real settled Gated picks were "
            "performing worse than unfiltered Research) to each title's own already-"
            "validated confidence_threshold: LOL/DOTA2/VALORANT 0.05, CS2/Rainbow Six 0.03.",
            "ledger": "Flat Research + Gated Research",
        },
        {
            "sport": "CoD / Rocket League / Overwatch", "market": "Moneyline", "wired": False,
            "model": "Polymarket lists these leagues and real BBO is captured daily, but "
            "bo3.gg (the only esports data source) has no discipline for any of them -- not buildable",
            "ledger": "--",
        },
        {
            "sport": "KBO / NPB", "market": "Moneyline", "wired": True,
            "model": "international_baseball.py -- tie-aware home-field Elo "
            f"(KBO {active_version('kbo')}; NPB {active_version('npb')}; "
            "result/margin only, no starters/park/weather)",
            "ledger": "Flat Research + Gated Research",
        },
    ]
    return {
        "rows": rows,
        "note": (
            "Wiring and features, not validation stats -- see docs/PROJECT_STATUS.md's "
            "operating note."
        ),
    }


# ── SECTION: Backtests & Odds ───────────────────────────────────────


def backtests() -> list[dict]:
    items = []
    if OUTPUTS.exists():
        for path in sorted(OUTPUTS.glob("*.json")):
            payload = _read_json(path) or {}
            items.append(
                {
                    "file": path.name,
                    "modified": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()[
                        :16
                    ],
                    "status": payload.get("status"),
                    "sport": payload.get("sport") or ",".join(payload.get("sports_scope", [])[:4]),
                    "keys": sorted(payload)[:12],
                    "size_kb": round(path.stat().st_size / 1024, 1),
                }
            )
    return items


def odds_summary(sport: str | None = None) -> dict:
    """Per-sport summary of stored Polymarket odds snapshots for today."""
    today = _today()
    sports = [sport] if sport else list(SPORTS)
    result: dict = {}
    for s in sports:
        path = DATA / "odds" / s / today / "polymarket_snapshots.jsonl"
        if not path.exists():
            result[s] = {"snapshots": 0, "status": "no_data"}
            continue
        snaps = []
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                try:
                    snaps.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
        moneyline = [s for s in snaps if s.get("market_type") == "moneyline"]
        spread = [s for s in snaps if s.get("market_type") == "spread"]
        total = [s for s in snaps if s.get("market_type") == "total"]
        with_bbo = [
            s
            for s in snaps
            if (s.get("long") or {}).get("ask") is not None and (s.get("short") or {}).get("ask") is not None
        ]
        result[s] = {
            "snapshots": len(snaps),
            "moneyline": len(moneyline),
            "spread": len(spread),
            "total": len(total),
            "with_executable_bbo": len(with_bbo),
            "date": today,
        }
    return result if sport is None else result.get(sport, {})


def market_snapshots(sport: str, day: str) -> dict:
    path = DATA / "odds" / sport / day / "polymarket_snapshots.jsonl"
    latest: dict[str, dict] = {}
    first_ask: dict[str, float] = {}
    if path.exists():
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                try:
                    snap = json.loads(line)
                except json.JSONDecodeError:
                    continue
                slug = snap.get("market_slug")
                if not slug:
                    continue
                # Historical dashboard prices are pregame-only. Once an event
                # starts, later BBOs must never replace the last valid quote.
                # Legacy snapshots captured before these fields existed lack
                # them entirely — fall back to the pre-filter behavior for
                # those rather than dropping their price history outright.
                if snap.get("timestamp_valid") is False:
                    continue
                observed_raw, event_start_raw = snap.get("observed_at_utc"), snap.get("event_start_utc")
                if observed_raw and event_start_raw:
                    try:
                        observed_at = datetime.fromisoformat(str(observed_raw).replace("Z", "+00:00"))
                        event_start = datetime.fromisoformat(str(event_start_raw).replace("Z", "+00:00"))
                    except ValueError:
                        observed_at = event_start = None
                    if observed_at is not None and observed_at >= event_start:
                        continue
                ask = (snap.get("long") or {}).get("ask")
                if slug not in first_ask and ask is not None:
                    first_ask[slug] = ask
                latest[slug] = snap
    rows = []
    for slug, snap in sorted(latest.items()):
        long_side = snap.get("long") or {}
        ask, bid = long_side.get("ask"), long_side.get("bid")
        rows.append(
            {
                "market_slug": slug,
                # Market tables can contain hundreds of contracts. Keep this
                # endpoint local and deterministic instead of doing one public
                # metadata lookup per total/team-total slug.
                "market_name": _human_market_name(str(slug), allow_lookup=False),
                "market_type": snap.get("market_type"),
                "line": snap.get("line"),
                "league": snap.get("league"),
                "event_start_utc": snap.get("event_start_utc"),
                "description": long_side.get("description"),
                "bid": bid,
                "ask": ask,
                "spread": round(ask - bid, 4) if ask is not None and bid is not None else None,
                "ask_size": long_side.get("ask_size"),
                "bid_size": long_side.get("bid_size"),
                "move": round(ask - first_ask[slug], 4) if ask is not None and slug in first_ask else None,
                "observed_at_utc": snap.get("observed_at_utc"),
            }
        )
    return {"sport": sport, "date": day, "markets": rows, "count": len(rows)}


def _team_matches(team_name: str, side_description: str) -> bool:
    team = " ".join(team_name.casefold().split())
    description = " ".join(side_description.casefold().split())
    if not team or not description:
        return False
    if team == description:
        return True
    shorter, longer = (description, team) if len(description) <= len(team) else (team, description)
    return f" {shorter} " in f" {longer} "


def _lines_match(a: float | None, b: float | None) -> bool:
    return a is not None and b is not None and abs(a - b) < 1e-6


def _row_selected_team(row: dict) -> str | None:
    selection = str(row.get("selection") or "").casefold()
    if selection == "home":
        return str(row.get("home_team") or "") or None
    if selection == "away":
        return str(row.get("away_team") or "") or None
    return None


def _row_line(row: dict) -> float | None:
    try:
        return float(row.get("line"))  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _row_matches_snapshot_event(row: dict, snapshot: dict) -> bool:
    """Confirm a spread/total snapshot belongs to this row's actual game.

    Unlike moneyline (long/short descriptions ARE the two team names, so
    matching both against the row's away/home teams already confirms the
    right game), a spread snapshot's "team" is only ONE team (the long
    side), and a total snapshot has no team field at all -- only a line.
    Real bug found and fixed 2026-08-04 while writing this matching (never
    shipped): without this check, _spread_side_for_row's line-negation
    branch matches ANY OTHER game's spread whose line happens to be the
    exact negation of this row's line (e.g. every one of a day's ~15 "-1.5"
    MLB spread markets satisfies "not is_market_team and line == -1.5" for
    a row picking the away side of an unrelated game) -- same risk for
    total's line-only match. `event_title` carries both team names for
    every market type, so requiring both to match closes this."""
    title = str(snapshot.get("event_title") or "")
    away = str(row.get("away_team") or "")
    home = str(row.get("home_team") or "")
    return bool(away) and bool(home) and _team_matches(away, title) and _team_matches(home, title)


def _spread_side_for_row(row: dict, snapshot: dict) -> str | None:
    """Which side ("long"/"short") of a spread snapshot this row's team+line
    picks, or None if it doesn't match at all (including belonging to a
    different game -- see _row_matches_snapshot_event).

    Duplicated from polymarket_execute.py's _resolve_spread_side (same "zero
    imports from that package" rule _team_matches's docstring already
    follows) -- a spread market's own line/team always describe its LONG
    side; SHORT is always the exact negation. A row's line is selection-
    relative, so it only matches the market's long side when the picked
    team IS the market's own team; for the opponent, the row's line must
    equal the negation instead. Found 2026-08-04: this matching never
    existed in dashboard_server.py at all -- _pick_quote returned None for
    every spread row unconditionally, so a real, sized MLB spread pick could
    never be previewed or ordered even when the exact market genuinely
    existed live on Polymarket."""
    if not _row_matches_snapshot_event(row, snapshot):
        return None
    selected_team = _row_selected_team(row)
    row_line = _row_line(row)
    market_team = snapshot.get("team")
    market_line = snapshot.get("line")
    if not selected_team or row_line is None or market_team is None or market_line is None:
        return None
    is_market_team = _team_matches(selected_team, str(market_team))
    if is_market_team and _lines_match(row_line, float(market_line)):
        return "long"
    if not is_market_team and _lines_match(row_line, -float(market_line)):
        return "short"
    return None


def _total_side_for_row(row: dict, snapshot: dict) -> str | None:
    """Which side ("long"/"short") of a total snapshot this row's
    over/under+line picks, or None if it doesn't match. Same "found
    2026-08-04, never existed" note as _spread_side_for_row -- and the same
    cross-game collision risk is even sharper here: a total snapshot has NO
    team field at all, only a line, so without _row_matches_snapshot_event
    ANY other game's total market with the same line (e.g. exactly 8.5)
    would match a row that has nothing to do with that game."""
    if not _row_matches_snapshot_event(row, snapshot):
        return None
    selection = str(row.get("selection") or "").casefold().strip()
    if selection not in ("over", "under"):
        return None
    row_line = _row_line(row)
    market_line = snapshot.get("line")
    if row_line is None or market_line is None or not _lines_match(row_line, float(market_line)):
        return None
    long_desc = str((snapshot.get("long") or {}).get("description") or "").casefold().strip()
    short_desc = str((snapshot.get("short") or {}).get("description") or "").casefold().strip()
    matches_long = long_desc == selection
    matches_short = short_desc == selection
    if matches_long == matches_short:
        return None  # ambiguous or neither side literally says "over"/"under"
    return "long" if matches_long else "short"


def _pick_quote(row: dict) -> dict | None:
    """Last valid pregame executable side quote for a moneyline, spread, or
    total pick.

    Moneyline-only until 2026-08-04 -- MLB spread/total became real, sized
    Main-ledger picks (LEDGER_ROUTING.md) without this function ever being
    extended, so `quote is None` unconditionally for every spread/total row
    made every one of them permanently unexecutable in the dashboard ("no
    exact executable Polymarket US market mapping"), even when the live
    Polymarket market genuinely existed (verified live: a real -1.5 spread
    and a real 8.5 total both existed for a real logged Main pick that was
    showing this exact error). polymarket_execute.py's deeper, real
    execution-time re-verification (_verify_live_side_and_timing, P0-1
    2026-08-03) already supported spread/total/btts -- it was simply
    unreachable from the dashboard order flow because this earlier,
    ticket-building stage always refused first.
    """
    market_type = row.get("market_type")
    if market_type not in ("moneyline", "spread", "total"):
        return None
    sport = str(row.get("league") or "").lower()
    if sport not in SPORTS:
        return None
    try:
        event_start = datetime.fromisoformat(str(row.get("event_start_utc") or "").replace("Z", "+00:00"))
    except ValueError:
        return None
    day = event_start.astimezone(EASTERN).date().isoformat()
    path = DATA / "odds" / sport / day / "polymarket_snapshots.jsonl"
    if not path.exists():
        return None
    latest: dict[str, dict] = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                snapshot = json.loads(line)
            except json.JSONDecodeError:
                continue
            if snapshot.get("timestamp_valid") is False:
                continue
            observed_raw = snapshot.get("observed_at_utc")
            if observed_raw:
                try:
                    snapshot_observed = datetime.fromisoformat(str(observed_raw).replace("Z", "+00:00"))
                except ValueError:
                    snapshot_observed = None
                if snapshot_observed is not None and snapshot_observed >= event_start:
                    continue
            if snapshot.get("market_type") != market_type:
                continue
            slug = str(snapshot.get("market_slug") or "")
            if not slug or any(
                marker in slug.casefold() for marker in ("-f5-", "-f3-", "-f7-", "-1st-", "-h1-", "-h2-")
            ):
                continue
            if market_type == "moneyline":
                long_description = str((snapshot.get("long") or {}).get("description") or "")
                short_description = str((snapshot.get("short") or {}).get("description") or "")
                away = str(row.get("away_team") or "")
                home = str(row.get("home_team") or "")
                if not (
                    (_team_matches(away, long_description) and _team_matches(home, short_description))
                    or (_team_matches(home, long_description) and _team_matches(away, short_description))
                ):
                    continue
            elif market_type == "spread":
                # Multiple alternate-line markets exist per event -- this
                # must match the row's exact team+line, not just the game,
                # or every alternate line would collide in `latest` and trip
                # the doubleheader guard below.
                if _spread_side_for_row(row, snapshot) is None:
                    continue
            else:  # total
                if _total_side_for_row(row, snapshot) is None:
                    continue
            latest[slug] = snapshot
    if not latest:
        return None
    if len(latest) > 1:
        return None  # doubleheader: two contracts matched; never guess which game
    snapshot = max(latest.values(), key=lambda item: str(item.get("observed_at_utc") or ""))
    if market_type == "moneyline":
        selected_team = (
            str(row.get("home_team") or "")
            if str(row.get("selection") or "").casefold() == "home"
            else str(row.get("away_team") or "")
        )
        matches_long = _team_matches(selected_team, str((snapshot.get("long") or {}).get("description") or ""))
        matches_short = _team_matches(selected_team, str((snapshot.get("short") or {}).get("description") or ""))
        if matches_long == matches_short:
            return None  # ambiguous side (same-city names) — never default to long
        side_name = "long" if matches_long else "short"
    elif market_type == "spread":
        side_name = _spread_side_for_row(row, snapshot)
    else:  # total
        side_name = _total_side_for_row(row, snapshot)
    if side_name is None:
        return None
    side = snapshot.get(side_name) or {}
    ask = _number(side.get("ask"), -1)
    if not 0 < ask < 1:
        return None
    observed = str(snapshot.get("observed_at_utc") or "")
    try:
        observed_at = datetime.fromisoformat(observed.replace("Z", "+00:00"))
        age_seconds = max(0, int((datetime.now(timezone.utc) - observed_at).total_seconds()))
    except ValueError:
        age_seconds = 10**9
    return {
        "market_slug": snapshot.get("market_slug"),
        "side": side_name,
        "description": side.get("description"),
        "bid": side.get("bid"),
        "ask": ask,
        "ask_size": side.get("ask_size"),
        "observed_at_utc": observed,
        "age_seconds": age_seconds,
        "fresh": age_seconds <= 300,
        "market_state": snapshot.get("market_state"),
        "price_role": "pregame_close",
        "seconds_before_start": max(0, int((event_start - observed_at).total_seconds())),
    }


# ── SECTION: Orders & Execution ─────────────────────────────────────


def _load_orders() -> dict:
    payload = _read_json(ORDERS_FILE) or {}
    orders = payload.get("orders") if isinstance(payload, dict) else None
    rows = list(orders) if isinstance(orders, list) else []
    repaired = False
    # Older dashboard builds mixed the CLI confirmation prompt into stdout.
    # The exchange could accept an order and return an ID, while json.loads()
    # rejected the combined prompt + JSON and locally recorded it as refused.
    # Recover those durable exchange acknowledgements so a refresh cannot offer
    # the same model order a second time.
    for row in rows:
        if row.get("status") != "refused" or not isinstance(row.get("error"), str):
            continue
        decoded = _decode_command_output(row["error"])
        if decoded.get("status") == "submitted" and decoded.get("order_id"):
            row.update(
                status="submitted",
                order_id=str(decoded["order_id"]),
                order_state=decoded.get("order_state"),
                error=None,
            )
            repaired = True
    result = {"orders": rows}
    if repaired:
        _save_orders(result)
    return result


def _save_orders(payload: dict) -> None:
    DASH_DIR.mkdir(exist_ok=True)
    temporary = ORDERS_FILE.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")
    os.replace(temporary, ORDERS_FILE)


def _decode_command_output(raw: str) -> dict:
    """Decode a CLI JSON result even when an interactive prompt precedes it."""
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else {}
    except (json.JSONDecodeError, TypeError):
        pass
    decoder = json.JSONDecoder()
    best: dict = {}
    for index, character in enumerate(str(raw)):
        if character != "{":
            continue
        try:
            value, end = decoder.raw_decode(str(raw)[index:])
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict) and not str(raw)[index + end :].strip():
            best = value
            break
    return best


def _latest_order_for_pick(row: dict, quote: dict | None, orders: dict | None = None) -> dict | None:
    """Find an order across equivalent model-version rows for the same contract side."""
    orders = (orders or _load_orders())["orders"]
    pick_id = str(row.get("pick_id") or "")
    direct = [order for order in orders if str(order.get("pick_id") or "") == pick_id]
    if direct:
        return direct[-1]
    if quote is None:
        return None
    equivalent = [
        order
        for order in orders
        if order.get("status") == "submitted"
        and order.get("market_slug") == quote.get("market_slug")
        and order.get("side") == quote.get("side")
    ]
    return equivalent[-1] if equivalent else None


def _filled_entry_for_pick(
    row: dict, orders: dict | None = None, portfolio_history: dict | None = None
) -> dict | None:
    """Exchange-backed entry price for a filled dashboard BUY, in pick-side terms."""
    pick_id = str(row.get("pick_id") or "")
    filled = [
        order
        for order in (orders or _load_orders())["orders"]
        if str(order.get("pick_id") or "") == pick_id
        and order.get("action", "buy") == "buy"
        and (order.get("status") == "filled" or _number(order.get("cum_quantity"), 0) > 0)
    ]
    if not filled:
        return None
    order = filled[-1]
    limit_price = _number(order.get("price"), None)
    side = str(order.get("side") or "")
    slug = str(order.get("market_slug") or "")
    submitted = str(order.get("submitted_at_utc") or "")
    quantity = _number(order.get("cum_quantity") or order.get("size_shares"), None)

    # Portfolio trades use the exchange's YES/long coordinate even for a NO
    # fill. Match the fill and convert it back to the outcome actually bought.
    candidates = []
    for activity in (portfolio_history or _load_portfolio_history())["activities"]:
        if activity.get("type") != "trade" or activity.get("market_slug") != slug:
            continue
        occurred = str(activity.get("occurred_at_utc") or "")
        if submitted and occurred and occurred < submitted:
            continue
        trade_quantity = _number(activity.get("quantity"), None)
        if quantity is not None and trade_quantity is not None and abs(quantity - trade_quantity) > 0.01:
            continue
        raw_price = _number(activity.get("exchange_price", activity.get("price")), None)
        if raw_price is None or not 0 < raw_price < 1:
            continue
        selected_price = 1 - raw_price if side == "short" else raw_price
        if limit_price is not None and selected_price > limit_price + 0.0001:
            continue
        candidates.append((occurred, selected_price, str(activity.get("activity_id") or "")))
    if candidates:
        _, price, activity_id = sorted(candidates)[0]
        return {
            "price": round(price, 6),
            "basis": "exchange_trade",
            "side": side,
            "market_slug": slug,
            "activity_id": activity_id,
        }
    if limit_price is None:
        return None
    return {
        "price": limit_price,
        "basis": "filled_order_limit",
        "side": side,
        "market_slug": slug,
    }


def _dashboard_order_status(exchange_state: str | None) -> str:
    state = str(exchange_state or "").upper()
    return {
        "ORDER_STATE_FILLED": "filled",
        "ORDER_STATE_CANCELED": "canceled",
        "ORDER_STATE_REPLACED": "replaced",
        "ORDER_STATE_REJECTED": "rejected",
        "ORDER_STATE_EXPIRED": "expired",
    }.get(state, "submitted")


def _reconcile_orders() -> None:
    """Replace local submission state with the exchange's current order state.

    Real race fixed 2026-08-02: this used to read+write orders.json without
    holding _ORDER_LOCK, unlike submit_order/preview_position_sell/etc.
    Called from dashboard_picks() on essentially every /api/picks request,
    so it could run concurrently with a real order submission -- read a
    stale snapshot (before the new order was appended), then write that
    stale snapshot back after submit_order's own locked append completed,
    silently erasing the just-submitted order record. Held under the lock
    now, matching every other read-modify-write of this file.
    """
    with _ORDER_LOCK:
        _reconcile_orders_locked()


def _reconcile_orders_locked() -> None:
    payload = _load_orders()
    active = [
        order for order in payload["orders"] if order.get("status") == "submitted" and order.get("order_id")
    ]
    if not active:
        return
    order_ids = sorted({str(order["order_id"]) for order in active})

    def fetch_order_states() -> dict:
        try:
            command = _resolve_runner() + ["order-status"]
            for order_id in order_ids:
                command.extend(("--order-id", order_id))
            process = subprocess.run(
                command,
                cwd=ROOT,
                capture_output=True,
                text=True,
                timeout=20,
                env=_runner_env(),
            )
        except (OSError, subprocess.TimeoutExpired):
            return {}
        raw = process.stdout if process.returncode == 0 else process.stderr
        return _decode_command_output(raw)

    result = _cached("order-states:" + ",".join(order_ids), 10, fetch_order_states)
    if result.get("status") != "live":
        return
    by_id = {str(item.get("order_id")): item for item in result.get("orders", []) if item.get("order_id")}
    changed = False
    for order in active:
        snapshot = by_id.get(str(order["order_id"]))
        if snapshot is None:
            continue
        exchange_state = snapshot.get("order_state")
        updates = {
            "order_state": exchange_state,
            "status": _dashboard_order_status(exchange_state),
            "cum_quantity": snapshot.get("cum_quantity"),
            "leaves_quantity": snapshot.get("leaves_quantity"),
            "last_checked_at_utc": result.get("observed_at_utc"),
        }
        if any(order.get(key) != value for key, value in updates.items()):
            order.update(updates)
            changed = True
    if changed:
        _save_orders(payload)


def _event_already_started(row: dict) -> bool:
    event_start = str(row.get("event_start_utc") or "")
    try:
        game_time = datetime.fromisoformat(event_start.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return False
    return game_time < datetime.now(timezone.utc)


def _order_readiness(row: dict, quote: dict | None) -> tuple[bool, str]:
    if row.get("status") != "open":
        return False, "pick is not open"
    if row.get("record_type") == "QUALIFIED_SHADOW_CALL":
        if float(row.get("units") or 0) <= 0:
            return False, "qualified pick has no authorized units"
    elif row.get("record_type") == "RESEARCH_OBSERVATION":
        eligible, reason = _manual_research_eligibility(row)
        if not eligible:
            return False, reason
        if not _suggested_units(row):
            return False, "manual order has no authorized unit cap"
    else:
        return False, "unsupported ledger record type"
    if quote is None:
        return False, "no exact executable Polymarket US market mapping"
    if not quote.get("fresh"):
        return False, "market quote is older than 5 minutes; scan prices first"
    if quote.get("market_state") != "MARKET_STATE_OPEN":
        return False, "market is not open"
    # Block orders on past games
    if _event_already_started(row):
        return False, "game has already started"
    missing = [name for name in ("POLYMARKET_KEY_ID", "POLYMARKET_SECRET_KEY") if not os.environ.get(name)]
    if missing:
        return False, f"missing {' and '.join(missing)}"
    return True, "ready"


def _net_position_quantity(slug: str, portfolio_history: dict) -> float | None:
    """Net shares still held on a market, from cached (no live call) activity history.

    Prefers the exchange's own settlement record when one exists: a
    "settlement" activity's after_quantity is ground truth for a resolved
    market, and holds even when a position went through several buy/sell
    round-trips beforehand. Only when the market hasn't settled does this
    fall back to summing trades -- the exchange reports cost_basis_usd/
    realized_pnl_usd on trades that close or reduce a position (null on
    pure opens), so opening-quantity minus closing-quantity approximates net
    exposure for a still-open, never-settled position. That fallback is a
    heuristic, not ground truth: multiple round-trips on the same market
    (sell then rebuy) can make it wrong, which is exactly why the settlement
    record is checked first.
    """
    activities = [a for a in portfolio_history.get("activities", []) if a.get("market_slug") == slug]
    if not activities:
        return None
    settlements = [a for a in activities if a.get("type") == "settlement"]
    if settlements:
        latest = max(settlements, key=lambda a: str(a.get("occurred_at_utc") or ""))
        after = latest.get("after_quantity")
        if after is not None:
            return _number(after, 0.0)
    trades = [a for a in activities if a.get("type") == "trade"]
    if not trades:
        return None
    net = 0.0
    for trade in trades:
        quantity = _number(trade.get("quantity"), 0.0)
        closing = trade.get("cost_basis_usd") is not None or trade.get("realized_pnl_usd") is not None
        net += -quantity if closing else quantity
    return net


def _decorate_pick(row: dict, orders: dict | None = None, portfolio_history: dict | None = None) -> dict:
    quote = _pick_quote(row)
    ready, reason = _order_readiness(row, quote)
    order = _latest_order_for_pick(row, quote, orders)
    manual, _ = _manual_research_eligibility(row)
    filled_entry = _filled_entry_for_pick(row, orders, portfolio_history)
    position_closed = False
    if order and order.get("action", "buy") == "buy" and order.get("status") == "filled":
        slug = str(order.get("market_slug") or "")
        if slug:
            net = _net_position_quantity(slug, portfolio_history or _load_portfolio_history())
            position_closed = net is not None and abs(net) < 1e-6
    display_units = (
        _number(row.get("units"))
        or _number(row.get("research_score_units"))
        or _suggested_units(row)
        or 0
    )
    display_pnl = _number(row.get("pnl_units")) or _number(row.get("research_pnl_units"))
    # Fallback: compute P&L from american_odds when research_pnl_units is
    # absent. Confirmed 2026-08-01: this never fires against real data --
    # every row settle() ever touches already has a real pnl_units/
    # research_pnl_units -- it's a defensive net for malformed/legacy rows
    # only. Deliberately NOT importing pricing.profit_units here (this file
    # has zero dependencies on the model_prediction package by design, kept
    # runnable standalone); this formula must instead be kept in exact sync
    # with pricing.profit_units by hand -- see
    # tests/test_dashboard_server.py's
    # test_pnl_fallback_formula_matches_pricing_profit_units, which fails
    # loudly if the two ever diverge.
    if display_pnl == 0 and row.get("result") in ("win", "loss") and row.get("american_odds"):
        try:
            odds = int(row["american_odds"])
            if odds > 0:
                display_pnl = display_units * odds / 100
            else:
                display_pnl = display_units * 100 / abs(odds)
            if row["result"] == "loss":
                display_pnl = -display_units
        except (ValueError, TypeError):
            pass
    return {
        **row,
        # Preserve ledger facts in the API. A Research NO_CALL must remain
        # units=0 instead of looking like a sized Gated call merely because
        # the dashboard can calculate a hypothetical display size.
        "display_units": display_units,
        "display_pnl_units": display_pnl,
        "quote": quote,
        "order": order,
        "filled_entry": filled_entry,
        "position_closed": position_closed,
        "buy_ready": ready,
        "buy_block_reason": reason,
        "unit_value_usd": _unit_value_usd(),
        "order_authorization": ("manual_research_override" if manual else "qualified_model"),
    }


def _pick_identity(row: dict) -> tuple[str, ...]:
    """Canonical dashboard identity, independent of model version or logged price."""
    event = str(row.get("event_id") or "").strip()
    if not event:
        event = "|".join(
            str(row.get(key) or "").strip().casefold()
            for key in ("event_start_utc", "away_team", "home_team")
        )
    line = row.get("line")
    try:
        line_value = f"{float(line):g}" if line not in (None, "") else ""
    except (TypeError, ValueError):
        line_value = str(line or "").strip().casefold()
    return (
        str(row.get("league") or "").strip().casefold(),
        event,
        str(row.get("market_type") or "").strip().casefold(),
        str(row.get("selection") or "").strip().casefold(),
        line_value,
        str(row.get("period") or row.get("horizon") or "").strip().casefold(),
    )


def _dedupe_picks(rows: list[dict]) -> list[dict]:
    """Keep the latest ledger observation for each actual bet shown in the UI."""
    latest: dict[tuple[str, ...], tuple[str, int, dict]] = {}
    for index, row in enumerate(rows):
        rank = (str(row.get("created_at_utc") or ""), index, row)
        key = _pick_identity(row)
        if key not in latest or rank[:2] >= latest[key][:2]:
            latest[key] = rank
    return [item[2] for item in sorted(latest.values(), key=lambda item: item[1])]


def dashboard_picks() -> list[dict]:
    """Latest unique picks with persistent local-clear and order state attached."""
    _reconcile_orders()
    archived = set(_load_archive()["pick_ids"])
    orders, portfolio_history = _load_orders(), _load_portfolio_history()
    # No RESEARCH_ONLY_LEAGUES filter needed here anymore: read_picks() only
    # ever sources from data/main/<sport>.xlsx for _MAIN_LEDGER_SPORTS
    # (mlb/wnba/soccer/tennis), so it's now physically impossible for a
    # research-only league's row to appear here. The old filter was
    # actively wrong for soccer and tennis specifically -- both have real
    # Main-ledger rows since their 2026-08-02/08-03 promotion, but
    # RESEARCH_ONLY_LEAGUES was never updated and was silently hiding them.
    return [
        {
            **_decorate_pick(row, orders, portfolio_history),
            "archived": str(row.get("pick_id")) in archived,
            "suggested_paper_units": _suggested_units(row),
        }
        for row in _dedupe_picks(read_picks())
    ]


def preview_order(payload: dict) -> dict:
    action = str(payload.get("action") or "buy").lower()
    if action not in ("buy", "sell"):
        return {"status": "refused", "error": "action must be buy or sell"}
    pick_id = str(payload.get("pick_id") or "")
    row = _find_pick_by_id(pick_id)
    if row is None:
        return {"status": "refused", "error": "unknown pick id"}
    decorated = _decorate_pick(row)
    quote = decorated["quote"]
    if quote is None:
        return {"status": "refused", "error": "no executable quote for this contract"}
    # Buys require the buy-readiness gate. Sells are exits and only require an
    # executable quote (you can always try to close a position you hold).
    if action == "buy" and not decorated["buy_ready"]:
        return {"status": "refused", "error": decorated["buy_block_reason"]}
    try:
        raw_price = float(payload.get("price"))
        size_shares = float(payload.get("size_shares"))
    except (TypeError, ValueError):
        return {"status": "refused", "error": "price and shares must be numeric"}
    # Validate the tick on the RAW input; rounding first would silently accept
    # (and change) a sub-cent price the user never confirmed.
    if not 0.01 <= raw_price <= 0.99 or abs(raw_price * 100 - round(raw_price * 100)) > 1e-8:
        return {"status": "refused", "error": "limit price must be a 0.01 tick from 0.01 to 0.99"}
    price = round(raw_price, 2)
    if not 0 < size_shares <= 100000:
        return {"status": "refused", "error": "shares must be greater than 0 and at most 100,000"}
    estimated_cost = round(price * size_shares, 2)
    manual = row.get("record_type") == "RESEARCH_OBSERVATION"

    if action == "sell":
        # A resting SELL limit must sit AT OR ABOVE the current bid (post-only:
        # do not cross into the bid). No dollar cost cap — a sell returns
        # capital. Proceeds are informational. Sells are otherwise less
        # restricted than buys (you can always try to close a position you
        # hold) -- but _pick_quote never returns a snapshot observed at or
        # after event_start_utc, so once a game has started the only
        # available quote is a frozen pregame snapshot that can never
        # update again. Validating a resting sell's "don't cross the bid"
        # check against that frozen number would be actively misleading,
        # not just imprecise, so this one case is still blocked.
        if _event_already_started(row):
            return {"status": "refused", "error": "game has already started; quote can no longer update"}
        bid = quote.get("bid")
        if bid is not None and price <= float(bid):
            return {
                "status": "refused",
                "error": (
                    f"resting sell price must be above the current bid {float(bid):.2f}; "
                    "crossing orders are blocked"
                ),
            }
        maximum_cost = None
    else:
        # Buy path: limit below model probability for manual research and keep
        # the authorized unit cap. A limit at/above the ask becomes an IOC
        # marketable limit; a lower limit remains post-only GTC.
        execution_config = (_config_payload().get("execution") or {}) if manual else {}
        if (
            manual
            and execution_config.get("manual_research_require_positive_edge", True)
            and price >= float(row.get("model_probability") or 0)
        ):
            return {
                "status": "refused",
                "error": (
                    f"manual buy limit {price:.2f} must stay below the model probability "
                    f"{float(row.get('model_probability') or 0):.2f}"
                ),
            }
        authorized_units = _suggested_units(row) if manual else float(row.get("units") or 0)
        maximum_cost = round(float(authorized_units or 0) * _unit_value_usd(), 2)
        if estimated_cost > maximum_cost + 0.005:
            return {
                "status": "refused",
                "error": (
                    f"order cost ${estimated_cost:.2f} exceeds this pick's "
                    f"{float(row.get('units') or 0):g}U cap (${maximum_cost:.2f})"
                ),
            }
    ask = float(quote["ask"]) if action == "buy" else None
    marketable = action == "buy" and price >= ask
    order_type = "limit_ioc" if marketable else "limit_gtc"
    nonce = secrets.token_urlsafe(24)
    ticket = {
        "nonce": nonce,
        "pick_id": pick_id,
        "action": action,
        "market_slug": quote["market_slug"],
        "side": quote["side"],
        "price": price,
        "size_shares": size_shares,
        "units": round(estimated_cost / _unit_value_usd(), 4),
        "unit_value_usd": _unit_value_usd(),
        "estimated_cost_usd": estimated_cost,
        "estimated_proceeds_usd": estimated_cost if action == "sell" else None,
        "maximum_cost_usd": maximum_cost,
        "order_type": order_type,
        "execution_mode": "marketable_limit" if marketable else "resting_limit",
        "reference_ask": ask,
        "manual_research_order": manual,
        "created_at": time.time(),
        "expires_at": time.time() + 300,
    }
    with _ORDER_LOCK:
        _ORDER_PREVIEWS[nonce] = ticket
    return {"status": "preview", **ticket}


def submit_order(payload: dict) -> dict:
    nonce = str(payload.get("nonce") or "")
    with _ORDER_LOCK:
        ticket = _ORDER_PREVIEWS.pop(nonce, None)
    if ticket is None or time.time() > float(ticket["expires_at"]):
        return {"status": "refused", "error": "order preview expired; preview it again"}
    row = _find_pick_by_id(ticket["pick_id"])
    if row is None:
        return {"status": "refused", "error": "pick disappeared before submission"}
    quote = _pick_quote(row)
    if quote is None or quote["market_slug"] != ticket["market_slug"]:
        return {"status": "refused", "error": "market changed; preview the order again"}
    action = ticket.get("action", "buy")
    if action == "sell":
        # See the matching comment in preview_order: a game already in
        # progress has no live quote to validate against (_pick_quote only
        # ever returns pregame snapshots), so re-check here too rather than
        # trust that preview-time state still holds at submission time.
        if _event_already_started(row):
            return {"status": "refused", "error": "game has already started; quote can no longer update"}
        bid = quote.get("bid")
        if bid is not None and ticket["price"] <= float(bid):
            return {"status": "refused", "error": "bid moved above your limit; preview the sell again"}
    else:
        ready, reason = _order_readiness(row, quote)
        if not ready:
            return {"status": "refused", "error": reason}
        ask = float(quote["ask"])
        if ticket.get("order_type") == "limit_ioc":
            if ticket["price"] < ask:
                return {
                    "status": "refused",
                    "error": (
                        f"current ask moved to {ask:.2f}, above your {ticket['price']:.2f} "
                        "buy cap; preview the order again"
                    ),
                }
        elif ticket["price"] >= ask:
            return {
                "status": "refused",
                "error": "ask moved down through your resting limit; preview the order again",
            }
    command = _resolve_runner() + [
        "execute",
        "--pick-id",
        ticket["pick_id"],
        "--size-shares",
        str(ticket["size_shares"]),
        "--price",
        str(ticket["price"]),
        "--side",
        ticket["side"],
        "--action",
        action,
        "--order-type",
        ticket.get("order_type", "limit_gtc"),
        "--market-slug",
        ticket["market_slug"],
        "--execute",
    ]
    if ticket.get("manual_research_order"):
        command.append("--manual-research-order")
    process = subprocess.run(
        command,
        cwd=ROOT,
        input="Y\n",
        capture_output=True,
        text=True,
        timeout=30,
        env=_runner_env(),
    )
    raw = process.stdout if process.returncode == 0 else process.stderr
    result = _decode_command_output(raw)
    if not result:
        result = {"status": "refused", "error": raw[-1000:] or "order command failed"}
    record = {
        **ticket,
        "nonce": None,
        "status": result.get("status", "refused"),
        "order_id": result.get("order_id"),
        "order_state": result.get("order_state"),
        "exchange_price": result.get("exchange_price"),
        "price_basis": "selected_outcome_probability",
        "exchange_price_basis": "long_side_probability",
        "submitted_at_utc": datetime.now(timezone.utc).isoformat(),
        "error": result.get("error"),
    }
    with _ORDER_LOCK:
        orders = _load_orders()
        orders["orders"].append(record)
        _save_orders(orders)
    with _CACHE_LOCK:
        _CACHE.clear()
    return {**result, "pick_id": ticket["pick_id"]}


def _live_bbo(market_slug: str) -> dict | None:
    """Fetch a fresh BBO for one market slug from the public gateway."""
    try:
        from model_prediction.data_sources.polymarket_us import PolymarketUSClient

        return PolymarketUSClient().snapshot(market_slug)
    except Exception:  # noqa: BLE001 - any failure => no quote, caller handles
        return None


def preview_position_sell(payload: dict) -> dict:
    """Preview a resting SELL limit against a held live exchange position."""
    slug = str(payload.get("market_slug") or "")
    side = str(payload.get("side") or "long")
    if not slug or side not in ("long", "short"):
        return {"status": "refused", "error": "market_slug and side (long|short) are required"}
    try:
        raw_price = float(payload.get("price"))
        size_shares = float(payload.get("size_shares"))
    except (TypeError, ValueError):
        return {"status": "refused", "error": "price and shares must be numeric"}
    if not 0.01 <= raw_price <= 0.99 or abs(raw_price * 100 - round(raw_price * 100)) > 1e-8:
        return {"status": "refused", "error": "limit price must be a 0.01 tick from 0.01 to 0.99"}
    price = round(raw_price, 2)
    if not 0 < size_shares <= 1_000_000:
        return {"status": "refused", "error": "shares must be greater than 0"}
    portfolio = live_portfolio_view()
    position = next(
        (
            item
            for item in (portfolio.get("open") or {}).get("positions", [])
            if item.get("market_slug") == slug and item.get("side") == side
        ),
        None,
    )
    if portfolio.get("status") != "live" or position is None:
        return {"status": "refused", "error": "live position could not be verified"}
    held = _number(position.get("available_quantity"), 0.0)
    if size_shares > held + 1e-9:
        return {
            "status": "refused",
            "error": f"cannot sell {size_shares:g} shares; only {held:g} are available",
        }
    snapshot = _live_bbo(slug)
    bid = None
    if snapshot:
        bid = (snapshot.get(side) or {}).get("bid")
    if bid is not None and price <= float(bid):
        return {
            "status": "refused",
            "error": (
                f"resting sell price must be above the current {side} bid {float(bid):.2f}; "
                "crossing orders are blocked"
            ),
        }
    nonce = secrets.token_urlsafe(24)
    ticket = {
        "nonce": nonce,
        "kind": "position_sell",
        "market_slug": slug,
        "side": side,
        "price": price,
        "size_shares": size_shares,
        "estimated_proceeds_usd": round(price * size_shares, 2),
        "current_bid": bid,
        "verified_available_quantity": held,
        "created_at": time.time(),
        "expires_at": time.time() + 300,
    }
    with _ORDER_LOCK:
        _ORDER_PREVIEWS[nonce] = ticket
    return {"status": "preview", **ticket}


def submit_position_sell(payload: dict) -> dict:
    nonce = str(payload.get("nonce") or "")
    with _ORDER_LOCK:
        ticket = _ORDER_PREVIEWS.pop(nonce, None)
    if ticket is None or ticket.get("kind") != "position_sell" or time.time() > float(ticket["expires_at"]):
        return {"status": "refused", "error": "sell preview expired; preview it again"}
    portfolio = live_portfolio_view()
    position = next(
        (
            item
            for item in (portfolio.get("open") or {}).get("positions", [])
            if item.get("market_slug") == ticket["market_slug"] and item.get("side") == ticket["side"]
        ),
        None,
    )
    held = _number((position or {}).get("available_quantity"), 0.0)
    if portfolio.get("status") != "live" or position is None or ticket["size_shares"] > held + 1e-9:
        return {"status": "refused", "error": "available live shares changed; preview the sell again"}
    # Re-check the bid moved-through condition against a fresh quote.
    snapshot = _live_bbo(ticket["market_slug"])
    if snapshot:
        bid = (snapshot.get(ticket["side"]) or {}).get("bid")
        if bid is not None and ticket["price"] <= float(bid):
            return {"status": "refused", "error": "bid moved above your limit; preview the sell again"}
    command = _resolve_runner() + [
        "sell-position",
        "--market-slug",
        ticket["market_slug"],
        "--side",
        ticket["side"],
        "--price",
        str(ticket["price"]),
        "--size-shares",
        str(ticket["size_shares"]),
        "--execute",
    ]
    process = subprocess.run(
        command,
        cwd=ROOT,
        input="Y\n",
        capture_output=True,
        text=True,
        timeout=30,
        env=_runner_env(),
    )
    raw = process.stdout if process.returncode == 0 else process.stderr
    result = _decode_command_output(raw)
    if not result:
        result = {"status": "refused", "error": raw[-1000:] or "sell command failed"}
    record = {
        **ticket,
        "nonce": None,
        "status": result.get("status", "refused"),
        "order_id": result.get("order_id"),
        "order_state": result.get("order_state"),
        "exchange_price": result.get("exchange_price"),
        "price_basis": "selected_outcome_probability",
        "exchange_price_basis": "long_side_probability",
        "submitted_at_utc": datetime.now(timezone.utc).isoformat(),
        "error": result.get("error"),
    }
    with _ORDER_LOCK:
        orders = _load_orders()
        orders["orders"].append(record)
        _save_orders(orders)
    with _CACHE_LOCK:
        _CACHE.clear()
    return result


def live_gateway_slate(sport: str, day: str) -> dict:
    """Read-only live discovery quotes from the public gateway (indicative)."""
    league = {"mlb": "mlb", "nba": "nba", "wnba": "wnba", "nfl": "nfl"}.get(sport)
    if league is None:
        return {
            "events": [],
            "observed_at_utc": datetime.now(UTC).isoformat(),
            "error": "live indicative gateway is unavailable for this sport",
        }
    url = f"{GATEWAY}/v2/leagues/{league}/events?limit=50&section=general&type=sport"
    try:
        with urllib.request.urlopen(url, timeout=15) as response:
            payload = json.loads(response.read().decode())
    except Exception as error:  # noqa: BLE001
        return {
            "events": [],
            "observed_at_utc": datetime.now(UTC).isoformat(),
            "error": type(error).__name__,
        }
    events = []
    for event in payload.get("events", []):
        start = str(event.get("startTime") or "")
        if not start:
            continue
        try:
            start_et = datetime.fromisoformat(start.replace("Z", "+00:00")).astimezone(EASTERN)
        except ValueError:
            continue
        if start_et.date().isoformat() != day:
            continue
        markets = []
        for market in event.get("markets", []):
            sides = []
            for side in market.get("marketSides", []):
                quote = side.get("quote")
                if isinstance(quote, dict):
                    quote = quote.get("value")
                try:
                    quote = float(quote)
                except (TypeError, ValueError):
                    quote = None
                sides.append({"description": side.get("description"), "quote": quote})
            markets.append(
                {
                    "slug": market.get("slug"),
                    "type": market.get("sportsMarketTypeV2") or market.get("sportsMarketType"),
                    "line": market.get("line"),
                    "sides": sides,
                }
            )
        events.append(
            {"title": event.get("title"), "start_utc": start, "slug": event.get("slug"), "markets": markets}
        )
    return {
        "events": events,
        "observed_at_utc": datetime.now(UTC).isoformat(),
        "note": "indicative discovery quotes; decision prices come from stored BBO asks",
    }


def _all_ledger_rows_for_price_scan() -> list[dict]:
    """Every row from all four ledgers (Main, Flat, Research, Gated
    Research), for the "Scan Open Ledger Prices" action.

    read_picks() only ever parsed picks.xlsx (Main) -- confirmed a real gap
    (2026-07-31): every open Flat/Research/Gated Research pick's price was
    permanently stale, since nothing else ever refreshed them. Pulled into
    its own function (rather than inlined in _action_command) so tests can
    monkeypatch this one seam instead of four separate parse calls. The
    caller's (sport, game_day, slug) dedup already collapses the same real
    contract appearing in more than one ledger (e.g. an MLB game open in
    both Main and Flat) into a single --contract entry.
    """
    return (
        read_picks()
        + read_flat_picks()
        + _parse_research_picks(gated=False)
        + _parse_research_picks(gated=True)
    )


def _action_command(name: str, payload: dict) -> list[str]:
    runner = _resolve_runner()
    if name == "run_tests":
        # pytest lives next to whatever python the runner resolved to.
        python = runner[0] if runner[0].endswith(("python", "python.exe", "python3")) else sys.executable
        return [python, "-m", "pytest", "tests/", "-q", "--no-header"]
    cli = runner if len(runner) > 1 else runner  # module or console-script form
    if name == "daily":
        # One scheduling authority: the run supervisor. Executing
        # run_daily.sh directly here would make the dashboard a second
        # scheduler — two paths capable of acting like the control plane
        # fighting over the daily lock, with a legitimate scheduled run
        # showing up as a failed dashboard job. Route through the
        # supervisor so a busy lease comes back as exit 75 (the daily_lock
        # convention) and is recorded as skipped, not failed.
        return [sys.executable, "-m", "model_prediction.run_supervisor", "run", "daily"]
    if name == "flat_forecast":
        return cli + ["flat-forecast", "--all", "--date", str(payload.get("date") or _today()), "--log"]
    if name == "main_forecast":
        # Same command as Step 3 of run_daily.sh: MLB/WNBA -> picks.xlsx,
        # esports/soccer/KBO/NPB -> separate per-sport research and gated
        # workbooks. One shared forecast pass, not three independent ones —
        # the Ledger, Research, and Gated Research tab buttons all trigger it.
        return cli + [
            "forecast", "--all", "--date", str(payload.get("date") or _today()),
            "--log", "--replace-today", "--model", "learned",
        ]
    if name == "refresh_prices":
        day = str(payload.get("date") or _today())
        command = cli + ["polymarket-ledger-prices", "--date", day]
        seen: set[tuple[str, str, str]] = set()
        archived = set(_load_archive()["pick_ids"])
        for row in _dedupe_picks(_all_ledger_rows_for_price_scan()):
            if row.get("status") != "open" or str(row.get("pick_id")) in archived:
                continue
            quote = _pick_quote(row) or {}
            sport = str(row.get("league") or "").strip().lower()
            slug = str(quote.get("market_slug") or "").strip()
            try:
                game_day = (
                    datetime.fromisoformat(str(row.get("event_start_utc") or "").replace("Z", "+00:00"))
                    .astimezone(EASTERN)
                    .date()
                    .isoformat()
                )
            except ValueError:
                continue
            target = (sport, game_day, slug)
            if sport not in SPORTS or not slug or target in seen:
                continue
            seen.add(target)
            command += ["--contract", f"{sport}@{game_day}={slug}"]
        return command
    if name == "settle":
        return cli + ["settle", "--all-unsettled"]
    if name == "bootstrap":
        command = cli + [
            "bootstrap",
            "--sport",
            _safe_sport(payload.get("sport")),
            "--from",
            str(payload.get("from_date") or _today()),
        ]
        if payload.get("to_date"):
            command += ["--to", str(payload["to_date"])]
        return command
    raise ValueError(f"unknown action: {name}")


def _today() -> str:
    return datetime.now(timezone.utc).astimezone(EASTERN).date().isoformat()


class InvalidSportError(ValueError):
    """Raised when a client requests an unsupported sport slug."""


def _safe_sport(value) -> str:
    if value not in SPORTS:
        raise InvalidSportError(f"unsupported sport: {value}")
    return str(value)


# ── SECTION: Jobs & Actions ─────────────────────────────────────────


def _job_status_for_returncode(returncode: int) -> str:
    """Map a job's exit code to a dashboard status.

    75 is the daily_lock convention (LOCK_BUSY_EXIT): the supervisor
    records the run as skipped when another run holds the lease. A manual
    trigger landing during the scheduled run is a coalesced skip, not a
    failure — lock refusal must not be routine scheduling behavior.
    """
    if returncode == 75:
        return "skipped"
    return "ok" if returncode == 0 else "failed"


def start_action(name: str, payload: dict) -> dict:
    """Launch a whitelisted action as a background job; return its id at once.

    Actions like `daily` legitimately run for minutes (slate discovery plus
    ~200 BBO snapshots). Holding the HTTP request open that long is what made
    the browser show "Failed to fetch" — so the POST returns immediately and
    the page polls /api/job.
    """
    if not _ACTION_LOCK.acquire(blocking=False):
        running = next((j for j in _JOBS.values() if j["status"] == "running"), None)
        return {
            "status": "busy",
            "error": "another action is already running",
            "job_id": running["job_id"] if running else None,
        }
    try:
        command = _action_command(name, payload)
    except (ValueError, RuntimeError) as error:
        _ACTION_LOCK.release()
        return {"status": "failed", "error": str(error)}
    job_id = f"{name}-{int(time.time())}"
    job = {
        "job_id": job_id,
        "action": name,
        "status": "running",
        "command": " ".join(command[-8:]),
        "output_tail": "",
        "started_at": datetime.now(timezone.utc).isoformat()[:19],
        "started_monotonic": time.time(),
        "seconds": 0.0,
    }
    with _JOBS_LOCK:
        _JOBS[job_id] = job
        while len(_JOBS) > 20:
            _JOBS.pop(next(iter(_JOBS)))
    _log(f"job started: {job_id} :: {job['command']}")
    _persist_jobs()

    def _work() -> None:
        started = time.time()
        chunks: list[str] = []
        process = None
        try:
            process = subprocess.Popen(
                command,
                cwd=ROOT,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                env=_runner_env(),
            )
            assert process.stdout is not None
            for line in process.stdout:
                chunks.append(line)
                if len(chunks) > 4000:
                    chunks = chunks[-2000:]
                with _JOBS_LOCK:
                    job["output_tail"] = "".join(chunks)[-12000:]
                    job["seconds"] = round(time.time() - started, 1)
            returncode = process.wait(timeout=3600)
            status = _job_status_for_returncode(returncode)
        except Exception as error:  # noqa: BLE001
            status, returncode = "failed", -1
            chunks.append(f"\n{type(error).__name__}: {error}\n")
        finally:
            # A finished (or failed) job must leave NOTHING behind: kill any
            # still-running child, reap it, and close the pipe.
            if process is not None:
                if process.poll() is None:
                    process.kill()
                    try:
                        process.wait(timeout=10)
                    except subprocess.TimeoutExpired:
                        pass
                if process.stdout is not None:
                    process.stdout.close()
        with _JOBS_LOCK:
            job.update(
                {
                    "status": status,
                    "returncode": returncode,
                    "seconds": round(time.time() - started, 1),
                    "output_tail": "".join(chunks)[-12000:],
                }
            )
        _LAST_ACTION[name] = dict(job)
        _log(f"job finished: {job_id} :: {status} in {job['seconds']}s")
        _persist_jobs()
        with _CACHE_LOCK:
            _CACHE.clear()
        _ACTION_LOCK.release()

    threading.Thread(target=_work, daemon=True).start()
    return {"status": "started", "job_id": job_id, "command": job["command"]}


def job_status(job_id: str) -> dict:
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            # Server may have restarted mid-job; answer from the on-disk record
            # instead of a bare 404 so the page can explain what happened.
            disk = _load_persisted_jobs().get(job_id)
            if disk:
                if disk.get("status") == "running":
                    disk["status"] = "interrupted"
                    disk["error"] = (
                        "the dashboard server restarted while this job was running; "
                        "the underlying CLI run may still have completed — check the ledger/summary"
                    )
                return disk
            return {"status": "unknown", "error": "no such job"}
        snapshot = {key: value for key, value in job.items() if key != "started_monotonic"}
        if snapshot["status"] == "running":
            snapshot["seconds"] = round(time.time() - job["started_monotonic"], 1)
        return snapshot


def _persist_jobs() -> None:
    try:
        DASH_DIR.mkdir(exist_ok=True)
        with _JOBS_LOCK:
            payload = {
                job_id: {k: v for k, v in job.items() if k != "started_monotonic"}
                for job_id, job in _JOBS.items()
            }
        JOBS_FILE.write_text(json.dumps(payload, default=str), encoding="utf-8")
    except OSError:
        pass


def _load_persisted_jobs() -> dict:
    try:
        return json.loads(JOBS_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _latest_persisted_action(action: str) -> dict | None:
    matches = [
        job
        for job in _load_persisted_jobs().values()
        if job.get("action") == action and job.get("status") != "running"
    ]
    if not matches:
        return None
    return max(matches, key=lambda job: str(job.get("started_at") or ""))


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------

_REBUILD_VIEWS = {"status", "sports", "benchmark", "economics", "runs", "health", "shadow-picks"}


class Handler(BaseHTTPRequestHandler):
    def _send(self, payload, content_type="application/json", code=200) -> None:
        body = payload if isinstance(payload, bytes) else json.dumps(payload, default=str).encode()
        try:
            self.send_response(code)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionResetError):
            pass  # client disconnected — nothing to do

    def _send_head(self, payload, content_type="application/json", code=200) -> None:
        body = payload if isinstance(payload, bytes) else json.dumps(payload, default=str).encode()
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()

    def log_message(self, fmt, *args):  # quiet
        pass

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        query = {key: values[0] for key, values in parse_qs(parsed.query).items()}
        route = parsed.path
        try:
            if route in ("/", "/dashboard.html"):
                page = ROOT / "dashboard.html"
                if page.exists():
                    self._send(_inject_dashboard_token(page.read_bytes()), "text/html; charset=utf-8")
                else:
                    self._send({"error": "dashboard.html missing"}, code=404)
            elif route == "/api/status":
                self._send(_cached("status", 30, status))
            elif route.startswith("/api/data/"):
                # SQL-backed read-only data service (consolidation C):
                # paginated predictions, counts, runs, promotions, health,
                # and the cheap change fingerprint. No Excel, no mutation.
                data_route = route.removeprefix("/api/data/")
                try:
                    self._send(data_service_handle(data_route, parse_qs(parsed.query)))
                except KeyError:
                    self._send({"error": f"unknown data route: {data_route}"}, code=404)
            elif route == "/api/matrix":
                self._send(_cached("matrix", 60, matrix))
            elif route == "/api/production-evidence":
                self._send(_cached("production-evidence", 30, production_evidence))
            elif route == "/api/production-canary":
                self._send(_cached("production-canary", 15, _production_canary_status))
            elif route == "/api/model-ledgers":
                self._send(_cached("model-ledgers", 30, model_ledger_comparison))
            elif route == "/api/picks":
                self._send(_cached("picks", 30, dashboard_picks))
            elif route == "/api/flat-picks":

                def _flat_picks_decorated():
                    # Flat is now split per sport (data/flat/<sport>.xlsx),
                    # populated only for the sports that actually pair with
                    # Main -- esports/KBO/NPB physically can't appear here
                    # anymore (see main_ledgers.py), so the old
                    # FLAT_HIDDEN_LEAGUES filter is redundant for them and
                    # was actively wrong for tennis (real flat rows exist for
                    # it, same as soccer, but the set never got updated when
                    # tennis was promoted alongside soccer on 2026-08-03).
                    flat = read_flat_picks()
                    orders = _load_orders()
                    portfolio = _load_portfolio_history()
                    return [_decorate_pick(row, orders, portfolio) for row in flat]

                self._send(_cached("flat-picks", 30, _flat_picks_decorated))
            elif route == "/api/performance":
                sport = str(query.get("sport") or "").strip()
                self._send(
                    _cached(
                        f"performance:{sport.casefold() or 'all'}",
                        30,
                        lambda: performance_for_sport(read_picks(), sport),
                    )
                )
            elif route == "/api/flat-performance":
                sport = str(query.get("sport") or "").strip()
                flat = read_flat_picks()
                self._send(
                    _cached(
                        f"flat-performance:{sport.casefold() or 'all'}",
                        30,
                        lambda: performance_for_sport(flat, sport),
                    )
                )
            elif route == "/api/research-performance":
                sport = str(query.get("sport") or "").strip()
                research = _parse_research_picks()
                self._send(
                    _cached(
                        f"research-performance:{sport.casefold() or 'all'}",
                        30,
                        lambda: performance_for_sport(research, sport),
                    )
                )
            elif route == "/api/research-picks":
                self._send(_cached("research-picks", 60, lambda: [_decorate_pick(r) for r in _parse_research_picks()]))
            elif route == "/api/gated-research-performance":
                sport = str(query.get("sport") or "").strip()
                self._send(
                    _cached(
                        f"gated-research-performance:{sport.casefold() or 'all'}",
                        60,
                        lambda: performance_for_sport(_parse_research_picks(gated=True), sport),
                    )
                )
            elif route == "/api/gated-research-picks":
                self._send(_cached("gated-research-picks", 60, lambda: [_decorate_pick(r) for r in _parse_research_picks(gated=True)]))
            elif route == "/api/backtests":
                self._send(_cached("backtests", 60, backtests))
            elif route == "/api/backtest":
                name = Path(query.get("file", "")).name
                path = OUTPUTS / name
                if path.exists() and path.suffix == ".json":
                    self._send(path.read_bytes())
                else:
                    self._send({"error": "not found"}, code=404)
            elif route == "/api/market":
                sport = _safe_sport(query.get("sport", "mlb"))
                day = query.get("date") or _today()
                self._send(_cached(f"market:{sport}:{day}", 60, lambda: market_snapshots(sport, day)))
            elif route == "/api/live":
                sport = _safe_sport(query.get("sport", "mlb"))
                day = query.get("date") or _today()
                self._send(_cached(f"live:{sport}:{day}", 120, lambda: live_gateway_slate(sport, day)))
            elif route == "/api/audit":
                self._send(_cached("audit", 60, _audit_tail))
            elif route == "/api/job":
                self._send(job_status(str(query.get("id", ""))))
            elif route == "/api/today":
                day = query.get("date") or _today()
                self._send(_cached(f"today:{day}", 20, lambda: today_picks(day)))
            elif route == "/api/odds":
                sport = query.get("sport")
                self._send(
                    _cached(f"odds:{sport or 'all'}", 30, lambda: odds_summary(sport if sport else None))
                )
            elif route == "/api/open":
                self._send(_cached("open", 15, open_picks))
            elif route == "/api/history":
                days = int(query.get("days", "30"))
                sport = query.get("sport")
                self._send(
                    _cached(f"history:{days}:{sport or 'all'}", 30, lambda: history_picks(days, sport))
                )
            elif route == "/api/bets":
                self._send(_cached("bets", 15, bets_view))
            elif route == "/api/orders":
                self._send(_load_orders())
            elif route == "/api/health":
                self._send({"ok": True, "at": datetime.now(timezone.utc).isoformat()[:19]})
            elif route.startswith("/api/rebuild/"):
                view = route.removeprefix("/api/rebuild/")
                if view in _REBUILD_VIEWS:
                    self._send(_cached(f"rebuild:{view}", 30, lambda: rebuild_view(view)))
                else:
                    self._send({"error": "unknown route"}, code=404)
            else:
                self._send({"error": "unknown route"}, code=404)
        except InvalidSportError as error:
            self._send({"error": str(error)}, code=400)
        except Exception as error:  # noqa: BLE001
            self._send({"error": f"{type(error).__name__}: {error}"}, code=500)

    def do_HEAD(self) -> None:  # noqa: N802
        route = urlparse(self.path).path
        if route.startswith("/api/data/"):
            self._send_head({"ok": True})
            return
        if route.startswith("/api/rebuild/"):
            view = route.removeprefix("/api/rebuild/")
            if view in _REBUILD_VIEWS:
                try:
                    self._send_head(_cached(f"rebuild:{view}", 30, lambda: rebuild_view(view)))
                except Exception as error:  # noqa: BLE001
                    self._send_head({"error": f"{type(error).__name__}: {error}"}, code=500)
                return
        self._send_head({"error": "unknown route"}, code=404)

    def _reject_rebuild_mutation(self) -> bool:
        if urlparse(self.path).path.startswith("/api/rebuild/"):
            self._send(
                {"error": "method not allowed", "allowed_methods": ["GET", "HEAD"]},
                code=405,
            )
            return True
        return False

    def _local_origin_ok(self) -> bool:
        """Reject cross-site (CSRF) POSTs: browser requests from any web page
        carry an Origin header; only same-host origins (or none — curl, CLI)
        are allowed to hit state-changing routes on this local server."""
        origin = str(self.headers.get("Origin") or "")
        if not origin:
            return True
        host = str(self.headers.get("Host") or "")
        return origin in (f"http://{host}", f"https://{host}") or origin.startswith(
            ("http://127.0.0.1:", "http://localhost:")
        )

    def do_POST(self) -> None:  # noqa: N802
        if self._reject_rebuild_mutation():
            return
        if not self._local_origin_ok():
            self._send({"status": "refused", "error": "cross-origin request rejected"}, code=403)
            return
        if not secrets.compare_digest(str(self.headers.get("X-Dashboard-Token") or ""), _DASHBOARD_TOKEN):
            self._send({"status": "refused", "error": "missing or invalid dashboard session token"}, code=401)
            return
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length") or 0)
        try:
            payload = json.loads(self.rfile.read(length) or b"{}")
        except json.JSONDecodeError:
            payload = {}
        if payload.get("confirm") is not True:
            self._send(
                {"status": "refused", "error": "confirmation required: resend with confirm=true"}, code=400
            )
            return
        if parsed.path == "/api/archive":
            action = str(payload.get("action"))
            scope = payload.get("pick_ids") if action == "clear_ids" else str(payload.get("scope", ""))
            self._send(archive_action(action, scope or []))
        elif parsed.path == "/api/dedupe":
            self._send(dedupe_ledger())
        elif parsed.path == "/api/model-ledgers/decision":
            self._send(record_model_ledger_decision(payload))
        elif parsed.path == "/api/action":
            self._send(start_action(str(payload.get("action")), payload))
        elif parsed.path == "/api/order/preview":
            self._send(preview_order(payload))
        elif parsed.path == "/api/order/preview-position":
            self._send(preview_position_sell(payload))
        elif parsed.path == "/api/order/submit-position":
            self._send(submit_position_sell(payload))
        elif parsed.path == "/api/order/submit":
            self._send(submit_order(payload))
        elif parsed.path == "/api/settings/unit-value":
            try:
                self._send(_set_unit_value_usd(payload.get("unit_value_usd")))
            except (OSError, RuntimeError, ValueError, yaml.YAMLError) as error:
                self._send({"status": "refused", "error": str(error)}, code=400)
        elif parsed.path == "/api/settings/auto-unit-value":
            pct = float(payload.get("pct", 10))
            self._send(_auto_adjust_unit_value(pct))
        else:
            self._send({"error": "unknown route"}, code=404)

    def do_PUT(self) -> None:  # noqa: N802
        if not self._reject_rebuild_mutation():
            self._send({"error": "unknown route"}, code=404)

    def do_PATCH(self) -> None:  # noqa: N802
        if not self._reject_rebuild_mutation():
            self._send({"error": "unknown route"}, code=404)

    def do_DELETE(self) -> None:  # noqa: N802
        if not self._reject_rebuild_mutation():
            self._send({"error": "unknown route"}, code=404)


def _auto_adjust_unit_value(pct: float = 10.0) -> dict:
    """Set 1U to pct% of the LIVE exchange USD balance, moving up or down.

    The exchange balance is the only honest bankroll number: revaluing
    historical unit P&L at the current unit value (the previous approach)
    was circular and compounded the unit on every win. When the live
    account is unreachable the unit value is left unchanged.
    """
    fraction = max(0.01, min(0.50, pct / 100.0))  # clamp 1-50%
    portfolio = live_portfolio_view()
    if portfolio.get("status") != "live":
        return {
            "status": "unavailable",
            "error": (
                "live exchange balance unreachable: "
                + str(portfolio.get("error") or "authentication required")
            ),
            "note": "Unit value unchanged; auto-sizing requires the real account balance.",
        }
    balance = _number((portfolio.get("balance") or {}).get("current_usd"), None)
    if balance is None or balance <= 0:
        return {
            "status": "unavailable",
            "error": "exchange returned no positive USD balance",
            "note": "Unit value unchanged.",
        }
    suggested = round(balance * fraction, 2)
    current = _unit_value_usd()
    if abs(suggested - current) < 0.01:
        return {
            "status": "no_change",
            "balance_usd": round(balance, 2),
            "current_unit": current,
            "suggested_unit": suggested,
            "note": f"{pct:.1f}% of ${balance:,.2f} = ${suggested:.2f}, already current.",
        }
    try:
        result = _set_unit_value_usd(suggested)
        result["balance_usd"] = round(balance, 2)
        result["action"] = "raised" if suggested > current else "lowered"
        return result
    except (OSError, RuntimeError, ValueError, yaml.YAMLError) as error:
        return {"status": "error", "error": str(error)}


def today_picks(day: str) -> dict:
    """Latest unique, locally visible picks played on a US-Eastern date."""
    rows = []
    archived = set(_load_archive()["pick_ids"])
    orders, portfolio_history = _load_orders(), _load_portfolio_history()
    for row in _dedupe_picks(read_picks()):
        if str(row.get("pick_id")) in archived:
            continue
        start = row.get("event_start_utc")
        if not start:
            continue
        try:
            start_dt = datetime.fromisoformat(str(start).replace("Z", "+00:00"))
        except ValueError:
            continue
        start_et = start_dt.astimezone(EASTERN)
        if start_et.date().isoformat() != day:
            continue
        rows.append(
            {
                **_decorate_pick(row, orders, portfolio_history),
                "start_et": start_et.strftime("%I:%M %p ET"),
                "start_sort": start_dt.isoformat(),
                "suggested_paper_units": _suggested_units(row),
            }
        )
    rows.sort(key=lambda r: (r["start_sort"], str(r.get("league")), str(r.get("market_type"))))
    return {
        "date": day,
        "picks": rows,
        "count": len(rows),
        "open": sum(1 for r in rows if r.get("status") == "open"),
        "settled": sum(1 for r in rows if r.get("status") == "settled"),
    }


def open_picks() -> dict:
    """All open ledger picks with model probability, market odds, and edge."""
    rows = []
    for row in read_picks():
        if row.get("status") != "open":
            continue
        model_p = _number(row.get("model_probability"))
        market_p = _number(row.get("market_implied_probability"))
        edge = model_p - market_p if model_p and market_p else None
        rows.append(
            {
                "pick_id": str(row.get("pick_id", "")),
                "league": str(row.get("league", "")),
                "away_team": str(row.get("away_team", "")),
                "home_team": str(row.get("home_team", "")),
                "selection": str(row.get("selection", "")),
                "market_type": str(row.get("market_type", "")),
                "event_start_utc": str(row.get("event_start_utc", "")),
                "model_probability": round(model_p, 4) if model_p else None,
                "market_implied_probability": round(market_p, 4) if market_p else None,
                "edge": round(edge, 4) if edge is not None else None,
                "american_odds": row.get("american_odds"),
                "units": _number(row.get("units")),
                "record_type": str(row.get("record_type", "")),
                "model_version": str(row.get("model_version", "")),
                "reason_code": str(row.get("reason_code", "")),
            }
        )
    # Sort: earliest game first
    rows.sort(key=lambda r: r["event_start_utc"])
    qualified = [r for r in rows if r["record_type"] == "QUALIFIED_SHADOW_CALL"]
    research = [r for r in rows if r["record_type"] == "RESEARCH_OBSERVATION"]
    return {
        "open": rows,
        "count": len(rows),
        "qualified_count": len(qualified),
        "research_count": len(research),
        "total_units": round(sum(r["units"] for r in qualified), 2),
    }


def history_picks(days: int = 30, sport: str | None = None) -> dict:
    """Settled picks within the last N days, optionally filtered by sport."""
    cutoff = datetime.now(timezone.utc)
    rows = []
    for row in read_picks():
        if row.get("status") != "settled":
            continue
        if sport and str(row.get("league", "")).lower() != sport.lower():
            continue
        settled_at = row.get("settled_at_utc")
        if settled_at:
            try:
                settled_dt = datetime.fromisoformat(str(settled_at).replace("Z", "+00:00"))
            except ValueError:
                continue
            if (cutoff - settled_dt).days > days:
                continue
        model_p = _number(row.get("model_probability"))
        market_p = _number(row.get("market_implied_probability"))
        rows.append(
            {
                "pick_id": str(row.get("pick_id", "")),
                "league": str(row.get("league", "")),
                "away_team": str(row.get("away_team", "")),
                "home_team": str(row.get("home_team", "")),
                "selection": str(row.get("selection", "")),
                "market_type": str(row.get("market_type", "")),
                "result": str(row.get("result", "")),
                "away_score": row.get("away_score"),
                "home_score": row.get("home_score"),
                "model_probability": round(model_p, 4) if model_p else None,
                "market_implied_probability": round(market_p, 4) if market_p else None,
                "pnl_units": _number(row.get("pnl_units")),
                "units": _number(row.get("units")) or _number(row.get("research_score_units")),
                "settled_at_utc": str(row.get("settled_at_utc", "")),
                "event_start_utc": str(row.get("event_start_utc", "")),
                "record_type": str(row.get("record_type", "")),
                "american_odds": row.get("american_odds"),
            }
        )
    rows.sort(key=lambda r: r["settled_at_utc"], reverse=True)
    wins = sum(1 for r in rows if r["result"] == "win")
    losses = sum(1 for r in rows if r["result"] == "loss")
    pushes = sum(1 for r in rows if r["result"] == "push")
    total_pnl = sum(r["pnl_units"] for r in rows)
    return {
        "history": rows,
        "count": len(rows),
        "wins": wins,
        "losses": losses,
        "pushes": pushes,
        "hit_rate": round(wins / (wins + losses), 4) if (wins + losses) else None,
        "total_pnl": round(total_pnl, 4),
        "days": days,
        "sport": sport,
    }


def _amount_value(value) -> float | None:
    if isinstance(value, dict):
        value = value.get("value")
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _live_model_links() -> dict[tuple[str, str], dict]:
    """Connect exchange contracts to model rows without treating picks as positions."""
    links: dict[tuple[str, str], dict] = {}
    all_rows = read_picks()
    rows_by_id = {str(row.get("pick_id") or ""): row for row in all_rows}

    def _link(row: dict, side: str) -> dict:
        quote = _pick_quote(row) or {}
        bid = _number(quote.get("bid"), None)
        ask = _number(quote.get("ask"), None)
        return {
            "pick_id": str(row.get("pick_id") or ""),
            "league": str(row.get("league") or ""),
            "away_team": str(row.get("away_team") or ""),
            "home_team": str(row.get("home_team") or ""),
            "selection": str(row.get("selection") or ""),
            "market_type": str(row.get("market_type") or ""),
            "model_probability": _number(row.get("model_probability"), None),
            "model_version": str(row.get("model_version") or ""),
            "side": side,
            "decision_price": ask,
            "decision_bid": bid,
            "decision_spread": (
                round(ask - bid, 6) if ask is not None and bid is not None else None
            ),
            "quote_observed_at_utc": quote.get("observed_at_utc"),
        }

    for row in _dedupe_picks(all_rows):
        quote = _pick_quote(row)
        if quote is None:
            continue
        side = str(quote["side"])
        links[(str(quote["market_slug"]), side)] = _link(row, side)
    # An exchange-acknowledged dashboard order links later fills back to the
    # model pick, including a partial fill followed by cancellation.
    for order in _load_orders()["orders"]:
        if order.get("status") not in {"submitted", "filled", "canceled", "replaced"}:
            continue
        row = rows_by_id.get(str(order.get("pick_id") or ""))
        slug = str(order.get("market_slug") or "")
        side = str(order.get("side") or "")
        if row is not None and slug and side:
            links[(slug, side)] = _link(row, side)
    return links


def _activity_outcome_side(payload: dict) -> str | None:
    """Return long/short only when an exchange activity states its outcome side."""
    for raw in (
        payload.get("outcomeSide"),
        payload.get("positionSide"),
        payload.get("intent"),
        payload.get("side"),
    ):
        value = str(raw or "").upper()
        if value.endswith("_SHORT") or value in {"SHORT", "NO"}:
            return "short"
        if value.endswith("_LONG") or value in {"LONG", "YES"}:
            return "long"
    return None


def _activity_link(
    slug: str,
    explicit_side: str | None,
    links: dict[tuple[str, str], dict],
) -> tuple[str | None, dict | None]:
    """Resolve a side, inferring it only when exactly one linked side exists."""
    if explicit_side:
        return explicit_side, links.get((slug, explicit_side))
    candidates = [(side, link) for (market_slug, side), link in links.items() if market_slug == slug]
    if len(candidates) == 1:
        return candidates[0]
    return None, None


def _selected_short_pnl(exchange_price: float | None, exchange_pnl: float | None) -> float | None:
    """Correct terminal synthetic-NO P&L without rewriting ordinary trade P&L."""
    if exchange_pnl is None:
        return None
    if exchange_price is not None and exchange_price <= 0.01:
        return abs(exchange_pnl)  # YES lost, so the held NO side won.
    if exchange_price is not None and exchange_price >= 0.99:
        return -abs(exchange_pnl)  # YES won, so the held NO side lost.
    return exchange_pnl


def _normalize_live_activity(item: dict, links: dict[tuple[str, str], dict]) -> dict | None:
    trade = item.get("trade") if isinstance(item.get("trade"), dict) else None
    resolution = item.get("positionResolution") if isinstance(item.get("positionResolution"), dict) else None
    if trade:
        slug = str(trade.get("marketSlug") or "")
        occurred = str(trade.get("updateTime") or trade.get("createTime") or "")
        outcome_side, linked = _activity_link(slug, _activity_outcome_side(trade), links)
        exchange_price = _amount_value(trade.get("price"))
        selected_price = exchange_price
        if exchange_price is not None and outcome_side == "short":
            selected_price = round(1 - exchange_price, 6)
        exchange_pnl = _amount_value(trade.get("realizedPnl"))
        selected_pnl = exchange_pnl
        if exchange_pnl is not None and outcome_side == "short":
            selected_pnl = _selected_short_pnl(exchange_price, exchange_pnl)
        return {
            "activity_id": f"trade:{trade.get('id') or slug + ':' + occurred}",
            "type": "trade",
            "market_slug": slug,
            "title": str((trade.get("marketMetadata") or {}).get("title") or slug),
            "occurred_at_utc": occurred,
            "price": selected_price,
            "exchange_price": exchange_price,
            "price_basis": ("selected_short_probability" if outcome_side == "short" else "long_probability"),
            "quantity": _number(trade.get("qtyDecimal") or trade.get("qty"), None),
            "cost_basis_usd": _amount_value(trade.get("costBasis")),
            "fee_usd": _amount_value(
                trade.get("fee")
                or trade.get("fees")
                or trade.get("feeAmount")
                or trade.get("feePaid")
            ),
            "realized_pnl_usd": selected_pnl,
            "exchange_realized_pnl_usd": exchange_pnl,
            "pnl_basis": (
                "terminal_short_outcome_adjustment" if selected_pnl != exchange_pnl else "exchange_reported"
            ),
            "state": str(trade.get("state") or ""),
            "is_aggressor": trade.get("isAggressor"),
            "outcome_side": outcome_side,
            "model_pick": linked,
        }
    if resolution:
        slug = str(resolution.get("marketSlug") or "")
        occurred = str(resolution.get("updateTime") or "")
        outcome_side = _activity_outcome_side(resolution)
        before = resolution.get("beforePosition") or {}
        after = resolution.get("afterPosition") or {}
        metadata = after.get("marketMetadata") or before.get("marketMetadata") or {}
        outcome_side, linked = _activity_link(slug, outcome_side, links)
        before_realized = _amount_value(before.get("realized"))
        after_realized = _amount_value(after.get("realized"))
        realized_delta = after_realized
        if before_realized is not None and after_realized is not None:
            realized_delta = round(after_realized - before_realized, 6)
        return {
            "activity_id": f"settlement:{resolution.get('tradeId') or slug + ':' + occurred}",
            "type": "settlement",
            "market_slug": slug,
            "title": str(metadata.get("title") or slug),
            "outcome": str(metadata.get("outcome") or ""),
            "occurred_at_utc": occurred,
            "resolution_side": str(resolution.get("side") or "").removeprefix("POSITION_RESOLUTION_SIDE_"),
            "outcome_side": outcome_side,
            "before_quantity": _number(before.get("netPositionDecimal") or before.get("netPosition"), None),
            "after_quantity": _number(after.get("netPositionDecimal") or after.get("netPosition"), None),
            "realized_pnl_usd": realized_delta,
            "cumulative_realized_pnl_usd": after_realized,
            "pnl_basis": "position_realized_delta",
            "model_pick": linked,
        }
    return None


def _load_portfolio_history() -> dict:
    payload = _read_json(PORTFOLIO_HISTORY_FILE) or {}
    activities = payload.get("activities") if isinstance(payload, dict) else None
    history_start = (
        str(payload.get("history_start_date") or _today()) if isinstance(payload, dict) else _today()
    )
    rows = [
        item
        for item in (list(activities) if isinstance(activities, list) else [])
        if _activity_on_or_after(item, history_start)
    ]
    return {
        "activities": rows,
        "last_synced_at_utc": payload.get("last_synced_at_utc") if isinstance(payload, dict) else None,
        "history_start_date": history_start,
    }


def _activity_on_or_after(item: dict, history_start: str) -> bool:
    try:
        occurred = datetime.fromisoformat(str(item.get("occurred_at_utc") or "").replace("Z", "+00:00"))
        return occurred.astimezone(EASTERN).date().isoformat() >= history_start
    except ValueError:
        return False


def _save_portfolio_history(activities: list[dict], observed_at: str) -> list[dict]:
    existing = _load_portfolio_history()
    prior = existing["activities"]
    history_start = existing["history_start_date"]
    merged = {
        str(item.get("activity_id")): item
        for item in [*prior, *activities]
        if item.get("activity_id") and _activity_on_or_after(item, history_start)
    }
    rows = sorted(merged.values(), key=lambda item: str(item.get("occurred_at_utc") or ""), reverse=True)[
        :2000
    ]
    DASH_DIR.mkdir(exist_ok=True)
    temporary = PORTFOLIO_HISTORY_FILE.with_suffix(".json.tmp")
    temporary.write_text(
        json.dumps(
            {
                "history_start_date": history_start,
                "last_synced_at_utc": observed_at,
                "activities": rows,
            },
            indent=2,
            default=str,
        )
        + "\n",
        encoding="utf-8",
    )
    os.replace(temporary, PORTFOLIO_HISTORY_FILE)
    return rows


def _team_name_index() -> dict[tuple[str, str], str]:
    def _build() -> dict[tuple[str, str], str]:
        registry = _read_json(DATA / "entities" / "teams.json") or {}
        index: dict[tuple[str, str], str] = {}
        for team in registry.get("teams") or []:
            league = str(team.get("league") or "").casefold()
            name = str(team.get("canonical_name") or "").strip()
            candidates = {
                str(team.get("abbreviation") or ""),
                str(team.get("canonical_team_id") or "").rsplit("-", 1)[-1],
            }
            for alias in team.get("aliases") or []:
                alias_name = str(alias.get("source_name") or "").strip()
                candidates.add(alias_name)
                words = re.findall(r"[A-Za-z0-9]+", alias_name)
                if len(words) > 1:
                    candidates.add("".join(word[0] for word in words))
            for candidate in candidates:
                token = re.sub(r"[^a-z0-9]", "", candidate.casefold())
                if token and name:
                    index.setdefault((league, token), name)
        return index

    return _cached("team-name-index", 300, _build)


def _public_market_question(slug: str) -> str | None:
    with _MARKET_QUESTION_LOCK:
        if slug in _MARKET_QUESTION_CACHE:
            return _MARKET_QUESTION_CACHE[slug]
    question: str | None = None
    try:
        request = urllib.request.Request(
            f"{GATEWAY}/v1/market/slug/{quote(slug, safe='-')}",
            headers={"User-Agent": "model-prediction-dashboard/2.0"},
        )
        with urllib.request.urlopen(request, timeout=5) as response:  # noqa: S310
            payload = json.loads(response.read())
        question = str((payload.get("market") or {}).get("question") or "").strip() or None
    except (OSError, TimeoutError, ValueError, json.JSONDecodeError):
        question = None
    with _MARKET_QUESTION_LOCK:
        _MARKET_QUESTION_CACHE[slug] = question
    return question


def _human_market_name(slug: str, title: str = "", *, allow_lookup: bool = True) -> str:
    """Turn an exchange identifier into a compact, readable market name."""
    match = re.match(
        r"^(?P<prefix>[a-z]+)-(?P<league>[a-z0-9]+)-(?P<away>[a-z0-9]+)-"
        r"(?P<home>[a-z0-9]+)-(?P<date>\d{4}-\d{2}-\d{2})(?:-(?P<detail>.*))?$",
        slug.casefold(),
    )
    if not match:
        return title if title and title != slug and title != "None" else slug

    league = match.group("league")
    names = _team_name_index()
    away = names.get((league, match.group("away")), match.group("away").upper())
    home = names.get((league, match.group("home")), match.group("home").upper())
    prefix = match.group("prefix")
    detail = match.group("detail") or ""
    league_label = league.upper()
    matchup = f"{away} @ {home}"

    # These prefixes hide materially different contracts behind similar
    # slugs (team total vs match total, first half vs full game, 1X2, props).
    # The exchange question is the canonical, specific market name.
    if allow_lookup and prefix in {"tsc", "atc"}:
        question = _public_market_question(slug)
        if question:
            return question.removesuffix("?")

    if prefix == "aec":
        market = "First 5 moneyline" if detail.startswith("f5") else "Moneyline"
        return f"{league_label} · {matchup} · {market}"
    if prefix in {"tsc", "asc", "atc"}:
        line_match = re.search(r"(\d+)pt(\d+)", detail)
        line = f"{line_match.group(1)}.{line_match.group(2)}" if line_match else ""
        period = "First 5 " if "f5" in detail else ""
        market = {"tsc": "Total", "atc": "Team total", "asc": "Spread"}[prefix]
        suffix = f" {line}" if line else ""
        return f"{league_label} · {matchup} · {period}{market}{suffix}"
    if prefix == "astatc":
        question = _public_market_question(slug)
        if question:
            clean = question.removesuffix("?")
            clean = re.sub(r"\s+in\s+[A-Z0-9 .'-]+\s+vs\.?\s+[A-Z0-9 .'-]+$", "", clean)
            hrr = re.fullmatch(r"Will (.+?) record at least (\d+) hits \+ runs \+ RBIs", clean)
            if hrr:
                clean = f"{hrr.group(1)} · {hrr.group(2)}+ hits + runs + RBIs"
            return f"{clean} · {matchup}"
        return f"{league_label} · {matchup} · Player prop"
    return title if title and title != slug and title != "None" else f"{league_label} · {matchup}"


def _portfolio_history_summary(
    activities: list[dict],
    source: str,
    links: dict[tuple[str, str], dict] | None = None,
) -> dict:
    links = _live_model_links() if links is None else links

    def side_adjust(item: dict) -> dict:
        if item.get("type") != "trade":
            return item
        slug = str(item.get("market_slug") or "")
        outcome_side, linked = _activity_link(slug, str(item.get("outcome_side") or "") or None, links)
        if outcome_side != "short":
            return {**item, "outcome_side": outcome_side, "model_pick": linked}
        exchange_price = _number(item.get("exchange_price"), None)
        if exchange_price is None:
            stored_price = _number(item.get("price"), None)
            if stored_price is not None:
                exchange_price = (
                    1 - stored_price
                    if item.get("price_basis") == "selected_short_probability"
                    else stored_price
                )
        exchange_pnl = _number(item.get("exchange_realized_pnl_usd"), None)
        if exchange_pnl is None:
            stored_pnl = _number(item.get("realized_pnl_usd"), None)
            if stored_pnl is not None:
                exchange_pnl = (
                    -stored_pnl
                    if item.get("pnl_basis")
                    in {
                        "selected_short_inverse_of_exchange_long",
                        "terminal_short_outcome_adjustment",
                    }
                    else stored_pnl
                )
        return {
            **item,
            "price": round(1 - exchange_price, 6) if exchange_price is not None else None,
            "exchange_price": exchange_price,
            "price_basis": "selected_short_probability",
            "realized_pnl_usd": _selected_short_pnl(exchange_price, exchange_pnl),
            "exchange_realized_pnl_usd": exchange_pnl,
            "pnl_basis": (
                "terminal_short_outcome_adjustment"
                if _selected_short_pnl(exchange_price, exchange_pnl) != exchange_pnl
                else item.get("pnl_basis")
            ),
            "outcome_side": outcome_side,
            "model_pick": linked,
        }

    decorated = [
        {
            **side_adjust(item),
            "market_name": _human_market_name(
                str(item.get("market_slug") or ""), str(item.get("title") or "")
            ),
        }
        for item in activities
    ]
    trades = [item for item in decorated if item.get("type") == "trade"]
    settlements = [item for item in decorated if item.get("type") == "settlement"]
    realized = sum(
        value for item in decorated if (value := _number(item.get("realized_pnl_usd"), None)) is not None
    )
    return {
        "activities": decorated,
        "count": len(decorated),
        "trade_count": len(trades),
        "settlement_count": len(settlements),
        "realized_pnl_usd": round(realized, 2),
        "source": source,
    }


def live_portfolio_view() -> dict:
    """Exchange-confirmed positions and activity; model picks never count as exposure."""
    cached = _load_portfolio_history()
    empty_open = {
        "positions": [],
        "count": 0,
        "cost_basis_usd": 0.0,
        "cash_value_usd": 0.0,
        "realized_pnl_usd": 0.0,
    }
    missing = [name for name in ("POLYMARKET_KEY_ID", "POLYMARKET_SECRET_KEY") if not os.environ.get(name)]
    if missing:
        return {
            "status": "unavailable",
            "error": f"missing {' and '.join(missing)}",
            "open": empty_open,
            "recent_history": _portfolio_history_summary(cached["activities"], "cached"),
            "last_synced_at_utc": cached["last_synced_at_utc"],
            "history_start_date": cached["history_start_date"],
        }
    try:
        process = subprocess.run(
            _resolve_runner() + ["live-portfolio"],
            cwd=ROOT,
            capture_output=True,
            text=True,
            timeout=30,
            env=_runner_env(),
        )
        raw_text = process.stdout if process.returncode == 0 else process.stderr
        raw = json.loads(raw_text)
        if process.returncode != 0 or raw.get("status") != "live":
            raise RuntimeError(str(raw.get("error") or "authenticated portfolio request failed"))
    except (OSError, subprocess.TimeoutExpired, json.JSONDecodeError, RuntimeError) as error:
        return {
            "status": "unavailable",
            "error": str(error)[:300],
            "open": empty_open,
            "recent_history": _portfolio_history_summary(cached["activities"], "cached"),
            "last_synced_at_utc": cached["last_synced_at_utc"],
            "history_start_date": cached["history_start_date"],
        }

    links = _live_model_links()
    positions = []
    for slug, item in (raw.get("positions") or {}).items():
        net = _number(item.get("netPositionDecimal") or item.get("netPosition"), 0.0)
        if abs(net) < 1e-9:
            continue
        side = "long" if net > 0 else "short"
        metadata = item.get("marketMetadata") or {}
        cost = _amount_value(item.get("cost"))
        cash_value = _amount_value(item.get("cashValue"))
        quote = _live_bbo(str(slug)) or {}
        side_quote = quote.get(side) or {}
        bid = _number(side_quote.get("bid"), None)
        ask = _number(side_quote.get("ask"), None)
        mark = cash_value / abs(net) if cash_value is not None and abs(net) > 0 else None
        exit_default = (
            min(0.99, round(float(bid) + 0.01, 2))
            if bid is not None
            else min(0.99, max(0.01, round(float(mark or 0.5), 2)))
        )
        positions.append(
            {
                "market_slug": str(slug),
                "title": str(metadata.get("title") or slug),
                "market_name": _human_market_name(str(slug), str(metadata.get("title") or "")),
                "outcome": str(metadata.get("outcome") or ""),
                "side": side,
                "quantity": abs(net),
                "available_quantity": abs(
                    _number(item.get("qtyAvailableDecimal") or item.get("qtyAvailable"), 0.0)
                ),
                "cost_basis_usd": cost,
                "cash_value_usd": cash_value,
                "bid": bid,
                "ask": ask,
                "exit_limit_default": exit_default,
                "realized_pnl_usd": _amount_value(item.get("realized")),
                "unrealized_pnl_usd": (
                    round(cash_value - cost, 2) if cash_value is not None and cost is not None else None
                ),
                "expired": bool(item.get("expired")),
                "updated_at_utc": str(item.get("updateTime") or ""),
                "model_pick": links.get((str(slug), side)),
            }
        )
    positions.sort(key=lambda item: item["updated_at_utc"], reverse=True)
    normalized = [
        activity
        for item in (raw.get("activities") or [])
        if (activity := _normalize_live_activity(item, links)) is not None
    ]
    history = _save_portfolio_history(normalized, str(raw.get("observed_at_utc") or ""))
    balances = raw.get("balances") or []
    usd = next((item for item in balances if item.get("currency") == "USD"), None)
    return {
        "status": "live",
        "source": raw.get("source"),
        "observed_at_utc": raw.get("observed_at_utc"),
        "history_start_date": _load_portfolio_history()["history_start_date"],
        "open": {
            "positions": positions,
            "count": len(positions),
            "cost_basis_usd": round(sum(_number(item.get("cost_basis_usd")) for item in positions), 2),
            "cash_value_usd": round(sum(_number(item.get("cash_value_usd")) for item in positions), 2),
            "realized_pnl_usd": round(sum(_number(item.get("realized_pnl_usd")) for item in positions), 2),
        },
        "recent_history": _portfolio_history_summary(history, "exchange_and_persisted", links),
        # Every other USD amount this same authenticated API returns (trade
        # price/costBasis/realizedPnl/fee, position cost/cashValue/realized)
        # arrives as a {"value": ..., "currency": ...} envelope and is parsed
        # with _amount_value(), never bare _number() -- these four balance
        # fields were the one place still using _number(), which returns the
        # `default` (None here) on a dict instead of raising, so a real
        # envelope-shaped balance response would have silently rendered every
        # balance figure (and _auto_adjust_unit_value's bankroll-percent
        # sizing, which reads current_usd) as unavailable with no error
        # surfaced.
        # _amount_value() unwraps the envelope when present and still handles
        # a bare number, so this is a strict superset, not a behavior change,
        # for whichever shape the endpoint actually returns.
        "balance": {
            "current_usd": _amount_value((usd or {}).get("currentBalance")),
            "buying_power_usd": _amount_value((usd or {}).get("buyingPower")),
            "open_orders_usd": _amount_value((usd or {}).get("openOrders")),
            "unsettled_funds_usd": _amount_value((usd or {}).get("unsettledFunds")),
        },
    }


def bets_view() -> dict:
    """Backward-compatible route name for the authenticated live portfolio."""
    return live_portfolio_view()


def _model_version_rank(row: dict) -> tuple:
    """Sort key for choosing which duplicate to KEEP. Higher = keep.

    Prefers the numerically-newest model version (v3 > v2), then the most
    recently created row. Production models always outrank older ones.
    """
    version = str(row.get("model_version") or "")
    digits = "".join(ch for ch in version.split("-")[-1] if ch.isdigit())
    version_number = int(digits) if digits else 0
    return (version_number, str(row.get("created_at_utc") or ""))


def dedupe_ledger() -> dict:
    """Remove duplicate OPEN ledger rows through the audited ledger path.

    A duplicate = same contract identity (league/event/market/selection/line)
    logged under more than one model version or run. Keeps one row per
    identity — the newest model version — and removes the rest via
    ``PickLedger.remove_open_rows`` (ledger lock + ``pick_removed`` audit
    events). Settled rows are results and are never touched; staked open rows
    are never deleted. Every per-sport Main file (data/main/<sport>.xlsx)
    that exists is backed up first.
    """
    from model_prediction.main_ledgers import MultiSportPickLedger  # local: heavy import

    existing_main_paths = [path for path in _main_ledger_paths() if path.exists()]
    if not existing_main_paths:
        return {"status": "refused", "error": "no Main ledger files found under data/main/"}
    ledger = MultiSportPickLedger(DATA)
    try:
        rows = ledger.rows()
    except ValueError as error:
        return {"status": "refused", "error": str(error)}
    groups: dict[tuple, list[dict]] = {}
    for row in rows:
        if row.get("status") != "open":
            continue
        groups.setdefault(_pick_identity(row), []).append(row)
    to_remove: list[str] = []
    for members in groups.values():
        if len(members) == 1:
            continue
        unstaked = [m for m in members if _number(m.get("units")) <= 0]
        if not unstaked:
            continue
        survivor = max(unstaked, key=_model_version_rank)
        to_remove.extend(str(m.get("pick_id") or "") for m in unstaked if m is not survivor)
    if not to_remove:
        return {"status": "ok", "removed": 0, "kept": len(rows), "note": "No open duplicate contracts found."}
    import shutil

    stamp = int(time.time())
    backups = []
    for path in existing_main_paths:
        backup = path.with_suffix(f".xlsx.dedupe-bak-{stamp}")
        shutil.copy2(path, backup)
        backups.append(backup.name)
    removed_ids = ledger.remove_open_rows(to_remove, reason="dashboard duplicate-contract dedupe")
    # Prune archived ids that no longer exist so the counter stays honest.
    surviving = {str(r.get("pick_id")) for r in ledger.rows()}
    archive = _load_archive()
    archive["pick_ids"] = sorted(pid for pid in archive["pick_ids"] if pid in surviving)
    archive["history"].append(
        {"at": datetime.now(timezone.utc).isoformat()[:19], "action": "dedupe", "rows": len(removed_ids)}
    )
    _save_archive(archive)
    with _CACHE_LOCK:
        _CACHE.clear()
        _PICKS_CACHE["mtime"] = None
    _log(f"dedupe: removed {len(removed_ids)} open duplicate rows, backups {', '.join(backups)}")
    return {
        "status": "ok",
        "removed": len(removed_ids),
        "kept": len(surviving),
        "backups": backups,
        "removed_pick_ids": removed_ids[:50],
        "note": f"Removed {len(removed_ids)} open duplicates via the audited ledger path. Backups: {', '.join(backups)}.",
    }


def _load_archive() -> dict:
    try:
        payload = json.loads(ARCHIVE_FILE.read_text(encoding="utf-8"))
        return {"pick_ids": list(payload.get("pick_ids", [])), "history": list(payload.get("history", []))}
    except (OSError, json.JSONDecodeError):
        return {"pick_ids": [], "history": []}


def _save_archive(archive: dict) -> None:
    DASH_DIR.mkdir(exist_ok=True)
    ARCHIVE_FILE.write_text(json.dumps(archive, indent=1), encoding="utf-8")


def archive_action(action: str, scope: str) -> dict:
    """Persistently hide safe ledger rows from the dashboard table.

    picks.xlsx is never touched: archived rows keep feeding performance,
    calibration, backtests, and research. This is a display ledger-clear,
    not a data delete. Open rows with positive units are never archived.
    """
    archive = _load_archive()
    if action == "restore":
        restored = len(archive["pick_ids"])
        archive["pick_ids"] = []
        archive["history"].append(
            {"at": datetime.now(timezone.utc).isoformat()[:19], "action": "restore", "rows": restored}
        )
        _save_archive(archive)
        with _CACHE_LOCK:
            _CACHE.clear()
        return {"status": "ok", "action": "restore", "restored": restored}
    if action == "clear_ids":
        requested = {str(pick_id) for pick_id in scope if str(pick_id)}
        if not requested:
            return {"status": "refused", "error": "no pick ids supplied"}
        # A visible ledger/Today row is DEDUPED across model versions, so its one
        # pick_id stands in for every sibling row sharing the same contract
        # identity. Expand each requested id to its whole identity group,
        # otherwise dedup resurrects the row from an un-archived sibling.
        all_rows = read_picks()
        identity_to_ids: dict[tuple, set[str]] = {}
        id_to_identity: dict[str, tuple] = {}
        for row in all_rows:
            pid = str(row.get("pick_id") or "")
            if not pid:
                continue
            identity = _pick_identity(row)
            identity_to_ids.setdefault(identity, set()).add(pid)
            id_to_identity[pid] = identity
        expanded: set[str] = set()
        for pid in requested:
            identity = id_to_identity.get(pid)
            expanded |= identity_to_ids.get(identity, {pid}) if identity else {pid}
        exposed = {
            str(row.get("pick_id"))
            for row in all_rows
            if row.get("status") == "open"
            and row.get("record_type") == "QUALIFIED_SHADOW_CALL"
            and float(row.get("units") or 0) > 0
        }
        blocked = sorted(expanded & exposed)
        allowed = expanded - exposed
        existing = set(archive["pick_ids"]) | allowed
        archive["pick_ids"] = sorted(existing)
        archive["history"].append(
            {"at": datetime.now(timezone.utc).isoformat()[:19], "action": "clear_ids", "rows": len(allowed)}
        )
        _save_archive(archive)
        with _CACHE_LOCK:
            _CACHE.clear()
        return {
            "status": "ok",
            "action": "clear_ids",
            "archived_now": len(allowed),
            "rows_selected": len(requested),
            "blocked_open_staked": blocked,
            "archived_total": len(existing),
            "note": "View-only: rows remain in picks.xlsx and keep feeding research metrics.",
        }
    if action != "clear" or scope not in ("day", "week", "month", "all"):
        return {
            "status": "refused",
            "error": "action must be clear(day|week|month|all), clear_ids, or restore",
        }
    today = datetime.now(timezone.utc).astimezone(EASTERN).date()
    days = {"day": 0, "week": 6, "month": 29}.get(scope)
    existing = set(archive["pick_ids"])
    added = protected = 0
    for row in read_picks():
        if (
            row.get("status") == "open"
            and row.get("record_type") == "QUALIFIED_SHADOW_CALL"
            and float(row.get("units") or 0) > 0
        ):
            protected += 1
            continue
        pick_id = str(row.get("pick_id"))
        if pick_id in existing:
            continue
        if days is not None:
            start = str(row.get("event_start_utc") or "")
            try:
                game_day = datetime.fromisoformat(start.replace("Z", "+00:00")).astimezone(EASTERN).date()
            except ValueError:
                continue
            if (today - game_day).days > days or game_day > today:
                continue
        existing.add(pick_id)
        added += 1
    archive["pick_ids"] = sorted(existing)
    archive["history"].append(
        {"at": datetime.now(timezone.utc).isoformat()[:19], "action": f"clear:{scope}", "rows": added}
    )
    _save_archive(archive)
    with _CACHE_LOCK:
        _CACHE.clear()
    return {
        "status": "ok",
        "action": f"clear:{scope}",
        "archived_now": added,
        "protected_open_staked": protected,
        "archived_total": len(existing),
        "note": "View-only: all rows remain in picks.xlsx and keep feeding research metrics.",
    }


def _suggested_units(row: dict) -> float | None:
    """Decision-time model size, reconstructed for both open and settled rows.

    (0.5U base + |p-0.5| * 10, capped at 2.0U, nearest 0.25U — the sizing that
    beat flat staking +34.1U vs +13.3U on the MLB walk-forward.) The immutable
    decision probability lets the dashboard retain the same displayed size
    after settlement. Every actual ledger stake stays 0 until a model is
    promoted past research.
    """
    try:
        p = float(row.get("model_probability") or 0)
        market = float(row.get("market_implied_probability") or 0)
    except (TypeError, ValueError):
        return None
    if not (0 < p < 1) or not (0 < market < 1):
        return None
    # Edge-scaled from model confidence, including negative-edge research rows,
    # so the sizing shown before and after settlement remains identical. The
    # +EV badge carries the tail/no-tail context.
    raw = 0.5 + abs(p - 0.5) * (2.0 - 0.5) / 0.15
    units = max(0.5, min(2.0, raw))
    return round(units / 0.25) * 0.25


def _audit_tail() -> dict:
    path = DATA / "events.jsonl"
    events = []
    if path.exists():
        with path.open(encoding="utf-8") as handle:
            lines = [line for line in handle if line.strip()]
        for line in lines[-25:]:
            try:
                item = json.loads(line)
                events.append(
                    {
                        "at": str(item.get("occurred_at_utc", ""))[:19],
                        "type": item.get("event_type"),
                        "subject": str(item.get("subject_id", ""))[:24],
                    }
                )
            except json.JSONDecodeError:
                continue
        return {"total_events": len(lines), "tail": list(reversed(events))}
    return {"total_events": 0, "tail": []}


# ── SECTION: Main Entry Point ───────────────────────────────────────


def main() -> None:
    arguments = argparse.ArgumentParser()
    arguments.add_argument("--port", type=int, default=DASHBOARD_PORT)
    options = arguments.parse_args()

    DASH_DIR.mkdir(exist_ok=True)
    _hydrate_jobs()
    server = None
    my_pid = os.getpid()
    try:
        ThreadingHTTPServer.daemon_threads = True
        ThreadingHTTPServer.allow_reuse_address = True
        server = ThreadingHTTPServer(("127.0.0.1", options.port), Handler)
        PID_FILE.write_text(str(my_pid))  # Write only after successful bind
        print(f"dashboard: http://127.0.0.1:{options.port}/  (Ctrl-C to stop)")
        print(f"dashboard: session token (for direct API calls): {_DASHBOARD_TOKEN}")
        server.serve_forever()
    except OSError as exc:
        if exc.errno == 48:
            print(f"dashboard: port {options.port} busy — is another instance running?", file=sys.stderr)
        else:
            raise
    except KeyboardInterrupt:
        print("\ndashboard stopped")
    finally:
        if server is not None:
            server.server_close()
        if PID_FILE.exists():
            try:
                if PID_FILE.read_text().strip() == str(my_pid):
                    PID_FILE.unlink()
            except OSError:
                pass


if __name__ == "__main__":
    main()
