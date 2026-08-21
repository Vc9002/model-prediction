from __future__ import annotations

import os
import tempfile
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SHEET_NAME = "Picks"
TABLE_NAME = "PicksLedger"

_INTEGER_FIELDS = {
    "american_odds",
    "away_score",
    "closing_american_odds",
    "confidence_score",
    "decision_american_odds",
    "home_score",
    "ledger_schema_version",
}
_GENERAL_NUMBER_FIELDS = {
    "closing_consensus_line",
    "closing_line",
    "decision_consensus_line",
    "decision_line",
    "line",
}
_SIX_DECIMAL_FIELDS = {
    "calibrated_probability",
    "closing_consensus_probability",
    "closing_decimal_odds",
    "closing_implied_probability",
    "closing_no_vig_probability",
    "closing_raw_implied_probability",
    "decision_consensus_probability",
    "decision_decimal_odds",
    "decision_no_vig_probability",
    "decision_raw_implied_probability",
    "decimal_odds",
    "edge",
    "market_implied_probability",
    "model_probability",
    "model_uncertainty",
    "probability_clv",
}
_TWO_DECIMAL_FIELDS = {"legacy_units", "units"}
_FOUR_DECIMAL_FIELDS = {"pnl_units", "research_score_units"}
_SIX_DECIMAL_RESULT_FIELDS = {"research_pnl_units"}
_PERCENT_FIELDS = {
    "closing_consensus_probability",
    "closing_implied_probability",
    "closing_no_vig_probability",
    "closing_raw_implied_probability",
    "decision_consensus_probability",
    "decision_no_vig_probability",
    "decision_raw_implied_probability",
    "edge",
    "market_implied_probability",
    "model_probability",
    "model_uncertainty",
    "probability_clv",
}
_NARRATIVE_FIELDS = {
    "corrective_action",
    "loss_cause",
    "rationale",
    "research_scoring_note",
    "risks",
}
_TIMESTAMP_FIELDS = {
    "created_at_utc",
    "event_start_utc",
    "observed_at_utc",
    "research_scored_at_utc",
    "settled_at_utc",
}


def read_xlsx_rows(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Excel ledger is missing required worksheet {SHEET_NAME!r}")
        sheet = workbook[SHEET_NAME]
        values = sheet.iter_rows(values_only=True)
        first = next(values, None)
        if first is None:
            return [], []
        headers = ["" if value is None else str(value) for value in first]
        rows: list[dict[str, str]] = []
        for values_row in values:
            if all(value is None for value in values_row):
                continue
            rows.append(
                {
                    header: _cell_to_text(value, header)
                    for header, value in zip(headers, values_row, strict=False)
                }
            )
        return headers, rows
    finally:
        workbook.close()


def write_xlsx_rows_atomic(
    path: str | Path,
    fieldnames: list[str],
    rows: Iterable[dict[str, str]],
) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    materialized = list(rows)
    workbook = _build_workbook(fieldnames, materialized)
    descriptor, temporary = tempfile.mkstemp(prefix="picks-", suffix=".xlsx", dir=destination.parent)
    os.close(descriptor)
    try:
        workbook.save(temporary)
        with open(temporary, "rb") as handle:
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    except Exception:
        if os.path.exists(temporary):
            os.unlink(temporary)
        raise
    finally:
        workbook.close()


def _build_workbook(fieldnames: list[str], rows: list[dict[str, str]]) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    assert summary is not None, "a freshly constructed Workbook always has an active sheet"
    summary.title = "Summary"
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="1F2937")

    for column_index, field in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=column_index, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(field)
    sheet.row_dimensions[1].height = 32

    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            text = row.get(field, "") or ""
            cell.value = _excel_value(text, field)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal="left"
                if field in _NARRATIVE_FIELDS
                else "right"
                if _is_numeric(field)
                else "left",
                vertical="top",
                wrap_text=field in _NARRATIVE_FIELDS,
            )
            cell.number_format = _number_format(field)
        sheet.row_dimensions[row_index].height = 36 if any(row.get(f) for f in _NARRATIVE_FIELDS) else 18

    last_column = get_column_letter(len(fieldnames))
    last_row = max(1, len(rows) + 1)
    if rows:
        table = Table(displayName=TABLE_NAME, ref=f"A1:{last_column}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.auto_filter.ref = f"A1:{last_column}1"
    sheet.print_title_rows = "1:1"
    _build_summary(summary, fieldnames, len(rows))
    workbook.active = 0
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.title = "Model Prediction Picks Ledger"
    workbook.properties.subject = "Shadow-only sports forecasting research ledger"
    workbook.properties.creator = "Model Prediction"
    return workbook


def _build_summary(sheet, fieldnames: list[str], row_count: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "MODEL PREDICTION — PICKS LEDGER"
    sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34

    controls = [("J1", "METRIC CONTROL"), ("J2", "Minimum calibration sample"), ("L2", 30)]
    for coordinate, value in controls:
        sheet[coordinate] = value
    sheet["J1"].fill = PatternFill("solid", fgColor="D9EAF7")
    sheet["J1"].font = Font(name="Aptos", size=10, bold=True, color="17365D")
    sheet["J2"].font = Font(name="Aptos", size=10, color="374151")
    sheet["L2"].font = Font(name="Aptos", size=10, bold=True, color="17365D")
    sheet["L2"].number_format = "0"

    last_row = max(2, row_count + 1)

    def rng(field: str) -> str:
        column = get_column_letter(fieldnames.index(field) + 1)
        return f"'Picks'!${column}$2:${column}${last_row}"

    status = rng("status")
    result = rng("result")
    record_type = rng("record_type")
    market_type = rng("market_type")
    research_units = rng("research_score_units")
    research_pnl = rng("research_pnl_units")
    qualified_pnl = rng("pnl_units")
    clv = rng("probability_clv")
    pick_ids = rng("pick_id")

    labels_and_formulas = {
        "A3": ("Total records", f"=COUNTA({pick_ids})"),
        "C3": ("Open records", f'=COUNTIF({status},"open")'),
        "E3": ("Settled records", f'=COUNTIF({status},"settled")'),
        "G3": (
            "Calibration status",
            f'=IF(COUNTIFS({status},"settled",{result},"win")+COUNTIFS({status},"settled",{result},"loss")>=$L$2,"Ready","Insufficient sample")',
        ),
        "A6": ("Research scored", f'=COUNTIF({research_units},">0")'),
        "C6": (
            "Research record",
            f'=COUNTIFS({research_units},">0",{result},"win")&"-"&COUNTIFS({research_units},">0",{result},"loss")',
        ),
        "E6": (
            "Research accuracy",
            f'=IFERROR(COUNTIFS({research_units},">0",{result},"win")/(COUNTIFS({research_units},">0",{result},"win")+COUNTIFS({research_units},">0",{result},"loss")),0)',
        ),
        "G6": ("Research ROI", f"=IFERROR(SUM({research_pnl})/SUM({research_units}),0)"),
        "A9": ("Research P&L", f"=SUM({research_pnl})"),
        "C9": ("Mean raw-probability CLV", f'=IFERROR(AVERAGEIF({clv},"<>",{clv}),0)'),
        "E9": (
            "Qualified calls",
            f'=COUNTIFS({record_type},"QUALIFIED_SHADOW_CALL",{status},"settled")',
        ),
        "G9": ("Qualified P&L", f'=SUMIFS({qualified_pnl},{record_type},"QUALIFIED_SHADOW_CALL")'),
    }
    for anchor, (label, formula) in labels_and_formulas.items():
        label_cell = sheet[anchor]
        value_cell = sheet.cell(row=label_cell.row + 1, column=label_cell.column)
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        label_cell.font = Font(name="Aptos", size=10, bold=True, color="17365D")
        label_cell.alignment = Alignment(vertical="center")
        value_cell.value = formula
        value_cell.fill = PatternFill("solid", fgColor="F7FAFC")
        value_cell.font = Font(name="Aptos Display", size=15, bold=True, color="111827")
        value_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[label_cell.row].height = 20
        sheet.row_dimensions[value_cell.row].height = 28

    sheet["E7"].number_format = "0.0%"
    sheet["G7"].number_format = "0.0%"
    sheet["A10"].number_format = "0.0000"
    sheet["C10"].number_format = "0.00%"
    sheet["G10"].number_format = "0.0000"

    sheet.merge_cells("A12:G12")
    sheet["A12"] = "SETTLED ACCURACY BY MARKET"
    sheet["A12"].fill = PatternFill("solid", fgColor="17365D")
    sheet["A12"].font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    headers = ["Market", "Settled", "Wins", "Losses", "Accuracy", "Research P&L", "Mean CLV"]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=13, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(name="Aptos", size=10, bold=True, color="17365D")
        cell.alignment = Alignment(horizontal="center")
    for row_index, market in enumerate(("moneyline", "spread", "total"), start=14):
        sheet.cell(row=row_index, column=1, value=market)
        sheet.cell(
            row=row_index,
            column=2,
            value=f'=COUNTIFS({market_type},$A{row_index},{status},"settled")',
        )
        sheet.cell(
            row=row_index,
            column=3,
            value=f'=COUNTIFS({market_type},$A{row_index},{status},"settled",{result},"win")',
        )
        sheet.cell(
            row=row_index,
            column=4,
            value=f'=COUNTIFS({market_type},$A{row_index},{status},"settled",{result},"loss")',
        )
        sheet.cell(row=row_index, column=5, value=f"=IFERROR(C{row_index}/(C{row_index}+D{row_index}),0)")
        sheet.cell(
            row=row_index,
            column=6,
            value=f"=SUMIFS({research_pnl},{market_type},$A{row_index})",
        )
        sheet.cell(
            row=row_index,
            column=7,
            value=f'=IFERROR(AVERAGEIFS({clv},{market_type},$A{row_index},{status},"settled"),0)',
        )
        for column in range(1, 8):
            cell = sheet.cell(row=row_index, column=column)
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_index % 2 == 0 else "F3F6F9")
            cell.alignment = Alignment(horizontal="right" if column > 1 else "left")
        sheet.cell(row=row_index, column=5).number_format = "0.0%"
        sheet.cell(row=row_index, column=6).number_format = "0.0000"
        sheet.cell(row=row_index, column=7).number_format = "0.00%"

    sheet.merge_cells("A18:L19")
    sheet["A18"] = (
        "Descriptive only—not evidence of profitability. Calibration needs 30 settled binary picks."
    )
    sheet["A18"].fill = PatternFill("solid", fgColor="FFF4CE")
    sheet["A18"].font = Font(name="Aptos", size=10, italic=True, color="7A4E00")
    sheet["A18"].alignment = Alignment(wrap_text=True, vertical="center")

    sheet.print_area = "A1:L19"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    for column in range(1, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 21
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["J"].width = 27
    sheet.column_dimensions["K"].width = 4
    sheet.column_dimensions["L"].width = 12


def _excel_value(text: str, field: str) -> str | int | float | None:
    if text == "":
        return None
    if field in _INTEGER_FIELDS:
        return int(float(text))
    if _is_numeric(field):
        return float(text)
    return text


def _as_float(value: object) -> float:
    if isinstance(value, (int, float, str)):
        return float(value)
    raise TypeError(f"cannot convert {value!r} to float")


def _cell_to_text(value: object, field: str) -> str:
    if value is None:
        return ""
    if field in _INTEGER_FIELDS:
        return str(int(_as_float(value)))
    if field in _GENERAL_NUMBER_FIELDS:
        return format(_as_float(value), ".15g")
    if field in _SIX_DECIMAL_FIELDS:
        return f"{_as_float(value):.6f}"
    if field in _TWO_DECIMAL_FIELDS:
        return f"{_as_float(value):.2f}"
    if field in _FOUR_DECIMAL_FIELDS:
        return f"{_as_float(value):.4f}"
    if field in _SIX_DECIMAL_RESULT_FIELDS:
        return f"{_as_float(value):.6f}"
    return str(value)


def _is_numeric(field: str) -> bool:
    return field in (
        _INTEGER_FIELDS
        | _GENERAL_NUMBER_FIELDS
        | _SIX_DECIMAL_FIELDS
        | _TWO_DECIMAL_FIELDS
        | _FOUR_DECIMAL_FIELDS
        | _SIX_DECIMAL_RESULT_FIELDS
    )


def _number_format(field: str) -> str:
    if field in _PERCENT_FIELDS:
        return "0.00%"
    if field in _INTEGER_FIELDS:
        return "0"
    if field in _GENERAL_NUMBER_FIELDS:
        return "0.0###"
    if field in _SIX_DECIMAL_FIELDS:
        return "0.000000"
    if field in _TWO_DECIMAL_FIELDS:
        return "0.00"
    if field in _FOUR_DECIMAL_FIELDS:
        return "0.0000"
    if field in _SIX_DECIMAL_RESULT_FIELDS:
        return "0.000000"
    if field in _TIMESTAMP_FIELDS:
        return FORMAT_TEXT
    return FORMAT_TEXT


def _column_width(field: str) -> float:
    if field in _NARRATIVE_FIELDS:
        return 48
    if field in _TIMESTAMP_FIELDS:
        return 25
    if "hash" in field:
        return 20
    if field in {"away_team", "home_team", "sportsbook", "model_version", "reason_code"}:
        return 24
    if field.endswith("_name") or field.startswith("original_"):
        return 22
    if field.endswith("_id") or field in {"pick_id", "event_id"}:
        return 18
    if _is_numeric(field):
        return 15
    return min(22, max(12, len(field) + 2))
