"""Dashboard evidence module (extracted from dashboard_server.py, DD-5).

Part of the dashboard_server.py -> dashboard/ package split. See MASTER.md
DD-5 and dashboard/__init__.py for the re-export shim that keeps the old
`dashboard_server` import path and test-suite symbol references working.
"""

from __future__ import annotations

import hashlib
import json
import math
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]


from model_prediction.dashboard.common import (
    _CACHE,
    _CACHE_LOCK,
    CONFIG_FILE,
    DATA,
    FEATURE_REGISTRY_FILE,
    OUTPUTS,
    ROOT,
    _config_payload,
    _number,
    _read_json,
)
from model_prediction.dashboard.picks import (
    _flat_ledger_paths,
    _main_ledger_paths,
    _research_ledger_paths,
)

# ── SECTION: Production Evidence ───────────────────────────────────


def _artifact_hash(payload: dict) -> str:
    canonical = {key: value for key, value in payload.items() if key != "artifact_hash"}
    encoded = json.dumps(canonical, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(encoded).hexdigest()


def _artifact_evidence(path: Path, expected_version: str, expected_sport: str) -> tuple[dict, dict]:
    raw = _read_json(path)
    if not isinstance(raw, dict):
        return (
            {
                "path": str(path),
                "available": False,
                "valid": False,
                "health": "MISSING",
                "sha256": None,
                "hash_verified": False,
                "lineage": "UNVERIFIED",
                "declared_hash": None,
                "computed_hash": None,
                "hash_valid": False,
                "artifact_model_version": None,
                "version_matches_config": False,
                "artifact_identity": None,
                "lineage_matches_config": False,
                "mismatches": ["artifact_missing_or_invalid_json"],
            },
            {},
        )

    declared_hash = raw.get("artifact_hash")
    computed_hash = _artifact_hash(raw)
    artifact_version = str(raw.get("model_version") or "")
    artifact_identity = str(raw.get("sport") or raw.get("title") or raw.get("league") or "")
    hash_valid = bool(declared_hash) and declared_hash == computed_hash
    version_valid = bool(expected_version) and artifact_version == expected_version
    lineage_valid = artifact_identity.casefold() == expected_sport.casefold()
    mismatches = []
    if not declared_hash:
        mismatches.append("artifact_hash_missing")
    elif not hash_valid:
        mismatches.append("artifact_hash_mismatch")
    if not version_valid:
        mismatches.append("artifact_model_version_mismatch")
    if not lineage_valid:
        mismatches.append("artifact_sport_or_title_mismatch")
    return (
        {
            "path": str(path),
            "available": True,
            "valid": hash_valid and version_valid and lineage_valid,
            "health": "VERIFIED" if hash_valid and version_valid and lineage_valid else "FAILED",
            "sha256": declared_hash,
            "hash_verified": hash_valid,
            "lineage": "VERIFIED" if version_valid and lineage_valid else "MISMATCH",
            "declared_hash": declared_hash,
            "computed_hash": computed_hash,
            "hash_valid": hash_valid,
            "artifact_model_version": artifact_version or None,
            "version_matches_config": version_valid,
            "artifact_identity": artifact_identity or None,
            "lineage_matches_config": lineage_valid,
            "mismatches": mismatches,
        },
        raw,
    )


def _production_model_spec(raw: dict) -> dict:
    moneyline = (raw.get("market_models") or {}).get("moneyline") or {}
    if moneyline:
        names = list(moneyline.get("feature_names") or [])
        coefficients = list(moneyline.get("coefficients") or [])
        return {
            "kind": "logistic_regression",
            "feature_schema_status": "declared",
            "features": [
                {"name": name, "coefficient": coefficient}
                for name, coefficient in zip(names, coefficients, strict=False)
            ],
            "feature_names": names,
            "coefficients": coefficients,
            "coefficient_count_matches_features": len(names) == len(coefficients),
            "intercept": moneyline.get("intercept"),
            "confidence_threshold": moneyline.get("confidence_threshold"),
            "positive_class": moneyline.get("positive_class"),
        }
    if raw.get("league") and "tie_probability" in raw:
        return {
            "kind": "tie_aware_elo",
            "feature_schema_status": "model_family_declared",
            "features": [{"name": "tie_aware_elo_rating_difference", "coefficient": None}],
            "feature_names": ["tie_aware_elo_rating_difference"],
            "coefficients": [],
            "coefficient_count_matches_features": True,
            "parameters": {
                "initial_rating": raw.get("initial_rating"),
                "k": raw.get("k"),
                "home_advantage": raw.get("home_advantage"),
                "tie_probability": raw.get("tie_probability"),
                "target": raw.get("target"),
            },
        }
    return {
        "kind": "neutral_series_elo",
        "feature_schema_status": "model_family_declared",
        "features": [{"name": "neutral_elo_rating_difference", "coefficient": None}],
        "feature_names": ["neutral_elo_rating_difference"],
        "coefficients": [],
        "coefficient_count_matches_features": True,
        "parameters": {
            "initial_rating": raw.get("initial_rating"),
            "k": raw.get("k"),
            "home_or_order_advantage": raw.get("home_or_order_advantage"),
            "confidence_threshold": raw.get("confidence_threshold"),
            "target": raw.get("target"),
        },
    }


def _rolling_declared_hash(artifact_path: str | None) -> str | None:
    """Declared hash of the ROLLING artifact for a config path, if any.

    K split (2026-08-15): the external validation reports describe the
    rolling artifacts under the runtime root; the frozen config copy is
    the promoted snapshot. When a rolling copy exists, its hash is what
    the report must match.
    """
    if not artifact_path:
        return None
    from model_prediction.runtime_paths import rolling_models_root

    candidate = rolling_models_root() / Path(artifact_path).name
    if not candidate.is_file():
        return None
    try:
        return json.loads(candidate.read_text(encoding="utf-8")).get("artifact_hash")
    except (OSError, json.JSONDecodeError):
        return None


def _locked_backfill_evidence(
    sport: str,
    version: str,
    raw: dict,
    artifact: dict,
    esports_validation: dict,
    international_validation: dict,
) -> dict:
    if not artifact.get("valid"):
        return {
            "status": "rejected_artifact_integrity",
            "source": None,
            "model_version": version,
            "metrics": None,
            "pnl_label": None,
            "profitability_claim": False,
        }
    if (
        raw.get("method") in ("logistic_regression", "cfb_key_number_engine", "elo_trend_efficiency")
        or str(raw.get("sport")).casefold() == "ncaaf"
    ):
        metrics = raw.get("qualification") or {}
        if not metrics or (
            metrics.get("locked_holdout") is not True and metrics.get("qualified") is not True
        ):
            return {
                "status": "rejected_missing_locked_holdout_metrics",
                "source": artifact.get("path"),
                "model_version": version,
                "metrics": None,
                "pnl_label": None,
                "profitability_claim": False,
            }
        return {
            "status": "verified",
            "source": artifact.get("path"),
            "model_version": version,
            "metrics": metrics,
            "pnl_label": "hypothetical_at_minus_110",
            "profitability_claim": False,
        }

    if raw.get("league"):
        league = str(raw.get("league") or sport).casefold()
        report = (international_validation.get("leagues") or {}).get(league) or {}
        locked = report.get("locked_test") or {}
        report_version = str(report.get("model_version") or "")
        report_hash = str(report.get("artifact_hash") or "")
        exact = (
            bool(locked)
            and report_version == version
            and report_hash == (_rolling_declared_hash(artifact.get("path")) or artifact.get("declared_hash"))
        )
        if not exact:
            return {
                "status": "rejected_missing_or_mismatched_locked_metrics",
                "source": str(OUTPUTS / "international-baseball-baseline-validation.json"),
                "model_version": report_version or None,
                "artifact_hash": report_hash or None,
                "metrics": None,
                "pnl_label": None,
                "profitability_claim": False,
            }
        return {
            "status": "verified",
            "source": str(OUTPUTS / "international-baseball-baseline-validation.json"),
            "model_version": report_version,
            "artifact_hash": report_hash,
            "metrics": locked,
            "pnl_label": "hypothetical_at_minus_110",
            "profitability_claim": False,
        }

    report = (esports_validation.get("titles") or {}).get(sport.lower()) or {}
    report_version = str(report.get("model_version") or "")
    report_hash = str(report.get("artifact_hash") or "")
    exact_version = report_version == version
    exact_hash = bool(report_hash) and report_hash == (
        _rolling_declared_hash(artifact.get("path")) or artifact.get("declared_hash")
    )
    locked = report.get("locked_test") or {}
    if not exact_version or not exact_hash or not locked:
        reasons = []
        if not exact_version:
            reasons.append("validation_model_version_mismatch")
        if not exact_hash:
            reasons.append("validation_artifact_hash_mismatch_or_missing")
        if not locked:
            reasons.append("locked_test_metrics_missing")
        return {
            "status": "rejected_external_validation_mismatch",
            "source": str(OUTPUTS / "esports-baseline-validation.json"),
            "model_version": report_version or None,
            "metrics": None,
            "mismatches": reasons,
            "pnl_label": None,
            "profitability_claim": False,
        }
    return {
        "status": "verified",
        "source": str(OUTPUTS / "esports-baseline-validation.json"),
        "model_version": report_version,
        "artifact_hash": report_hash,
        "metrics": locked,
        "pnl_label": "hypothetical_at_minus_110",
        "profitability_claim": False,
    }


def _backfill_aliases(backfill: dict, raw: dict) -> dict:
    """Flatten locked metrics for UI consumers while retaining full detail."""
    metrics = backfill.get("metrics") or {}
    if raw.get("method") == "logistic_regression":
        locked_training = (raw.get("training") or {}).get("locked_holdout") or {}
        obs = metrics.get("total_predictions", locked_training.get("observations"))
        calls = metrics.get("calls")
        hr = metrics.get("hit_rate")
        br = metrics.get("brier_score")
        q = metrics.get("qualified")
        aliases = {
            "observations": obs,
            "calls": calls,
            "hit_rate": hr,
            "brier_score": br,
            "qualified": q,
        }
        units = metrics.get("units_at_minus_110")
        hb = {"observations": obs, "calls": calls, "hit_rate": hr, "brier": br, "units_at_minus_110": units}
        aliases["holdout_backfill"] = {"all_calls": hb, "selection_calls": hb}
    elif raw.get("league"):
        calls = metrics.get("calls")
        hits = metrics.get("hits")
        obs = metrics.get("observations")
        hr = hits / calls if calls and hits is not None else None
        br = metrics.get("brier_settlement")
        q = raw.get("qualified_for_betting")
        aliases = {
            "observations": obs,
            "calls": calls,
            "hit_rate": hr,
            "brier_score": br,
            "qualified": q,
        }
        units = metrics.get("units_at_minus_110")
        hb = {"observations": obs, "calls": calls, "hit_rate": hr, "brier": br, "units_at_minus_110": units}
        aliases["holdout_backfill"] = {"all_calls": hb, "selection_calls": hb}
    else:
        selected = metrics.get("selected_matches") or {}
        all_m = metrics.get("all_matches") or {}
        aliases = {
            "observations": selected.get("observations"),
            "calls": selected.get("calls"),
            "hit_rate": selected.get("accuracy"),
            "brier_score": selected.get("brier"),
            "qualified": raw.get("qualified_for_betting"),
        }
        aliases["holdout_backfill"] = {
            "all_calls": {
                "observations": all_m.get("observations"),
                "calls": all_m.get("calls"),
                "hit_rate": all_m.get("accuracy"),
                "brier": all_m.get("brier"),
                "units_at_minus_110": all_m.get("units_at_minus_110"),
            },
            "selection_calls": {
                "observations": selected.get("observations"),
                "calls": selected.get("calls"),
                "hit_rate": selected.get("accuracy"),
                "brier": selected.get("brier"),
                "units_at_minus_110": selected.get("units_at_minus_110"),
            },
        }
    return {**backfill, **aliases}


def _read_evidence_ledger(path: Path) -> list[dict]:
    if load_workbook is None or not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Picks"] if "Picks" in workbook.sheetnames else workbook.active
    if sheet is None:
        workbook.close()
        return []
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value) if value is not None else "" for value in next(values)]
    except StopIteration:
        workbook.close()
        return []
    rows = []
    for raw_row in values:
        if not raw_row or all(value is None for value in raw_row):
            continue
        row = {header: raw_row[index] for index, header in enumerate(headers) if header}
        if row.get("pick_id") or row.get("event_id"):
            rows.append(row)
    workbook.close()
    return rows


def _normalized_line(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value).strip().casefold()
    return f"{number:.12g}"


def _ledger_deduplication_key(row: dict) -> tuple[str, ...]:
    return (
        str(row.get("event_id") or "").strip(),
        str(row.get("league") or "").strip().casefold(),
        str(row.get("market_type") or "").strip().casefold(),
        str(row.get("selection") or "").strip().casefold(),
        _normalized_line(row.get("line")),
        str(row.get("model_version") or "").strip(),
        str(row.get("model_artifact_hash") or "").strip(),
    )


def _deduplicate_ledger_rows(rows: list[dict]) -> tuple[list[dict], int]:
    unique: dict[tuple[str, ...], dict] = {}
    for row in rows:
        unique.setdefault(_ledger_deduplication_key(row), row)
    return list(unique.values()), len(rows) - len(unique)


def _model_owns_row(sport: str, model_config: dict, row: dict) -> bool:
    version = str(row.get("model_version") or "").casefold()
    league = str(row.get("league") or "").casefold()
    configured_leagues = {str(value).casefold() for value in model_config.get("leagues") or []}
    return version.startswith(f"{sport.casefold()}-") or league in {
        sport.casefold(),
        *configured_leagues,
    }


def _feature_attribution(rows: list[dict], feature_names: list[str]) -> dict:
    if not rows:
        return {
            "status": "no_exact_version_rows",
            "rows_with_complete_values": 0,
            "rows_missing_values": 0,
            "missing_by_feature": {name: 0 for name in feature_names},
        }
    if not feature_names:
        return {
            "status": "unavailable_artifact_has_no_explicit_feature_schema",
            "rows_with_complete_values": 0,
            "rows_missing_values": len(rows),
            "missing_by_feature": {},
        }
    missing_by_feature = {name: 0 for name in feature_names}
    complete = 0
    for row in rows:
        feature_values = row.get("feature_values")
        if isinstance(feature_values, str):
            try:
                feature_values = json.loads(feature_values)
            except json.JSONDecodeError:
                feature_values = {}
        feature_values = feature_values if isinstance(feature_values, dict) else {}
        missing = []
        for name in feature_names:
            value = row.get(name, row.get(f"feature_{name}", feature_values.get(name)))
            if value in (None, ""):
                missing.append(name)
                missing_by_feature[name] += 1
        if not missing:
            complete += 1
    return {
        "status": "complete" if complete == len(rows) and rows else "missing",
        "rows_with_complete_values": complete,
        "rows_missing_values": len(rows) - complete,
        "missing_by_feature": missing_by_feature,
    }


def _pnl_evidence(rows: list[dict]) -> dict:
    if not rows:
        empty = {
            "rows": 0,
            "staked_units": None,
            "pnl_units": None,
            "roi": None,
        }
        return {
            "shadow": {"label": "shadow_not_executed", **empty},
            "hypothetical": {"label": "hypothetical_fixed_unit_research", **empty},
            "executed": {
                "label": "executed",
                **empty,
                "status": "not_available_no_execution_attribution_in_ledgers",
            },
        }
    shadow = [row for row in rows if row.get("record_type") == "QUALIFIED_SHADOW_CALL"]
    shadow_staked = sum(_number(row.get("units")) for row in shadow)
    shadow_pnl = sum(_number(row.get("pnl_units")) for row in shadow)
    hypothetical = [row for row in rows if _number(row.get("research_score_units")) > 0]
    hypothetical_staked = sum(_number(row.get("research_score_units")) for row in hypothetical)
    hypothetical_pnl = sum(_number(row.get("research_pnl_units")) for row in hypothetical)
    return {
        "shadow": {
            "label": "shadow_not_executed",
            "rows": len(shadow),
            "staked_units": round(shadow_staked, 6),
            "pnl_units": round(shadow_pnl, 6),
            "roi": round(shadow_pnl / shadow_staked, 6) if shadow_staked else None,
        },
        "hypothetical": {
            "label": "hypothetical_fixed_unit_research",
            "rows": len(hypothetical),
            "staked_units": round(hypothetical_staked, 6),
            "pnl_units": round(hypothetical_pnl, 6),
            "roi": round(hypothetical_pnl / hypothetical_staked, 6) if hypothetical_staked else None,
        },
        "executed": {
            "label": "executed",
            "rows": 0,
            "staked_units": None,
            "pnl_units": None,
            "roi": None,
            "status": "not_available_no_execution_attribution_in_ledgers",
        },
    }


def _version_ledger_evidence(
    version: str,
    rows: list[dict],
    pushes: int,
    source: str,
    source_rows_before_deduplication: int,
    duplicates_removed: int,
    artifact: dict,
    feature_names: list[str],
    predecessor_version_counts: dict[str, int],
) -> dict:
    wins = sum(str(row.get("result") or "").casefold() == "win" for row in rows)
    brier_values = []
    for row in rows:
        probability = _number(row.get("model_probability"), None)
        if probability is None or not 0 <= probability <= 1:
            continue
        outcome = 1.0 if str(row.get("result") or "").casefold() == "win" else 0.0
        brier_values.append((probability - outcome) ** 2)
    clv_values = [
        _number(row.get("probability_clv"), None)
        for row in rows
        if row.get("probability_clv") not in (None, "")
    ]
    clv_values = [value for value in clv_values if value is not None]
    expected_hash = artifact.get("declared_hash") if artifact.get("valid") else None
    row_hashes = [str(row.get("model_artifact_hash") or "") for row in rows]
    matching_hash_rows = sum(bool(expected_hash) and value == expected_hash for value in row_hashes)
    mismatching_hash_rows = sum(bool(value) and value != expected_hash for value in row_hashes)
    missing_hash_rows = sum(not value for value in row_hashes)
    pnl = _pnl_evidence(rows)
    if pnl["shadow"]["rows"]:
        pnl_basis = "shadow"
        pnl_units = pnl["shadow"]["pnl_units"]
    elif pnl["hypothetical"]["rows"]:
        pnl_basis = "hypothetical"
        pnl_units = pnl["hypothetical"]["pnl_units"]
    else:
        pnl_basis = None
        pnl_units = None
    clv_complete = bool(rows) and len(clv_values) == len(rows)
    profitability_allowed = bool(rows) and pnl["executed"]["roi"] is not None and clv_complete
    if not rows:
        lineage_status = "no_exact_version_rows"
    elif not expected_hash:
        lineage_status = "artifact_unverified"
    elif matching_hash_rows == len(rows):
        lineage_status = "exact"
    elif mismatching_hash_rows and missing_hash_rows:
        lineage_status = "mixed_mismatch_and_missing"
    elif mismatching_hash_rows:
        lineage_status = "mismatch"
    elif missing_hash_rows:
        lineage_status = "missing"
    else:
        lineage_status = "mixed"
    blockers = []
    if not rows:
        blockers.append("no_exact_model_version_settled_decisive_rows")
    if pnl["executed"]["roi"] is None:
        blockers.append("executed_roi_unavailable")
    if not clv_complete:
        blockers.append("clv_missing_or_incomplete")
    return {
        "source": source,
        "model_version": version,
        "exact_version_rows": len(rows),
        "settled": len(rows) + pushes if rows or pushes else None,
        "settled_decisive_rows": len(rows),
        "wins": wins if rows else None,
        "losses": len(rows) - wins if rows else None,
        "pushes": pushes if rows or pushes else None,
        "hit_rate": round(wins / len(rows), 6) if rows else None,
        "brier": round(sum(brier_values) / len(brier_values), 6) if brier_values else None,
        "brier_rows": len(brier_values),
        "source_rows_before_deduplication": source_rows_before_deduplication,
        "duplicates_removed": duplicates_removed,
        "predecessor_rows_excluded": sum(predecessor_version_counts.values()),
        "predecessor_version_counts": predecessor_version_counts,
        "artifact_lineage": {
            "artifact_path": artifact.get("path"),
            "expected_hash": expected_hash,
            "artifact_valid": bool(artifact.get("valid")),
            "matching_hash_rows": matching_hash_rows,
            "mismatching_hash_rows": mismatching_hash_rows,
            "missing_hash_rows": missing_hash_rows,
            "status": lineage_status,
        },
        "feature_value_attribution": _feature_attribution(rows, feature_names),
        "pnl_units": pnl_units,
        "pnl_basis": pnl_basis,
        "pnl": pnl,
        "clv": {
            "rows": len(clv_values),
            "total_exact_version_rows": len(rows),
            "coverage": round(len(clv_values) / len(rows), 6) if rows else None,
            "complete": clv_complete,
            "mean_probability_clv": round(sum(clv_values) / len(clv_values), 6) if clv_values else None,
        },
        "profitability_claim": {
            "allowed": profitability_allowed,
            "blockers": blockers,
        },
    }


def _ledger_evidence_for_source(
    sport: str,
    model_config: dict,
    version: str,
    source: str,
    source_rows: list[dict],
    artifact: dict,
    feature_names: list[str],
) -> dict:
    relevant = [
        row
        for row in source_rows
        if str(row.get("status") or "").casefold() == "settled"
        and str(row.get("result") or "").casefold() in {"win", "loss", "push"}
        and _model_owns_row(sport, model_config, row)
    ]
    deduplicated, _all_duplicates_removed = _deduplicate_ledger_rows(relevant)
    exact_settled = [row for row in deduplicated if str(row.get("model_version") or "") == version]
    exact_rows = [row for row in exact_settled if str(row.get("result") or "").casefold() in {"win", "loss"}]
    pushes = sum(str(row.get("result") or "").casefold() == "push" for row in exact_settled)
    exact_source_rows = sum(
        str(row.get("model_version") or "") == version
        and str(row.get("result") or "").casefold() in {"win", "loss"}
        for row in relevant
    )
    predecessor_counts: dict[str, int] = {}
    for row in deduplicated:
        row_version = str(row.get("model_version") or "")
        decisive = str(row.get("result") or "").casefold() in {"win", "loss"}
        if row_version and row_version != version and decisive:
            predecessor_counts[row_version] = predecessor_counts.get(row_version, 0) + 1
    return _version_ledger_evidence(
        version,
        exact_rows,
        pushes,
        source,
        exact_source_rows,
        exact_source_rows - len(exact_rows),
        artifact,
        feature_names,
        dict(sorted(predecessor_counts.items())),
    )


def _feature_registry_evidence() -> dict:
    """Return a validated, read-only view of the durable feature registry."""
    path = FEATURE_REGISTRY_FILE
    raw = _read_json(path)
    errors: list[str] = []
    if not isinstance(raw, dict):
        return {
            "status": "missing_or_invalid",
            "valid": False,
            "path": str(path.relative_to(ROOT)) if path.is_relative_to(ROOT) else str(path),
            "sha256": None,
            "errors": ["registry_missing_or_invalid_json"],
            "features": [],
            "production_ablation_summary": [],
            "counts_by_verdict": {},
        }
    features = raw.get("features")
    if not isinstance(features, list):
        errors.append("features_not_list")
        features = []
    normalized: list[dict] = []
    names: set[str] = set()
    for index, item in enumerate(features):
        if not isinstance(item, dict):
            errors.append(f"feature_{index}_not_object")
            continue
        name = str(item.get("name") or "").strip()
        if not name:
            errors.append(f"feature_{index}_missing_name")
            continue
        if name in names:
            errors.append(f"duplicate_feature:{name}")
            continue
        names.add(name)
        if not str(item.get("verdict") or "").strip():
            errors.append(f"feature_missing_verdict:{name}")
        normalized.append(dict(item))
    ablations = raw.get("production_ablation_summary")
    if not isinstance(ablations, list):
        errors.append("production_ablation_summary_not_list")
        ablations = []
    valid_ablations: list[dict] = []
    seen_ablations: set[tuple[str, str, str]] = set()
    for index, item in enumerate(ablations):
        if not isinstance(item, dict):
            errors.append(f"ablation_{index}_not_object")
            continue
        identity = (
            str(item.get("sport") or "").casefold(),
            str(item.get("model_version") or ""),
            str(item.get("feature") or ""),
        )
        if not all(identity):
            errors.append(f"ablation_{index}_missing_identity")
            continue
        if identity in seen_ablations:
            errors.append("duplicate_ablation:" + ":".join(identity))
            continue
        seen_ablations.add(identity)
        if identity[2] not in names:
            errors.append(f"ablation_feature_not_registered:{identity[2]}")
        valid_ablations.append(dict(item))
    counts: dict[str, int] = {}
    for item in normalized:
        verdict = str(item.get("verdict") or "missing")
        counts[verdict] = counts.get(verdict, 0) + 1
    return {
        "status": "verified" if not errors else "invalid",
        "valid": not errors,
        "path": str(path.relative_to(ROOT)) if path.is_relative_to(ROOT) else str(path),
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "schema_version": raw.get("schema_version"),
        "last_updated": raw.get("last_updated"),
        "authoritative_evidence": raw.get("authoritative_evidence") or {},
        "retention_policy": raw.get("retention_policy") or {},
        "evidence_grades": raw.get("evidence_grades") or {},
        "errors": errors,
        "features": normalized,
        "production_ablation_summary": valid_ablations,
        "counts_by_verdict": dict(sorted(counts.items())),
    }


def _read_model_ledger_rows(path: Path) -> list[dict]:
    """Plain read of a data/model_ledgers/<model-id>.xlsx file -- same
    shape as _parse_picks but for the new per-model schema's "Predictions"
    sheet. Deliberately not an import from the model_prediction package
    (dashboard_server.py has zero imports from it) -- a small, duplicated
    reader here, matching how _parse_picks already duplicates rather than
    imports PickLedger's own reading logic."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb["Predictions"] if "Predictions" in wb.sheetnames else wb.active
        if sheet is None:
            return []
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        rows = []
        for values in rows_iter:
            if values is None or all(value is None for value in values):
                continue
            row = {
                header: ("" if value is None else str(value))
                for header, value in zip(headers, values, strict=False)
            }
            if row.get("prediction_id"):
                rows.append(row)
        return rows
    finally:
        wb.close()


def _is_model_ledger_backup(path: Path) -> bool:
    """Keep timestamped safety copies out of the active model registry."""
    stem = path.stem.casefold()
    return any(marker in stem for marker in (".bak-", ".backup", "-backup-"))


def _model_evidence_from_rows(model_id: str, rows: list[dict]) -> dict:
    """Lighter-weight duplicate of model_ledger.compute_model_evidence's
    Brier/log-loss/sample-size math for dashboard display -- no calibration
    buckets/ECE/logistic fit here, just what the design spec's evidence
    columns actually need. Same push-exclusion rule as ledger.py's own
    calibration_rows filter: settled win/loss only, a push is never folded
    in as a loss."""
    open_rows = [r for r in rows if r.get("status") == "open"]
    settled = [r for r in rows if r.get("status") == "settled"]
    failed = [r for r in rows if r.get("status") == "failed"]
    calibration_rows = [
        r for r in settled if r.get("result") in ("win", "loss") and r.get("model_probability")
    ]
    brier = log_loss = None
    if len(calibration_rows) >= 30:
        pairs = [
            (min(1 - 1e-12, max(1e-12, float(r["model_probability"]))), 1 if r["result"] == "win" else 0)
            for r in calibration_rows
        ]
        brier = sum((p - y) ** 2 for p, y in pairs) / len(pairs)
        log_loss = -sum(y * math.log(p) + (1 - y) * math.log(1 - p) for p, y in pairs) / len(pairs)
    clv_values = [float(r["probability_clv"]) for r in settled if r.get("probability_clv")]
    pnl_values = [float(r["pnl_units"]) for r in settled if r.get("pnl_units")]
    missing_input_rows = [r for r in rows if r.get("missing_inputs")]
    observed = sorted(r["observed_at_utc"] for r in rows if r.get("observed_at_utc"))
    return {
        "model_id": model_id,
        "total": len(rows),
        "open": len(open_rows),
        "settled": len(settled),
        "failed": len(failed),
        "wins": sum(1 for r in calibration_rows if r["result"] == "win"),
        "losses": sum(1 for r in calibration_rows if r["result"] == "loss"),
        "pushes": sum(1 for r in settled if r.get("result") not in ("win", "loss")),
        "pnl_units": round(sum(pnl_values), 4),
        "brier_score": brier,
        "log_loss": log_loss,
        "clv_coverage": round(len(clv_values) / len(settled), 4) if settled else None,
        "mean_clv": round(sum(clv_values) / len(clv_values), 6) if clv_values else None,
        "missing_input_rate": round(len(missing_input_rows) / len(rows), 4) if rows else None,
        "latest_observed_at_utc": observed[-1] if observed else None,
    }


def model_ledger_comparison() -> dict:
    """One row per event, one column per applicable model -- the
    operator's dashboard design spec. No qualified/research classification
    badges: evidence only (sample size, Brier, log loss, CLV, ROI,
    missing-input rate, data age), read straight from data/model_ledgers/.
    """
    ledgers_dir = DATA / "model_ledgers"
    if not ledgers_dir.exists():
        return {"generated_at": datetime.now(UTC).isoformat(), "events": [], "models": {}}
    evidence_by_model: dict[str, dict] = {}
    predictions_by_event: dict[str, list[dict]] = {}
    for path in sorted(ledgers_dir.glob("*.xlsx")):
        if _is_model_ledger_backup(path):
            continue
        model_id = path.stem
        rows = _read_model_ledger_rows(path)
        evidence_by_model[model_id] = _model_evidence_from_rows(model_id, rows)
        for row in rows:
            if row.get("status") != "open":
                continue  # open predictions only -- settled history lives in the model's own evidence
            predictions_by_event.setdefault(row["event_id"], []).append(
                {
                    "model_id": model_id,
                    "prediction_id": row.get("prediction_id"),
                    "market_type": row.get("market_type"),
                    "selection": row.get("selection"),
                    "line": row.get("line") or None,
                    "model_probability": _number(row.get("model_probability"), None),
                    "decision_price": _number(row.get("decision_price"), None),
                    "model_market_difference": _number(row.get("model_market_difference"), None),
                    "input_availability": row.get("input_availability") or None,
                    "event_start_utc": row.get("event_start_utc"),
                    "operator_decision": row.get("operator_decision") or None,
                }
            )
    events = [
        {"event_id": event_id, "predictions": predictions}
        for event_id, predictions in sorted(
            predictions_by_event.items(),
            key=lambda item: min((p["event_start_utc"] or "" for p in item[1]), default=""),
        )
    ]
    return {
        "generated_at": datetime.now(UTC).isoformat(),
        "events": events,
        "models": evidence_by_model,
    }


def record_model_ledger_decision(payload: dict) -> dict:
    """Store the operator's own event-level decision for one prediction,
    separate from the model's own output. "Not model promotion. It is an
    event-level decision... must not change the model's ledger,
    classification, historical statistics, or dashboard evidence."

    Reuses the real, lock-protected ModelLedger.record_operator_decision
    (a local, function-scoped import -- see dedupe_ledger's identical
    "heavy import" precedent -- rather than duplicating a second,
    potentially unlocked write path for the same file here).
    """
    model_id = str(payload.get("model_id") or "")
    prediction_id = str(payload.get("prediction_id") or "")
    decision = str(payload.get("decision") or "")
    if not model_id or not prediction_id or not decision:
        return {"status": "refused", "error": "model_id, prediction_id, and decision are required"}
    path = DATA / "model_ledgers" / f"{model_id}.xlsx"
    if not path.exists() or _is_model_ledger_backup(path):
        return {"status": "refused", "error": f"unknown model_id {model_id!r}"}

    from model_prediction.model_ledger import ModelLedger  # local: heavy import

    ledger = ModelLedger(path)
    units = payload.get("units")
    try:
        row = ledger.record_operator_decision(
            prediction_id,
            decision=decision,
            selected_model=payload.get("selected_model") or None,
            selected_market=payload.get("selected_market") or None,
            units=None if units is None or units == "" else float(units),
            note=payload.get("note") or None,
        )
    except KeyError:
        return {"status": "refused", "error": f"unknown prediction_id {prediction_id!r} in {model_id}"}
    except (TypeError, ValueError) as error:
        return {"status": "refused", "error": str(error)}
    with _CACHE_LOCK:
        _CACHE.pop("model-ledgers", None)
    return {"status": "ok", "row": row}


def _production_canary_status() -> dict:
    """Production canary dashboard card — model health, timestamps, prediction counts."""
    try:
        from dashboard.production import get_production_status

        return get_production_status()
    except Exception as e:  # noqa: BLE001 - any failure surfaces as a DOWN card, not a 500
        return {"status": "DOWN", "error": str(e)}


def production_evidence() -> dict:
    """Read-only, fail-closed evidence for every configured production artifact."""
    config = _config_payload()
    configured_models = config.get("models") or {}
    esports_validation = _read_json(OUTPUTS / "esports-baseline-validation.json") or {}
    international_validation = _read_json(OUTPUTS / "international-baseball-baseline-validation.json") or {}
    ledger_paths = (
        *_main_ledger_paths(),
        *_flat_ledger_paths(),
        *_research_ledger_paths(),
    )
    rows_by_source = {str(path.relative_to(ROOT)): _read_evidence_ledger(path) for path in ledger_paths}
    feature_registry = _feature_registry_evidence()
    registry_by_name = {str(item.get("name")): item for item in feature_registry["features"]}
    ablation_by_identity = {
        (
            str(item.get("sport") or "").casefold(),
            str(item.get("model_version") or ""),
            str(item.get("feature") or ""),
        ): item
        for item in feature_registry["production_ablation_summary"]
    }
    models = []
    for sport, model_config in configured_models.items():
        if not isinstance(model_config, dict) or not model_config.get("production_artifact"):
            continue
        version = str(model_config.get("active_production_version") or "")
        configured_path = Path(str(model_config["production_artifact"]))
        artifact_path = configured_path if configured_path.is_absolute() else ROOT / configured_path
        artifact, raw = _artifact_evidence(artifact_path, version, str(sport))
        spec = _production_model_spec(raw)
        locked = _backfill_aliases(
            _locked_backfill_evidence(
                str(sport),
                version,
                raw,
                artifact,
                esports_validation,
                international_validation,
            ),
            raw,
        )

        feature_names = list(spec.get("feature_names") or [])
        registry_features = []
        for feature_name in feature_names:
            registry_record = registry_by_name.get(str(feature_name))
            registry_features.append(
                {
                    "name": str(feature_name),
                    "registered": registry_record is not None,
                    "verdict": registry_record.get("verdict") if registry_record else None,
                    "status": registry_record.get("status") if registry_record else None,
                    "evidence_grade": registry_record.get("evidence_grade") if registry_record else None,
                    "sport_evidence": ablation_by_identity.get(
                        (str(sport).casefold(), version, str(feature_name))
                    ),
                }
            )
        # Main/Flat are per-sport files now (data/main/<sport>.xlsx,
        # data/flat/<sport>.xlsx) -- only sports in _MAIN_LEDGER_SPORTS
        # (mlb/wnba/soccer/tennis) have one; every other configured sport
        # (nba/nfl/esports/kbo/npb) correctly gets empty evidence here since
        # rows_by_source has no entry for a file that was never created.
        main_source = f"data/main/{str(sport).casefold()}.xlsx"
        main_ledger = _ledger_evidence_for_source(
            str(sport),
            model_config,
            version,
            main_source,
            rows_by_source.get(main_source, []),
            artifact,
            feature_names,
        )
        flat_source = f"data/flat/{str(sport).casefold()}.xlsx"
        flat_ledger = _ledger_evidence_for_source(
            str(sport),
            model_config,
            version,
            flat_source,
            rows_by_source.get(flat_source, []),
            artifact,
            feature_names,
        )

        sport_research_source = f"data/research/{str(sport).casefold()}.xlsx"
        legacy_research_source = "data/research.xlsx"
        research_source = (
            sport_research_source if sport_research_source in rows_by_source else legacy_research_source
        )
        research_ledger = _ledger_evidence_for_source(
            str(sport),
            model_config,
            version,
            research_source,
            rows_by_source.get(research_source, []),
            artifact,
            feature_names,
        )

        warnings = [{"code": code, "scope": "artifact"} for code in artifact["mismatches"]]
        if locked.get("status") != "verified":
            warnings.append({"code": locked["status"], "scope": "backfill"})
        if spec.get("coefficient_count_matches_features") is False:
            warnings.append({"code": "feature_coefficient_length_mismatch", "scope": "features"})
        if not feature_registry["valid"]:
            warnings.append({"code": "feature_registry_invalid", "scope": "features"})
        for feature in registry_features:
            if not feature["registered"]:
                warnings.append(
                    {
                        "code": "active_feature_not_registered",
                        "scope": "features",
                        "feature": feature["name"],
                    }
                )

        config_qualified = str(model_config.get("status") or "").casefold() in {
            "qualified",
            "shadow_qualified",
            "production",
        }
        artifact_qualified = (
            (raw.get("qualification") or {}).get("qualified")
            if raw.get("method") == "logistic_regression"
            else raw.get("qualified_for_betting")
        )
        qualification_overridden = model_config.get("qualification_override") is True
        if (
            isinstance(artifact_qualified, bool)
            and config_qualified != artifact_qualified
            and not qualification_overridden
        ):
            warnings.append(
                {
                    "code": "config_artifact_qualification_mismatch",
                    "scope": "qualification",
                    "config_status": model_config.get("status"),
                    "artifact_qualified": artifact_qualified,
                }
            )

        for ledger_name, ledger in (
            ("main_ledger", main_ledger),
            ("flat_ledger", flat_ledger),
            ("research_ledger", research_ledger),
        ):
            if not ledger["exact_version_rows"]:
                warnings.append(
                    {
                        "code": "no_exact_version_settled_decisive_rows",
                        "scope": ledger_name,
                    }
                )
                continue
            if ledger["artifact_lineage"]["status"] != "exact":
                warnings.append(
                    {
                        "code": "ledger_artifact_lineage_missing_or_mismatched",
                        "scope": ledger_name,
                    }
                )
            if ledger["feature_value_attribution"]["status"] != "complete":
                warnings.append({"code": "feature_value_attribution_missing", "scope": ledger_name})

        definition_valid = (
            artifact["valid"]
            and locked.get("status") == "verified"
            and spec.get("coefficient_count_matches_features") is not False
        )
        performance_complete = (
            main_ledger["profitability_claim"]["allowed"] and flat_ledger["profitability_claim"]["allowed"]
        )
        profitability = {
            "claim_allowed": False,
            "status": "not_established",
            "requires": [
                "exact_model_version_settled_decisive_rows",
                "executed_roi",
                "complete_clv",
            ],
            "blockers": sorted(
                {
                    blocker
                    for ledger in (main_ledger, flat_ledger, research_ledger)
                    for blocker in ledger["profitability_claim"]["blockers"]
                }
            ),
            "main_ledger_claim_allowed": main_ledger["profitability_claim"]["allowed"],
            "flat_ledger_claim_allowed": flat_ledger["profitability_claim"]["allowed"],
        }

        models.append(
            {
                "sport": str(sport).lower(),
                "model_version": version or None,
                "status": model_config.get("status"),
                "features": spec.get("features") or [],
                "feature_registry": registry_features,
                "backfill": locked,
                "main_ledger": main_ledger,
                "flat_ledger": flat_ledger,
                "research_ledger": research_ledger,
                "profitability": profitability,
                "warnings": warnings,
                "configured_status": model_config.get("status"),
                "active_model_version": version or None,
                "production_artifact": str(model_config["production_artifact"]),
                "artifact": artifact,
                "model_spec": spec,
                "locked_backfill": locked,
                "ledger_evidence": {
                    "main_ledger": main_ledger,
                    "flat_ledger": flat_ledger,
                },
                "model_definition_and_backfill_valid": definition_valid,
                "production_performance_evidence_complete": performance_complete,
                "evidence_valid": definition_valid and performance_complete,
                "issues": warnings,
            }
        )

    generated_at = datetime.now(UTC).isoformat()
    return {
        "generated_at": generated_at,
        "generated_at_utc": generated_at,
        "read_only": True,
        "claim_policy": {
            "profitability_requires": [
                "exact_model_version_settled_decisive_rows",
                "executed_roi",
                "complete_clv",
            ],
            "shadow_and_hypothetical_pnl_are_not_profitability_evidence": True,
        },
        "sources": {
            "config": str(CONFIG_FILE.relative_to(ROOT)),
            "feature_registry": feature_registry["path"],
            "ledgers": list(rows_by_source),
            "esports_validation": "outputs/latest/esports-baseline-validation.json",
            "international_validation": "outputs/latest/international-baseball-baseline-validation.json",
        },
        "configured_production_models": len(models),
        "feature_registry": feature_registry,
        "all_model_definitions_and_backfills_valid": bool(models)
        and all(model["model_definition_and_backfill_valid"] for model in models),
        "all_production_performance_evidence_complete": bool(models)
        and all(model["production_performance_evidence_complete"] for model in models),
        "all_production_evidence_valid": bool(models) and all(model["evidence_valid"] for model in models),
        "models": models,
    }
