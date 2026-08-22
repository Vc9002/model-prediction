"""Dashboard picks module (extracted from dashboard_server.py, DD-5).

Part of the dashboard_server.py -> dashboard/ package split. See MASTER.md
DD-5 and dashboard/__init__.py for the re-export shim that keeps the old
`dashboard_server` import path and test-suite symbol references working.
"""

from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


from model_prediction.dashboard.common import (
    _CACHE_LOCK,
    _FLAT_PICKS_CACHE,
    _MAIN_LEDGER_SPORTS,
    _PICKS_CACHE,
    DATA,
    ROOT,
    _get_dashboard_cache,
    _number,
)

# ── SECTION: Picks & Performance ────────────────────────────────────


def _main_ledger_paths() -> list[Path]:
    return [DATA / "main" / f"{sport}.xlsx" for sport in _MAIN_LEDGER_SPORTS]


def _flat_ledger_paths() -> list[Path]:
    return [DATA / "flat" / f"{sport}.xlsx" for sport in _MAIN_LEDGER_SPORTS]


def _read_split_picks(paths: list[Path], cache: dict[str, object]) -> list[dict]:
    """Aggregate rows across one tier's per-sport files (Main or Flat),
    keyed on the combined mtime of every file that actually exists so any
    single sport's file changing invalidates the cache -- same "only
    re-parse on change" principle read_picks() always used, just extended
    to N files instead of one.

    When the SQLite dashboard cache is available, reads from it instead of
    parsing Excel files (~100x faster).  Falls back to Excel parsing if the
    cache hasn't been built yet.
    """
    dc = _get_dashboard_cache()

    # Try SQLite cache first if paths are under the standard repo data root
    if dc is not None and paths and str(paths[0]).startswith(str(ROOT / "data")):
        first_path = str(paths[0])
        if "/flat/" in first_path or "\\flat\\" in first_path:
            tier = "flat"
        elif "/research/" in first_path or "\\research\\" in first_path:
            tier = "research"
        elif "/gated_research/" in first_path:
            tier = "gated_research"
        else:
            tier = "main"

        try:
            dc.refresh()  # no-op if mtimes unchanged, fast SQLite otherwise
            if tier in ("flat", "main", "research", "gated_research"):
                rows = dc.read_picks(tier)
            else:
                rows = dc.read_picks(tier)
            if rows:
                return rows
        except Exception:  # noqa: BLE001, S110 - cache read is best-effort, Excel is the fallback
            pass

    # Fallback to Excel parsing (original path)
    if load_workbook is None:
        return []
    existing = [path for path in paths if path.exists()]
    if not existing:
        return []
    mtime_key = tuple(sorted((str(path), path.stat().st_mtime) for path in existing))
    with _CACHE_LOCK:
        if cache["mtime"] == mtime_key:
            return cache["rows"]  # type: ignore[return-value]
    rows_out = [row for path in existing for row in _parse_picks(path)]
    with _CACHE_LOCK:
        cache["mtime"] = mtime_key
        cache["rows"] = rows_out
    return rows_out


def read_picks() -> list[dict]:
    """Parse every sport's Main ledger (data/main/<sport>.xlsx), only
    re-parsing a file whose mtime changed -- parsing is the single most
    expensive repeated operation the dashboard performs."""
    return _read_split_picks(_main_ledger_paths(), _PICKS_CACHE)


def read_flat_picks() -> list[dict]:
    """Parse every sport's Flat ledger (data/flat/<sport>.xlsx), same
    change-detection caching as read_picks()."""
    return _read_split_picks(_flat_ledger_paths(), _FLAT_PICKS_CACHE)


def _find_pick_by_id(pick_id: str) -> dict | None:
    """Search both main and flat ledgers for a pick by pick_id."""
    row = next((item for item in read_picks() if str(item.get("pick_id")) == pick_id), None)
    if row is not None:
        return row
    return next((item for item in read_flat_picks() if str(item.get("pick_id")) == pick_id), None)


def _parse_picks(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb["Picks"] if "Picks" in wb.sheetnames else wb.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    keep = [
        "pick_id",
        "created_at_utc",
        "event_start_utc",
        "event_id",
        "league",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "american_odds",
        "market_implied_probability",
        "model_probability",
        "model_uncertainty",
        "edge",
        "trade_candidate",
        "confidence_score",
        "units",
        "model_version",
        "status",
        "result",
        "away_score",
        "home_score",
        "probability_clv",
        "pnl_units",
        "settled_at_utc",
        "record_type",
        "decision",
        "reason_code",
        "research_score_units",
        "research_pnl_units",
        "sportsbook",
        "decision_no_vig_probability",
        "rationale",
        "risks",
        "unavailable_features",
        "elo_probability",
        "trend_gap",
        "defensive_trend_gap",
        "park_factor",
        "weather_factor",
        "pitcher_era_gap",
        "probable_starter_era_gap",
        "market_residual_probability",
    ]
    index = {name: headers.index(name) for name in keep if name in headers}
    picks = []
    for values in rows:
        if values is None or all(value is None for value in values):
            continue
        row = {name: values[position] for name, position in index.items()}
        if not row.get("pick_id"):
            continue
        picks.append(row)
    wb.close()
    return picks


def _research_ledger_paths(*, gated: bool = False) -> list[Path]:
    directory = DATA / ("gated_research" if gated else "research")
    paths = sorted(directory.glob("*.xlsx")) if directory.exists() else []
    if paths:
        return paths
    legacy = DATA / ("gated_research.xlsx" if gated else "research.xlsx")
    return [legacy] if legacy.exists() else []


_RESEARCH_PICKS_CACHE: dict[str, object] = {"mtime": None, "rows": []}
_GATED_RESEARCH_PICKS_CACHE: dict[str, object] = {"mtime": None, "rows": []}


def _parse_research_picks(*, gated: bool = False) -> list[dict]:
    cache = _GATED_RESEARCH_PICKS_CACHE if gated else _RESEARCH_PICKS_CACHE
    return _read_split_picks(_research_ledger_paths(gated=gated), cache)


def _pick_probability(row) -> float | None:
    value = _number(row.get("model_probability"), -1)
    return value if 0 < value < 1 else None


def _pick_pnl(row) -> float:
    """Return recorded unit P&L for a row that actually has a scoring basis."""
    units = _number(row.get("units"))
    if units > 0:
        return _number(row.get("pnl_units"))
    if _number(row.get("research_score_units")) > 0:
        return _number(row.get("research_pnl_units"))
    return 0.0


def _pick_is_scored(row: dict) -> bool:
    return _number(row.get("units")) > 0 or _number(row.get("research_score_units")) > 0


def _performance_game_key(row: dict) -> str:
    """Stable-enough event identity for dashboard-only game coverage metrics."""
    event_id = str(row.get("event_id") or "").strip()
    if event_id:
        return f"{str(row.get('league') or '').casefold()}:{event_id}"
    parts = (
        row.get("league"),
        row.get("event_start_utc"),
        row.get("away_team"),
        row.get("home_team"),
    )
    return "|".join(str(part or "").strip().casefold() for part in parts)


def _performance_breakdown(
    rows: list[dict],
    key_fn,
) -> dict[str, dict]:
    """Build the same result/risk/game metrics for any ledger-derived dimension."""
    groups: dict[str, dict] = {}
    game_keys: dict[str, set[str]] = {}
    for row in rows:
        key = str(key_fn(row) or "unknown").strip() or "unknown"
        won = row.get("result") == "win"
        entry = groups.setdefault(
            key,
            {"wins": 0, "calls": 0, "scored_calls": 0, "pnl": 0.0, "risked": 0.0},
        )
        entry["calls"] += 1
        entry["wins"] += won
        if _pick_is_scored(row):
            entry["scored_calls"] += 1
            entry["pnl"] += _pick_pnl(row)
            entry["risked"] += max(
                _number(row.get("units")),
                _number(row.get("research_score_units")),
            )
        game_keys.setdefault(key, set()).add(_performance_game_key(row))
    for key, entry in groups.items():
        entry["losses"] = entry["calls"] - entry["wins"]
        entry["win_rate"] = round(entry["wins"] / entry["calls"], 4) if entry["calls"] else None
        entry["pnl"] = round(entry["pnl"], 4) if entry["scored_calls"] else None
        entry["risked"] = round(entry["risked"], 4) if entry["scored_calls"] else None
        entry["roi"] = (
            round(entry["pnl"] / entry["risked"], 4) if entry["pnl"] is not None and entry["risked"] else None
        )
        entry["games"] = len(game_keys.get(key, set()))
    return dict(
        sorted(
            groups.items(),
            key=lambda item: (-item[1]["calls"], item[0].casefold()),
        )
    )


def performance(picks: list[dict]) -> dict:
    settled = [
        row for row in picks if row.get("status") == "settled" and row.get("result") in ("win", "loss")
    ]
    settled.sort(key=lambda row: str(row.get("settled_at_utc") or ""))
    wins = sum(1 for row in settled if row["result"] == "win")
    cumulative, curve = 0.0, []
    for row in settled:
        if not _pick_is_scored(row):
            continue
        cumulative += _pick_pnl(row)
        curve.append(
            {
                "t": str(row.get("settled_at_utc") or "")[:16],
                "pnl": round(cumulative, 4),
            }
        )
    by_sport = _performance_breakdown(settled, lambda row: row.get("league"))
    by_market = _performance_breakdown(settled, lambda row: row.get("market_type"))
    probability_rows = [row for row in settled if _pick_probability(row) is not None]

    def probability_bucket(row: dict) -> str:
        probability = _pick_probability(row)
        if probability is None:
            return "unknown"
        index = min(19, int(probability * 20))
        lower = index / 20
        upper = (index + 1) / 20
        return f"{lower:.2f}-{upper:.2f}"

    by_bucket = dict(sorted(_performance_breakdown(probability_rows, probability_bucket).items()))
    by_month = {}
    for row in settled:
        won = row["result"] == "win"
        month = str(row.get("settled_at_utc") or "")[:7]
        if month:
            entry = by_month.setdefault(
                month,
                {"pnl": 0.0, "calls": 0, "scored_calls": 0, "wins": 0},
            )
            entry["calls"] += 1
            entry["wins"] += won
            if _pick_is_scored(row):
                entry["scored_calls"] += 1
                entry["pnl"] = round(entry["pnl"] + _pick_pnl(row), 4)
    for entry in by_month.values():
        entry["losses"] = entry["calls"] - entry["wins"]
        entry["win_rate"] = round(entry["wins"] / entry["calls"], 4) if entry["calls"] else None
        if not entry["scored_calls"]:
            entry["pnl"] = None

    sport_market_groups: dict[tuple[str, str], list[dict]] = {}
    for row in settled:
        key = (
            str(row.get("league") or "unknown"),
            str(row.get("market_type") or "unknown"),
        )
        sport_market_groups.setdefault(key, []).append(row)
    by_sport_market = []
    for (sport, market_type), rows in sorted(
        sport_market_groups.items(),
        key=lambda item: (-len(item[1]), item[0][0].casefold(), item[0][1].casefold()),
    ):
        summary = next(iter(_performance_breakdown(rows, lambda _row: "all").values()))
        by_sport_market.append({"sport": sport, "market_type": market_type, **summary})

    market_contexts = []
    for market_type, market_rows in sorted(
        (
            (market_type, [row for row in settled if str(row.get("market_type") or "unknown") == market_type])
            for market_type in by_market
        ),
        key=lambda item: (-len(item[1]), item[0].casefold()),
    ):
        game_ids = {_performance_game_key(row) for row in market_rows}
        lines = [_number(row.get("line"), None) for row in market_rows if row.get("line") not in (None, "")]
        lines = [line for line in lines if line is not None]
        market_contexts.append(
            {
                "market_type": market_type,
                "calls": len(market_rows),
                "games": len(game_ids),
                "sports": sorted(
                    {str(row.get("league") or "unknown") for row in market_rows},
                    key=str.casefold,
                ),
                "selections": _performance_breakdown(
                    market_rows,
                    lambda row: row.get("selection"),
                ),
                "line": {
                    "count": len(lines),
                    "mean": round(sum(lines) / len(lines), 3) if lines else None,
                    "mean_abs": round(sum(abs(line) for line in lines) / len(lines), 3) if lines else None,
                    "min": round(min(lines), 3) if lines else None,
                    "max": round(max(lines), 3) if lines else None,
                },
            }
        )

    sport_contexts = []
    for sport, metrics in by_sport.items():
        sport_rows = [row for row in settled if str(row.get("league") or "unknown") == sport]
        sport_contexts.append(
            {
                "sport": sport,
                **metrics,
                "markets": sorted(
                    {str(row.get("market_type") or "unknown") for row in sport_rows},
                    key=str.casefold,
                ),
            }
        )

    game_market_types: dict[str, set[str]] = {}
    for row in settled:
        game_market_types.setdefault(_performance_game_key(row), set()).add(
            str(row.get("market_type") or "unknown")
        )
    unique_games = len(game_market_types)
    multi_market_games = sum(1 for markets in game_market_types.values() if len(markets) > 1)

    # calibration: 10 buckets predicted vs actual
    calibration = []
    for index in range(10):
        lo, hi = index / 10, (index + 1) / 10
        members = [
            row
            for row in settled
            if (p := _pick_probability(row)) is not None and (lo <= p < hi or (hi == 1.0 and p == 1.0))
        ]
        if members:
            calibration.append(
                {
                    "bucket": f"{lo:.1f}-{hi:.1f}",
                    "mean_p": round(sum(_pick_probability(m) for m in members) / len(members), 4),
                    "hit_rate": round(sum(1 for m in members if m["result"] == "win") / len(members), 4),
                    "count": len(members),
                }
            )
    # streaks
    current = longest_w = longest_l = 0
    run_w = run_l = 0
    for row in settled:
        if row["result"] == "win":
            run_w, run_l = run_w + 1, 0
        else:
            run_l, run_w = run_l + 1, 0
        longest_w, longest_l = max(longest_w, run_w), max(longest_l, run_l)
    current = run_w if run_w else -run_l
    clv = [
        _number(row.get("probability_clv"), None)
        for row in settled
        if row.get("probability_clv") not in (None, "")
    ]
    clv = [value for value in clv if value is not None]
    brier_values = [
        (probability - (1.0 if row["result"] == "win" else 0.0)) ** 2
        for row in settled
        if (probability := _pick_probability(row)) is not None
    ]
    scored = [row for row in settled if _pick_is_scored(row)]
    staked = sum(max(_number(row.get("units")), _number(row.get("research_score_units"))) for row in scored)
    total_pnl = round(sum(_pick_pnl(row) for row in scored), 4) if scored else None
    event_dates = sorted(
        str(row.get("event_start_utc") or "")[:10] for row in settled if row.get("event_start_utc")
    )
    return {
        "total_picks": len(picks),
        "settled": len(settled),
        "open": sum(1 for row in picks if row.get("status") == "open"),
        "wins": wins,
        "losses": len(settled) - wins,
        "scored_calls": len(scored),
        "win_rate": round(wins / len(settled), 4) if settled else None,
        "total_pnl": total_pnl,
        "roi": round(total_pnl / staked, 4) if total_pnl is not None and staked else None,
        "pnl_curve": curve,
        "by_sport": by_sport,
        "by_market": by_market,
        "by_sport_market": by_sport_market,
        "by_confidence": {name: table for name, table in by_bucket.items()},
        "by_month": dict(sorted(by_month.items())),
        "calibration": calibration,
        "streaks": {"longest_win": longest_w, "longest_loss": longest_l, "current": current},
        "mean_clv": round(sum(clv) / len(clv), 6) if clv else None,
        "brier": round(sum(brier_values) / len(brier_values), 6) if brier_values else None,
        "context": {
            "sports": sport_contexts,
            "markets": market_contexts,
            "unique_games": unique_games,
            "multi_market_games": multi_market_games,
            "max_markets_per_game": max(
                (len(markets) for markets in game_market_types.values()),
                default=0,
            ),
            "probability_rows": len(brier_values),
            "clv_rows": len(clv),
            "period_start": event_dates[0] if event_dates else None,
            "period_end": event_dates[-1] if event_dates else None,
        },
    }


def performance_for_sport(picks: list[dict], sport: str | None = None) -> dict:
    """Return one ledger's performance, optionally scoped to an actual ledger sport."""
    available = sorted(
        {
            str(row.get("league") or "").strip()
            for row in picks
            if row.get("status") == "settled"
            and row.get("result") in ("win", "loss")
            and str(row.get("league") or "").strip()
        },
        key=str.casefold,
    )
    requested = str(sport or "").strip()
    selected = next(
        (candidate for candidate in available if candidate.casefold() == requested.casefold()),
        None,
    )
    if selected:
        scoped = [
            row for row in picks if str(row.get("league") or "").strip().casefold() == selected.casefold()
        ]
    else:
        scoped = [] if requested else picks
    payload = performance(scoped)
    payload["available_sports"] = available
    payload["selected_sport"] = selected or (requested if requested else None)
    return payload
