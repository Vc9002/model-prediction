"""model-prediction CLI.

Daily loop: polymarket-slate -> forecast --log -> settle --all-unsettled ->
summary (or `daily` for all of it). Everything is shadow/paper by default;
the only real-money path is the `execute` subcommand behind the hard gate in
``data_sources/polymarket_execute.py``.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path

from .audit import AuditLog
from .backtester import walk_forward_backtest
from .bans import TeamBanList
from .config import (
    PROJECT_ROOT,
    audit_path,
    config_path,
    entity_registry_path,
    ledger_path,
    load_config,
    market_odds_snapshot_path,
    polymarket_snapshot_path,
    unit_policy,
)
from .data_sources.espn import ESPNClient, ESPNMLBClient, SPORT_LEAGUES
from .data_sources.espn_wnba_injuries import capture_espn_event_injuries
from .data_sources.kalshi import DEFERRED_MESSAGE as KALSHI_DEFERRED_MESSAGE
from .data_sources.mlb_market_odds import MarketOddsSnapshotStore, MLBMarketOddsFeed
from .data_sources.polymarket_execute import (
    ExecutionGateError,
    OrderTicket,
    PolymarketExecutor,
)
from .data_sources.polymarket_us import (
    POLYMARKET_SPORT_LEAGUES,
    PolymarketSnapshotStore,
    PolymarketUSClient,
    capture_slate_snapshots,
    probability_to_american,
    refresh_contract_snapshots,
)
from .data_sources.the_odds_api import TheOddsAPIClient
from .data_sources.wnba_injuries import capture_latest_report
from .domain import (
    EASTERN,
    LOSS_CLASSIFICATIONS,
    League,
    MarketType,
    ModelOrigin,
    ModelState,
    PickRequest,
    RecordType,
    eastern_today,
    parse_utc,
    utc_now,
)
from .eligibility import EligibilityResult, evaluate_eligibility
from .entities import EntityRegistry, EntityResolutionError
from .esports import (
    TITLE_SPECS,
    backfill_esports,
    forecast_esports_slate,
    validate_all_esports_baselines,
)
from .features.base import FeatureStore
from .forward import build_mlb_slate
from .ingest import Ingestor
from .international_baseball import (
    LEAGUE_SPECS as INTERNATIONAL_BASEBALL_LEAGUE_SPECS,
    backfill_international_baseball,
    forecast_international_baseball_slate,
    validate_all_international_baseball_baselines,
)
from .ledger import LEDGER_SCHEMA_VERSION, DuplicatePickError, PickLedger
from .learned_forward import build_learned_moneyline_slate, match_executable_quote
from .models import MODEL_SPECS
from .models.market_residual import MarketResidualModel, ResidualTrainingRow
from .models.mlb import load_formula_spec
from .total_score import validate_all_total_score_models
from .units import edge_scaled_units
from .validation import run_validation_audit, write_production_artifacts


SPORTS = tuple(POLYMARKET_SPORT_LEAGUES)
ESPN_SPORTS = tuple(SPORT_LEAGUES)
LEARNED_PRODUCTION_SPORTS = ("mlb", "nba", "wnba", "nfl", "soccer", "lol", "cs2", "dota2", "valorant")
ESPORTS_TITLES = ("lol", "cs2", "dota2", "valorant")

_LEDGER_LOCK = threading.Lock()

# League value on a ledger row -> ESPN league key(s) to search for results.
_LEDGER_LEAGUE_TO_ESPN = {
    "MLB": ("MLB",),
    "NBA": ("NBA",),
    "WNBA": ("WNBA",),
    "NFL": ("NFL",),
    "SOCCER": ("EPL", "LA_LIGA", "BUNDESLIGA", "SERIE_A", "MLS", "UCL", "WORLD_CUP"),
    "WORLD_CUP": ("WORLD_CUP",),
}


def parser() -> argparse.ArgumentParser:
    root = argparse.ArgumentParser(prog="model-prediction")
    commands = root.add_subparsers(dest="command", required=True)

    commands.add_parser("init-ledger")

    report = commands.add_parser("report")
    report.add_argument(
        "--record-type", choices=["qualified", "research", "QUALIFIED_SHADOW_CALL", "RESEARCH_OBSERVATION"]
    )
    report.add_argument("--origin", choices=[item.value for item in ModelOrigin])
    report.add_argument("--model-version")
    report.add_argument("--calibration-version")
    report.add_argument("--league", choices=[item.value for item in League])
    report.add_argument("--market", choices=[item.value for item in MarketType])
    report.add_argument("--by-odds-range", action="store_true")

    commands.add_parser("exposure")
    commands.add_parser("models")
    commands.add_parser("summary")
    commands.add_parser(
        "live-portfolio",
        help="read authenticated exchange positions, activities, and balances",
    )
    order_status = commands.add_parser(
        "order-status",
        help="read authoritative state for submitted exchange orders",
    )
    order_status.add_argument("--order-id", action="append", required=True)

    slate = commands.add_parser(
        "polymarket-slate", help="read dated sports slates from the public Polymarket US API"
    )
    slate.add_argument("--sport", choices=SPORTS)
    slate.add_argument("--league", help="single gateway league key (e.g. MLB, EPL, WTA)")
    slate.add_argument("--all", action="store_true", help="every supported sport")
    slate.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    slate.add_argument("--timezone", default="America/New_York")
    slate.add_argument("--provider", default="polymarket", choices=["polymarket", "kalshi"])
    slate.add_argument(
        "--no-snapshot-bbo",
        action="store_true",
        help="disable the default prospective BBO capture for discovered contracts",
    )

    snapshot = commands.add_parser(
        "polymarket-snapshot", help="freeze a read-only BBO snapshot for later pregame CLV"
    )
    snapshot.add_argument("--slug", required=True)
    snapshot.add_argument("--sport", help="store under data/odds/{sport}/{date}/ instead of the flat file")

    ledger_prices = commands.add_parser(
        "polymarket-ledger-prices",
        help="refresh BBOs only for exact contracts selected from the open ledger",
    )
    ledger_prices.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    ledger_prices.add_argument(
        "--contract",
        action="append",
        default=[],
        metavar="SPORT[@GAME_DATE]=MARKET_SLUG",
        help="repeat for each unique open ledger contract; GAME_DATE is Eastern time",
    )

    clv = commands.add_parser("polymarket-clv", help="probability CLV from the final stored pregame snapshot")
    clv.add_argument("--slug", required=True)
    clv.add_argument("--side", required=True, choices=["long", "short"])
    clv.add_argument("--decision-price", required=True, type=float)
    clv.add_argument("--sport")
    clv.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")

    forecast = commands.add_parser(
        "forecast", help="pregame learned LR + confidence-gate moneyline slate"
    )
    forecast.add_argument("--sport", choices=SPORTS + ESPORTS_TITLES)
    forecast.add_argument("--all", action="store_true")
    forecast.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    forecast.add_argument("--log", action="store_true", help="log only rows with exact executable prices")
    forecast.add_argument(
        "--model",
        choices=("learned", "legacy-measured-edge"),
        default="learned",
        help="default learned production path; legacy option is MLB-only research rollback",
    )
    forecast.add_argument(
        "--replace-today",
        action="store_true",
        help="clear existing open picks for today before re-forecasting (default on daily)",
    )

    flat_forecast = commands.add_parser(
        "flat-forecast", help="forecast every game with no edge gate → flat_picks.xlsx"
    )
    flat_forecast.add_argument("--sport", choices=SPORTS + ESPORTS_TITLES)
    flat_forecast.add_argument("--all", action="store_true")
    flat_forecast.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    flat_forecast.add_argument("--log", action="store_true", help="log all calls to flat ledger")

    log_cmd = commands.add_parser("log", help="alias for forecast --log")
    log_cmd.add_argument("--sport", choices=SPORTS, default="mlb")
    log_cmd.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    log_cmd.add_argument(
        "--model", choices=("learned", "legacy-measured-edge"), default="learned"
    )

    settle = commands.add_parser("settle")
    settle.add_argument("--pick-id")
    settle.add_argument("--away-score", type=int)
    settle.add_argument("--home-score", type=int)
    settle.add_argument(
        "--all-unsettled", action="store_true", help="grade every started open pick from ESPN"
    )
    settle.add_argument("--void-postponed", action="store_true")
    settle.add_argument("--closing-line", type=float)
    settle.add_argument("--closing-american-odds", type=int)
    settle.add_argument("--closing-no-vig-probability", type=float)
    settle.add_argument("--closing-consensus-probability", type=float)
    settle.add_argument("--closing-consensus-line", type=float)

    daily = commands.add_parser("daily", help="slate + forecast + log + settle + summary in one run")
    daily.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")

    ingest = commands.add_parser("ingest", help="cache one date of ESPN scores locally")
    ingest.add_argument("--sport", required=True, choices=ESPN_SPORTS)
    ingest.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")

    availability = commands.add_parser(
        "wnba-availability-capture",
        help="archive the latest official WNBA injury PDF and normalized point-in-time rows",
    )
    availability.add_argument(
        "--observed-at",
        help="UTC-aware ISO timestamp; defaults to now and never selects a future report",
    )
    availability.add_argument(
        "--event-id",
        action="append",
        default=[],
        help="also archive ESPN event injury statuses; repeat for each WNBA game",
    )

    bootstrap = commands.add_parser("bootstrap", help="idempotent historical backfill from ESPN")
    bootstrap.add_argument("--sport", choices=ESPN_SPORTS)
    bootstrap.add_argument("--all", action="store_true")
    bootstrap.add_argument("--from", dest="from_date", required=True)
    bootstrap.add_argument("--to", dest="to_date")

    esports_backfill = commands.add_parser(
        "esports-backfill",
        help="backfill no-key series-level results for isolated LoL and CS2 research models",
    )
    esports_backfill.add_argument("--title", choices=tuple(TITLE_SPECS))
    esports_backfill.add_argument("--all", action="store_true")
    esports_backfill.add_argument("--from", dest="from_date", required=True)
    esports_backfill.add_argument("--to", dest="to_date")

    esports_validate = commands.add_parser(
        "validate-esports",
        help="select separate LoL/CS2 Elo baselines and grade chronological locked tests",
    )
    esports_validate.add_argument(
        "--titles", nargs="+", choices=tuple(TITLE_SPECS), default=tuple(TITLE_SPECS)
    )
    esports_validate.add_argument(
        "--output", default="outputs/latest/esports-baseline-validation.json"
    )
    esports_validate.add_argument("--write-artifacts", action="store_true")

    esports_forecast = commands.add_parser(
        "esports-forecast",
        help="zero-unit exact-identity LoL/CS2 prices for Polymarket US match-winner contracts",
    )
    esports_forecast.add_argument("--title", required=True, choices=tuple(TITLE_SPECS))
    esports_forecast.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    esports_forecast.add_argument("--timezone", default="America/New_York")

    international_backfill = commands.add_parser(
        "international-baseball-backfill",
        help="backfill official no-key KBO and NPB regular-season results",
    )
    international_backfill.add_argument(
        "--league", choices=tuple(INTERNATIONAL_BASEBALL_LEAGUE_SPECS)
    )
    international_backfill.add_argument("--all", action="store_true")
    international_backfill.add_argument("--from", dest="from_date", required=True)
    international_backfill.add_argument("--to", dest="to_date")

    international_validate = commands.add_parser(
        "validate-international-baseball",
        help="select separate tie-aware KBO/NPB Elo baselines and grade locked tests",
    )
    international_validate.add_argument(
        "--leagues",
        nargs="+",
        choices=tuple(INTERNATIONAL_BASEBALL_LEAGUE_SPECS),
        default=tuple(INTERNATIONAL_BASEBALL_LEAGUE_SPECS),
    )
    international_validate.add_argument(
        "--output", default="outputs/latest/international-baseball-baseline-validation.json"
    )
    international_validate.add_argument("--write-artifacts", action="store_true")

    international_forecast = commands.add_parser(
        "international-baseball-forecast",
        help="zero-unit tie-aware KBO/NPB fair values using exact Polymarket US BBOs",
    )
    international_forecast.add_argument(
        "--league", required=True, choices=tuple(INTERNATIONAL_BASEBALL_LEAGUE_SPECS)
    )
    international_forecast.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    international_forecast.add_argument("--timezone")

    entities = commands.add_parser("bootstrap-entities", help="merge ESPN team lists into the registry")
    entities.add_argument("--league", required=True)

    features = commands.add_parser("features", help="compute point-in-time feature snapshots")
    features.add_argument("--sport", required=True, choices=SPORTS)
    features.add_argument("--date", default=eastern_today().isoformat(),
        help="ISO date; defaults to today in US-Eastern time")
    features.add_argument("--refresh", action="store_true")

    backtest = commands.add_parser("backtest", help="walk-forward chronological backtest")
    backtest.add_argument("--sport", required=True, choices=SPORTS)
    backtest.add_argument("--start", required=True)
    backtest.add_argument("--end", required=True)
    backtest.add_argument(
        "--market-lines",
        help="optional historical two-sided quote JSONL; absent quotes disable ROI/CLV claims",
    )
    backtest.add_argument(
        "--confidence-threshold",
        type=float,
        help="selective call threshold; required to evaluate model qualification",
    )
    backtest.add_argument(
        "--locked-holdout",
        action="store_true",
        help="affirm this untouched evaluation window was locked before model selection",
    )
    backtest.add_argument("--output", help="optional JSON report destination under the project root")

    validate = commands.add_parser(
        "validate-models",
        help="learn thresholds on chronological validation data and grade a locked holdout",
    )
    validate.add_argument(
        "--sports",
        nargs="+",
        choices=("mlb", "nba", "wnba", "nfl", "soccer"),
        default=("mlb", "nba", "wnba", "nfl", "soccer"),
    )
    validate.add_argument(
        "--output",
        default="outputs/latest/learned-model-validation.json",
        help="JSON report destination under the project root",
    )
    validate.add_argument(
        "--write-artifacts",
        action="store_true",
        help="write hash-verified learned artifacts under config/models",
    )

    total_validate = commands.add_parser(
        "validate-totals",
        help="validate research-only point-in-time combined-score models",
    )
    total_validate.add_argument(
        "--sports",
        nargs="+",
        choices=("mlb", "nba", "wnba", "nfl"),
        default=("mlb", "nba", "wnba", "nfl"),
    )
    total_validate.add_argument(
        "--output",
        default="outputs/latest/total-score-validation.json",
    )
    total_validate.add_argument("--write-artifacts", action="store_true")

    reconstruct = commands.add_parser(
        "reconstruct-mlb-markets",
        help="cache postgame ESPN summaries for diagnostic-only historical market testing",
    )
    reconstruct.add_argument("--start", required=True)
    reconstruct.add_argument("--end", required=True)
    reconstruct.add_argument("--output", default="data/historical/mlb_market_lines_reconstructed.jsonl")

    residual = commands.add_parser(
        "train-residual", help="train the market-residual layer on the rolling settled window"
    )
    residual.add_argument("--output", default="config/models/market-residual-v1.json")

    execute = commands.add_parser(
        "execute",
        help="REAL MONEY: place an order for a qualified pick (hard gate; requires --execute)",
    )
    execute.add_argument("--pick-id", required=True)
    execute.add_argument("--size-shares", type=float, required=True)
    execute.add_argument("--price", type=float, required=True, help="executable ask/bid, 0-1")
    execute.add_argument("--side", default="long", choices=["long", "short"])
    execute.add_argument("--action", default="buy", choices=["buy", "sell"])
    execute.add_argument(
        "--order-type",
        default="limit_gtc",
        choices=["limit_gtc", "limit_ioc"],
    )
    execute.add_argument("--market-slug", required=True)
    execute.add_argument(
        "--execute",
        dest="execute_flag",
        action="store_true",
        help="actually place the order; without it this is a dry-run preview",
    )
    execute.add_argument(
        "--manual-research-order",
        action="store_true",
        help="explicitly authorize an active-model positive-edge research row as a manual order",
    )

    sell_position = commands.add_parser(
        "sell-position",
        help="place a resting SELL limit against a live exchange position (no model pick)",
    )
    sell_position.add_argument("--market-slug", required=True)
    sell_position.add_argument("--side", required=True, choices=["long", "short"])
    sell_position.add_argument("--price", type=float, required=True, help="limit price 0-1")
    sell_position.add_argument("--size-shares", type=float, required=True)
    sell_position.add_argument("--execute", dest="execute_flag", action="store_true")

    research_score = commands.add_parser("score-research")
    research_score.add_argument("--pick-id", action="append", dest="pick_ids")
    research_score.add_argument("--all-research", action="store_true")
    research_score.add_argument("--units", type=float, required=True)
    research_score.add_argument("--note", default="retrospective fixed-stake research scoring")

    call = commands.add_parser("call", help="freeze one pre-game prediction manually")
    call.add_argument("--league", required=True, choices=[item.value for item in League])
    call.add_argument("--event-id", required=True)
    call.add_argument("--start", required=True)
    call.add_argument("--away", required=True)
    call.add_argument("--home", required=True)
    call.add_argument("--market", required=True, choices=[item.value for item in MarketType])
    call.add_argument("--selection", required=True)
    call.add_argument("--line", type=float)
    call.add_argument("--american-odds", type=int, required=True)
    call.add_argument("--sportsbook", required=True)
    call.add_argument("--probability", type=float, required=True)
    call.add_argument("--model-uncertainty", type=float)
    call.add_argument("--model-version", required=True)
    call.add_argument(
        "--origin", choices=[item.value for item in ModelOrigin], default=ModelOrigin.ANALYST_ESTIMATE.value
    )
    call.add_argument(
        "--model-state", choices=[item.value for item in ModelState], default=ModelState.RESEARCH.value
    )
    call.add_argument("--baseline-id")
    call.add_argument("--observed-at")
    call.add_argument("--model-artifact-hash", default="")
    call.add_argument("--calibration-method", default="identity")
    call.add_argument("--calibration-version", default="identity-v1")
    call.add_argument("--calibration-artifact-hash", default="")
    call.add_argument("--feature-schema-version", default="1")
    call.add_argument("--code-revision", default="unknown")
    call.add_argument("--decision-no-vig-probability", type=float)
    call.add_argument("--decision-consensus-probability", type=float)
    call.add_argument("--decision-consensus-line", type=float)
    call.add_argument("--rationale", required=True)
    call.add_argument("--risks", default="")

    closing = commands.add_parser("update-closing")
    closing.add_argument("--pick-id", required=True)
    closing.add_argument("--closing-line", type=float)
    closing.add_argument("--closing-american-odds", type=int, required=True)
    closing.add_argument("--closing-no-vig-probability", type=float)
    closing.add_argument("--closing-consensus-probability", type=float)
    closing.add_argument("--closing-consensus-line", type=float)

    void = commands.add_parser("void")
    void.add_argument("--pick-id", required=True)
    void.add_argument("--reason", required=True)

    review = commands.add_parser("review-loss")
    review.add_argument("--pick-id", required=True)
    review.add_argument("--classification", required=True, choices=sorted(LOSS_CLASSIFICATIONS))
    review.add_argument("--cause", required=True)
    review.add_argument("--action", required=True)

    ban = commands.add_parser("ban-team")
    ban_commands = ban.add_subparsers(dest="ban_command", required=True)
    for action in ("add", "remove", "check"):
        child = ban_commands.add_parser(action)
        child.add_argument("--league", required=True, choices=[item.value for item in League])
        child.add_argument("--team", required=True)
        if action == "add":
            child.add_argument("--reason", default="manual_governance")
            child.add_argument("--review-after")
    ban_commands.add_parser("list")

    collect = commands.add_parser("collect-scores", help="pull recent soccer scores from The Odds API (free tier, last 3 days)")
    collect.add_argument("--days", type=int, default=3, help="days to look back (max 3 on free tier)")

    checklist = commands.add_parser(
        "verify-checklist",
        help="run the model_improvements.md section-13 verification checklist for a sport",
    )
    checklist.add_argument("--sport", required=True)
    return root


# ---------------------------------------------------------------------------
# Command implementations
# ---------------------------------------------------------------------------


def _polymarket_slate(args, config) -> dict:
    if args.provider == "kalshi":
        return {"provider": "kalshi", "status": "deferred", "note": KALSHI_DEFERRED_MESSAGE}
    client = PolymarketUSClient()
    game_date = date.fromisoformat(args.date)
    if args.league:
        events = {args.league.upper(): client.slate(args.league, game_date, args.timezone)}
    elif args.all:
        events = {}
        for sport in SPORTS:
            for league, slate in client.sport_slate(sport, game_date, args.timezone).items():
                events[league] = slate
    elif args.sport:
        events = client.sport_slate(args.sport, game_date, args.timezone)
    else:
        raise ValueError("provide --sport, --league, or --all")
    bbo_capture = (
        {"status": "disabled"}
        if getattr(args, "no_snapshot_bbo", False)
        else capture_slate_snapshots(
            client,
            events,
            Path(ledger_path(config)).parent,
            args.date,
        )
    )
    return {
        "provider": "polymarket_us",
        "game_date": args.date,
        "timezone": args.timezone,
        "events_by_league": events,
        "event_count": sum(len(items) for items in events.values()),
        "prospective_bbo_capture": bbo_capture,
        "note": (
            "Public market discovery plus prospective executable-BBO storage. "
            "Event-list quotes remain indicative."
        ),
    }


def _forecast_mlb(args_date: str, log: bool, config, registry, bans, ledger, audit) -> dict:
    """Legacy Measured Edge research path retained as an explicit rollback."""
    spec = load_formula_spec(PROJECT_ROOT / "config/models/mlb-analyst-poisson-trend-v0.2.yaml")
    observed_at = utc_now()
    odds_api_key = os.getenv("THE_ODDS_API_KEY")
    odds_feed = MLBMarketOddsFeed(
        registry,
        MarketOddsSnapshotStore(market_odds_snapshot_path(config)),
        odds_api=TheOddsAPIClient(odds_api_key) if odds_api_key else None,
        observed_at=observed_at,
    )
    candidates, skipped, scheduled = build_mlb_slate(
        args_date,
        ESPNMLBClient(),
        spec,
        PROJECT_ROOT / "config/models/measured-edge-margin-v1.json",
        PROJECT_ROOT / "config/models/measured-edge-totals-v1.json",
        observed_at,
        odds_feed,
    )
    for item in skipped:
        if "NO_CALL_MARKET_UNAVAILABLE" in item["reason"]:
            audit.append(
                "forecast_no_call",
                item["event_id"],
                {
                    "reason_code": "NO_CALL_MARKET_UNAVAILABLE",
                    "detail": item["reason"],
                    "game_date": args_date,
                },
            )
    logged, duplicates = [], []
    if log:
        for candidate in candidates:
            request = PickRequest(
                event_start_utc=candidate.event_start_utc,
                event_id=candidate.event_id,
                league=League.MLB,
                away_team=candidate.away_team,
                home_team=candidate.home_team,
                market_type=candidate.market_type,
                selection=candidate.selection,
                line=candidate.line,
                sportsbook=candidate.sportsbook,
                american_odds=candidate.american_odds,
                model_probability=candidate.shrunk_probability,
                model_uncertainty=candidate.uncertainty,
                model_version=candidate.model_version,
                rationale=candidate.rationale,
                risks=candidate.risks,
                model_origin=ModelOrigin.STATISTICAL_MODEL,
                model_state=ModelState.RESEARCH,
                observed_at_utc=candidate.observed_at_utc,
                model_artifact_hash=candidate.model_artifact_hash,
                calibration_method="flat_probability_shrinkage_toward_half",
                calibration_version=candidate.calibration_version,
                calibration_artifact_hash=candidate.model_artifact_hash,
                feature_schema_version=candidate.feature_schema_version,
                entity_map_version=registry.version,
                code_revision="measured-edge-paired-v1",
                decision_no_vig_probability=candidate.no_vig_probability,
            )
            request.validate(now=observed_at)
            away = registry.resolve(request.league, request.away_team, request.event_start_utc)
            home = registry.resolve(request.league, request.home_team, request.event_start_utc)
            eligibility = evaluate_eligibility(
                request,
                registry,
                bans,
                ledger.exposure(
                    request,
                    now=observed_at,
                    canonical_team_ids=(away.canonical_team_id, home.canonical_team_id),
                ),
                unit_policy(config),
                now=observed_at,
            )
            try:
                logged.append(ledger.append_evaluated(request, eligibility, now=observed_at))
            except DuplicatePickError as error:
                duplicates.append(error.pick_id)
    return {
        "sport": "mlb",
        "model_name": "Measured Edge Paired Models",
        "model_versions": ["measured-edge-margin-v1", "measured-edge-totals-v1"],
        "game_date": args_date,
        "scheduled_games": scheduled,
        "market_calls_created": len(candidates),
        "logged": len(logged),
        "duplicate_pick_ids": duplicates,
        "skipped": skipped,
        "candidates": [asdict(candidate) for candidate in candidates],
        "note": "All entries are zero-unit research; closing odds are attached only after start.",
    }


def _forecast_learned_sport(
    sport: str,
    args_date: str,
    log: bool,
    config,
    registry=None,
    bans=None,
    ledger=None,
    *,
    maximum_data_age_hours: float | None = None,
    maximum_unreviewed_disagreement: float | None = None,
    flat_mode: bool = False,
) -> dict:
    """Default production forecast path for audited learned moneyline models."""
    model_config = config["models"][sport.upper()]
    artifact_value = model_config.get("production_artifact")
    if not artifact_value:
        return {
            "sport": sport,
            "status": "no_production_artifact",
            "logged": 0,
            "candidates": [],
            "note": "Fail closed: no hash-verified learned artifact is configured.",
        }
    artifact_path = Path(artifact_value)
    if not artifact_path.is_absolute():
        artifact_path = PROJECT_ROOT / artifact_path
    try:
        candidates, skipped, scheduled = build_learned_moneyline_slate(
            sport=sport,
            game_date=args_date,
            store=FeatureStore(Path(ledger_path(config)).parent),
            client=ESPNClient(),
            artifact_path=artifact_path,
            observed_at=utc_now(),
        )
    except ValueError as error:
        return {
            "sport": sport,
            "status": "forecast_unavailable",
            "reason": str(error),
            "logged": 0,
            "candidates": [],
        }
    calls = [candidate for candidate in candidates if candidate.call]
    qualified_calls = [candidate for candidate in calls if candidate.model_qualified]
    research_calls = [candidate for candidate in calls if not candidate.model_qualified]
    # Flat mode: log every game regardless of confidence threshold.
    to_log = candidates if flat_mode else calls
    logged: list[dict] = []
    duplicates: list[str] = []
    unmatched: list[dict] = []
    edge_blocked: list[dict] = []
    if log and to_log and registry is not None and bans is not None and ledger is not None:
        data_root = Path(ledger_path(config)).parent
        observed_at = utc_now()
        # Main ledger: units come from evaluate_eligibility (fail-closed).
        # Flat ledger: every game gets diagnostic edge-scaled units.
        configured_state = str(model_config.get("status", "research"))
        for candidate in to_log:
            quote = match_executable_quote(data_root, sport, args_date, candidate)
            if quote is None:
                unmatched.append(
                    {"event_id": candidate.event_id,
                     "reason": "no stored executable moneyline BBO matched this matchup"}
                )
                continue
            # Gate: require minimum edge over executable Polymarket ask.
            # Per-sport minimum edge from model.yaml; defaults to 2% absolute.
            # Flat mode bypasses the edge gate — log every call regardless.
            if not flat_mode:
                min_edge = float(model_config.get("min_edge", 0.02))
                model_edge = candidate.model_probability - quote["executable_ask"]
                if model_edge < min_edge:
                    edge_pct = f"{min_edge*100:.0f}%"
                    edge_blocked.append(
                        {"event_id": candidate.event_id,
                         "reason": f"model edge {model_edge:.4f} below {edge_pct} minimum over executable ask {quote['executable_ask']:.4f}"}
                    )
                    continue
            # Convert UTC event time to Eastern for consistent ledger display
            try:
                event_et = datetime.fromisoformat(candidate.event_start_utc.replace('Z','+00:00')).astimezone(EASTERN).strftime('%Y-%m-%dT%H:%M:%S%z')
            except (ValueError, TypeError):
                event_et = candidate.event_start_utc
            request = PickRequest(
                event_start_utc=event_et,
                event_id=candidate.event_id,
                league=League(sport.upper()),
                away_team=candidate.away_team,
                home_team=candidate.home_team,
                market_type=MarketType.MONEYLINE,
                selection=candidate.selection,
                line=None,
                sportsbook="polymarket_us",
                american_odds=probability_to_american(quote["executable_ask"]),
                model_probability=candidate.model_probability,
                model_uncertainty=None,
                model_version=candidate.model_version,
                rationale=(
                    f"Learned LR call at threshold {candidate.confidence_threshold:.4f}; "
                    f"executable ask {quote['executable_ask']:.4f} "
                    f"({quote['market_slug']})."
                ),
                risks="Learned model; promotion gate not yet open; zero-unit research until review.",
                model_origin=ModelOrigin.STATISTICAL_MODEL,
                model_state=ModelState(configured_state),
                observed_at_utc=str(quote.get("observed_at_utc") or ""),
                model_artifact_hash=candidate.model_artifact_hash,
                calibration_method="learned_lr",
                calibration_version=candidate.model_version,
                calibration_artifact_hash=candidate.model_artifact_hash,
                feature_schema_version=candidate.feature_snapshot_hash[:16],
                entity_map_version=registry.version,
                code_revision=candidate.model_version,
                decision_no_vig_probability=quote.get("no_vig_probability"),
            )
            try:
                request.validate(now=observed_at)
                away = registry.resolve(request.league, request.away_team, request.event_start_utc)
                home = registry.resolve(request.league, request.home_team, request.event_start_utc)
                eligibility_kwargs: dict = {"now": observed_at}
                if maximum_data_age_hours is not None:
                    eligibility_kwargs["maximum_age_hours"] = maximum_data_age_hours
                if maximum_unreviewed_disagreement is not None:
                    eligibility_kwargs["maximum_unreviewed_disagreement"] = maximum_unreviewed_disagreement
                eligibility = evaluate_eligibility(
                    request, registry, bans,
                    ledger.exposure(
                        request, now=observed_at,
                        canonical_team_ids=(away.canonical_team_id, home.canonical_team_id),
                    ),
                    unit_policy(config), **eligibility_kwargs,
                )
                # Flat mode only: assign diagnostic edge-scaled units so the
                # separate flat ledger's one-unit accounting stays comparable.
                # The MAIN ledger honors evaluate_eligibility exactly — a gate
                # that returned 0 units did so for a reason (staleness,
                # disagreement, exposure caps) and must never be overridden.
                if flat_mode and eligibility.units <= 0:
                    forced_units = edge_scaled_units(
                        candidate.model_probability,
                        request.model_uncertainty or 0.05,
                        request.american_odds,
                        unit_policy(config),
                    )
                    eligibility = EligibilityResult(
                        eligibility.record_type,
                        eligibility.decision,
                        eligibility.reason_code,
                        forced_units,
                        eligibility.confidence_score,
                        eligibility.edge,
                        eligibility.adjusted_edge,
                        eligibility.away_team,
                        eligibility.home_team,
                    )
                with _LEDGER_LOCK:
                    logged.append(ledger.append_evaluated(request, eligibility, now=observed_at))
            except DuplicatePickError as error:
                duplicates.append(error.pick_id)
            except (EntityResolutionError, ValueError) as error:
                unmatched.append({"event_id": candidate.event_id, "reason": str(error)[:200]})
    if not log:
        logging_note = "Logging not requested."
    elif flat_mode:
        logging_note = (
            f"Flat mode: logged {len(logged)} of {len(candidates)} games "
            f"({len(calls)} above threshold); "
            f"{len(duplicates)} duplicates, {len(unmatched)} without a matched quote, "
            f"{len(edge_blocked)} blocked by edge gate."
        )
    elif not calls:
        logging_note = "No calls above the learned confidence threshold."
    else:
        logging_note = (
            f"Logged {len(logged)} of {len(calls)} calls against stored executable asks; "
            f"{len(duplicates)} duplicates, {len(unmatched)} without a matched quote, "
            f"{len(edge_blocked)} blocked by edge gate."
        )
    return {
        "sport": sport,
        "status": "learned_forecast_complete",
        "model_version": candidates[0].model_version if candidates else model_config["active_production_version"],
        "artifact": str(artifact_path),
        "game_date": args_date,
        "scheduled_games": scheduled,
        "calls": len(calls),
        "qualified_shadow_calls": len(qualified_calls),
        "zero_unit_research_calls": len(research_calls),
        "logged": len(logged),
        "logged_pick_ids": [row["pick_id"] for row in logged],
        "duplicate_pick_ids": duplicates,
        "unmatched_quotes": unmatched,
        "edge_blocked": edge_blocked,
        "skipped": skipped,
        "candidates": [candidate.to_dict() for candidate in candidates],
        "note": logging_note,
    }


def _log_esports_forecast(
    forecast: dict,
    config: dict,
    ledger: Any,
    flat_mode: bool = False,
) -> int:
    """Log esports contracts through the real eligibility gates.

    Esports promotion to shadow_qualified units is a DELIBERATE config
    decision (models.<TITLE>.status). This path enforces the same gates as
    every other sport — staleness, model/market disagreement, exposure caps,
    and unit-engine sizing — via ``evaluate_esports_eligibility``. Entity and
    ban resolution is name-based because esports teams are not yet in the
    canonical registry. Returns the count of logged rows.
    """
    from .data_sources.polymarket_us import probability_to_american
    from .eligibility import evaluate_esports_eligibility

    logged = 0
    model_config = config["models"].get(forecast["title"].upper(), {})
    min_edge = float(model_config.get("min_edge", 0.02))
    configured_state = str(model_config.get("status", "research"))
    title = forecast["title"].upper()
    league = League(title)
    observed_now = utc_now()

    for contract in forecast.get("priced_contracts", []):
        # Only log the model's pick: the side with higher model probability
        sides = contract.get("sides", [])
        if len(sides) != 2:
            continue
        best_side = max(sides, key=lambda s: float(s["model_probability"]))
        model_prob = float(best_side["model_probability"])
        ask = float(best_side["executable_ask"])
        edge = model_prob - ask

        # Edge gate for main ledger
        if not flat_mode and edge < min_edge:
            continue

        selected_team = str(best_side["team"])
        # Polymarket side ordering is arbitrary for venue-neutral esports;
        # teams[0]/teams[1] map to ledger home/away consistently with
        # settlement, which reconstructs the selected team the same way.
        teams = list(contract["teams"])
        home_team = teams[0]
        away_team = teams[1]
        pick_is_home = selected_team == home_team

        american_odds = probability_to_american(ask)
        request = PickRequest(
            event_start_utc=str(contract["event_start_utc"]),
            event_id=str(contract["event_id"]),
            league=league,
            away_team=away_team,
            home_team=home_team,
            market_type=MarketType.MONEYLINE,
            selection="home" if pick_is_home else "away",
            line=None,
            sportsbook="polymarket_us",
            american_odds=american_odds,
            model_probability=round(model_prob, 6),
            model_uncertainty=None,
            model_version=str(forecast["model_version"]),
            rationale=(
                f"Neutral Elo baseline; executable ask {ask:.4f} "
                f"(market_slug={contract['market_slug']})."
            ),
            risks="Config-promoted esports baseline; gates enforced at log time.",
            model_origin=ModelOrigin.STATISTICAL_MODEL,
            model_state=ModelState(configured_state),
            observed_at_utc=str(contract.get("observed_at_utc") or "") or None,
            model_artifact_hash=str(contract.get("artifact_hash", "")),
            calibration_method="neutral_elo",
            calibration_version=str(forecast["model_version"]),
            calibration_artifact_hash=str(contract.get("artifact_hash", "")),
            code_revision=str(forecast["model_version"]),
        )
        try:
            request.validate(now=observed_now)
            exposure = ledger.exposure(request, now=observed_now)
            eligibility = evaluate_esports_eligibility(
                request,
                exposure,
                unit_policy(config),
                now=observed_now,
                maximum_age_hours=float(config["project"].get("maximum_data_age_hours", 12)),
                maximum_unreviewed_disagreement=float(
                    config["project"].get("maximum_unreviewed_market_disagreement", 0.10)
                ),
            )
            if flat_mode and eligibility.units <= 0:
                # Flat ledger: diagnostic edge-scaled sizing, same as the
                # learned-sport flat path.
                from .eligibility import EligibilityResult as _ER
                eligibility = _ER(
                    eligibility.record_type,
                    eligibility.decision,
                    eligibility.reason_code,
                    edge_scaled_units(model_prob, 0.05, american_odds, unit_policy(config)),
                    eligibility.confidence_score,
                    eligibility.edge,
                    eligibility.adjusted_edge,
                    eligibility.away_team,
                    eligibility.home_team,
                )
            with _LEDGER_LOCK:
                ledger.append_evaluated(request, eligibility, now=observed_now)
            logged += 1
        except DuplicatePickError:
            continue
        except (ValueError, KeyError):
            continue

    return logged


def _forecast_research_sport(sport: str, args_date: str, config) -> dict:
    """Research-only preview for non-MLB sports from cached data. Never logs."""
    store = FeatureStore(Path(ledger_path(config)).parent)
    games = store.games_before(sport, args_date)
    if len(games) < 50:
        return {
            "sport": sport,
            "status": "insufficient_local_history",
            "cached_games_before_date": len(games),
            "note": (
                f"Run `model-prediction bootstrap --sport {sport} --from <season-start>` to build "
                "the local dataset. Research models never log qualified calls until they reach "
                "60% hit rate on 50+ locked-holdout calls with every called month positive."
            ),
        }
    from .models.registry import get_model

    model = get_model(sport)
    return {
        "sport": sport,
        "status": "research_preview_available",
        "model_version": model.version,
        "cached_games_before_date": len(games),
        "note": (
            "Model is RESEARCH state: predictions available programmatically; no ledger writes "
            "until backtest validation and lifecycle promotion."
        ),
    }


def _settle_all_unsettled(args, config, ledger) -> dict:
    """Grade every started open pick from ESPN scoreboards or Polymarket resolution."""
    now = utc_now()
    espn = ESPNClient()
    market_store = MarketOddsSnapshotStore(market_odds_snapshot_path(config))
    settled, voided, pending, failures = [], [], [], []
    for row in ledger.rows():
        if row["status"] != "open":
            continue
        try:
            start = parse_utc(row["event_start_utc"])
        except ValueError:
            failures.append({"pick_id": row["pick_id"], "reason": "bad event_start_utc"})
            continue
        if start > now:
            pending.append(row["pick_id"])
            continue
        # Esports: settle via Polymarket contract resolution
        if row["league"] in ("LOL", "CS2", "DOTA2", "VALORANT"):
            result = _settle_esports_pick(row, ledger)
            if result is None:
                pending.append(row["pick_id"])
            elif result.get("voided"):
                voided.append(result["pick_id"])
            elif result.get("settled"):
                settled.append(result)
            else:
                failures.append(result)
            continue
        leagues = _LEDGER_LEAGUE_TO_ESPN.get(row["league"], ())
        game_day = start.astimezone(EASTERN).date().isoformat()
        match = _find_espn_result(espn, leagues, game_day, row)
        if match is None:
            pending.append(row["pick_id"])
            continue
        status = match.get("status_name", "")
        if status in {"STATUS_POSTPONED", "STATUS_CANCELED"}:
            if args.void_postponed:
                voided.append(ledger.void(row["pick_id"], f"event {status.lower()}")["pick_id"])
            else:
                pending.append(row["pick_id"])
            continue
        if not match.get("completed"):
            pending.append(row["pick_id"])
            continue
        closing_line = closing_odds = closing_probability = None
        quote = market_store.closing_quote(
            row["event_id"], row["event_start_utc"], row["market_type"], row["selection"]
        )
        if quote is not None:
            closing_odds = int(quote["american_odds"])
            closing_probability = float(quote["decision_probability"])
            if quote.get("line") is not None:
                closing_line = float(quote["line"])
        try:
            result = ledger.settle(
                row["pick_id"],
                int(match["away_score"]),
                int(match["home_score"]),
                closing_line,
                closing_odds,
                closing_raw_probability=closing_probability,
            )
            settled.append({"pick_id": row["pick_id"], "result": result["result"]})
        except (KeyError, ValueError) as error:
            failures.append({"pick_id": row["pick_id"], "reason": str(error)})
    return {
        "settled": settled,
        "voided": voided,
        "still_open": pending,
        "failures": failures,
        "note": "Results pulled from ESPN scoreboards; closing prices from stored pregame BBO asks.",
    }


def _find_espn_result(espn: ESPNClient, leagues, game_day: str, row) -> dict | None:
    """Find a completed-game record matching a ledger row by id or team names."""
    away_names = {row["away_team"].casefold(), row["original_away_team"].casefold()}
    home_names = {row["home_team"].casefold(), row["original_home_team"].casefold()}
    for league in leagues:
        try:
            scoreboard = espn.scoreboard(league, game_day)
        except Exception:  # noqa: BLE001 - one bad league feed must not stop settlement
            continue
        for event in scoreboard.get("events", []):
            competition = (event.get("competitions") or [{}])[0]
            status = competition.get("status", {}).get("type", {})
            by_side = {item.get("homeAway"): item for item in competition.get("competitors", [])}
            away, home = by_side.get("away"), by_side.get("home")
            if not away or not home:
                continue
            id_match = str(event.get("id")) == row["event_id"]
            name_match = (
                away["team"].get("displayName", "").casefold() in away_names
                and home["team"].get("displayName", "").casefold() in home_names
            )
            # Prefer exact event_id match; name_match is a fallback only for
            # legacy/research rows with fake event_ids. Never use name_match
            # for rows with real (numeric) ESPN event IDs — it can mis-score
            # doubleheaders where both games share the same team names.
            row_eid = str(row.get("event_id", "") or "").strip()
            if id_match:
                pass  # exact match — always use this
            elif name_match and not row_eid.isdigit():
                pass  # fallback for legacy rows with fake event IDs
            else:
                continue
            record = {
                "status_name": status.get("name", ""),
                "completed": bool(status.get("completed")),
            }
            if record["completed"]:
                try:
                    record["away_score"] = int(float(away.get("score", 0) or 0))
                    record["home_score"] = int(float(home.get("score", 0) or 0))
                except (TypeError, ValueError):
                    record["completed"] = False
            return record
    return None


def _settle_esports_pick(row: dict, ledger) -> dict | None:
    """Settle an esports pick via Polymarket contract resolution. Returns None if pending."""
    import re
    from .data_sources.polymarket_us import PolymarketUSClient
    # Extract market slug from rationale: "... (slug)."
    rationale = str(row.get("rationale", ""))
    match = re.search(r"\(([a-z0-9\-]+)\)", rationale)
    if not match:
        return {"pick_id": row["pick_id"], "reason": "no market slug in rationale"}
    slug = match.group(1)
    try:
        snap = PolymarketUSClient().snapshot(slug)
    except Exception as e:
        return None  # API unavailable, leave pending
    resolution = snap.get("resolution")
    if not resolution:
        return None  # not resolved yet
    # Determine win/loss: did the team we picked win (contract resolved YES)?
    outcome = snap.get("outcome", "")
    # For moneyline: if resolution matches our selection side, it's a win
    # Contract resolves to the winning team's side
    selected_team = row["home_team"] if row["selection"] == "home" else row["away_team"]
    # Get closing price from resolution
    try:
        closing_price = float(snap.get("close_price", 0) or 0)
    except (ValueError, TypeError):
        closing_price = 0
    # Determine result
    if outcome == "Yes":
        # Contract resolved YES = the listed outcome team won
        resolved_team = snap.get("question", "").split("?")[0].strip() if "?" in snap.get("question","") else ""
        is_win = _identity_key(str(selected_team)) in _identity_key(str(resolved_team)) or \
                 _identity_key(str(resolved_team)) in _identity_key(str(selected_team))
    else:
        is_win = False
    try:
        result = ledger.settle(
            row["pick_id"],
            int(is_win), 0 if is_win else 1,  # away_score, home_score
            None,  # closing_line
            None,  # closing_american_odds
            closing_raw_probability=closing_price if closing_price else None,
        )
        return {"pick_id": row["pick_id"], "result": result["result"], "settled": True}
    except Exception as e:
        return {"pick_id": row["pick_id"], "reason": str(e)}


def _identity_key(value: str) -> str:
    return "".join(c.lower() for c in value if c.isalnum())


def _clear_today_open(ledger, date_str: str, by_event_date: bool = False) -> None:
    """Remove all open picks created on the given date before re-forecasting.
    
    When by_event_date is True, also removes open picks whose event_start_utc
    matches date_str — used for flat ledger to prevent duplicate forecast runs.
    """
    rows = ledger.rows()
    to_remove = []
    for row in rows:
        if row.get("status") != "open":
            continue
        created = str(row.get("created_at_utc", "") or "")
        event = str(row.get("event_start_utc", "") or "")
        if created.startswith(date_str):
            to_remove.append(row["pick_id"])
        elif by_event_date and event.startswith(date_str):
            to_remove.append(row["pick_id"])
    if to_remove:
        # Use openpyxl directly since ledger has no remove method
        import openpyxl as _xl
        wb = _xl.load_workbook(ledger.path)
        ws = wb[wb.sheetnames[0] if wb.sheetnames else "Picks"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        pid_col = headers.index("pick_id") + 1 if "pick_id" in headers else 1
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, pid_col).value or "") in to_remove:
                rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
        wb.save(ledger.path)


def _drift_check(settled_qualified: list, config: dict) -> dict:
    """Compare live settled hit rate against model holdout for each sport."""
    from collections import defaultdict
    import json
    import math
    from pathlib import Path as _Path

    by_sport = defaultdict(lambda: {"wins": 0, "losses": 0})
    for row in settled_qualified:
        sport = str((row.get("league") or row.get("sport") or "?")).upper()
        if row.get("result") == "win":
            by_sport[sport]["wins"] += 1
        elif row.get("result") == "loss":
            by_sport[sport]["losses"] += 1

    drift = {}
    for sport_name, counts in by_sport.items():
        n = counts["wins"] + counts["losses"]
        if n < 10:
            drift[sport_name] = {"status": "insufficient_sample", "n": n}
            continue

        live_hr = counts["wins"] / n
        holdout_hr = None
        model_path = _Path(f"config/models/{sport_name.lower()}-elo-trend-lr-v3.json")
        if model_path.exists():
            try:
                artifact = json.loads(model_path.read_text())
                holdout_hr = artifact.get("qualification", {}).get("hit_rate")
            except (json.JSONDecodeError, KeyError):
                pass

        if holdout_hr is None:
            drift[sport_name] = {"status": "no_holdout_reference", "live_hr": round(live_hr, 4), "n": n}
            continue

        # 2-sigma check
        se = math.sqrt(holdout_hr * (1 - holdout_hr) / n)
        z = (live_hr - holdout_hr) / se if se > 0 else 0
        status = "drifting" if z < -2.0 else ("excelling" if z > 2.0 else "on_track")

        drift[sport_name] = {
            "status": status,
            "live_hr": round(live_hr, 4),
            "holdout_hr": round(holdout_hr, 4),
            "n": n,
            "z_score": round(z, 2),
        }

    return drift


def _summary(config, ledger) -> dict:
    rows = ledger.rows()
    today = utc_now().astimezone(EASTERN).date().isoformat()

    def _et_day(value: str) -> str:
        try:
            return parse_utc(value).astimezone(EASTERN).date().isoformat()
        except (KeyError, ValueError):
            return ""

    created_today = [row for row in rows if _et_day(row["created_at_utc"]) == today]
    settled_today = [row for row in rows if row["settled_at_utc"] and _et_day(row["settled_at_utc"]) == today]
    open_rows = [row for row in rows if row["status"] == "open"]
    settled_qualified = [
        row for row in rows if row["status"] == "settled" and row["record_type"] == "QUALIFIED_SHADOW_CALL"
    ]
    pnl = sum(float(row["pnl_units"] or 0) for row in settled_qualified)
    clv_values = [float(row["probability_clv"]) for row in rows if row["probability_clv"]]
    measured_edge_rows = [row for row in rows if row["model_version"].startswith("measured-edge")]
    measured_edge_settled = [row for row in measured_edge_rows if row["status"] == "settled"]
    return {
        "date_et": today,
        "picks_logged_today": len(created_today),
        "picks_settled_today": len(settled_today),
        "open_picks": len(open_rows),
        "open_units": round(
            sum(
                float(row["units"] or 0) for row in open_rows if row["record_type"] == "QUALIFIED_SHADOW_CALL"
            ),
            2,
        ),
        "qualified_pnl_units_all_time": round(pnl, 4),
        "mean_probability_clv": round(sum(clv_values) / len(clv_values), 6) if clv_values else None,
        "bankroll_reference_units": config["bankroll"]["reference_units"],
        "research_progress": {
            "measured_edge_forward_picks": len(measured_edge_rows),
            "measured_edge_settled": len(measured_edge_settled),
            "iteration_policy": "continuous_no_parameter_freezes",
            "promotion_requires": "versioned walk-forward ablation and locked holdout",
        },
        "model_drift": _drift_check(settled_qualified, config),
        "note": "Shadow research accounting only; no real-money authorization.",
    }


def main(argv: list[str] | None = None) -> None:
    args = parser().parse_args(argv)
    config = load_config()
    audit = AuditLog(audit_path(config))
    registry = EntityRegistry.from_json(entity_registry_path(config))
    bans = TeamBanList(config_path(), registry, audit)
    research_scoring = config.get("research_scoring", {})
    research_score_units = (
        float(research_scoring["default_units"]) if research_scoring.get("auto_score_on_settlement") else None
    )
    ledger = PickLedger(
        ledger_path(config),
        audit_path(config),
        research_score_units=research_score_units,
        research_scoring_mode=str(research_scoring.get("sizing", "fixed")),
        research_scoring_note=research_scoring.get("note", "fixed-stake hypothetical research scoring"),
    )
    data_root = Path(ledger_path(config)).parent
    try:
        if args.command == "init-ledger":
            ledger.initialize()
            output = {"ledger": str(ledger.path), "status": "ready", "schema_version": LEDGER_SCHEMA_VERSION}
        elif args.command == "report":
            filters = {
                key: value
                for key, value in {
                    "record_type": args.record_type,
                    "origin": args.origin,
                    "model_version": args.model_version,
                    "calibration_version": args.calibration_version,
                    "league": args.league,
                    "market": args.market,
                }.items()
                if value is not None
            }
            output = ledger.report(filters, by_odds_range=args.by_odds_range)
        elif args.command == "models":
            output = {league.value: asdict(spec) for league, spec in MODEL_SPECS.items()}
        elif args.command == "summary":
            output = _summary(config, ledger)
        elif args.command == "live-portfolio":
            output = PolymarketExecutor(audit).portfolio_snapshot()
        elif args.command == "order-status":
            output = PolymarketExecutor(audit).order_snapshots(args.order_id)
        elif args.command == "polymarket-slate":
            output = _polymarket_slate(args, config)
        elif args.command == "polymarket-snapshot":
            snapshot = PolymarketUSClient().snapshot(args.slug)
            if args.sport:
                event_day = utc_now().astimezone(EASTERN).date().isoformat()
                store = PolymarketSnapshotStore.for_sport_date(data_root, args.sport, event_day)
            else:
                store = PolymarketSnapshotStore(polymarket_snapshot_path(config))
            store.append(snapshot)
            output = snapshot
        elif args.command == "polymarket-ledger-prices":
            contracts = []
            for value in args.contract:
                sport_day, separator, slug = value.partition("=")
                sport, day_separator, contract_day = sport_day.partition("@")
                if not separator or not sport or not slug:
                    raise ValueError("contract must use SPORT[@GAME_DATE]=MARKET_SLUG")
                if sport not in SPORTS:
                    raise ValueError(f"unsupported contract sport: {sport}")
                if day_separator:
                    date.fromisoformat(contract_day)
                contracts.append(
                    {
                        "sport": sport,
                        "market_slug": slug,
                        "game_date": contract_day if day_separator else args.date,
                    }
                )
            output = refresh_contract_snapshots(
                PolymarketUSClient(), contracts, data_root, args.date
            )
        elif args.command == "polymarket-clv":
            if not 0 < args.decision_price < 1:
                raise ValueError("decision price must be between 0 and 1")
            market = PolymarketUSClient().market(args.slug)
            if utc_now() < parse_utc(market["gameStartTime"]):
                raise ValueError("event has not started; a final pregame closing snapshot does not exist yet")
            if args.sport and args.date:
                store = PolymarketSnapshotStore.for_sport_date(data_root, args.sport, args.date)
            else:
                store = PolymarketSnapshotStore(polymarket_snapshot_path(config))
            closing = store.closing_snapshot(args.slug, market["gameStartTime"])
            if closing is None:
                raise ValueError("no stored pregame snapshot exists for this market")
            closing_price = closing[args.side]["price"]
            if closing_price is None:
                raise ValueError("stored pregame snapshot has no executable price for this side")
            output = {
                "market_slug": args.slug,
                "side": args.side,
                "decision_price_probability": args.decision_price,
                "decision_american_odds": probability_to_american(args.decision_price),
                "closing_price_probability": closing_price,
                "closing_american_odds": probability_to_american(closing_price),
                "probability_clv": round(closing_price - args.decision_price, 6),
                "closing_snapshot_observed_at_utc": closing["observed_at_utc"],
                "definition": "closing executable probability minus decision executable probability",
            }
        elif args.command in {"forecast", "log", "flat-forecast"}:
            log = args.command == "log" or getattr(args, "log", False) or args.command == "flat-forecast"
            replace_today = getattr(args, "replace_today", False) or args.command == "flat-forecast"
            is_flat = args.command == "flat-forecast"
            sports = list(SPORTS) + list(ESPORTS_TITLES) if getattr(args, "all", False) else [args.sport or "mlb"]
            if is_flat:
                # Flat forecast: separate ledger, no edge gate
                flat_ledger_path = Path(ledger_path(config)).parent / "flat_picks.xlsx"
                flat_ledger = PickLedger(flat_ledger_path)
                if replace_today and log:
                    _clear_today_open(flat_ledger, args.date, by_event_date=True)
            elif replace_today and log:
                _clear_today_open(ledger, args.date)
            results = {}
            for sport in sports:
                if sport == "esports":
                    continue  # handled individually as lol/cs2
                selected_model = getattr(args, "model", "learned")
                if selected_model == "legacy-measured-edge":
                    if sport != "mlb":
                        raise ValueError("legacy-measured-edge is available only for MLB")
                    results[sport] = _forecast_mlb(args.date, log, config, registry, bans, ledger, audit)
                elif sport in ESPORTS_TITLES:
                    results[sport] = forecast_esports_slate(
                        data_root=Path(ledger_path(config)).parent,
                        artifact_dir=PROJECT_ROOT / "config/models",
                        title=sport,
                        game_date=args.date,
                    )
                    if log and ledger is not None:
                        use_ledger = flat_ledger if is_flat else ledger
                        _log_esports_forecast(results[sport], config, use_ledger, flat_mode=is_flat)
                elif sport in LEARNED_PRODUCTION_SPORTS:
                    use_ledger = flat_ledger if is_flat else ledger
                    results[sport] = _forecast_learned_sport(
                        sport, args.date, log, config, registry, bans, use_ledger,
                        maximum_data_age_hours=float(config["project"].get("maximum_data_age_hours", 12)),
                        maximum_unreviewed_disagreement=float(
                            config["project"].get("maximum_unreviewed_market_disagreement", 0.10)
                        ),
                        flat_mode=is_flat,
                    )
                else:
                    results[sport] = _forecast_research_sport(sport, args.date, config)
            output = results[sports[0]] if len(sports) == 1 else results
        elif args.command == "settle":
            if args.all_unsettled:
                output = _settle_all_unsettled(args, config, ledger)
            else:
                if not args.pick_id or args.away_score is None or args.home_score is None:
                    raise ValueError("provide --pick-id with --away-score/--home-score, or --all-unsettled")
                closing_line = args.closing_line
                closing_odds = args.closing_american_odds
                closing_probability = None
                if closing_odds is None:
                    row = next((r for r in ledger.rows() if r["pick_id"] == args.pick_id), None)
                    if row is None:
                        raise KeyError(f"unknown pick id: {args.pick_id}")
                    quote = MarketOddsSnapshotStore(market_odds_snapshot_path(config)).closing_quote(
                        row["event_id"], row["event_start_utc"], row["market_type"], row["selection"]
                    )
                    if quote is not None:
                        closing_odds = int(quote["american_odds"])
                        closing_probability = float(quote["decision_probability"])
                        if quote.get("line") is not None:
                            closing_line = float(quote["line"])
                output = ledger.settle(
                    args.pick_id,
                    args.away_score,
                    args.home_score,
                    closing_line,
                    closing_odds,
                    closing_no_vig_probability=args.closing_no_vig_probability,
                    closing_consensus_probability=args.closing_consensus_probability,
                    closing_consensus_line=args.closing_consensus_line,
                    closing_raw_probability=closing_probability,
                )
        elif args.command == "daily":
            slate_args = argparse.Namespace(
                provider="polymarket",
                league=None,
                all=True,
                sport=None,
                date=args.date,
                timezone="America/New_York",
                no_snapshot_bbo=False,
            )
            slate = _polymarket_slate(slate_args, config)
            # Run WNBA availability capture, prior build, and soccer collection
            # in parallel with the slate (all I/O-bound).
            from .data_sources.odds_soccer_scores import collect_soccer_scores
            wnba_priors_result = {"status": "skipped"}
            soccer_collection = {}
            def _capture_wnba():
                try:
                    from .data_sources.espn import ESPNClient
                    wnba_scoreboard = ESPNClient().scoreboard("WNBA", args.date)
                    wnba_event_ids = [
                        str(event["id"]) for event in wnba_scoreboard.get("events", [])
                    ]
                    if wnba_event_ids:
                        capture_latest_report(data_root, observed_at=utc_now())
                        for event_id in wnba_event_ids:
                            try:
                                capture_espn_event_injuries(
                                    data_root, event_id=event_id,
                                    client=ESPNClient(), observed_at=utc_now(),
                                )
                            except Exception:
                                pass
                except Exception:
                    pass
            def _build_priors():
                nonlocal wnba_priors_result
                try:
                    from .features.base import FeatureStore
                    from .data_sources.espn import ESPNClient
                    from .wnba_availability_evaluation import build_and_save_priors
                    wnba_priors_result = build_and_save_priors(
                        store=FeatureStore(data_root),
                        client=ESPNClient(),
                        game_date=args.date,
                        data_root=data_root,
                    )
                except Exception:
                    pass
            def _collect_soccer():
                nonlocal soccer_collection
                try:
                    soccer_collection = collect_soccer_scores(days_from=3)
                except Exception:
                    pass
            with ThreadPoolExecutor(max_workers=3) as io_pool:
                f1 = io_pool.submit(_capture_wnba)
                f2 = io_pool.submit(_build_priors)
                f3 = io_pool.submit(_collect_soccer)
                for f in (f1, f2, f3):
                    f.result()  # Wait for all, surface exceptions
            _clear_today_open(ledger, args.date, by_event_date=True)
            # Also clear and forecast for flat ledger
            flat_ledger_path = Path(ledger_path(config)).parent / "flat_picks.xlsx"
            flat_ledger = PickLedger(flat_ledger_path)
            _clear_today_open(flat_ledger, args.date, by_event_date=True)
            LEARNED_SPORTS = ("mlb", "nba", "wnba", "nfl", "soccer")
            max_data_age = float(config["project"].get("maximum_data_age_hours", 12))
            max_disagreement = float(config["project"].get("maximum_unreviewed_market_disagreement", 0.10))
            # Compute forecasts once, log to both ledgers (compute > log main > log flat)
            forecast_result = {}
            workers = min(len(LEARNED_SPORTS), 5)
            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = {}
                for sport in LEARNED_SPORTS:
                    futures[pool.submit(
                        _forecast_learned_sport,
                        sport, args.date, True, config, registry, bans, ledger,
                        maximum_data_age_hours=max_data_age,
                        maximum_unreviewed_disagreement=max_disagreement,
                    )] = sport
                for future in as_completed(futures):
                    sport = futures[future]
                    try:
                        forecast_result[sport] = future.result()
                    except Exception as exc:
                        forecast_result[sport] = {
                            "sport": sport, "status": "error", "reason": str(exc),
                            "logged": 0, "candidates": [],
                        }
            # Re-log computed candidates to flat ledger (flat mode, no edge gate)
            with ThreadPoolExecutor(max_workers=workers) as pool:
                flat_futures = {}
                for sport in LEARNED_SPORTS:
                    result = forecast_result.get(sport, {})
                    candidates = result.get("candidates", [])
                    if candidates:
                        flat_futures[pool.submit(
                            _forecast_learned_sport,
                            sport, args.date, True, config, registry, bans, flat_ledger,
                            maximum_data_age_hours=max_data_age,
                            maximum_unreviewed_disagreement=max_disagreement,
                            flat_mode=True,
                        )] = sport
                for future in as_completed(flat_futures):
                    sport = flat_futures[future]
                    try:
                        forecast_result[f"_flat_{sport}"] = future.result()
                    except Exception:
                        pass
            # Esports run serially
            for title in ESPORTS_TITLES:
                forecast_result[title] = forecast_esports_slate(
                    data_root=Path(ledger_path(config)).parent,
                    artifact_dir=PROJECT_ROOT / "config/models",
                    title=title,
                    game_date=args.date,
                )
                _log_esports_forecast(forecast_result[title], config, ledger, flat_mode=False)
                _log_esports_forecast(forecast_result[title], config, flat_ledger, flat_mode=True)
            flat_result = {sport: forecast_result.get(f"_flat_{sport}", forecast_result.get(sport, {})) for sport in LEARNED_SPORTS}
            for title in ESPORTS_TITLES:
                flat_result[title] = forecast_result[title]
            for result in forecast_result.values():
                result.pop("candidates", None)
            # Read back stored Polymarket odds snapshots for per-sport summaries.
            odds_by_sport = {}
            for sport in LEARNED_PRODUCTION_SPORTS:
                odds_sport = "esports" if sport in ESPORTS_TITLES else sport
                snap_path = (
                    Path(ledger_path(config)).parent
                    / "odds" / odds_sport / args.date / "polymarket_snapshots.jsonl"
                )
                if snap_path.exists():
                    snaps = [
                        json.loads(line)
                        for line in snap_path.read_text(encoding="utf-8").strip().split("\n")
                        if line.strip()
                    ]
                    odds_by_sport[sport] = {
                        "snapshots": len(snaps),
                        "moneyline_snapshots": sum(
                            1 for s in snaps if s.get("market_type") == "moneyline"
                        ),
                        "spread_snapshots": sum(
                            1 for s in snaps if s.get("market_type") == "spread"
                        ),
                        "total_snapshots": sum(
                            1 for s in snaps if s.get("market_type") == "total"
                        ),
                        "with_executable_bbo": sum(
                            1 for s in snaps
                            if s.get("long", {}).get("ask") is not None
                            and s.get("short", {}).get("ask") is not None
                        ),
                    }
            settle_args = argparse.Namespace(all_unsettled=True, void_postponed=False)
            settlement = _settle_all_unsettled(settle_args, config, ledger)

            for result in flat_result.values():
                result.pop("candidates", None)
            flat_settlement = _settle_all_unsettled(settle_args, config, flat_ledger)

            output = {
                "date": args.date,
                "step1_polymarket_search": {
                    "event_count": slate["event_count"],
                    "leagues_with_events": [
                        league for league, items in slate["events_by_league"].items() if items
                    ],
                    "bbo_capture": slate.get("prospective_bbo_capture", {}),
                },
                "step2_3_forecast_and_log": forecast_result,
                "step3b_polymarket_odds_snapshots": odds_by_sport,
                "step4_settlement": settlement,
                "step1b_soccer_scores": soccer_collection,
                "step5_summary": _summary(config, ledger),
                "step5b_wnba_availability": {
                    "priors": wnba_priors_result,
                },
                "step6_flat_forecast_and_log": flat_result,
                "step7_flat_settlement": flat_settlement,
            }
        elif args.command == "ingest":
            output = Ingestor(data_root, audit=audit).ingest_scores(args.sport, args.date)
        elif args.command == "wnba-availability-capture":
            availability_observed_at = (
                parse_utc(args.observed_at) if args.observed_at else utc_now()
            )
            output = {
                "official_report": capture_latest_report(
                    data_root,
                    observed_at=availability_observed_at,
                ),
                "espn_events": [
                    capture_espn_event_injuries(
                        data_root,
                        event_id=event_id,
                        client=ESPNClient(),
                        observed_at=availability_observed_at,
                    )
                    for event_id in args.event_id
                ],
            }
        elif args.command == "bootstrap":
            ingestor = Ingestor(data_root, audit=audit)
            if args.all:
                output = {
                    sport: ingestor.bootstrap(sport, args.from_date, args.to_date)
                    for sport in ESPN_SPORTS
                }
            elif args.sport:
                output = ingestor.bootstrap(args.sport, args.from_date, args.to_date)
            else:
                raise ValueError("provide --sport or --all")
        elif args.command == "esports-backfill":
            if args.all:
                titles = tuple(TITLE_SPECS)
            elif args.title:
                titles = (args.title,)
            else:
                raise ValueError("provide --title or --all")
            output = {
                title: backfill_esports(data_root, title, args.from_date, args.to_date)
                for title in titles
            }
        elif args.command == "validate-esports":
            output = validate_all_esports_baselines(
                data_root,
                args.titles,
                PROJECT_ROOT / "config/models" if args.write_artifacts else None,
            )
            destination = PROJECT_ROOT / args.output
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(
                json.dumps(output, indent=2, sort_keys=True, default=str) + "\n",
                encoding="utf-8",
            )
            output["report_path"] = str(destination)
        elif args.command == "esports-forecast":
            output = forecast_esports_slate(
                data_root,
                PROJECT_ROOT / "config/models",
                args.title,
                args.date,
                args.timezone,
            )
        elif args.command == "international-baseball-backfill":
            if args.all:
                leagues = tuple(INTERNATIONAL_BASEBALL_LEAGUE_SPECS)
            elif args.league:
                leagues = (args.league,)
            else:
                raise ValueError("provide --league or --all")
            output = {
                league: backfill_international_baseball(
                    data_root, league, args.from_date, args.to_date
                )
                for league in leagues
            }
        elif args.command == "validate-international-baseball":
            output = validate_all_international_baseball_baselines(
                data_root,
                args.leagues,
                PROJECT_ROOT / "config/models" if args.write_artifacts else None,
            )
            destination = PROJECT_ROOT / args.output
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(
                json.dumps(output, indent=2, sort_keys=True, default=str) + "\n",
                encoding="utf-8",
            )
            output["report_path"] = str(destination)
        elif args.command == "international-baseball-forecast":
            output = forecast_international_baseball_slate(
                data_root,
                PROJECT_ROOT / "config/models",
                args.league,
                args.date,
                args.timezone,
            )
        elif args.command == "bootstrap-entities":
            output = Ingestor(data_root, audit=audit).bootstrap_entities(
                args.league, entity_registry_path(config)
            )
        elif args.command == "features":
            store = FeatureStore(data_root)
            snapshots = store.compute_all(args.sport, args.date)
            output = {
                name: {"computation_hash": snap["computation_hash"], "input_games": snap["input_games"]}
                for name, snap in snapshots.items()
            }
        elif args.command == "backtest":
            output = walk_forward_backtest(
                FeatureStore(data_root),
                args.sport,
                args.start,
                args.end,
                market_lines_path=args.market_lines,
                confidence_threshold=args.confidence_threshold,
                locked_holdout=args.locked_holdout,
            )
            if args.output:
                destination = PROJECT_ROOT / args.output
                destination.parent.mkdir(parents=True, exist_ok=True)
                destination.write_text(
                    json.dumps(output, indent=2, sort_keys=True, default=str) + "\n",
                    encoding="utf-8",
                )
                output["report_path"] = str(destination)
        elif args.command == "validate-models":
            output = run_validation_audit(
                FeatureStore(data_root),
                args.sports,
                PROJECT_ROOT / "data/historical/mlb_market_lines_reconstructed.jsonl",
            )
            destination = PROJECT_ROOT / args.output
            destination.parent.mkdir(parents=True, exist_ok=True)
            if args.write_artifacts:
                output["production_artifacts"] = write_production_artifacts(
                    output, PROJECT_ROOT / "config/models"
                )
            destination.write_text(
                json.dumps(output, indent=2, sort_keys=True, default=str) + "\n",
                encoding="utf-8",
            )
            output["report_path"] = str(destination)
        elif args.command == "validate-totals":
            output = validate_all_total_score_models(
                FeatureStore(data_root),
                args.sports,
                PROJECT_ROOT / "config/models" if args.write_artifacts else None,
            )
            destination = PROJECT_ROOT / args.output
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(
                json.dumps(output, indent=2, sort_keys=True, default=str) + "\n",
                encoding="utf-8",
            )
            output["report_path"] = str(destination)
        elif args.command == "reconstruct-mlb-markets":
            output = Ingestor(data_root, audit=audit).reconstruct_mlb_markets(
                args.start,
                args.end,
                PROJECT_ROOT / args.output,
            )
        elif args.command == "train-residual":
            rows = [
                ResidualTrainingRow(
                    settled_at_utc=row["settled_at_utc"],
                    model_probability=float(row["model_probability"]),
                    market_probability=float(
                        row["decision_no_vig_probability"]
                        or row["decision_raw_implied_probability"]
                        or row["market_implied_probability"]
                    ),
                    outcome=1 if row["result"] == "win" else 0,
                )
                for row in ledger.rows()
                if row["status"] == "settled"
                and row["result"] in {"win", "loss"}
                and row["settled_at_utc"]
                and row["model_probability"]
            ]
            model = MarketResidualModel.train(rows)
            destination = PROJECT_ROOT / args.output
            model.save(destination)
            output = {
                "model_version": model.version,
                "sample_size": model.sample_size,
                "identity_fallback": model.is_identity,
                "artifact": str(destination),
            }
        elif args.command == "execute":
            row = next((r for r in ledger.rows() if r["pick_id"] == args.pick_id), None)
            if row is None:
                raise KeyError(f"unknown pick id: {args.pick_id}")
            manual = bool(args.manual_research_order)
            execution_config = config.get("execution", {})
            if manual:
                if not execution_config.get("allow_manual_research_orders", False):
                    raise ExecutionGateError("REFUSED: manual research orders are disabled in config.")
                league = League(row["league"])
                active_version = (config.get("models", {}).get(league.value, {}) or {}).get(
                    "active_production_version"
                )
                if (
                    execution_config.get("manual_research_require_active_model", True)
                    and row.get("model_version") != active_version
                ):
                    raise ExecutionGateError(
                        "REFUSED: manual order row was not produced by the active model version."
                    )
                model_probability = float(row.get("model_probability") or 0)
                market_probability = float(row.get("market_implied_probability") or 0)
                edge = model_probability - market_probability
                if (
                    execution_config.get("manual_research_require_positive_edge", True)
                    and edge <= 0
                ):
                    raise ExecutionGateError("REFUSED: manual research order has no positive edge.")
                # The "limit below model probability" rule is a BUY guard (do not
                # pay more than fair value). A SELL is an exit at a target price
                # and is exempt.
                if (
                    args.action == "buy"
                    and execution_config.get("manual_research_require_positive_edge", True)
                    and args.price >= model_probability
                ):
                    raise ExecutionGateError(
                        "REFUSED: manual buy limit must remain below the model probability."
                    )
                for team_key in ("away_team", "home_team"):
                    _, banned = bans.check(league, row[team_key])
                    if banned:
                        raise ExecutionGateError(
                            f"REFUSED: {row[team_key]} is on the permanent team ban list."
                        )
            policy = unit_policy(config)
            if manual:
                maximum_units = edge_scaled_units(
                    float(row["model_probability"]),
                    float(row.get("model_uncertainty") or 0),
                    int(row["american_odds"]),
                    policy,
                )
            else:
                maximum_units = float(row.get("units") or 0)
            # A SELL is an exit — it returns capital, so no dollar cost cap
            # applies. Buys keep the unit cap.
            maximum_cost = None if args.action == "sell" else round(maximum_units * policy.unit_value_usd, 2)
            ticket = OrderTicket(
                market_slug=args.market_slug,
                token_side=args.side,
                action=args.action,
                order_type=args.order_type,
                price=args.price,
                size_shares=args.size_shares,
                pick_id=args.pick_id,
                estimated_cost_usd=round(args.price * args.size_shares, 2),
                maximum_cost_usd=maximum_cost,
                authorization_type=("manual_research_override" if manual else "qualified_model"),
            )
            executor = PolymarketExecutor(audit)
            output = executor.execute(
                ticket,
                row,
                execute_flag=args.execute_flag,
                user_command=True,
                manual_research_order=manual,
            )
        elif args.command == "sell-position":
            # Close a live exchange position you already hold. There is no
            # model pick and no unit cap (a sell returns capital); the hard
            # gates (credentials, post-only, confirmation, audit) still apply.
            ticket = OrderTicket(
                market_slug=args.market_slug,
                token_side=args.side,
                action="sell",
                order_type="limit_gtc",
                price=args.price,
                size_shares=args.size_shares,
                pick_id=f"live-position:{args.market_slug}:{args.side}",
                estimated_cost_usd=round(args.price * args.size_shares, 2),
                maximum_cost_usd=None,
                authorization_type="live_position_exit",
            )
            synthetic_row = {
                "record_type": "LIVE_POSITION_EXIT",
                "status": "open",
                "reason_code": "LIVE_POSITION_EXIT",
            }
            output = PolymarketExecutor(audit).execute(
                ticket,
                synthetic_row,
                execute_flag=args.execute_flag,
                user_command=True,
                manual_research_order=True,
            )
        elif args.command == "exposure":
            rows = ledger.rows()
            qualified_open = [
                row
                for row in rows
                if row["record_type"] == "QUALIFIED_SHADOW_CALL" and row["status"] == "open"
            ]
            output = {
                "open_units": round(sum(float(row["units"] or 0) for row in qualified_open), 2),
                "open_qualified_calls": len(qualified_open),
                "zero_unit_research_observations": sum(
                    row["record_type"] == "RESEARCH_OBSERVATION" for row in rows
                ),
                "starting_bankroll_units": config["bankroll"]["reference_units"],
                "note": "Shadow research accounting only; no real-money authorization",
            }
        elif args.command == "ban-team":
            output = _handle_ban(args, bans)
        elif args.command == "collect-scores":
            from .data_sources.odds_soccer_scores import collect_soccer_scores
            output = collect_soccer_scores(days_from=args.days)
        elif args.command == "verify-checklist":
            from .source_policy import DEFAULT_SOURCES
            from .verification_checklist import format_checklist, run_checklist

            sport = args.sport.lower()
            source_keys = [key for key, spec in DEFAULT_SOURCES.items() if sport in spec.leagues]
            artifact_hashes: list[str] = []
            model_config = config["models"].get(sport.upper())
            if model_config and model_config.get("production_artifact"):
                artifact_path = PROJECT_ROOT / model_config["production_artifact"]
                if artifact_path.exists():
                    artifact = json.loads(artifact_path.read_text(encoding="utf-8"))
                    if artifact.get("artifact_hash"):
                        artifact_hashes.append(artifact["artifact_hash"])
            research_units = [
                float(row["units"] or 0)
                for row in ledger.rows()
                if row["league"].lower() == sport and row["record_type"] == RecordType.RESEARCH_OBSERVATION.value
            ]
            results = run_checklist(
                source_keys=source_keys or None,
                artifact_hashes=artifact_hashes or None,
                research_units=research_units or None,
            )
            output = {
                "sport": sport,
                "results": [vars(result) for result in results],
                "formatted": format_checklist(results),
            }
        elif args.command == "score-research":
            if bool(args.pick_ids) == bool(args.all_research):
                raise ValueError("provide either --pick-id or --all-research")
            pick_ids = args.pick_ids or [
                row["pick_id"]
                for row in ledger.rows()
                if row["record_type"] == "RESEARCH_OBSERVATION" and row["status"] == "settled"
            ]
            rows = ledger.score_research(pick_ids, args.units, args.note)
            output = {
                "scored_records": len(rows),
                "research_score_units_each": args.units,
                "research_pnl_units": round(sum(float(row["research_pnl_units"]) for row in rows), 6),
                "note": args.note,
            }
        elif args.command == "call":
            request = PickRequest(
                event_start_utc=args.start,
                event_id=args.event_id,
                league=League(args.league),
                away_team=args.away,
                home_team=args.home,
                market_type=MarketType(args.market),
                selection=args.selection,
                line=args.line,
                sportsbook=args.sportsbook,
                american_odds=args.american_odds,
                model_probability=args.probability,
                model_uncertainty=args.model_uncertainty,
                model_version=args.model_version,
                rationale=args.rationale,
                risks=args.risks,
                model_origin=ModelOrigin(args.origin),
                model_state=ModelState(args.model_state),
                baseline_identifier=args.baseline_id,
                observed_at_utc=args.observed_at,
                model_artifact_hash=args.model_artifact_hash,
                calibration_method=args.calibration_method,
                calibration_version=args.calibration_version,
                calibration_artifact_hash=args.calibration_artifact_hash,
                feature_schema_version=args.feature_schema_version,
                entity_map_version=registry.version,
                code_revision=args.code_revision,
                decision_no_vig_probability=args.decision_no_vig_probability,
                decision_consensus_probability=args.decision_consensus_probability,
                decision_consensus_line=args.decision_consensus_line,
            )
            request.validate()
            away = registry.resolve(request.league, request.away_team, request.event_start_utc)
            home = registry.resolve(request.league, request.home_team, request.event_start_utc)
            exposure = ledger.exposure(
                request, canonical_team_ids=(away.canonical_team_id, home.canonical_team_id)
            )
            eligibility = evaluate_eligibility(
                request,
                registry,
                bans,
                exposure,
                unit_policy(config),
                maximum_age_hours=float(config["project"].get("maximum_data_age_hours", 12)),
                maximum_unreviewed_disagreement=float(
                    config["project"].get("maximum_unreviewed_market_disagreement", 0.10)
                ),
            )
            row = ledger.append_evaluated(request, eligibility)
            output = {
                "decision": eligibility.decision,
                "record_type": eligibility.record_type.value,
                "reason_code": eligibility.reason_code,
                "units": eligibility.units,
                "pick": row,
            }
        elif args.command == "update-closing":
            output = ledger.update_closing(
                args.pick_id,
                args.closing_line,
                args.closing_american_odds,
                closing_no_vig_probability=args.closing_no_vig_probability,
                closing_consensus_probability=args.closing_consensus_probability,
                closing_consensus_line=args.closing_consensus_line,
            )
        elif args.command == "void":
            output = ledger.void(args.pick_id, args.reason)
        elif args.command == "review-loss":
            output = ledger.review_loss(args.pick_id, args.classification, args.cause, args.action)
        else:
            raise ValueError(f"unknown command: {args.command}")
        print(json.dumps(output, indent=2, sort_keys=True, default=str))
    except ExecutionGateError as error:
        print(json.dumps({"status": "refused", "error": str(error)}, indent=2), file=sys.stderr)
        raise SystemExit(3)
    except DuplicatePickError as error:
        audit.append("pick_rejected", error.pick_id, {"reason_code": "NO_CALL_DUPLICATE"})
        print(
            json.dumps(
                {
                    "decision": "NO_CALL",
                    "reason_code": "NO_CALL_DUPLICATE",
                    "existing_pick_id": error.pick_id,
                },
                indent=2,
            )
        )
    except EntityResolutionError as error:
        audit.append(
            "pick_rejected", "entity", {"reason_code": "NO_CALL_ENTITY_UNRESOLVED", "detail": str(error)}
        )
        _fail(str(error), "NO_CALL_ENTITY_UNRESOLVED")
    except (KeyError, ValueError) as error:
        if args.command == "polymarket-clv":
            reason = "POLYMARKET_CLV_NOT_AVAILABLE"
        else:
            reason = "NO_CALL_EVENT_STARTED" if "started" in str(error) else "NO_CALL_INVALID_MARKET"
        _fail(str(error), reason)


def _handle_ban(args: argparse.Namespace, bans: TeamBanList) -> object:
    if args.ban_command == "list":
        return [asdict(entry) for entry in bans.list()]
    team, banned = bans.check(args.league, args.team)
    if args.ban_command == "check":
        return {
            "league": team.league.value,
            "canonical_team_id": team.canonical_team_id,
            "canonical_name": team.canonical_name,
            "banned": banned,
        }
    entry, changed = (
        bans.add(args.league, args.team, args.reason, args.review_after)
        if args.ban_command == "add"
        else bans.remove(args.league, args.team)
    )
    return {**asdict(entry), "changed": changed, "banned": args.ban_command == "add"}


def _fail(message: str, reason_code: str) -> None:
    print(
        json.dumps({"decision": "NO_CALL", "reason_code": reason_code, "error": message}, indent=2),
        file=sys.stderr,
    )
    raise SystemExit(2)


if __name__ == "__main__":
    main()
