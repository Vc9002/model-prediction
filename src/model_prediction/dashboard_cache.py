"""SQLite-backed dashboard cache — mirrors Excel ledger data for fast reads.

Replaces the expensive openpyxl.load_workbook() path in dashboard_server.py
with a SQLite mirror that refreshes only when source files change (mtime).

Usage:
    cache = DashboardCache("/path/to/data")
    picks = cache.read_picks("flat")  # reads from SQLite, ~50x faster
    cache.refresh()                    # re-parses Excel if mtimes changed
"""

from __future__ import annotations

import json
import sqlite3
import threading
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

# ── column mapping ──────────────────────────────────────────────────────

# Columns the dashboard actually uses from parsed pick rows
DASHBOARD_COLUMNS = [
    "pick_id", "created_at_utc", "event_start_utc", "event_id",
    "league", "away_team", "home_team", "market_type", "selection",
    "line", "american_odds", "market_implied_probability",
    "model_probability", "model_uncertainty", "edge", "trade_candidate",
    "confidence_score", "units", "model_version", "status", "result",
    "away_score", "home_score", "probability_clv", "pnl_units",
    "settled_at_utc", "record_type", "decision", "reason_code",
    "research_score_units", "research_pnl_units",
]


class DashboardCache:
    """SQLite mirror of all dashboard-relevant Excel ledger data."""

    def __init__(self, data_root: str | Path, *, db_path: Path | None = None) -> None:
        self.data_root = Path(data_root)
        # The cache DB is mutable runtime state: callers that run
        # operationally pass the runtime root's path; the repo-local
        # default survives only for dev/test convenience.
        self.db_path = Path(db_path) if db_path is not None else self.data_root / "dashboard_cache.db"
        self._lock = threading.Lock()

    # ── public API ──────────────────────────────────────────────────────

    def refresh(self, *, force: bool = False) -> dict[str, int]:
        """Re-parse Excel ledgers if mtimes changed. Returns counts per tier.

        Set ``force=True`` to re-parse even if mtimes are unchanged.
        """
        with self._lock:
            conn = sqlite3.connect(str(self.db_path))
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            try:
                self._init_schema(conn)
                counts: dict[str, int] = {}

                for tier, paths in self._all_tiers().items():
                    mtimes = self._compute_mtime_key(paths)
                    if not force and self._is_fresh(conn, tier, mtimes):
                        continue
                    rows = self._parse_all(paths)
                    self._replace_tier(conn, tier, rows, mtimes)
                    counts[tier] = len(rows)

                conn.commit()
                return counts
            finally:
                conn.close()

    def read_picks(self, tier: str) -> list[dict[str, Any]]:
        """Read all picks for a tier from SQLite. Falls back to empty list."""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                "SELECT * FROM picks WHERE tier = ? ORDER BY row_index",
                (tier,),
            ).fetchall()
            return [_row_to_dict(r) for r in rows]
        except sqlite3.OperationalError:
            return []
        finally:
            conn.close()

    def read_model_ledger(self, model_id: str) -> list[dict[str, Any]]:
        """Read model-ledger rows from SQLite."""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                "SELECT * FROM model_picks WHERE model_id = ? "
                "ORDER BY row_index",
                (model_id,),
            ).fetchall()
            return [_row_to_dict(r) for r in rows]
        except sqlite3.OperationalError:
            return []
        finally:
            conn.close()

    def read_research(self, gated: bool = False) -> list[dict[str, Any]]:
        """Read research/gated-research picks from SQLite."""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        try:
            tier = "gated_research" if gated else "research"
            rows = conn.execute(
                "SELECT * FROM picks WHERE tier = ? ORDER BY row_index",
                (tier,),
            ).fetchall()
            return [_row_to_dict(r) for r in rows]
        except sqlite3.OperationalError:
            return []
        finally:
            conn.close()

    # ── schema ──────────────────────────────────────────────────────────

    def _init_schema(self, conn: sqlite3.Connection) -> None:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS _meta (
                tier TEXT PRIMARY KEY,
                mtime_key TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                refreshed_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS picks (
                tier TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                payload_json TEXT NOT NULL,
                PRIMARY KEY (tier, row_index)
            );

            CREATE TABLE IF NOT EXISTS model_picks (
                model_id TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                payload_json TEXT NOT NULL,
                PRIMARY KEY (model_id, row_index)
            );

            CREATE INDEX IF NOT EXISTS idx_picks_tier ON picks(tier);
            CREATE INDEX IF NOT EXISTS idx_model_picks_id ON model_picks(model_id);
        """)

    # ── file discovery ──────────────────────────────────────────────────

    def _all_tiers(self) -> dict[str, list[Path]]:
        """Return all tier → file-paths mappings."""
        flat_dir = self.data_root / "flat"
        research_dir = self.data_root / "research"
        gated_dir = self.data_root / "gated_research"
        model_dir = self.data_root / "model_ledgers"

        tiers: dict[str, list[Path]] = {}

        if flat_dir.exists():
            tiers["flat"] = sorted(flat_dir.glob("*.xlsx"))
        if research_dir.exists():
            tiers["research"] = sorted(research_dir.glob("*.xlsx"))
        if gated_dir.exists():
            tiers["gated_research"] = sorted(gated_dir.glob("*.xlsx"))
        if model_dir.exists():
            for path in sorted(model_dir.glob("*.xlsx")):
                model_id = path.stem
                tiers[f"model:{model_id}"] = [path]

        return tiers

    # ── mtime tracking ─────────────────────────────────────────────────

    @staticmethod
    def _compute_mtime_key(paths: list[Path]) -> str:
        existing = sorted(
            (str(p), p.stat().st_mtime) for p in paths if p.exists()
        )
        return json.dumps(existing, sort_keys=True)

    @staticmethod
    def _is_fresh(
        conn: sqlite3.Connection, tier: str, mtime_key: str
    ) -> bool:
        row = conn.execute(
            "SELECT mtime_key FROM _meta WHERE tier = ?", (tier,)
        ).fetchone()
        return row is not None and row[0] == mtime_key

    def _replace_tier(
        self,
        conn: sqlite3.Connection,
        tier: str,
        rows: list[dict[str, Any]],
        mtime_key: str,
    ) -> None:
        """Atomically replace all rows for a tier."""
        conn.execute("DELETE FROM picks WHERE tier = ?", (tier,))
        conn.execute("DELETE FROM _meta WHERE tier = ?", (tier,))

        from datetime import UTC, datetime

        now = datetime.now(UTC).isoformat()

        if tier.startswith("model:"):
            model_id = tier.split(":", 1)[1]
            conn.execute(
                "DELETE FROM model_picks WHERE model_id = ?", (model_id,)
            )
            conn.executemany(
                "INSERT INTO model_picks(model_id, row_index, payload_json) "
                "VALUES(?, ?, ?)",
                [(model_id, i, json.dumps(r)) for i, r in enumerate(rows)],
            )
        else:
            conn.executemany(
                "INSERT INTO picks(tier, row_index, payload_json) "
                "VALUES(?, ?, ?)",
                [(tier, i, json.dumps(r)) for i, r in enumerate(rows)],
            )

        conn.execute(
            "INSERT INTO _meta(tier, mtime_key, row_count, refreshed_at) "
            "VALUES(?, ?, ?, ?)",
            (tier, mtime_key, len(rows), now),
        )

    # ── Excel parsing ───────────────────────────────────────────────────

    def _parse_all(self, paths: list[Path]) -> list[dict[str, Any]]:
        """Parse all Excel files into rows, keeping only dashboard columns."""
        rows: list[dict[str, Any]] = []
        for path in paths:
            rows.extend(self._parse_one(path))
        return rows

    @staticmethod
    def _parse_one(path: Path) -> list[dict[str, Any]]:
        """Parse a single Excel ledger file."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        # A corrupt/foreign xlsx fails in exactly these ways (bad zip, missing
        # sheet internals, unreadable file) -- anything else is a real bug and
        # should surface instead of being swallowed as "no rows".
        except (OSError, ValueError, KeyError, BadZipFile):
            return []

        try:
            sheet = wb["Picks"] if "Picks" in wb.sheetnames else wb.active
            row_iter = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(row_iter)
            except StopIteration:
                wb.close()
                return []

            headers = [
                str(h) if h is not None else "" for h in raw_headers
            ]
            rows: list[dict[str, Any]] = []
            for values in row_iter:
                if all(v is None for v in values):
                    continue
                row = {}
                for h, v in zip(headers, values):
                    if h in DASHBOARD_COLUMNS:
                        row[h] = str(v) if v is not None else ""
                rows.append(row)
            return rows
        finally:
            wb.close()


def _row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    """Convert a SQLite row to a dict, unpacking payload_json."""
    d = dict(row)
    payload = d.pop("payload_json", "{}")
    if isinstance(payload, str):
        try:
            return json.loads(payload)
        except json.JSONDecodeError:
            return {}
    return d


# ── module-level singleton ──────────────────────────────────────────────

_cache: DashboardCache | None = None
_cache_lock = threading.Lock()


def get_cache(
    data_root: str | Path | None = None, *, db_path: Path | None = None
) -> DashboardCache:
    """Return the module-level DashboardCache singleton."""
    global _cache
    if _cache is None:
        with _cache_lock:
            if _cache is None:
                root = Path(data_root) if data_root else (
                    Path(__file__).resolve().parents[1] / "data"
                )
                _cache = DashboardCache(root, db_path=db_path)
    return _cache
