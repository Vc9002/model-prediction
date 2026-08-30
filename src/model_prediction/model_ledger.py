"""Per-model-identity prediction ledger.

Operator directive, 2026-08-02: "every model runs in production... no
benchmark, challenger, research, shadow, qualified, or unqualified
classifications... no model-state field controlling routing... I decide
what to use." This is the new ledger this directive calls for -- one file
per MODEL IDENTITY (not one shared file per sport), no eligibility gates, no
record_type/model_state classification deciding whether a row gets written.
Every model writes a prediction (or an explicit failure record) for every
event it can evaluate; the operator's own decision about what to act on is
tracked separately, in the same row, never mutating the model's own output.

This module is additive: `PickLedger` (ledger.py) is untouched and remains
what the live forecast pipeline writes through today. Migrating existing
historical data into this new schema and cutting the live pipeline over to
write here are separate, later steps -- see DEBUG.md/docs for status.
"""

from __future__ import annotations

import fcntl
import os
import tempfile
import time
import uuid
from collections.abc import Iterator
from contextlib import contextmanager, suppress
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .domain import iso_utc, utc_now

MODEL_LEDGER_SCHEMA_VERSION = "1"
SHEET_NAME = "Predictions"

# fcntl.flock(LOCK_EX) blocks indefinitely with no built-in timeout -- a hung
# holder (not crashed; POSIX auto-releases on process death) would otherwise
# wedge every other appender forever with no diagnostic. Same pattern as
# ledger.py/audit.py's own lock helpers.
LOCK_TIMEOUT_SECONDS = 30
LOCK_POLL_INTERVAL_SECONDS = 0.1


class ModelLedgerLockTimeout(TimeoutError):
    pass


def _acquire_exclusive_lock(fileno: int, path: Path, timeout: float = LOCK_TIMEOUT_SECONDS) -> None:
    deadline = time.monotonic() + timeout
    while True:
        try:
            fcntl.flock(fileno, fcntl.LOCK_EX | fcntl.LOCK_NB)
            return
        except BlockingIOError:
            if time.monotonic() >= deadline:
                raise ModelLedgerLockTimeout(
                    f"could not acquire lock on {path} within {timeout}s -- "
                    "another process may be hung while holding it"
                ) from None
            time.sleep(LOCK_POLL_INTERVAL_SECONDS)


# The model's own output. Never touched by an operator action.
MODEL_FIELDS = [
    "prediction_id",
    "ledger_schema_version",
    "model_id",
    "model_version",
    "artifact_hash",
    "code_revision",
    "feature_schema_version",
    "event_id",
    "market_type",
    "selection",
    "line",
    "model_probability",
    "model_projection",
    "model_uncertainty",
    "decision_price",
    "market_no_vig_probability",
    "model_market_difference",
    "observed_at_utc",
    "event_start_utc",
    "input_cutoff_utc",
    "input_availability",
    "missing_inputs",
    "source_lineage",
]

# Model-written but mutated by settle().
RESULT_FIELDS = [
    "status",
    "result",
    "closing_price",
    "probability_clv",
    "pnl_units",
    "settled_at_utc",
    "failure_reason",
]

# Written only by record_operator_decision(); never affects the model's own
# fields above, and never decides what gets logged in the first place --
# "this is not model promotion, it is an event-level decision."
OPERATOR_FIELDS = [
    "operator_decision",
    "operator_selected_model",
    "operator_selected_market",
    "operator_units",
    "operator_timestamp",
    "operator_note",
]

FIELDNAMES = [*MODEL_FIELDS, *RESULT_FIELDS, *OPERATOR_FIELDS]

# The only things allowed to stop a numeric prediction (operator directive):
# objective integrity failures, not classification. Callers pass one of
# these as `reason` to append_failure -- kept here as the canonical list
# rather than a free-text convention, so a typo doesn't silently invent a
# new, unrecognized failure reason.
INTEGRITY_FAILURE_REASONS = frozenset(
    {
        "event_already_started",
        "required_identity_unresolved",
        "artifact_hash_invalid",
        "feature_timestamp_after_cutoff",
        "wrong_market_or_settlement_horizon",
        "market_sides_unmapped",
        "model_calculation_failed",
        "required_input_unavailable_no_fallback_defined",
    }
)


# (league, market_type) -> model identity. Shared with
# scripts/migrate_to_model_ledgers.py so the historical backfill and the
# live pipeline agree on the same identity for the same model. SOCCER
# intentionally maps every market_type to the same identity: one
# Poisson/Dixon-Coles model produces both moneyline and total predictions,
# not two separate models.
MODEL_ID_BY_LEAGUE_AND_MARKET: dict[tuple[str, str], str] = {
    ("MLB", "moneyline"): "mlb-moneyline-elo-trend-lr",
    ("MLB", "spread"): "mlb-spread-measured-edge",
    ("MLB", "total"): "mlb-total-measured-edge",
    ("NBA", "moneyline"): "nba-moneyline-elo-trend-lr",
    ("NBA", "spread"): "nba-spread-baseline",
    ("NBA", "total"): "nba-total-score-ridge",
    ("WNBA", "moneyline"): "wnba-moneyline-elo-trend-lr",
    ("WNBA", "spread"): "wnba-spread-margin",
    ("WNBA", "total"): "wnba-total-score-ridge",
    ("NFL", "moneyline"): "nfl-moneyline-elo-trend-lr",
    ("NFL", "spread"): "nfl-spread-baseline",
    ("NFL", "total"): "nfl-total-score-ridge",
    ("SOCCER", "moneyline"): "soccer-poisson-dc",
    ("SOCCER", "total"): "soccer-poisson-dc",
    ("SOCCER", "btts"): "soccer-poisson-dc",
    ("TENNIS", "moneyline"): "tennis-surface-elo",
    ("LOL", "moneyline"): "lol-tiered-elo",
    ("CS2", "moneyline"): "cs2-tiered-elo",
    ("DOTA2", "moneyline"): "dota2-tiered-elo",
    ("VALORANT", "moneyline"): "valorant-tiered-elo",
    ("RAINBOW_SIX", "moneyline"): "rainbow-six-tiered-elo",
    ("KBO", "moneyline"): "kbo-tie-aware-elo",
    ("NPB", "moneyline"): "npb-tie-aware-elo",
}


def model_id_for(league: str, market_type: str) -> str | None:
    return MODEL_ID_BY_LEAGUE_AND_MARKET.get((str(league).upper(), str(market_type)))


def _normalized_line(row: dict[str, str]) -> str:
    line = str(row.get("line") or "")
    with suppress(ValueError):
        line = str(abs(float(line))) if line else ""
    return line


def _signed_line(row: dict[str, str]) -> str:
    """The line as a SIGNED number, representation-normalized.

    "+5.5" and "5.5" are the same contract and must match; "+1.5" and
    "-1.5" are OPPOSITE contracts (the selected side receives vs. gives
    points) and must never collapse. abs()-based normalization, which is
    right for the append-side dedupe key, would silently grade a
    sign-flipped re-forecast backwards -- found by code review
    2026-08-19."""
    raw = str(row.get("line") or "")
    try:
        return f"{float(raw):g}"
    except ValueError:
        return raw


def _event_settlement_key(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    """The contract a settlement grades: event/market/line/model/selection.

    Deliberately WITHOUT `observed_at_utc`. An outcome is a property of the
    event, not of when we happened to forecast it, so every open row for
    this contract grades identically off one real result. See
    `_prediction_dedupe_key` for why the append side needs the timestamp
    and this side must not have it.

    `selection` IS included: away +5.5 and home -5.5 are opposite sides of
    one game and resolve oppositely, so a re-forecast that crossed 0.5 and
    flipped sides must not inherit the other side's result. No live row has
    flipped yet, so this changes nothing today -- it makes a future flip
    stay honestly open instead of being graded backwards. The line itself
    is kept SIGNED for the same reason: away +1.5 and away -1.5 are
    different contracts that resolve oppositely."""
    return (
        str(row.get("event_id") or ""),
        str(row.get("market_type") or ""),
        _signed_line(row),
        str(row.get("model_version") or ""),
        str(row.get("selection") or ""),
    )


def _prediction_dedupe_key(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    """One real decision, regardless of which old destination ledger(s)
    logged it. No `sportsbook` component: unlike ledger.py's
    `_market_duplicate_key`, this schema has no sportsbook field at all
    (every row in this system is polymarket_us; see the operator's own
    common-schema spec) -- including it here would compare a caller-
    supplied value against a stored row that can never have it, which
    would never match and silently defeat the dedupe check entirely.

    `observed_at_utc` IS included -- cli.py's Main/Flat/Research/Gated
    calls for one real decision all share the exact same `PickRequest`
    object (and therefore an identical `observed_at_utc`), so they still
    collapse into one row here. But an event whose still-open pick later
    gets replaced by a fresh forecast (a real, distinct decision with new
    probabilities/prices, same event/market/line/model_version) gets a new
    `observed_at_utc` and must NOT be silently dropped as a duplicate --
    found live 2026-08-02: a re-forecast WNBA pick with updated
    model_probability/decision_price was skipped entirely because only the
    stale, previously-migrated row's key was checked.

    This is an APPEND-side key only. Settlement must use
    `_event_settlement_key`: reusing this one there left every re-forecast
    row of a graded event open forever (found 2026-08-18, WNBA spread --
    42 of 67 rows stuck open, evidence computed on 9 rows instead of 51)."""
    return (
        str(row.get("event_id") or ""),
        str(row.get("market_type") or ""),
        _normalized_line(row),
        str(row.get("model_version") or ""),
        str(row.get("observed_at_utc") or ""),
    )


def record_from_pick_request(
    model_ledgers_dir: str | Path,
    request: Any,
    eligibility: Any,
    created: datetime | None = None,
) -> dict[str, str] | None:
    """Live-pipeline hook: write the equivalent ModelLedger row for a
    PickRequest/EligibilityResult pair the old PickLedger just logged.

    Returns None (no-op) when the (league, market_type) has no model_id
    mapping yet -- some sports/markets aren't wired into the new system
    yet, which must degrade gracefully, not raise, since this always runs
    alongside the real, working PickLedger write it must never break.

    Dedupes against the target ledger's existing rows by the same
    market-identity notion the migration script uses: the same real
    decision can get logged to more than one old destination ledger (e.g.
    both Main and Flat), and each of those calls this same hook -- without
    this check every one of them would create a separate duplicate row
    here, even though old-ledger routing (Main/Flat/Research/Gated) is
    exactly the distinction this new schema doesn't have anymore.

    ``observed_at_utc`` falls back to ``created`` (the pick's own append
    timestamp) exactly like `_append_record`'s own `row["observed_at_utc"]`
    does when `request.observed_at_utc` is unset -- found 2026-08-02 while
    wiring `observed_at_utc` into the dedupe key: without matching that
    fallback here, every request with no explicit `observed_at_utc` wrote a
    blank value, so every re-forecast of that event collapsed into the
    *same* blank-keyed row again, defeating the very fix that added
    `observed_at_utc` to the key in the first place.
    """
    model_id = model_id_for(request.league.value, request.market_type.value)
    if model_id is None:
        return None
    ledger = ModelLedger(Path(model_ledgers_dir) / f"{model_id}.xlsx")
    try:
        from .pricing import implied_probability

        decision_price = implied_probability(request.american_odds)
    except (ImportError, ValueError, ZeroDivisionError):
        decision_price = None
    operator_units = float(getattr(eligibility, "units", 0.0) or 0.0)
    if operator_units <= 0:
        operator_units = 1.0
    record = {
        "model_id": model_id,
        "model_version": request.model_version,
        "artifact_hash": request.model_artifact_hash,
        "code_revision": request.code_revision,
        "feature_schema_version": request.feature_schema_version,
        "event_id": request.event_id,
        "market_type": request.market_type.value,
        "selection": request.selection,
        "line": request.line,
        "model_probability": request.model_probability,
        "model_uncertainty": request.model_uncertainty,
        "decision_price": decision_price,
        "market_no_vig_probability": request.decision_no_vig_probability,
        "model_market_difference": getattr(eligibility, "edge", None),
        "observed_at_utc": request.observed_at_utc or iso_utc(created or utc_now()),
        "event_start_utc": request.event_start_utc,
        "status": "open",
        "operator_decision": "CALL",
        "operator_units": operator_units,
        "operator_timestamp": request.observed_at_utc or iso_utc(created or utc_now()),
        "operator_note": "automatic universal paper call",
    }
    key = _prediction_dedupe_key(record)
    existing_keys = {_prediction_dedupe_key(row) for row in ledger.rows()}
    if key in existing_keys:
        return None
    return ledger.append_prediction(record)


def settle_from_pick_row(model_ledgers_dir: str | Path, row: dict[str, Any]) -> list[dict[str, str]]:
    """Live-pipeline settlement hook, mirroring record_from_pick_request.

    Found 2026-08-02 during a live-run verification: `ModelLedger.settle()`
    existed but nothing ever called it -- every model ledger row stayed
    "open" forever, even after PickLedger.settle() graded the real,
    equivalent pick, so compute_model_evidence's hit-rate/Brier/calibration
    numbers (the whole point of "operator decides using real evidence")
    never actually populated. `PickLedger.settle()` calls this with the
    just-settled row (result/pnl_units/probability_clv/closing_* already
    set); this settles EVERY still-open ModelLedger row for the same
    event/market/line/model. Never raises: a missing match (unmapped
    league/market, or the append-side hook never fired for this pick, or
    it's already settled) degrades to a silent no-op, same contract as the
    append hook.

    Matching is by `_event_settlement_key`, NOT the append-side
    `_prediction_dedupe_key`. Using the append key here was a real bug
    (found 2026-08-18): it carries `observed_at_utc`, so only the single
    row whose forecast timestamp equalled the settled pick's ever graded,
    and every re-forecast row for the same finished game stayed open
    forever. Live WNBA spread ledger: 42 of 67 rows stuck open, so
    hit-rate/Brier/calibration ran on 9 rows instead of 51 -- the same
    silent-evidence-starvation failure this hook was added to fix in the
    first place, one layer down.
    """
    model_id = model_id_for(str(row.get("league", "")), str(row.get("market_type", "")))
    if model_id is None:
        return []
    path = Path(model_ledgers_dir) / f"{model_id}.xlsx"
    if not path.exists():
        return []
    ledger = ModelLedger(path)
    key = _event_settlement_key(
        {
            "event_id": row.get("event_id", ""),
            "market_type": row.get("market_type", ""),
            "line": row.get("line", ""),
            "model_version": row.get("model_version", ""),
            "selection": row.get("selection", ""),
        }
    )
    closing_price_raw = row.get("closing_raw_implied_probability") or row.get("closing_implied_probability")
    probability_clv_raw = row.get("probability_clv")
    pnl_units_raw = row.get("pnl_units")
    return ledger.settle_event(
        key,
        result=str(row.get("result", "")),
        closing_price=float(closing_price_raw)
        if closing_price_raw is not None and closing_price_raw != ""
        else None,
        pnl_units=float(pnl_units_raw) if pnl_units_raw is not None and pnl_units_raw != "" else None,
        probability_clv=float(probability_clv_raw)
        if probability_clv_raw is not None and probability_clv_raw != ""
        else None,
    )


class ModelLedger:
    """One file per model identity (e.g. mlb-moneyline-elo-trend-lr.xlsx).

    No eligibility gates: every prediction this model can compute for an
    event gets written, or an explicit failure record does (append_failure)
    -- never a silently skipped event. Missing *optional* inputs are the
    model's own concern (input_availability/missing_inputs describe them);
    only the INTEGRITY_FAILURE_REASONS above justify no numeric row at all.
    """

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if self.path.suffix.casefold() != ".xlsx":
            raise ValueError("a model ledger must be an .xlsx workbook")
        self.lock_path = self.path.with_suffix(self.path.suffix + ".lock")

    @contextmanager
    def _lock(self) -> Iterator[None]:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.lock_path.open("a+") as lock:
            _acquire_exclusive_lock(lock.fileno(), self.lock_path)
            try:
                yield
            finally:
                fcntl.flock(lock.fileno(), fcntl.LOCK_UN)

    def _read_unlocked(self) -> list[dict[str, str]]:
        if not self.path.exists():
            return []
        return _read_predictions(self.path)

    def _write_unlocked(self, rows: list[dict[str, str]]) -> None:
        _write_predictions_atomic(self.path, rows)

    def rows(self) -> list[dict[str, str]]:
        with self._lock():
            return self._read_unlocked()

    def append_prediction(
        self, record: dict[str, Any], *, prediction_id: str | None = None
    ) -> dict[str, str]:
        """Write one model prediction. `record` should cover the MODEL_FIELDS
        this prediction has real values for; anything omitted is left blank,
        not guessed. `status` defaults to "open" -- callers settle() later.

        `prediction_id`: a fresh uuid4 by default. Migration code passes the
        original ledger row's own pick_id here instead, so a row moved into
        this new schema keeps a stable, traceable identity rather than
        getting a disconnected new one.
        """
        with self._lock():
            rows = self._read_unlocked()
            existing_ids = {row["prediction_id"] for row in rows}
            row = {field: "" for field in FIELDNAMES}
            row.update(
                {key: "" if value is None else str(value) for key, value in record.items() if key in row}
            )
            chosen_id = prediction_id or uuid.uuid4().hex[:16]
            if chosen_id in existing_ids:
                raise ValueError(f"prediction_id {chosen_id!r} already exists in {self.path}")
            row["prediction_id"] = chosen_id
            row["ledger_schema_version"] = MODEL_LEDGER_SCHEMA_VERSION
            if not row["status"]:
                row["status"] = "open"
            rows.append(row)
            self._write_unlocked(rows)
            return row

    def append_failure(
        self,
        *,
        model_id: str,
        model_version: str,
        event_id: str,
        reason: str,
        **extra: Any,
    ) -> dict[str, str]:
        """Explicit failure record -- an integrity gate blocked a numeric
        prediction. Never a silently dropped event."""
        if reason not in INTEGRITY_FAILURE_REASONS:
            raise ValueError(
                f"unrecognized integrity failure reason {reason!r}; "
                f"expected one of {sorted(INTEGRITY_FAILURE_REASONS)}"
            )
        record = {
            "model_id": model_id,
            "model_version": model_version,
            "event_id": event_id,
            "status": "failed",
            "failure_reason": reason,
            **extra,
        }
        return self.append_prediction(record)

    def settle(
        self,
        prediction_id: str,
        *,
        result: str,
        closing_price: float | None = None,
        pnl_units: float | None = None,
        probability_clv: float | None = None,
        settled_at: datetime | None = None,
    ) -> dict[str, str]:
        with self._lock():
            rows = self._read_unlocked()
            row = next((r for r in rows if r["prediction_id"] == prediction_id), None)
            if row is None:
                raise KeyError(prediction_id)
            row["status"] = "settled"
            row["result"] = result
            if closing_price is not None:
                row["closing_price"] = f"{closing_price:.6f}"
            if pnl_units is not None:
                row["pnl_units"] = f"{pnl_units:.4f}"
            if probability_clv is not None:
                row["probability_clv"] = f"{probability_clv:.6f}"
            row["settled_at_utc"] = iso_utc(settled_at or utc_now())
            self._write_unlocked(rows)
            return row

    def settle_event(
        self,
        key: tuple[str, str, str, str, str],
        *,
        result: str,
        closing_price: float | None = None,
        pnl_units: float | None = None,
        probability_clv: float | None = None,
        settled_at: datetime | None = None,
    ) -> list[dict[str, str]]:
        """Settle EVERY still-open row for one event/market/line/model.

        One lock/read/write cycle rather than N calls to `settle()`: a
        graded event routinely has several re-forecast rows, and settling
        them one at a time would re-read and rewrite the workbook once per
        row (and leave a partially-settled event behind on any failure).

        Operator decision 2026-08-18: `pnl_units` is written to every
        matching row, not only the row that carried the real stake. Note
        that `compute_model_evidence` SUMS this field, so a model's
        reported `pnl_units` scales with the re-forecast count -- it is a
        per-prediction economic record here, not a bankroll total."""
        settled_rows: list[dict[str, str]] = []
        with self._lock():
            rows = self._read_unlocked()
            stamp = iso_utc(settled_at or utc_now())
            for row in rows:
                # Only OPEN rows are gradeable. A `failed` row is an
                # integrity-gate record (append_failure) that says the model
                # never made a call -- settling it with a win/loss and pnl
                # would fabricate a bet that never existed. Found by code
                # review 2026-08-19.
                if row["status"] != "open" or _event_settlement_key(row) != key:
                    continue
                row["status"] = "settled"
                row["result"] = result
                if closing_price is not None:
                    row["closing_price"] = f"{closing_price:.6f}"
                if pnl_units is not None:
                    row["pnl_units"] = f"{pnl_units:.4f}"
                if probability_clv is not None:
                    row["probability_clv"] = f"{probability_clv:.6f}"
                row["settled_at_utc"] = stamp
                settled_rows.append(row)
            if settled_rows:
                self._write_unlocked(rows)
        return settled_rows

    def repair_events_from_canonical(
        self,
        corrections: dict[tuple[str, str, str, str, str], dict[str, Any]],
        *,
        correction_reason: str,
    ) -> list[dict[str, str]]:
        """Synchronize open or already-settled rows from canonical evidence.

        This is an explicit repair boundary, not part of normal settlement.
        A model ledger may contain several forecast-time observations of the
        same contract; all share one realized outcome. The caller must back up
        the workbook and provide a nonblank reason before this method can
        correct stranded open rows or stale zero-P&L settlements in one atomic
        workbook rewrite.
        """
        if not correction_reason.strip():
            raise ValueError("a nonblank correction reason is required")
        repaired: list[dict[str, str]] = []
        with self._lock():
            rows = self._read_unlocked()
            for row in rows:
                if row["status"] == "failed":
                    continue
                correction = corrections.get(_event_settlement_key(row))
                if correction is None:
                    continue
                result = str(correction.get("result") or "")
                if result not in {"win", "loss", "push"}:
                    raise ValueError("canonical correction requires win/loss/push result")
                row.update(
                    {
                        "status": "settled",
                        "result": result,
                        "settled_at_utc": str(correction.get("settled_at_utc") or iso_utc(utc_now())),
                    }
                )
                for field in ("closing_price", "probability_clv", "pnl_units"):
                    value = correction.get(field)
                    if value is not None and value != "":
                        decimals = 4 if field == "pnl_units" else 6
                        row[field] = f"{float(value):.{decimals}f}"
                repaired.append(row)
            if repaired:
                self._write_unlocked(rows)
        return repaired

    def record_operator_decision(
        self,
        prediction_id: str,
        *,
        decision: str,
        selected_model: str | None = None,
        selected_market: str | None = None,
        units: float | None = None,
        note: str | None = None,
    ) -> dict[str, str]:
        """Store the operator's own event-level decision, separate from the
        model's output -- "not model promotion... must not change the
        model's ledger, classification, historical statistics, or dashboard
        evidence." Only the OPERATOR_FIELDS block is touched."""
        with self._lock():
            rows = self._read_unlocked()
            row = next((r for r in rows if r["prediction_id"] == prediction_id), None)
            if row is None:
                raise KeyError(prediction_id)
            row["operator_decision"] = decision
            row["operator_selected_model"] = selected_model or ""
            row["operator_selected_market"] = selected_market or ""
            row["operator_units"] = "" if units is None else str(units)
            row["operator_timestamp"] = iso_utc(utc_now())
            row["operator_note"] = note or ""
            self._write_unlocked(rows)
            return row


def _read_predictions(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"model ledger is missing required worksheet {SHEET_NAME!r}")
        sheet = workbook[SHEET_NAME]
        values = sheet.iter_rows(values_only=True)
        first = next(values, None)
        if first is None:
            return []
        headers = ["" if value is None else str(value) for value in first]
        rows: list[dict[str, str]] = []
        for values_row in values:
            if all(value is None for value in values_row):
                continue
            row = {
                header: "" if value is None else str(value)
                for header, value in zip(headers, values_row, strict=False)
            }
            rows.append(row)
        return rows
    finally:
        workbook.close()


def _write_predictions_atomic(path: Path, rows: list[dict[str, str]]) -> None:
    """Plain single-sheet writer -- deliberately not xlsx_ledger.py's
    write_xlsx_rows_atomic, which hardcodes PickLedger's own field names
    (record_type, pick_id, research_score_units, ...) into a formula-driven
    summary sheet and cannot be reused for a differently-shaped schema."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="1F2937")
    for column_index, field in enumerate(FIELDNAMES, start=1):
        cell = sheet.cell(row=1, column=column_index, value=field)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[get_column_letter(column_index)].width = max(14, min(40, len(field) + 2))
    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(FIELDNAMES, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row.get(field, "") or "")
            cell.font = body_font
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{max(1, len(rows) + 1)}"
    descriptor, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(descriptor)
    try:
        workbook.save(temporary)
        os.chmod(temporary, 0o644)
        os.replace(temporary, path)
    except BaseException:
        with suppress(FileNotFoundError):
            os.unlink(temporary)
        raise


def compute_model_evidence(ledger: ModelLedger) -> dict[str, Any]:
    """Real evidence for one model: sample size, Brier, log loss,
    calibration, CLV coverage/mean, missing-input rate, data age. Per the
    operator's dashboard design spec -- evidence, not a qualified/research
    classification badge.

    Reuses calibration.py's calibration_metrics (Brier/log loss/calibration
    intercept-slope/ECE/reliability buckets) rather than reimplementing it,
    and matches ledger.py's own existing calibration_rows/CLV filtering
    exactly: settled win/loss only for calibration (pushes excluded
    entirely, never folded in as a loss).
    """
    from .calibration import calibration_metrics

    rows = ledger.rows()
    open_rows = [row for row in rows if row["status"] == "open"]
    settled = [row for row in rows if row["status"] == "settled"]
    failed = [row for row in rows if row["status"] == "failed"]
    calibration_rows = [
        row for row in settled if row["result"] in {"win", "loss"} and row["model_probability"]
    ]
    calibration = calibration_metrics(
        [float(row["model_probability"]) for row in calibration_rows],
        [1 if row["result"] == "win" else 0 for row in calibration_rows],
    )
    clv_values = [float(row["probability_clv"]) for row in settled if row["probability_clv"]]
    pnl_values = [float(row["pnl_units"]) for row in settled if row["pnl_units"]]
    missing_input_rows = [row for row in rows if row["missing_inputs"]]
    observed_dates = sorted(row["observed_at_utc"] for row in rows if row["observed_at_utc"])
    return {
        "model_id": rows[0]["model_id"] if rows else None,
        "total": len(rows),
        "open": len(open_rows),
        "settled": len(settled),
        "failed": len(failed),
        "wins": sum(1 for row in calibration_rows if row["result"] == "win"),
        "losses": sum(1 for row in calibration_rows if row["result"] == "loss"),
        "pushes": sum(1 for row in settled if row["result"] not in ("win", "loss")),
        "pnl_units": round(sum(pnl_values), 4),
        "calibration": calibration,
        "clv_coverage": round(len(clv_values) / len(settled), 4) if settled else None,
        "mean_clv": round(sum(clv_values) / len(clv_values), 6) if clv_values else None,
        "missing_input_rate": round(len(missing_input_rows) / len(rows), 4) if rows else None,
        "latest_observed_at_utc": observed_dates[-1] if observed_dates else None,
        "oldest_observed_at_utc": observed_dates[0] if observed_dates else None,
    }
